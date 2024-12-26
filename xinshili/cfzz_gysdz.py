import re
from openpyxl import load_workbook
import xlwings as xw
import os
import utils
from datetime import datetime


def load_excel_file(file_path):
    """
    加载 Excel 文件
    """
    try:
        return load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"未找到文件: {file_path}")
    except Exception as e:
        raise Exception(f"加载文件 {file_path} 时发生错误: {e}")


def get_active_sheet(workbook):
    """
    获取活动表
    """
    try:
        return workbook.active
    except Exception as e:
        raise Exception(f"获取活动表时发生错误: {e}")


def find_columns(sheet, column_names):
    """
    动态查找指定列名的列号
    """
    try:
        column_map = {}
        for col in sheet.iter_cols(1, sheet.max_column):
            header = col[0].value
            if header in column_names:
                column_map[header] = col[0].column
                if len(column_map) == len(column_names):
                    break
        if len(column_map) < len(column_names):
            missing = set(column_names) - set(column_map.keys())
            raise ValueError(f"未找到以下列名: {missing}")
        return column_map
    except Exception as e:
        raise Exception(f"查找列时发生错误: {e}")


def write_data_to_sheet(sheet1, sheet2, columns1, columns2, exchange_rate):
    """
    从表1复制数据到表2
    """
    try:
        target_row = 2  # 表2从第2行开始写入
        for row in range(3, sheet1.max_row + 1):  # 跳过表1的第2行
            # 从表1获取数据
            paid_time_value = sheet1.cell(row, columns1["Paid Time"]).value
            order_id_value = sheet1.cell(row, columns1["Order ID"]).value
            quantity_value = sheet1.cell(row, columns1["Quantity"]).value
            seller_sku_value = sheet1.cell(row, columns1["Seller SKU"]).value

            # 写入表2
            sheet2.cell(target_row, columns2["日期"]).value = paid_time_value
            sheet2.cell(target_row, columns2["订单单号"]).value = order_id_value
            sheet2.cell(target_row, columns2["数量"]).value = quantity_value
            sheet2.cell(target_row, columns2["款号"]).value = seller_sku_value

            # 写入公式到运费和汇率列
            sheet2.cell(target_row, columns2["运费"]).value = f"=2.99+IF(E{target_row}=1,0,(E{target_row}-1)*0.5)"
            sheet2.cell(target_row, columns2["汇率"]).value = exchange_rate

            target_row += 1
    except Exception as e:
        raise Exception(f"写入数据到表2时发生错误: {e}")


def match_and_write_prices(sheet2, sheet4, columns2, sku_col_4, cost_col_4):
    """
    匹配表2和表4的款号，并提取成本价写入表2
    """
    try:
        for row_2 in range(2, sheet2.max_row + 1):  # 从表2的第2行开始
            sku_value_2 = sheet2.cell(row_2, columns2["款号"]).value
            if sku_value_2:
                matched_cost = None
                for row_4 in range(2, sheet4.max_row + 1):  # 从表4的第2行开始
                    sku_value_4 = sheet4.cell(row_4, sku_col_4).value
                    if sku_value_4 == sku_value_2:  # 匹配款号
                        cost_value_4 = sheet4.cell(row_4, cost_col_4).value
                        # 提取括号中的数值
                        match = re.search(r"（([\d.]+)", cost_value_4)
                        if match:
                            matched_cost = float(match.group(1))
                        break
                # 将匹配到的成本价写入表2的采购价列
                sheet2.cell(row_2, columns2["采购价"]).value = matched_cost
    except Exception as e:
        raise Exception(f"匹配和写入成本价时发生错误: {e}")


def calculate_rmb_prices(file_path):
    """
    计算采购价（RMB）并写入表格
    """
    try:
        app = xw.App(visible=False)
        wb = xw.Book(file_path)
        sheet = wb.sheets[0]

        # 假设表中的列号
        purchase_price_col = "D"
        quantity_col = "E"
        freight_col = "F"
        exchange_rate_col = "G"
        purchase_price_rmb_col = "H"

        for row in range(2, sheet.used_range.last_cell.row + 1):
            purchase_price = sheet.range(f"{purchase_price_col}{row}").value
            quantity = sheet.range(f"{quantity_col}{row}").value
            freight = sheet.range(f"{freight_col}{row}").value
            exchange_rate = sheet.range(f"{exchange_rate_col}{row}").value

            if purchase_price and quantity and freight and exchange_rate:
                purchase_price_rmb = ((purchase_price * float(quantity)) + freight) * exchange_rate
                sheet.range(f"{purchase_price_rmb_col}{row}").value = round(purchase_price_rmb, 2)

        last_row = sheet.range('H' + str(sheet.cells.last_cell.row)).end('up').row
        h_values = sheet.range(f'H2:H{last_row}').value
        h_total = round(sum(value for value in h_values if isinstance(value, (int, float))), 2)

        sheet.range("L1").value = f"总计:¥ {h_total} 元"

        # 自动调整所有单元格的列宽和行高
        sheet.used_range.columns.autofit()  # 自动调整所有列宽
        sheet.used_range.rows.autofit()  # 自动调整所有行高

        # 设置所有单元格内容居中
        used_range = sheet.used_range
        used_range.api.HorizontalAlignment = -4108  # xlCenter，水平居中
        used_range.api.VerticalAlignment = -4108  # xlCenter，垂直居中

        wb.save(file_path)
        wb.close()
        app.quit()

        return h_total
    except Exception as e:
        raise Exception(f"计算和写入采购价（RMB）时发生错误: {e}")


def main():
    try:
        source_file = input("请输入源表文件的绝对路径：")
        exchange_rate = round(float(utils.get_usd_to_cny_rate()), 2) + 0.01

        # 加载文件
        wb1 = load_excel_file(source_file)
        wb2 = load_excel_file("/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF供应商对账表模版.xlsx")
        wb4 = load_excel_file("/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF产品核对表1114.xlsx")

        # 获取活动表
        sheet1 = get_active_sheet(wb1)
        sheet2 = get_active_sheet(wb2)
        sheet4 = get_active_sheet(wb4)

        # 查找列
        columns1 = find_columns(sheet1, ["Paid Time", "Order ID", "Quantity", "Seller SKU"])
        columns2 = find_columns(sheet2, ["日期", "订单单号", "数量", "运费", "款号", "汇率", "采购价"])

        # 写入数据
        write_data_to_sheet(sheet1, sheet2, columns1, columns2, exchange_rate)

        # 匹配款号和成本价
        match_and_write_prices(sheet2, sheet4, columns2, sku_col_4=8, cost_col_4=5)

        # 保存文件
        output_path = "/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF供应商对账表_temp.xlsx"
        wb2.save(output_path)

        # 计算采购价（RMB）
        prices = calculate_rmb_prices(output_path)

        old_file_path = output_path
        new_file_path = "/Users/zkp/Desktop/B&Y/CZFF供应商对账/" + "CZFF供应商对账表" + str(
            prices) + "元-" + utils.get_yd() + ".xlsx"
        os.rename(old_file_path, new_file_path)

        print(f"文件已重命名为：{new_file_path}")
    except Exception as e:
        print(f"程序运行时发生错误: {e}")


if __name__ == "__main__":
    main()
