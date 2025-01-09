from openpyxl import load_workbook, Workbook


def aggregate_and_save(input_file, output_file):
    """
    将输入文件的 AR 列相同内容对应的 AZ 列的值进行汇总，并生成新的 Excel 文件。
    :param input_file: 输入 Excel 文件路径
    :param output_file: 输出 Excel 文件路径
    """
    try:
        # 加载输入文件
        wb = load_workbook(input_file)
        sheet = wb.active

        # 获取列索引
        data = sheet.iter_rows(min_row=2, values_only=True)  # 跳过表头
        header = [cell.value for cell in sheet[1]]  # 获取表头
        ar_index = header.index("SKU")  # AR 列索引
        az_index = header.index("Outbound Qty/出库数量")  # AZ 列索引
        b_index = header.index("Client/客户")  # AZ 列索引

        # 汇总数据到字典
        result_dict = {}
        for row in data:
            ar_value = row[ar_index]
            az_value = row[az_index]
            b_value = row[b_index]

            # if ar_value is not None:
            #     result_dict[ar_value] = result_dict.get(ar_value, 0) + (az_value if isinstance(az_value, (int, float)) else 0)
            if ar_value is not None and b_value is not None:
                key = (ar_value, b_value)
                result_dict[key] = result_dict.get(key, 0) + (az_value if isinstance(az_value, (int, float)) else 0)

        # 创建新的工作簿
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "汇总结果"

        # 写入表头
        new_sheet.append(["SKU", "Outbound Qty/出库数量", "Client/客户"])

        # 写入汇总结果
        for (ar, b), total in result_dict.items():
            new_sheet.append([ar, b, total])

        # 保存到新文件
        new_wb.save(output_file)
        print(f"汇总结果已保存到文件: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


# # 示例使用
# input_file = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/ParcelOutbound_20250108160854.xlsx"
# output_file = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/ParcelOutbound_20250108160854_整理.xlsx"
# aggregate_and_save(input_file, output_file)
#
# # 示例使用
# input_file1 = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/ParcelOutbound_20250108143950.xlsx"
# output_file1 = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/ParcelOutbound_20250108143950_整理.xlsx"
# aggregate_and_save(input_file1, output_file1)
#
# # 示例使用
# input_file2 = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/ParcelOutbound_20250108110927.xlsx"
# output_file2 = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/ParcelOutbound_20250108110927_整理.xlsx"
# aggregate_and_save(input_file2, output_file2)


input_file2 = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/未发出总共185单.xlsx"
output_file2 = "/Users/zkp/Library/Containers/com.tencent.xinWeChat/Data/Library/Application Support/com.tencent.xinWeChat/2.0b4.0.9/aee968804ccf60699f2aada7c6e578a8/Message/MessageTemp/fd8511edf7e3e2cab96f63bd9b9d5f86/File/未发出总共185单_整理.xlsx"
aggregate_and_save(input_file2, output_file2)
