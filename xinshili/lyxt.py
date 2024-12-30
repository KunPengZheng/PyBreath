import openpyxl
import os
import utils


def copy_column(source_sheet, target_sheet, source_col, target_col):
    """
    将 source_file 中 source_col 列的数据，从 source_start_row 开始，
    复制到 target_file 中 target_col 列，从 target_start_row 开始。

    :param source_col: 源文件列名（如 'A'）
    :param target_col: 目标文件列名（如 'B'）
    """

    source_start_row = 3
    target_start_row = 2

    # 从源文件读取数据
    data_to_copy = []
    for row in range(source_start_row, source_sheet.max_row + 1):
        cell_value = source_sheet[f"{source_col}{row}"].value
        data_to_copy.append(cell_value)

    # 将数据写入目标文件
    for i, value in enumerate(data_to_copy):
        target_sheet[f"{target_col}{target_start_row + i}"] = value

    return len(data_to_copy)


# 示例参数
source_file_path = input("请输入源表文件的绝对路径：")  # 源文件路径
dir_path = "/Users/zkp/Desktop/B&Y/龙妍系统/"
target_file_path = dir_path + "批量下单模版.xlsx"  # 目标文件路径
out_file_path = dir_path + "龙妍系统-temp.xlsx"

# 打开源文件和目标文件
source_wb = openpyxl.load_workbook(source_file_path)
target_wb = openpyxl.load_workbook(target_file_path)
# 默认选中第一个工作表
source_sheet = source_wb.active
target_sheet = target_wb.active

source_start = 3  # 源文件的开始行
target_start = 2  # 目标文件的开始行

source_column = "A"  # 源文件的列
target_column = "B"  # 目标文件的列

source_column1 = "C"  # 源文件的列
target_column1 = "F"  # 目标文件的列
target_column1_1 = "G"  # 目标文件的列

source_column2 = "D"  # 源文件的列
target_column2 = "L"  # 目标文件的列

source_column3 = "E"  # 源文件的列
target_column3 = "M"  # 目标文件的列

source_column4 = "F"  # 源文件的列
target_column4 = "J"  # 目标文件的列

source_column5 = "G"  # 源文件的列
target_column5 = "I"  # 目标文件的列

source_column6 = "H"  # 源文件的列
target_column6 = "O"  # 目标文件的列

source_column7 = "I"  # 源文件的列
target_column7 = "P"  # 目标文件的列

source_column8 = "K"  # 源文件的列
target_column8 = "AJ"  # 目标文件的列

source_column9 = "L"  # 源文件的列
target_column9 = "Y"  # 目标文件的列

source_column10 = "M"  # 源文件的列
target_column10 = "Z"  # 目标文件的列

source_column11 = "N"  # 源文件的列
target_column11 = "AA"  # 目标文件的列

source_column12 = "O"  # 源文件的列
target_column12 = "AB"  # 目标文件的列

source_column13 = "T"  # 源文件的列
target_column13 = "AH"  # 目标文件的列

# 调用函数
copy_column(source_sheet, target_sheet, source_column, target_column)
copy_column(source_sheet, target_sheet, source_column1, target_column1)
copy_column(source_sheet, target_sheet, source_column1, target_column1_1)
copy_column(source_sheet, target_sheet, source_column2, target_column2)
copy_column(source_sheet, target_sheet, source_column3, target_column3)
copy_column(source_sheet, target_sheet, source_column4, target_column4)
copy_column(source_sheet, target_sheet, source_column5, target_column5)
copy_column(source_sheet, target_sheet, source_column6, target_column6)
copy_column(source_sheet, target_sheet, source_column7, target_column7)
copy_column(source_sheet, target_sheet, source_column8, target_column8)
copy_column(source_sheet, target_sheet, source_column9, target_column9)
copy_column(source_sheet, target_sheet, source_column10, target_column10)
copy_column(source_sheet, target_sheet, source_column11, target_column11)
copy_column(source_sheet, target_sheet, source_column12, target_column12)
count = copy_column(source_sheet, target_sheet, source_column13, target_column13)

for row in range(2, source_sheet.max_row):
    target_sheet[f"A{row}"].value = "usps-93"
    target_sheet[f"C{row}"].value = "包裹"
    target_sheet[f"D{row}"].value = "USD"
    target_sheet[f"H{row}"].value = "US"
    target_sheet[f"X{row}"].value = 1
    target_sheet[f"AC{row}"].value = "产品"
    target_sheet[f"AD{row}"].value = "Product"
    target_sheet[f"AE{row}"].value = "个"
    target_sheet[f"AF{row}"].value = 1
    target_sheet[f"AG{row}"].value = 20
    target_sheet[f"AX{row}"].value = "Huang Shan"
    target_sheet[f"AY{row}"].value = "Huang Shan"
    target_sheet[f"AZ{row}"].value = "US"
    target_sheet[f"BA{row}"].value = "PA"
    target_sheet[f"BB{row}"].value = "PHILADELPHIA"
    target_sheet[f"BD{row}"].value = "1014 BYBERRY RD"
    target_sheet[f"BG{row}"].value = "19116"
    target_sheet[f"BH{row}"].value = "(+1)2017377237"

# 保存目标文件
target_wb.save(out_file_path)

# 重命名
new_file_path = dir_path + "龙妍系统" "-费城-" + str(count) + "单-" + utils.get_yd() + ".xlsx"
os.rename(out_file_path, new_file_path)

print(f"完成${new_file_path}")

utils.open_dir(dir_path)
