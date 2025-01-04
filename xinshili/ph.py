from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re


def copy_columns_with_format(src_file, dst_file, output_path, start_row=2, cols_range='A:AG'):
    """
    将源文件的指定列复制到目标文件，保留格式。

    :param src_file: 源文件路径
    :param dst_file: 目标文件路径
    :param start_row: 起始行，从源文件的第 start_row 行开始
    :param cols_range: 要复制的列范围，例如 'A:AG'
    """
    # 打开源文件和目标文件
    src_wb = load_workbook(src_file)
    dst_wb = load_workbook(dst_file)

    src_ws = src_wb.active
    dst_ws = dst_wb.active

    # 获取列范围
    start_col, end_col = cols_range.split(':')
    start_col_idx = src_ws[start_col + '1'].column
    end_col_idx = src_ws[end_col + '1'].column

    # 遍历源文件中的指定列
    for col_idx in range(start_col_idx, end_col_idx + 1):
        col_letter = get_column_letter(col_idx)
        for row_idx, cell in enumerate(src_ws[col_letter][start_row - 1:], start=start_row):
            dst_cell = dst_ws.cell(row=row_idx, column=col_idx)

            # 复制值
            dst_cell.value = cell.value

            # 复制格式
            # if cell.has_style:
            #     dst_cell.font = cell.font
            #     dst_cell.border = cell.border
            #     dst_cell.fill = cell.fill
            #     dst_cell.number_format = cell.number_format
            #     dst_cell.protection = cell.protection
            #     dst_cell.alignment = cell.alignment

    # 保存目标文件
    dst_wb.save(output_path)
    print(f"数据已成功复制到 {output_path}")


def contains_chinese(content):
    """
    检查字符串中是否包含中文字符
    """
    if content and isinstance(content, str):
        return bool(re.search(r'[\u4e00-\u9fa5]', content))
    return False


def check_chinese_in_excel(file_path):
    """
    检查有效列中，除第一行外的单元格是否包含中文，打印列名而不是列数，并排除AG列。
    """
    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    sheet = wb.active

    # 获取有效列范围
    max_row = sheet.max_row
    max_column = sheet.max_column

    # 遍历有效列，排除第一行和AG列
    for col in range(1, max_column + 1):
        col_letter = sheet.cell(row=1, column=col).column_letter
        if col_letter == "AG":  # 跳过AG列
            continue

        for row in range(2, max_row + 1):  # 从第2行开始
            cell_value = sheet.cell(row=row, column=col).value
            if contains_chinese(cell_value):
                print(f"第{row}行，{col_letter} 列包含中文：{cell_value}")


def check_contains_2024(file_path):
    """
    检查 B 列第二行开始的每个单元格是否包含 '2024'，存在的打印所在的行数。
    :param file_path: Excel 文件路径
    """
    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    sheet = wb.active  # 默认第一个工作表

    # 遍历 B 列（第二列），从第二行开始
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=2)  # B 列是第2列
        if cell.value and '2024' in str(cell.value):  # 转为字符串后检查是否包含 '2024'
            cell.value = str(cell.value).replace('2024', '2025')  # 替换 '2024' 为 '2025'
            print(f"B列第 {row} 行包含 '2024'")


def check_price_min(file_path):
    """
    检查 O 列第二行开始的每个单元格的价格是否 <= 4，存在的打印所在的行数。
    :param file_path: Excel 文件路径
    """
    # 加载工作簿和工作表
    wb = load_workbook(file_path)
    sheet = wb.active  # 默认第一个工作表

    # O 列是第 15 列
    column_index = 15

    # 遍历 O 列从第二行开始的每一行
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        if cell.value is not None:  # 确保单元格不是空的
            try:
                value = float(cell.value)  # 转为浮点数以便比较
                if value <= 4:
                    print(f"第 {row} 行的 O 列单元格值为 {value}")
            except ValueError:
                print(f"第 {row} 行的 O 列单元格值 {cell.value} 不是数字")


def match_and_fill_with_openpyxl(file1_path, file2_path, output_path):
    """
    根据文件1的E列数据匹配文件2的B列，如果匹配成功，将文件2的I列内容填充到文件1的P列。

    :param file1_path: 文件1路径
    :param file2_path: 文件2路径
    :param output_path: 结果保存路径
    """
    # 加载文件1和文件2
    wb1 = load_workbook(file1_path)
    ws1 = wb1.active  # 默认第一个工作表

    wb2 = load_workbook(file2_path)
    ws2 = wb2.active  # 默认第一个工作表

    # 创建文件2的B列与I列的映射关系
    mapping = {ws2.cell(row=row, column=2).value: ws2.cell(row=row, column=9).value
               for row in range(2, ws2.max_row + 1)}  # B列是第2列，I列是第9列

    rows_to_delete = []  # 记录需要删除的行号
    # 遍历文件1的E列，从第二行开始
    for row in range(2, ws1.max_row + 1):
        value = ws1.cell(row=row, column=5).value  # E列是第5列
        if value in mapping:  # 如果匹配到
            matched_content = mapping[value]

            # 判断匹配内容是否满足条件
            if (isinstance(matched_content, str) and "预警" in matched_content) or \
                    (isinstance(matched_content, str) and "缺货" in matched_content) or \
                    (isinstance(matched_content, (int, float)) and matched_content < 10):
                ws1.cell(row=row, column=16).value = 0  # P列是第16列
            else:
                ws1.cell(row=row, column=16).value = matched_content
        else:
            rows_to_delete.append(row)  # 记录匹配不到的行号

    # 删除匹配不到的行，从后往前删除以避免行号变化影响
    for row in reversed(rows_to_delete):
        ws1.delete_rows(row)

    # 保存文件1为新的结果文件
    wb1.save(output_path)
    print(f"结果已保存到: {output_path}")


# 示例使用
src_file_path = "/Users/zkp/Desktop/B&Y/铺货/tiktok_global_product_20250103160430973_1573179.xlsx"  # 源文件路径
dst_file_path = "/Users/zkp/Desktop/B&Y/铺货/template_tiktokGlobal.xlsx"  # 目标文件路径
output_path = "/Users/zkp/Desktop/B&Y/铺货/template_tiktokGlobal_result.xlsx"
check_path = "/Users/zkp/Desktop/B&Y/铺货/table_1.xlsx"  # 问题：xls转换为xlsx

copy_columns_with_format(src_file_path, dst_file_path, output_path, start_row=2, cols_range='A:AG')
match_and_fill_with_openpyxl(output_path, check_path, output_path)
check_chinese_in_excel(output_path)
check_contains_2024(output_path)
check_price_min(output_path)
