from datetime import datetime, timezone, timedelta

import pandas as pd
import os
from openpyxl import load_workbook
from xinshili.fs_utils_plus import get_token, dimension_range, FsConstants, value_range
from xinshili.gjgz_plus333 import check_and_add_courier_column, extract_and_process_data, RowName, CourierStateMapKey, \
    update_courier_status


def convert_china_to_utc(china_time: datetime) -> datetime:
    """
    将中国时间转换为零时区（UTC/Zulu Time），忽略夏令时
    :param china_time: 中国时间 datetime 对象（无 tzinfo 或本地时间）
    :return: UTC 时间 datetime（无 tzinfo）
    """
    if not isinstance(china_time, datetime):
        raise ValueError("china_time 必须是 datetime 类型")

    china_tz = timezone(timedelta(hours=8))  # 中国时区 UTC+8
    utc_tz = timezone.utc  # UTC 零时区

    # 设置中国时区 → 转为 UTC → 去除 tzinfo
    return china_time.replace(tzinfo=china_tz).astimezone(utc_tz).replace(tzinfo=None)


def process_tracking_time1(file_path):
    df = pd.read_excel(file_path)

    # 先排除状态为 unpaid、delivered、irregular_no_tracking 的行
    df_filtered = df[~df['Courier/快递'].str.lower().isin(['unpaid', 'delivered', "irregular_no_tracking"])].copy()

    intervals = []
    states = []
    track_times = []

    # 当前中国时间转换为零时区时间，因为usps接口返回的时间是领时区的
    us_now = convert_china_to_utc(datetime.now())

    for idx, row in df_filtered.iterrows():
        try:
            outbound_time_str = str(row.get("发货时间", "")).strip()

            base_us_time = None
            if outbound_time_str and outbound_time_str.lower() != "nan":
                try:
                    creation_china_time = datetime.strptime(outbound_time_str, "%Y-%m-%d %H:%M:%S")
                    base_us_time = convert_china_to_utc(creation_china_time)
                except Exception:
                    base_us_time = None

            if base_us_time:
                base_diff = us_now - base_us_time
                base_hours = round(base_diff.total_seconds() / 3600, 2)

                if base_hours > 72:  # 如果 当前时间 - 发货时间 > 72 小时，默认不能再替换运单号。但是结果结果不一定卡死再72小时，而且不包含周末和节假日
                    intervals.append(base_hours)
                    states.append("无法替换")
                    track_times.append(us_now)
                    continue  # 跳过后续判断

            # 最新事件时间
            date_str = str(row.get("LatestEventSfDate/最新事件时间", "")).strip()
            time_str = str(row.get("LatestEventSfTime/最新事件时间", "")).strip()
            last_time_str = str(row.get("LastEventSfTime/上一条轨迹时间", "")).strip()

            if last_time_str and date_str and time_str and \
                    last_time_str.lower() != "nan" and date_str.lower() != "nan" and time_str.lower() != "nan":
                last_time = datetime.strptime(f"{last_time_str}", "%Y-%m-%d %H:%M")
                event_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                # 最新轨迹时间 - 上一条轨迹时间
                diff = event_time - last_time
            elif date_str and time_str and date_str.lower() != "nan" and time_str.lower() != "nan":
                event_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                # 跟踪时间 - 最新轨迹时间
                diff = us_now - event_time
            elif outbound_time_str and outbound_time_str.lower() != "nan":
                creation_time = datetime.strptime(outbound_time_str, "%Y-%m-%d %H:%M:%S")
                # 跟踪时间 - 发货时间
                diff = us_now - convert_china_to_utc(creation_time)
            else:
                diff = None

            if diff is not None:
                hours = round(diff.total_seconds() / 3600, 2)

                total_seconds = int(diff.total_seconds())
                hours_part = total_seconds // 3600
                minutes_part = (total_seconds % 3600) // 60
                interval_str = f"{hours_part:02}:{minutes_part:02}"
                intervals.append(interval_str)

                if hours >= 72:
                    states.append("无法替换")
                elif hours >= 48:
                    states.append("阳单替换")
                elif hours >= 24:
                    states.append("预备阳单")
                else:
                    states.append("轨迹正常")
            else:
                intervals.append(None)
                states.append("")

            track_times.append(us_now)

        except Exception as e:
            print(f"⚠️ 第 {idx} 行处理失败: {e}")
            intervals.append(None)
            states.append("")
            track_times.append(us_now)

    # 写入结果列
    df_filtered["TrackTimeInterval/跟踪时间间隔"] = intervals
    df_filtered["TrackTimeIntervalState/跟踪时间间隔状态"] = states
    df_filtered["Tacking_Time/追踪时间"] = [t.strftime("%Y-%m-%d %H:%M:%S") if isinstance(t, datetime) else "" for t in
                                            track_times]

    # 用处理后数据更新原始表
    df.update(df_filtered)
    df.to_excel(file_path, index=False)
    print(f"✅ 已处理并保存至：{file_path}")


def export_yd_data(source_file, target_file):
    # 读取文件1（源文件）
    df = pd.read_excel(source_file)

    def safe_concat_time(row):
        date_part = str(row.get("LatestEventSfDate/最新事件时间", "")).strip()
        time_part = str(row.get("LatestEventSfTime/最新事件时间", "")).strip()
        if date_part.lower() not in ["", "nan"] and time_part.lower() not in ["", "nan"]:
            return f"{date_part} {time_part}"
        else:
            return ""

    df["最新轨迹时间"] = df.apply(safe_concat_time, axis=1)

    # 构造目标列的数据
    export_df = pd.DataFrame({
        "付款时间": df["付款时间"],
        "发货时间": df["发货时间"],
        "订单号": df["订单号"],
        "运单号": df["运单号"],
        "轨迹状态": df["Courier/快递"],
        "追踪时间": df["Tacking_Time/追踪时间"],
        "上一条轨迹时间": df["LastEventSfTime/上一条轨迹时间"],
        "最新轨迹位置": df["LatestEventSfSite/最新事件地点"],
        "最新轨迹时间": df["最新轨迹时间"],
        "时间间隔": df["TrackTimeInterval/跟踪时间间隔"],
        "处理状态": df["TrackTimeIntervalState/跟踪时间间隔状态"],
    })

    # 写入到目标文件（如果存在则覆盖）
    export_df.to_excel(target_file, index=False)
    print(f"✅ 已成功导出 YD 数据至：{target_file}")

    return len(export_df)


def read_xlsx_as_nested_list(file_path):
    # 读取 Excel 文件，保留空值为 ""，禁用自动类型转换
    df = pd.read_excel(file_path, dtype=object).fillna("")

    # 获取列名作为第一行
    header = list(df.columns)

    # 获取数据行
    data_rows = df.values.tolist()

    # 将列头插入到数据最前面
    nested_list = [header] + data_rows
    return nested_list


def force_column_as_text(xlsx_path, column_names):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # 获取第一行列名
    headers = [cell.value for cell in ws[1]]
    target_indexes = [i + 1 for i, col in enumerate(headers) if col in column_names]

    for col_idx in target_indexes:
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = '@'
                if cell.value is not None:
                    cell.value = str(cell.value)  # 强制转为字符串

    wb.save(xlsx_path)


if __name__ == '__main__':
    # 订单号，运单号，发货时间，付款时间
    xlsx_path = "/Users/zkp/Downloads/order_120250624135226262_1573179_副本.xlsx"

    check_and_add_courier_column(xlsx_path)
    results = extract_and_process_data(xlsx_path, RowName.Courier, 100, wl_name='运单号')

    all_maps = {
        CourierStateMapKey.not_yet_map: results[CourierStateMapKey.not_yet_map],
        CourierStateMapKey.pre_ship_map: results[CourierStateMapKey.pre_ship_map],
        CourierStateMapKey.unpaid_map: results[CourierStateMapKey.unpaid_map],
        CourierStateMapKey.delivered_map: results[CourierStateMapKey.delivered_map],
        CourierStateMapKey.no_tracking_map: results[CourierStateMapKey.no_tracking_map],
        CourierStateMapKey.tracking_map: results[CourierStateMapKey.tracking_map],
        CourierStateMapKey.possession_sf_date_map: results[CourierStateMapKey.possession_sf_date_map],
        CourierStateMapKey.latest_event_sf_date_map: results[CourierStateMapKey.latest_event_sf_date_map],
        CourierStateMapKey.sf_date_equality_map: results[CourierStateMapKey.sf_date_equality_map],
        CourierStateMapKey.latest_event_sf_time_map: results[CourierStateMapKey.latest_event_sf_time_map],
        CourierStateMapKey.latest_event_sf_site_map: results[CourierStateMapKey.latest_event_sf_site_map],
        CourierStateMapKey.alert_map: results[CourierStateMapKey.alert_map],
    }

    column_mapping = {
        CourierStateMapKey.not_yet_map: RowName.Courier,
        CourierStateMapKey.pre_ship_map: RowName.Courier,
        CourierStateMapKey.unpaid_map: RowName.Courier,
        CourierStateMapKey.delivered_map: RowName.Courier,
        CourierStateMapKey.no_tracking_map: RowName.Courier,
        CourierStateMapKey.tracking_map: RowName.Courier,
        CourierStateMapKey.alert_map: RowName.Courier,
        CourierStateMapKey.possession_sf_date_map: RowName.PossessionSfDate,
        CourierStateMapKey.latest_event_sf_date_map: RowName.LatestEventSfDate,
        CourierStateMapKey.sf_date_equality_map: RowName.SfDateInterval,
        CourierStateMapKey.latest_event_sf_time_map: RowName.LatestEventSfTime,
        CourierStateMapKey.latest_event_sf_site_map: RowName.LatestEventSfSite,
    }

    update_courier_status(xlsx_path, all_maps, wl='运单号', column_map=column_mapping)

    process_tracking_time1(xlsx_path)
    # force_column_as_text(xlsx_path, ["订单号"])

    xlsx2 = "/Users/zkp/Desktop/B&Y/轨迹统计/xyl_track/xyl_track_merger_temp.xlsx"
    data_len = export_yd_data(xlsx_path, xlsx2)
    # force_column_as_text(xlsx2, ["订单号"])

    # 示例调用
    # result = read_xlsx_as_nested_list(xlsx2)
    # for row in result:
    #     print(row)

    # token = get_token()
    # dimension_range(token, FsConstants.gjgz_token, "yTIUrm", 1, 10, majorDimension="COLUMNS")
    # value_range(token, FsConstants.gjgz_token, "yTIUrm", f"A1:K{len(result)}", result)
