import pandas as pd
from openpyxl import load_workbook
import os


def remove_st_rows_excel(input_file, output_file=None):
    # 读取 Excel
    df = pd.read_excel(input_file, dtype=object)

    # 删除 "名称" 列中包含 "ST" 或 "st" 的行
    df = df[~df["名称"].str.contains("ST|st", na=False)]

    # 输出文件路径
    if output_file is None:
        output_file = input_file.replace(".xlsx", "_clean.xlsx")

    # 保存为 Excel
    df.to_excel(output_file, index=False)

    print(f"? 已处理完成，结果保存到: {output_file}")


def merge_excels_with_dynamic_columns(file1, file2, output_file):
    # 固定列（顺序必须固定）
    fixed_columns = [
        "代码", "名称",
        "2024-03(股东人数)", "2024-06(股东人数)",
        "2024-09(股东人数)", "2024-12(股东人数)",
        "2025-03(股东人数)", "2025-06(股东人数)"
    ]

    # 读取两个 Excel
    df1 = pd.read_excel(file1, dtype=object)
    df2 = pd.read_excel(file2, dtype=object)

    # 收集所有列名（固定列 + 动态列）
    all_columns = list(dict.fromkeys(fixed_columns + list(df1.columns) + list(df2.columns)))

    # 确保所有 DataFrame 都有这些列，缺失填 0
    for df in [df1, df2]:
        for col in all_columns:
            if col not in df.columns:
                df[col] = 0

    # 统一列顺序（固定列在前，动态列按字母/时间顺序排列）
    dynamic_columns = [c for c in all_columns if c not in fixed_columns]
    dynamic_columns = sorted(dynamic_columns)  # 按时间排序（如果列名规范化了）
    final_columns = fixed_columns + dynamic_columns

    # 合并
    merged_df = pd.concat([df1[final_columns], df2[final_columns]], ignore_index=True)

    # 保存结果
    merged_df.to_excel(output_file, index=False)
    print(f"? 合并完成，保存到: {output_file}")


def remove_zero_cells_and_shift_left(input_file, output_file):
    """
    规则：
    - 固定列（不会被修改/删除）:
        "代码", "名称",
        "2024-03(股东人数)", "2024-06(股东人数)",
        "2024-09(股东人数)", "2024-12(股东人数)",
        "2025-03(股东人数)"
    - 对于其它（动态）列：如果单元格值为 0（数字 0 或字符串 "0"），
      则删除该单元格，使右侧单元格左移一格（仅本行操作）。
    - 最终把所有动态列的表头清空（保留固定列表头）。
    """

    # 固定列（不要改）
    fixed_columns = [
        "代码", "名称",
        "2024-03(股东人数)", "2024-06(股东人数)",
        "2024-09(股东人数)", "2024-12(股东人数)",
        "2025-03(股东人数)", "2025-06(股东人数)"
    ]

    # 读取表头以确定列顺序（只读取表头也可以，但这里读全部以防万一）
    df = pd.read_excel(input_file)
    all_columns = list(df.columns)

    # 动态列 = 全部列中不属于固定列的
    dynamic_columns = [c for c in all_columns if c not in fixed_columns]
    if not dynamic_columns:
        print("?? 没有检测到动态列，未做任何修改。")
        # 直接拷贝保存原文件或退出
        if output_file != input_file:
            df.to_excel(output_file, index=False)
            print(f"已将原文件复制为 {output_file}")
        return

    # 打开 Excel 工作簿
    wb = load_workbook(input_file)
    ws = wb.active

    # 列名 -> 列号 映射（openpyxl 列号从1开始）
    col_index_map = {c: i + 1 for i, c in enumerate(all_columns)}
    dynamic_start = min(col_index_map[c] for c in dynamic_columns)
    dynamic_end = max(col_index_map[c] for c in dynamic_columns)

    max_row = ws.max_row

    # 遍历每一行，从第2行开始（第1行为表头）
    for row in range(2, max_row + 1):
        col = dynamic_start
        # 对当前行，在动态列范围内循环
        while col <= dynamic_end:
            cell = ws.cell(row=row, column=col)
            val = cell.value

            # 判断是否等于 0 （支持 int/float 或字符串 '0'，忽略 None/空字符串）
            is_zero = False
            if val is None:
                is_zero = False
            elif isinstance(val, (int, float)):
                is_zero = (val == 0)
            elif isinstance(val, str):
                is_zero = (val.strip() == "0")
            else:
                is_zero = False

            if is_zero:
                # 将该单元格右边所有动态列值左移一格
                for c in range(col, dynamic_end):
                    ws.cell(row=row, column=c).value = ws.cell(row=row, column=c + 1).value
                # 最右侧动态列位置置空
                ws.cell(row=row, column=dynamic_end).value = None
                # 注意：不要增加 col，这样会再次检查当前位置（因为已经左移了一个新值到当前位置）
                # 动态_end 不变（我们不删除列，只是左移数据）
            else:
                col += 1

    # 最后清空动态列的表头（保持固定列标题不变）
    for c in range(dynamic_start, dynamic_end + 1):
        ws.cell(row=1, column=c).value = ""

    # 保存
    # 如果输出目录不存在则创建
    out_dir = os.path.dirname(os.path.abspath(output_file))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_file)
    print(f"? 已处理并保存到: {output_file}")


def calculate_diff_ratio(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    max_row = ws.max_row

    # 动态获取最大数据列（假设前两列是“代码”“名称”，从第3列开始是数值列）
    max_col = 0
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=3):
        for idx, cell in enumerate(row, start=3):
            if isinstance(cell.value, (int, float)):
                max_col = max(max_col, idx)

    # 计算结果列从最大数据列向右跳 3~6 列
    p_col_index = max_col + 3
    y_col_index = max_col + 4
    z_col_index = max_col + 5
    aa_col_index = max_col + 6

    # 表头
    ws.cell(row=1, column=p_col_index).value = "2025-03 vs 2024-12 比率"
    ws.cell(row=1, column=y_col_index).value = "2025-06 vs 2025-03 比率"
    ws.cell(row=1, column=z_col_index).value = "倒数第一 vs 2025-03 比率"
    ws.cell(row=1, column=aa_col_index).value = "倒数第一 vs 倒数第二 比率"

    # 找到关键列索引
    header_row = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_2025_06 = header_row.get("2025-06(股东人数)")
    col_2025_03 = header_row.get("2025-03(股东人数)")
    col_2024_12 = header_row.get("2024-12(股东人数)")

    for row in range(2, max_row + 1):
        values = []
        for col in range(3, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                values.append(val)

        # P 列计算（2025-03 vs 2024-12）
        if col_2024_12 and col_2025_03:
            a_val = ws.cell(row=row, column=col_2025_03).value
            b_val = ws.cell(row=row, column=col_2024_12).value
            ws.cell(row=row, column=p_col_index).value = ((a_val - b_val) / b_val) if b_val else 0

        # Q 列计算（2025-06 vs 2025-03）
        if col_2025_06 and col_2025_03:
            a_val = ws.cell(row=row, column=col_2025_06).value
            b_val = ws.cell(row=row, column=col_2025_03).value
            ws.cell(row=row, column=y_col_index).value = ((a_val - b_val) / b_val) if b_val else 0

        # R 列计算（倒数第一 vs 2025-03）
        if col_2025_03 and values:
            b_val = ws.cell(row=row, column=col_2025_03).value
            last_val = values[-1]
            ws.cell(row=row, column=z_col_index).value = ((last_val - b_val) / b_val) if b_val else 0

        # S 列计算（倒数第一 vs 倒数第二）
        if len(values) >= 2:
            last_val = values[-1]
            second_last_val = values[-2]
            ws.cell(row=row, column=aa_col_index).value = (
                    (last_val - second_last_val) / second_last_val) if second_last_val else 0
        else:
            ws.cell(row=row, column=aa_col_index).value = 0

    wb.save(output_file)
    print(f"? 已完成，结果已保存到 {output_file}")


if __name__ == "__main__":
    sh_gdrs = "/Users/zkp/Documents/gdrs/副本sh3_gdrs(1).xlsx"
    sz_gdrs = "/Users/zkp/Documents/gdrs/副本sz3_gdrs(1).xlsx"
    remove_st_rows_excel(sh_gdrs)
    remove_st_rows_excel(sz_gdrs)

    sh_gdrs = "/Users/zkp/Documents/gdrs/副本sh3_gdrs(1)_clean.xlsx"
    sz_gdrs = "/Users/zkp/Documents/gdrs/副本sz3_gdrs(1)_clean.xlsx"
    merged = "/Users/zkp/Documents/gdrs/merged.xlsx"
    merge_excels_with_dynamic_columns(sz_gdrs, sh_gdrs, merged)
    remove_zero_cells_and_shift_left(merged, merged)
    calculate_diff_ratio(merged, merged)
