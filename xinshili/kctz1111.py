import os
import re
import pandas as pd

import xlwings as xw
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from xinshili.utils import dirname, get_filename_without_extension


def merge_and_sum_sku(input_file, output_file):
    """
    将 SKU 列相同内容的行合并，并将对应的产品总数列的数值相加，按产品总数降序排列
    :param input_file: 输入 Excel 文件路径
    :param output_file: 输出 Excel 文件路径
    """
    app = xw.App(visible=False)  # 隐藏 Excel 窗口
    try:
        # 打开输入文件
        wb = app.books.open(input_file)
        sheet = wb.sheets.active  # 获取活动工作表

        # 获取数据范围
        data = sheet.range("A1").expand("table").value  # 获取整个表格数据
        header = data[0]  # 表头
        rows = data[1:]  # 数据部分

        # 检查列名
        if "SKU" not in header or "产品总数" not in header:
            raise ValueError("表格中未找到 'sku' 或 '产品总数' 列，请检查数据格式。")

        # 获取列索引
        sku_index = header.index("SKU")
        product_total_index = header.index("产品总数")

        # 使用字典合并数据
        result_dict = {}
        for row in rows:
            sku = str(row[sku_index]).strip() if row[sku_index] else None
            product_total = row[product_total_index] if isinstance(row[product_total_index], (int, float)) else 0
            if sku:
                result_dict[sku] = result_dict.get(sku, 0) + product_total

        # 转换为列表形式并排序
        result_data = [["SKU", "产品总数", "预计剩余"]]  # 新表头
        sorted_data = sorted(result_dict.items(), key=lambda x: x[1], reverse=True)  # 按产品总数降序排序
        result_data.extend([[sku, total] for sku, total in sorted_data])

        # 写入新工作表
        new_sheet = wb.sheets.add(name="合并结果")
        new_sheet.range("A1").value = result_data

        # 保存输出文件
        wb.save(output_file)
        print(f"合并结果已保存到: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")
    finally:
        app.quit()


def match_and_fill_with_openpyxl(file1_path, file2_path, output_path):
    """
    根据文件1的E列数据匹配文件2的B列，如果匹配成功，将文件2的I列内容填充到文件1的P列。

    :param file1_path: 文件1路径
    :param file2_path: 文件2路径
    :param output_path: 结果保存路径
    """
    # 加载文件1和文件2
    wb1 = load_workbook(file1_path)
    ws1 = wb1.active  # 默认第一个工作表

    wb2 = load_workbook(file2_path)
    ws2 = wb2.active  # 默认第一个工作表

    # 创建文件2的B列与I列的映射关系
    mapping = {ws2.cell(row=row, column=2).value: ws2.cell(row=row, column=9).value
               for row in range(2, ws2.max_row + 1)}  # B列是第2列，I列是第9列

    # 设置字体颜色为红色
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色填充

    # 遍历文件1的E列，从第二行开始
    for row in range(2, ws1.max_row + 1):
        value = ws1.cell(row=row, column=1).value  # A列是第1列
        if value in mapping:  # 如果匹配到
            matched_content = mapping[value]

            if isinstance(matched_content, str) and "预警" in matched_content:
                match = re.search(r"\d+", matched_content)
                if match:
                    number = int(match.group())  # 提取并转换为整数
                    value_ = number - int(ws1.cell(row=row, column=2).value)
                    ws1.cell(row=row, column=3).value = value_
                    if value_ < 30:
                        ws1[f"C{row}"].fill = red_fill
                else:
                    print("未找到数字")
            elif isinstance(matched_content, str) and "缺货" in matched_content:
                ws1.cell(row=row, column=3).value = 0
                ws1[f"C{row}"].fill = red_fill
            else:
                value_ = int(matched_content) - int(ws1.cell(row=row, column=2).value)
                ws1.cell(row=row, column=3).value = value_
                if value_ < 30:
                    ws1[f"C{row}"].fill = red_fill

    for column_cells in ws1.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
                cell.alignment = Alignment(horizontal="center", vertical="center")
        adjusted_width = max_length + 2  # 增加一些宽度间隔
        ws1.column_dimensions[column_letter].width = adjusted_width

    # 保存文件1为新的结果文件
    wb1.save(output_path)
    print(f"结果已保存到: {output_path}")


def convert_xls_to_xlsx(input_file, output_file):
    try:
        # 打开 .xls 文件
        app = xw.App(visible=False)
        wb = app.books.open(input_file)

        # 保存为 .xlsx 格式
        wb.save(output_file)

        # 关闭工作簿和应用程序
        wb.close()
        app.quit()
        print(f"文件已成功转换为: {output_file}")
    except Exception as e:
        if 'app' in locals():
            app.quit()
        print(f"转换过程中发生错误: {e}")


def convert_csv_to_xlsx(csv_file, xlsx_file):
    """
    将 CSV 文件转换为 XLSX 文件格式。

    :param csv_file: 输入的 CSV 文件路径
    :param xlsx_file: 输出的 XLSX 文件路径
    """
    try:
        # 读取 CSV 文件
        data = pd.read_csv(csv_file)

        # 将数据写入 XLSX 文件
        data.to_excel(xlsx_file, index=False, engine="openpyxl")

        print(f"文件已成功转换为 XLSX 格式: {xlsx_file}")
    except Exception as e:
        print(f"转换过程中发生错误: {e}")


# order_input_file = input("请输入订单文件的绝对路径：")
# kc_input_file = input("请输入库存文件的绝对路径：")
# kc_file_extension = os.path.splitext(kc_input_file)[-1].lower()
# kc_xlsx_path = dirname(kc_input_file) + "/" + get_filename_without_extension(kc_input_file) + "_cover.xlsx"
# if kc_file_extension == ".csv":
#     print("检测到 CSV 文件，开始处理...")
#     convert_csv_to_xlsx(kc_input_file, kc_xlsx_path)
# elif kc_file_extension == ".xls":
#     print("检测到 XLSX 文件，开始处理...")
#     convert_xls_to_xlsx(kc_input_file, kc_xlsx_path)
# else:
#     raise ValueError("不支持的文件格式，仅支持 .csv 和 .xlsx")
# output_file = dirname(kc_input_file) + "/" + "result.xlsx"
# merge_and_sum_sku(order_input_file, output_file)
# match_and_fill_with_openpyxl(output_file, kc_xlsx_path, output_file)


input_file = "/Users/zkp/Desktop/B&Y/kctz/order_120250113092529444_1573179.xlsx"  # 输入文件路径
output_file = "/Users/zkp/Desktop/B&Y/kctz/xsxsxs.xlsx"  # 输出文件路径
table_1 = "/Users/zkp/Desktop/B&Y/kctz/table_1_cover.xlsx"  # 输出文件路径
merge_and_sum_sku(input_file, output_file)
match_and_fill_with_openpyxl(output_file, table_1, output_file)
