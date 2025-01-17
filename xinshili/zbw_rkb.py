from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


def expand_sku_and_dimensions_rows(file1, file2, output_file,
                                   sku_column_file1="SKU", box_column_file1="总共箱数",
                                   qty_column_file1="单箱数量", size_column_file1="箱子尺寸",
                                   weight_column_file1="单箱重量",
                                   sku_column_file2="SKU/SKU", qty_column_file2="Qty per Package/每箱数量",
                                   length_column_file2="Package Length/箱子长度 (cm/inch)",
                                   width_column_file2="Package Width/箱子宽度 (cm/inch)",
                                   height_column_file2="Package Height/箱子高度 (cm/inch)",
                                   weight_column_file2="Package Weight/箱子重量 (kg/lb)",
                                   id_column_file2="Identification No./标识号",
                                   package_qty_column_file2="Package Qty/箱子数量",
                                   unit_column_file2="Unit (Metric/Imperial)/单位 (公制/英制)",
                                   default_unit="公制 (cm/kg)"):
    """
    根据文件 1 的 SKU、总共箱数、单箱数量、单箱重量和箱子尺寸，将数据扩展并复制到文件 2 的指定列中。
    并为文件 2 添加箱子尺寸的长度、宽度、高度和重量，同时配置标识号、箱子数量和单位。

    :param file1: 输入的 Excel 文件 1，包含 SKU、总共箱数、单箱数量、单箱重量和箱子尺寸
    :param file2: 输入的 Excel 文件 2，目标列
    :param output_file: 输出的 Excel 文件路径
    :param sku_column_file1: 文件 1 中的 SKU 列名
    :param box_column_file1: 文件 1 中的箱数列名
    :param qty_column_file1: 文件 1 中的单箱数量列名
    :param size_column_file1: 文件 1 中的箱子尺寸列名
    :param weight_column_file1: 文件 1 中的单箱重量列名
    :param sku_column_file2: 文件 2 中的 SKU 列名
    :param qty_column_file2: 文件 2 中的单箱数量列名
    :param length_column_file2: 文件 2 中的箱子长度列名
    :param width_column_file2: 文件 2 中的箱子宽度列名
    :param height_column_file2: 文件 2 中的箱子高度列名
    :param weight_column_file2: 文件 2 中的箱子重量列名
    :param id_column_file2: 文件 2 中的标识号列名
    :param package_qty_column_file2: 文件 2 中的箱子数量列名
    :param unit_column_file2: 文件 2 中的单位列名
    :param default_unit: 单位列的默认值，默认为 "公制 (cm/kg)"
    """
    try:
        # 加载文件 1
        wb1 = load_workbook(file1, data_only=True)
        ws1 = wb1.active

        # 获取文件 1 的列名到索引的映射
        headers_file1 = {cell.value: idx for idx, cell in enumerate(ws1[1])}

        # 确保列名存在
        required_columns = [sku_column_file1, box_column_file1, qty_column_file1, size_column_file1,
                            weight_column_file1]
        for column in required_columns:
            if column not in headers_file1:
                raise ValueError(f"文件 1 中未找到列: {column}")

        # 提取文件 1 的数据
        sku_idx_file1 = headers_file1[sku_column_file1]
        box_idx_file1 = headers_file1[box_column_file1]
        qty_idx_file1 = headers_file1[qty_column_file1]
        size_idx_file1 = headers_file1[size_column_file1]
        weight_idx_file1 = headers_file1[weight_column_file1]

        expanded_rows = []
        for row in ws1.iter_rows(min_row=2, values_only=True):
            sku = row[sku_idx_file1]
            box_count = row[box_idx_file1]
            qty_per_box = row[qty_idx_file1]
            size = row[size_idx_file1]
            weight = row[weight_idx_file1]
            if sku and isinstance(box_count, (int, float)) and isinstance(qty_per_box,
                                                                          (int, float)) and size and weight:
                dimensions = size.split("*")
                if len(dimensions) == 3:  # 确保尺寸格式正确
                    length, width, height = dimensions
                    for _ in range(int(box_count)):
                        expanded_rows.append((sku, qty_per_box, length, width, height, weight))

        # 加载文件 2
        wb2 = load_workbook(file2)
        ws2 = wb2.active

        # 获取目标列的索引
        headers_file2 = {cell.value: idx for idx, cell in enumerate(ws2[1])}
        required_columns_file2 = [sku_column_file2, qty_column_file2,
                                  length_column_file2, width_column_file2, height_column_file2,
                                  weight_column_file2, id_column_file2, package_qty_column_file2, unit_column_file2]
        for column in required_columns_file2:
            if column not in headers_file2:
                raise ValueError(f"文件 2 中未找到列: {column}")

        sku_idx_file2 = headers_file2[sku_column_file2]
        qty_idx_file2 = headers_file2[qty_column_file2]
        length_idx_file2 = headers_file2[length_column_file2]
        width_idx_file2 = headers_file2[width_column_file2]
        height_idx_file2 = headers_file2[height_column_file2]
        weight_idx_file2 = headers_file2[weight_column_file2]
        id_idx_file2 = headers_file2[id_column_file2]
        package_qty_idx_file2 = headers_file2[package_qty_column_file2]
        unit_idx_file2 = headers_file2[unit_column_file2]

        # 清空目标列的数据并填充扩展数据
        for row_idx in range(2, ws2.max_row + 1):
            for idx in [sku_idx_file2, qty_idx_file2, length_idx_file2, width_idx_file2,
                        height_idx_file2, weight_idx_file2, id_idx_file2, package_qty_idx_file2, unit_idx_file2]:
                ws2.cell(row=row_idx, column=idx + 1).value = None

        for idx, (sku, qty, length, width, height, weight) in enumerate(expanded_rows, start=2):
            ws2.cell(row=idx, column=sku_idx_file2 + 1).value = sku
            ws2.cell(row=idx, column=qty_idx_file2 + 1).value = qty
            ws2.cell(row=idx, column=length_idx_file2 + 1).value = float(length)
            ws2.cell(row=idx, column=width_idx_file2 + 1).value = float(width)
            ws2.cell(row=idx, column=height_idx_file2 + 1).value = float(height)
            ws2.cell(row=idx, column=weight_idx_file2 + 1).value = float(weight)
            ws2.cell(row=idx, column=id_idx_file2 + 1).value = idx - 1  # 自然数字从 1 开始
            ws2.cell(row=idx, column=package_qty_idx_file2 + 1).value = 1  # 箱子数量统一为 1
            ws2.cell(row=idx, column=unit_idx_file2 + 1).value = default_unit  # 设置默认单位

        # 调整单元格宽度并居中
        for column_cells in ws2.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = max_length + 2
            ws2.column_dimensions[column_letter].width = adjusted_width

        # 保存输出文件
        wb2.save(output_file)
        print(f"结果已保存到文件: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


# 示例调用
file1_path = "/Users/zkp/Desktop/B&Y/入库单/zbw/zyl美西改美中仓库.xlsx"  # 替换为 Excel 文件 1 的路径
file2_path = "/Users/zkp/Desktop/B&Y/入库单/zbw/入库模版.xlsx"  # 替换为 Excel 文件 2 的路径
output_file_path = "/Users/zkp/Desktop/B&Y/入库单/zbw/output.xlsx"  # 替换为输出文件路径

expand_sku_and_dimensions_rows(file1_path, file2_path, output_file_path)
