from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

from xinshili import utils

"""
领星oms入库单，用于给仓库管理人员核准需要入库的产品的类别，箱数，总数 是否正确。
"""


def expand_sku_and_dimensions_rows(file1, file2,
                                   sku_column_file1="SKU", total_qty_column_file1="总数量",
                                   box_column_file1="总共箱数", qty_column_file1="单箱数量",
                                   size_column_file1="箱子尺寸", weight_column_file1="单箱重量",
                                   sku_column_file2="SKU/SKU", qty_column_file2="Qty per Package/每箱数量",
                                   length_column_file2="Package Length/箱子长度 (cm/inch)",
                                   width_column_file2="Package Width/箱子宽度 (cm/inch)",
                                   height_column_file2="Package Height/箱子高度 (cm/inch)",
                                   weight_column_file2="Package Weight/箱子重量 (kg/lb)",
                                   id_column_file2="Identification No./标识号",
                                   package_qty_column_file2="Package Qty/箱子数量",
                                   unit_column_file2="Unit (Metric/Imperial)/单位 (公制/英制)",
                                   default_unit="公制 (cm/kg)"):
    """
    根据文件 1 的 SKU、总共箱数、单箱数量、单箱重量和箱子尺寸，将数据扩展并复制到文件 2 的指定列中。
    并为文件 2 添加箱子尺寸的长度、宽度、高度和重量，同时配置标识号、箱子数量和单位。

    :param file1: 输入的 Excel 文件 1，包含 SKU、总共箱数、单箱数量、单箱重量和箱子尺寸
    :param file2: 输入的 Excel 文件 2，目标列
    :param sku_column_file1: 文件 1 中的 SKU 列名
    :param box_column_file1: 文件 1 中的箱数列名
    :param qty_column_file1: 文件 1 中的单箱数量列名
    :param size_column_file1: 文件 1 中的箱子尺寸列名
    :param weight_column_file1: 文件 1 中的单箱重量列名
    :param sku_column_file2: 文件 2 中的 SKU 列名
    :param qty_column_file2: 文件 2 中的单箱数量列名
    :param length_column_file2: 文件 2 中的箱子长度列名
    :param width_column_file2: 文件 2 中的箱子宽度列名
    :param height_column_file2: 文件 2 中的箱子高度列名
    :param weight_column_file2: 文件 2 中的箱子重量列名
    :param id_column_file2: 文件 2 中的标识号列名
    :param package_qty_column_file2: 文件 2 中的箱子数量列名
    :param unit_column_file2: 文件 2 中的单位列名
    :param default_unit: 单位列的默认值，默认为 "公制 (cm/kg)"
    """
    try:
        # 加载文件 1
        wb1 = load_workbook(file1, data_only=True)
        ws1 = wb1.active

        # 获取文件 1 的列名到索引的映射
        headers_file1 = {cell.value: idx for idx, cell in enumerate(ws1[1])}

        # 确保列名存在
        required_columns = [sku_column_file1, total_qty_column_file1, box_column_file1, qty_column_file1,
                            size_column_file1, weight_column_file1]
        for column in required_columns:
            if column not in headers_file1:
                raise ValueError(f"文件 1 中未找到列: {column}")

        sku_idx_file1 = headers_file1[sku_column_file1]
        total_qty_idx_file1 = headers_file1[total_qty_column_file1]
        box_idx_file1 = headers_file1[box_column_file1]
        qty_idx_file1 = headers_file1[qty_column_file1]
        size_idx_file1 = headers_file1[size_column_file1]
        weight_idx_file1 = headers_file1[weight_column_file1]

        expanded_rows = []
        for row in ws1.iter_rows(min_row=2, values_only=True):
            sku = row[sku_idx_file1]
            total_qty = row[total_qty_idx_file1]
            box_count = row[box_idx_file1]
            qty_per_box = row[qty_idx_file1]
            size = row[size_idx_file1]
            weight = row[weight_idx_file1]

            if sku and total_qty and box_count and qty_per_box and size and weight:
                # 校验并调整最后一箱的数量
                calculated_total_qty = box_count * qty_per_box
                remaining_qty = total_qty % qty_per_box if total_qty != calculated_total_qty else 0

                for i in range(int(box_count)):
                    if i == int(box_count) - 1 and remaining_qty != 0:  # 最后一箱校准
                        expanded_rows.append((sku, remaining_qty, size, weight))
                    else:
                        expanded_rows.append((sku, qty_per_box, size, weight))

        # 加载文件 2
        wb2 = load_workbook(file2)
        ws2 = wb2.active

        # 获取文件 2 的列名到索引的映射
        headers_file2 = {cell.value: idx for idx, cell in enumerate(ws2[1])}
        required_columns_file2 = [sku_column_file2, qty_column_file2, length_column_file2, width_column_file2,
                                  height_column_file2, weight_column_file2, id_column_file2, package_qty_column_file2,
                                  unit_column_file2]
        for column in required_columns_file2:
            if column not in headers_file2:
                raise ValueError(f"文件 2 中未找到列: {column}")

        sku_idx_file2 = headers_file2[sku_column_file2]
        qty_idx_file2 = headers_file2[qty_column_file2]
        length_idx_file2 = headers_file2[length_column_file2]
        width_idx_file2 = headers_file2[width_column_file2]
        height_idx_file2 = headers_file2[height_column_file2]
        weight_idx_file2 = headers_file2[weight_column_file2]
        id_idx_file2 = headers_file2[id_column_file2]
        package_qty_idx_file2 = headers_file2[package_qty_column_file2]
        unit_idx_file2 = headers_file2[unit_column_file2]

        # 清空文件 2 数据
        ws2.delete_rows(2, ws2.max_row)

        # 填充文件 2 的数据
        for idx, (sku, qty, size, weight) in enumerate(expanded_rows, start=2):
            dimensions = size.split("*")
            if len(dimensions) != 3:
                raise ValueError(f"SKU {sku} 的箱子尺寸格式不正确: {size}")
            length, width, height = dimensions

            ws2.cell(row=idx, column=sku_idx_file2 + 1).value = sku
            ws2.cell(row=idx, column=qty_idx_file2 + 1).value = qty
            ws2.cell(row=idx, column=length_idx_file2 + 1).value = float(length)
            ws2.cell(row=idx, column=width_idx_file2 + 1).value = float(width)
            ws2.cell(row=idx, column=height_idx_file2 + 1).value = float(height)
            ws2.cell(row=idx, column=weight_idx_file2 + 1).value = float(weight)
            ws2.cell(row=idx, column=id_idx_file2 + 1).value = idx - 1
            ws2.cell(row=idx, column=package_qty_idx_file2 + 1).value = 1
            ws2.cell(row=idx, column=unit_idx_file2 + 1).value = default_unit

            ws2.cell(row=idx, column=id_idx_file2 + 1).value = idx - 1  # 自然数字从 1 开始
            ws2.cell(row=idx, column=package_qty_idx_file2 + 1).value = 1  # 箱子数量统一为 1
            ws2.cell(row=idx, column=unit_idx_file2 + 1).value = default_unit  # 设置默认单位

        # 调整单元格宽度并居中
        for column_cells in ws2.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = max_length + 2
            ws2.column_dimensions[column_letter].width = adjusted_width

        # 求和 Package Qty/箱子数量
        package_qty_sum = sum(
            cell for row in
            ws2.iter_cols(min_col=package_qty_idx_file2 + 1, max_col=package_qty_idx_file2 + 1, min_row=2,
                          values_only=True)
            for cell in row if isinstance(cell, (int, float))
        )
        # 求和 Qty per Package/每箱数量
        qty_per_package_sum = sum(
            cell for row in
            ws2.iter_cols(min_col=qty_idx_file2 + 1, max_col=qty_idx_file2 + 1, min_row=2, values_only=True)
            for cell in row if isinstance(cell, (int, float))
        )
        # print(f"Package Qty/箱子数量总和: {package_qty_sum}")
        # print(f"Qty per Package/每箱数量总和: {qty_per_package_sum}")

        # 输出路径
        output_file = "/Users/zkp/Desktop/B&Y/oms_rkd/" + utils.get_filename_without_extension(file1_path) \
                      + f"_{package_qty_sum}箱" + f"_{qty_per_package_sum}件" + ".xlsx"

        wb2.save(output_file)
        print(f"结果已保存至: {output_file}")
        utils.open_dir(utils.get_file_dir(output_file))

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


file1_path = input("请输入源表文件的绝对路径：")  # eg: "/Users/zkp/PycharmProjects/PyBreath/xinshili/xlsx/rkd/demo.xlsx"
file2_path = utils.current_dir() + "/xlsx/rkd/入库单模版.xlsx"
expand_sku_and_dimensions_rows(file1_path, file2_path)
