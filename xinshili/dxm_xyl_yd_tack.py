import re
from datetime import datetime, timezone, timedelta

import pandas as pd
import os
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from xinshili.flld_gjgz import get_unpaid_tracking_data, get_alert_intercepted_data
from xinshili.fs_utils_plus import get_token, dimension_range, FsConstants, value_range, ClientMapConstants, \
    ClientConstants, brief_sheet_value, brief_sheet_bg, khhz_sheet_value, khhz_sheet_bg, fs_msg, FsUserID
from xinshili.gjgz_plus333 import check_and_add_courier_column, extract_and_process_data, RowName, CourierStateMapKey, \
    update_courier_status, is_time_difference_exceed, process_tracking_no, CourierStateMapValue, \
    find_irregular_tracking_numbers, update_courier_status1, filter_tracking_numbers, count_pattern_state, Pattern, \
    get_days_difference
from xinshili.pd_utils import remove_duplicates_by_column
from xinshili.utils import natural_key, getYmd, is_us_weekend, delete_file, round2
from xinshili.yd_to_dxm import clean_order_id


def convert_china_to_utc0(china_time: datetime) -> datetime:
    """
    将中国时间转换为零时区（UTC/Zulu Time），忽略夏令时
    :param china_time: 中国时间 datetime 对象（无 tzinfo 或本地时间）
    :return: UTC 时间 datetime（无 tzinfo）
    """
    if not isinstance(china_time, datetime):
        raise ValueError("china_time 必须是 datetime 类型")

    china_tz = timezone(timedelta(hours=8))  # 中国时区 UTC+8
    utc_tz = timezone.utc  # UTC 零时区

    # 设置中国时区 → 转为 UTC → 去除 tzinfo
    return china_time.replace(tzinfo=china_tz).astimezone(utc_tz).replace(tzinfo=None)


def process_tracking_time1(file_path):
    df = pd.read_excel(file_path, dtype=str)

    # 先排除状态为 unpaid、delivered、irregular_no_tracking 的行
    df_filtered = df[~df[RowName.Courier].str.lower().isin([
        # CourierStateMapValue.unpaid,
        # CourierStateMapValue.delivered,
        CourierStateMapValue.irregular_no_tracking])].copy()

    intervals = []
    states = []
    track_times = []

    # 当前中国时间转换为零时区时间，因为usps接口返回的时间是领时区的
    utc0_now = convert_china_to_utc0(datetime.now())

    for idx, row in df_filtered.iterrows():
        try:
            courier = str(row.get(RowName.Courier, "")).strip().lower()
            yd_number = str(row.get(RowName.YD_Number, "")).strip()

            # 优先判断是否已交付或已换阳单
            if courier == CourierStateMapValue.delivered:
                intervals.append("")
                states.append("已经交付")
                track_times.append(utc0_now)
                continue

            # 原来的运单号不能是unpaid，如果是unpaid要接着往下执行，进行补发提示
            if yd_number and yd_number.lower() != "nan" and courier != CourierStateMapValue.unpaid:
                intervals.append("")
                states.append("已换阳单")
                track_times.append(utc0_now)
                continue

            # 如果没有触发上面两个判断，则进入常规时间判断逻辑
            ship_time_str = str(row.get("发货时间", "")).strip()
            date_str = str(row.get(RowName.LatestEventSfDate, "")).strip()
            time_str = str(row.get(RowName.LatestEventSfTime, "")).strip()

            date_flag = date_str and date_str.lower() != "nan"
            time_flag = time_str and time_str.lower() != "nan"
            outbound_time_flag = ship_time_str and ship_time_str.lower() != "nan"

            diff = None
            if date_flag:  # 存在最新日期
                if time_flag:  # 存在最新时间
                    latest_date_time = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
                else:
                    latest_date_time = datetime.strptime(f"{date_str}", "%Y-%m-%d")
                # 现在时间 - 最新日期时间
                diff = utc0_now - latest_date_time
            elif outbound_time_flag:
                creation_time = datetime.strptime(ship_time_str, "%Y-%m-%d %H:%M:%S")
                # 现在时间 - 发货时间
                diff = utc0_now - convert_china_to_utc0(creation_time)

            if diff is not None:
                hours = round(diff.total_seconds() / 3600, 2)
                total_seconds = int(diff.total_seconds())
                hours_part = total_seconds // 3600
                minutes_part = (total_seconds % 3600) // 60
                interval_str = f"{hours_part:02}:{minutes_part:02}"
                intervals.append(interval_str)

                # 默认无法替换判断时间为 72 小时
                delay_hours = 72

                # ✅ 判断付款时间是否为美国周五、六、日，如果是则延迟 48 小时
                pay_time_str = str(row.get("付款时间", "")).strip()
                if pay_time_str and pay_time_str.lower() != "nan":
                    try:
                        pay_time_china = datetime.strptime(pay_time_str, "%Y-%m-%d %H:%M:%S")
                        pay_time_utc0 = convert_china_to_utc0(pay_time_china)
                        # 转为美国东部时间（或你实际使用的 USPS 时区）
                        pay_time_us = pay_time_utc0.replace(tzinfo=timezone.utc).astimezone(
                            ZoneInfo("America/New_York"))
                        if pay_time_us.weekday() in [4, 5, 6]:  # 星期五、六、日
                            delay_hours += 48
                    except Exception as e:
                        print(f"⚠️ 第 {idx} 行付款时间解析失败: {e}")

                # 优先判断是否unpaid
                if courier == CourierStateMapValue.unpaid:
                    if hours <= 192:  # <=8天
                        states.append("邮资未付<=8天")
                    else:
                        states.append("邮资未付>8天")
                    track_times.append(utc0_now)
                    continue

                if hours >= delay_hours:
                    states.append("无法替换")
                elif hours >= 48:
                    states.append("阳单替换")
                elif hours >= 24:
                    states.append("预备阳单")
                # elif hours >= 120:
                #     states.append("超5天无轨迹更新")
                else:
                    states.append("轨迹正常")
            else:
                intervals.append(None)
                states.append("")

            track_times.append(utc0_now)

        except Exception as e:
            print(f"⚠️ 第 {idx} 行处理失败: {e}")
            intervals.append(None)
            states.append("")
            track_times.append(utc0_now)

    df_filtered[RowName.TrackTimeInterval] = intervals
    df_filtered[RowName.TrackTimeIntervalState] = states
    df_filtered[RowName.Tacking_Time] = [
        t.strftime("%Y-%m-%d %H:%M:%S") if isinstance(t, datetime) else "" for t in track_times
    ]

    df.update(df_filtered)
    df.to_excel(file_path, index=False)
    print(f"✅ 已处理并保存至：{file_path}")


def create_fs_xlsx_file(file_path):
    # 定义表头
    columns = [
        RowName.Pay_Time, RowName.Ship_Time, RowName.Order_Num, RowName.Track_Num,
        RowName.Track_State, RowName.Analyse_State,
        # RowName.Last_Track_Time,
        RowName.Latest_Track_Site, RowName.Latest_Track_Time, RowName.Interval_Time, RowName.Process_Time,
        RowName.YD_Number2, RowName.YD_State2
    ]

    # 创建空的 DataFrame 并写入文件（覆盖或新建）
    df = pd.DataFrame(columns=columns)
    df.to_excel(file_path, index=False)
    print(f"✅ 文件已{'创建' if not os.path.exists(file_path) else '清空并重建'}：{file_path}")


def export_yd_data(source_file, target_file):
    # 读取源文件
    df = pd.read_excel(source_file, dtype=str)

    def safe_concat_time(row):
        date_part = str(row.get(RowName.LatestEventSfDate, "")).strip()
        time_part = str(row.get(RowName.LatestEventSfTime, "")).strip()
        if date_part.lower() not in ["", "nan"]:
            if time_part.lower() not in ["", "nan"]:
                return f"{date_part} {time_part}"
            else:
                return date_part
        else:
            return ""

    df[RowName.Latest_Track_Time] = df.apply(safe_concat_time, axis=1)

    # 构造目标列的数据（强制转换为字符串以防止科学计数法）
    export_df = pd.DataFrame({
        RowName.Pay_Time: df[RowName.Pay_Time],
        RowName.Ship_Time: df[RowName.Ship_Time],
        RowName.Order_Num: df[RowName.Order_Num].astype(str),
        RowName.Track_Num: df[RowName.Track_Num].astype(str),
        RowName.Track_State: df[RowName.Courier],
        RowName.Analyse_State: df[RowName.Tacking_Time],
        # RowName.Last_Track_Time: df[RowName.LastEventSfTime],
        RowName.Latest_Track_Site: df[RowName.LatestEventSfSite],
        RowName.Latest_Track_Time: df[RowName.Latest_Track_Time],
        RowName.Interval_Time: df[RowName.TrackTimeInterval],
        RowName.Process_Time: df[RowName.TrackTimeIntervalState],
        RowName.YD_Number2: df[RowName.YD_Number],
        RowName.YD_State2: df[RowName.YD_State],
    })

    # 文件已存在 → 打开并追加
    wb = load_workbook(target_file)
    ws = wb.active
    start_row = ws.max_row + 1

    for i, row in enumerate(dataframe_to_rows(export_df, index=False, header=False)):
        for j, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + i, column=j, value=value)

            # 设置为文本格式以防止科学计数法
            header = ws.cell(row=1, column=j).value
            if header in [RowName.Order_Num, RowName.Track_Num]:
                cell.number_format = "@"
                cell.value = str(value)  # 强制转为字符串

    wb.save(target_file)
    print(f"✅ 已追加导出 YD 数据至：{target_file}")
    return len(export_df)


def get_xlsx_data_len(file_path):
    # 读取 Excel 文件，保留空值为 ""，禁用自动类型转换
    df = pd.read_excel(file_path, dtype=object).fillna("")
    # 获取列名作为第一行
    header = list(df.columns)
    # 获取数据行
    data_rows = df.values.tolist()
    # 将列头插入到数据最前面
    nested_list = [header] + data_rows
    return nested_list


def usps_track(xlsx_path, column_name, wl_name):
    results = extract_and_process_data(xlsx_path, column_name, 100, wl_name=wl_name, dxm_xyl_yd_flag=True)
    all_maps = {}
    column_mapping = {}
    if wl_name == RowName.Track_Num:
        all_maps = {
            CourierStateMapKey.not_yet_map: results[CourierStateMapKey.not_yet_map],
            CourierStateMapKey.pre_ship_map: results[CourierStateMapKey.pre_ship_map],
            CourierStateMapKey.unpaid_map: results[CourierStateMapKey.unpaid_map],
            CourierStateMapKey.delivered_map: results[CourierStateMapKey.delivered_map],
            CourierStateMapKey.no_tracking_map: results[CourierStateMapKey.no_tracking_map],
            CourierStateMapKey.tracking_map: results[CourierStateMapKey.tracking_map],
            CourierStateMapKey.possession_sf_date_map: results[CourierStateMapKey.possession_sf_date_map],
            CourierStateMapKey.latest_event_sf_date_map: results[CourierStateMapKey.latest_event_sf_date_map],
            CourierStateMapKey.sf_date_equality_map: results[CourierStateMapKey.sf_date_equality_map],
            CourierStateMapKey.latest_event_sf_time_map: results[CourierStateMapKey.latest_event_sf_time_map],
            CourierStateMapKey.latest_event_sf_site_map: results[CourierStateMapKey.latest_event_sf_site_map],
            CourierStateMapKey.alert_map: results[CourierStateMapKey.alert_map],
        }
        column_mapping = {
            CourierStateMapKey.not_yet_map: RowName.Courier,
            CourierStateMapKey.pre_ship_map: RowName.Courier,
            CourierStateMapKey.unpaid_map: RowName.Courier,
            CourierStateMapKey.delivered_map: RowName.Courier,
            CourierStateMapKey.no_tracking_map: RowName.Courier,
            CourierStateMapKey.tracking_map: RowName.Courier,
            CourierStateMapKey.possession_sf_date_map: RowName.PossessionSfDate,
            CourierStateMapKey.latest_event_sf_date_map: RowName.LatestEventSfDate,
            CourierStateMapKey.sf_date_equality_map: RowName.SfDateInterval,
            CourierStateMapKey.latest_event_sf_time_map: RowName.LatestEventSfTime,
            CourierStateMapKey.latest_event_sf_site_map: RowName.LatestEventSfSite,
            CourierStateMapKey.alert_map: RowName.Courier,
        }
    else:
        all_maps = {
            CourierStateMapKey.not_yet_map: results[CourierStateMapKey.not_yet_map],
            CourierStateMapKey.pre_ship_map: results[CourierStateMapKey.pre_ship_map],
            CourierStateMapKey.unpaid_map: results[CourierStateMapKey.unpaid_map],
            CourierStateMapKey.delivered_map: results[CourierStateMapKey.delivered_map],
            CourierStateMapKey.no_tracking_map: results[CourierStateMapKey.no_tracking_map],
            CourierStateMapKey.tracking_map: results[CourierStateMapKey.tracking_map],
            # CourierStateMapKey.possession_sf_date_map: results[CourierStateMapKey.possession_sf_date_map],
            # CourierStateMapKey.latest_event_sf_date_map: results[CourierStateMapKey.latest_event_sf_date_map],
            # CourierStateMapKey.sf_date_equality_map: results[CourierStateMapKey.sf_date_equality_map],
            # CourierStateMapKey.latest_event_sf_time_map: results[CourierStateMapKey.latest_event_sf_time_map],
            # CourierStateMapKey.latest_event_sf_site_map: results[CourierStateMapKey.latest_event_sf_site_map],
            CourierStateMapKey.alert_map: results[CourierStateMapKey.alert_map],
        }
        column_mapping = {
            CourierStateMapKey.not_yet_map: RowName.YD_State,
            CourierStateMapKey.pre_ship_map: RowName.YD_State,
            CourierStateMapKey.unpaid_map: RowName.YD_State,
            CourierStateMapKey.delivered_map: RowName.YD_State,
            CourierStateMapKey.no_tracking_map: RowName.YD_State,
            CourierStateMapKey.tracking_map: RowName.YD_State,
            # CourierStateMapKey.possession_sf_date_map: RowName.PossessionSfDate,
            # CourierStateMapKey.latest_event_sf_date_map: RowName.LatestEventSfDate,
            # CourierStateMapKey.sf_date_equality_map: RowName.SfDateInterval,
            # CourierStateMapKey.latest_event_sf_time_map: RowName.LatestEventSfTime,
            # CourierStateMapKey.latest_event_sf_site_map: RowName.LatestEventSfSite,
            CourierStateMapKey.alert_map: RowName.YD_State,
        }

    update_courier_status(xlsx_path, all_maps, wl=wl_name, column_map=column_mapping)


def go(xlsx_path, dxm_xyl_track_merger, yd_flag=True):
    check_and_add_courier_column(xlsx_path)

    process_tracking_no(xlsx_path, RowName.Track_Num)
    if yd_flag:
        process_tracking_no(xlsx_path, RowName.YD_Number)

    irregular_number_map = find_irregular_tracking_numbers(xlsx_path, RowName.Track_Num)
    irregular_number_list = []
    if irregular_number_map:
        irregular_number_list = list(irregular_number_map.keys())
        update_courier_status1(xlsx_path, irregular_number_map, RowName.Track_Num)

    if yd_flag:
        usps_track(xlsx_path, RowName.Courier, RowName.Track_Num)
        usps_track(xlsx_path, RowName.YD_State, RowName.YD_Number)

    process_tracking_time1(xlsx_path)
    if yd_flag:
        export_yd_data(xlsx_path, dxm_xyl_track_merger)

    if not yd_flag:
        analyse_state(xlsx_path, irregular_number_list)


def auto(root_dir, yd_flag=True):
    dxm_xyl_track_merger = "/Users/zkp/Desktop/B&Y/轨迹统计/dxm_xyl_track/dxm_xyl_track_merger.xlsx"
    if yd_flag:
        create_fs_xlsx_file(dxm_xyl_track_merger)

    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    if yd_flag:
        track_day = 10
    else:
        track_day = 10
    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = re.search(r"发货时间(\d+)_", file)
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= track_day:
                        print(f"正在处理文件: {xlsx_path}")
                        go(xlsx_path, dxm_xyl_track_merger, False)

    if yd_flag:
        result = get_xlsx_data_len(dxm_xyl_track_merger)
        token = get_token()
        # dimension_range(token, FsConstants.gjgz_token,  ClientMapConstants[ClientConstants.dxm_xyl_yd], 1, 12, majorDimension=FsConstants.COLUMNS)
        value_range(token, FsConstants.gjgz_token, ClientMapConstants[ClientConstants.dxm_xyl_yd], f"A1:L{len(result)}",
                    result)


def update_yd_number(source_file, folder_path):
    # 读取源文件（文件1）
    df_source = pd.read_excel(source_file, dtype=str)
    df_source.fillna('', inplace=True)

    if "订单编号" not in df_source.columns or "平台回传单号" not in df_source.columns:
        print("❌ 文件1中缺少必要列：订单编号 或 平台回传单号")
        return

    # ✅ 仅保留“订单编号”和“平台回传单号”都不为空的有效记录
    df_valid = df_source[
        (df_source["订单编号"].str.strip() != "") &
        (df_source["平台回传单号"].str.strip() != "")
        ]

    print(f"📋 阳单源文件中有效记录数：{len(df_valid)} 条")

    # 构建映射：订单编号 → 平台回传单号
    mapping = dict(zip(df_valid["订单编号"].apply(clean_order_id), df_valid["平台回传单号"].str.strip()))

    # 遍历文件夹中的所有 xlsx 文件
    for dirpath, dirnames, filenames in os.walk(folder_path):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                file_path = os.path.join(sun_dir_path, file)
                try:
                    df = pd.read_excel(file_path, dtype=str)
                    df.fillna('', inplace=True)

                    if RowName.Order_Num in df.columns:
                        if RowName.YD_Number not in df.columns:
                            df[RowName.YD_Number] = ""

                        match_count = 0
                        for i, order_id in df[RowName.Order_Num].items():
                            order_id = str(order_id).strip()
                            if order_id in mapping:
                                yd_number = mapping[order_id]
                                df.at[i, RowName.YD_Number] = yd_number
                                print(f"✅ 匹配成功：文件：{file_path}，订单编号：{order_id}，平台回传单号：{yd_number}")
                                match_count += 1

                        if match_count > 0:
                            df.to_excel(file_path, index=False)
                            print(f"📊 文件 {file_path} 共匹配成功 {match_count} 条记录")
                    else:
                        print(f"⚠️ 文件 {file_path} 缺少“订单号”列，跳过")

                except Exception as e:
                    print(f"❌ 文件 {file_path} 处理失败: {e}")


def analyse_state(xlsx_path, irregular_number_list):
    output_file = os.path.splitext(xlsx_path)[0] + "_去重0.xlsx"
    all_total_count = remove_duplicates_by_column(xlsx_path, output_file, "运单号")  # 无筛选订单总数
    delete_file(output_file)

    output_file = os.path.splitext(xlsx_path)[0] + "_去重1.xlsx"
    filter_tracking_numbers(xlsx_path, output_file, "运单号")
    xlsx_path = output_file

    total_count, no_track_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.no_track)
    track_count = total_count - no_track_count
    total_count2, delivered_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.delivered)
    total_count3, unpaid_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.unpaid)
    total_count4, not_yet_count = count_pattern_state(xlsx_path, RowName.Courier, r"not_yet")
    total_count5, pre_ship_count = count_pattern_state(xlsx_path, RowName.Courier, r"pre_ship")
    total_count6, alert_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.alert)

    text = ""
    fs_text = ""
    ck_time = get_days_difference(xlsx_path, "发货时间")
    gz_time = getYmd()
    interval_time = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(ck_time, "%Y/%m/%d")).days
    is_usweekend = is_us_weekend(ck_time)
    date_obj = datetime.strptime(ck_time, "%Y/%m/%d").date()
    previous_day = date_obj - timedelta(days=1)
    actual_interval = 0
    if is_usweekend == 6:  # 6是中国周日，美国周六
        actual_interval = interval_time - 2
    elif is_usweekend == 0:  # 0是中国周一，美国周日
        actual_interval = interval_time - 1

    text += f"打单日期：{ck_time}"
    text += f"\n跟踪日期：{gz_time}"

    text += f"\n订单总数：{total_count}【{all_total_count}】"
    fs_text += f"\n订单总数：{total_count}【{all_total_count}】"

    swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))
    wswl = round2(100 - swl)
    text += f"\n上网：（{track_count}, {swl}%）"
    fs_text += f"\n上网：（{track_count}, {swl}%）"

    text += f"\n未上网：（{no_track_count}, {wswl}%）"
    fs_text += f"\n未上网：（{no_track_count}, {wswl}%）"

    irregular_no_tracking_countl = round2((len(irregular_number_list) / int(all_total_count)) * 100)
    text += f"\nirregular_number：（{len(irregular_number_list)}, {irregular_no_tracking_countl}%）"
    fs_text += f"\nirregular_number：（{len(irregular_number_list)}, {irregular_no_tracking_countl}%）"

    not_yet_countl = round2((int(not_yet_count) / int(total_count)) * 100)
    text += f"\nnot_yet：（{not_yet_count}, {not_yet_countl}%）"
    fs_text += f"\nnot_yet：（{not_yet_count}, {not_yet_countl}%）"

    pre_ship_countl = round2((int(pre_ship_count) / int(total_count)) * 100)
    text += f"\npre_ship：（{pre_ship_count}, {pre_ship_countl}%）"
    fs_text += f"\npre_ship：（{pre_ship_count}, {pre_ship_countl}%）"

    delivered_countl = round2((int(delivered_count) / int(total_count)) * 100)
    text += f"\ndelivered：（{delivered_count}, {delivered_countl}%）"
    fs_text += f"\ndelivered：（{delivered_count}, {delivered_countl}%）"

    unpaid_countl = round2((int(unpaid_count) / int(total_count)) * 100)
    text += f"\nunpaid：（{unpaid_count}, {unpaid_countl}%）"
    fs_text += f"\nunpaid：（{unpaid_count}, {unpaid_countl}%）"

    alert_countl = round2((int(alert_count) / int(total_count)) * 100)
    text += f"\nalert：（{alert_count}, {alert_countl}%）"
    fs_text += f"\nalert：（{alert_count}, {alert_countl}%）"

    unpaid_tracking_data = get_unpaid_tracking_data(xlsx_path, courier_column='Courier/快递',
                                                    waybill_column='订单号',
                                                    tracking_column='运单号',
                                                    date_column='UnpaidDate/unpaid记录时间',
                                                    key_value='unpaid')

    current_day_unpaid_text = ""
    current_day_unpaid_len = 0
    now_strftime = datetime.now().strftime('%Y-%m-%d')
    if (len(unpaid_tracking_data) > 0):
        text += f"\n-------unpaid详情-------"
        fs_text += f"\n-------unpaid详情-------"
        for times, value in unpaid_tracking_data.items():
            result = f"\n 🕒：{times}"
            for key, val in value.items():
                result += f"\n（单号：{key}, 快递单号：{val}）"
                if times == now_strftime:
                    current_day_unpaid_text += f"{val}\n"
                    current_day_unpaid_len += 1
            result += "\n"
            text += result
            fs_text += result

    alert_intercepted_tracking_data = get_alert_intercepted_data(xlsx_path, waybill_column="订单号",
                                                                 tracking_column="运单号", key_value="alert")

    if (len(alert_intercepted_tracking_data) > 0):
        text += f"\n-------alert详情-------"
        fs_text += f"\n-------alert详情-------"
        for key, value in alert_intercepted_tracking_data.items():
            text += f"\n（单号：{key}, 快递单号：{value}）"
            fs_text += f"\n（单号：{key}, 快递单号：{value}）"

    track_stop_72 = get_alert_intercepted_data(xlsx_path, courier_column="TrackTimeIntervalState/跟踪时间间隔状态",
                                               waybill_column="订单号",
                                               tracking_column="运单号", key_value="无法替换")
    track_stop_48 = get_alert_intercepted_data(xlsx_path, courier_column="TrackTimeIntervalState/跟踪时间间隔状态",
                                               waybill_column="订单号",
                                               tracking_column="运单号", key_value="阳单替换")
    track_stop_24 = get_alert_intercepted_data(xlsx_path, courier_column="TrackTimeIntervalState/跟踪时间间隔状态",
                                               waybill_column="订单号",
                                               tracking_column="运单号", key_value="预备阳单")
    # track_stop_120 = get_alert_intercepted_data(xlsx_path, courier_column="TrackTimeIntervalState/跟踪时间间隔状态",
    #                                             waybill_column="订单号",
    #                                             tracking_column="运单号", key_value="超5天无轨迹补发")
    track_stop_72_int = len(track_stop_72)
    track_stop_48_int = len(track_stop_48)
    track_stop_24_int = len(track_stop_24)
    # track_stop_120_int = len(track_stop_120)
    # track_stop_120l = round2((int(track_stop_120_int) / int(total_count)) * 100)
    track_stop_72l = round2((int(track_stop_72_int) / int(total_count)) * 100)
    track_stop_48l = round2((int(track_stop_48_int) / int(total_count)) * 100)
    track_stop_24l = round2((int(track_stop_24_int) / int(total_count)) * 100)
    text += (f"\n超24小时轨迹未更新：（{track_stop_24_int}, {track_stop_24l}%）"
             f"\n超48小时轨迹未更新：（{track_stop_48_int}, {track_stop_48l}%）"
             f"\n超72小时轨迹未更新：（{track_stop_72_int}, {track_stop_72l}%）")
             # f"\n超120小时轨迹未更新：（{track_stop_120_int}, {track_stop_120l}%）")
    fs_text += (f"\n超24小时轨迹未更新：（{track_stop_24_int}, {track_stop_24l}%）"
                f"\n超48小时轨迹未更新：（{track_stop_48_int}, {track_stop_48l}%）"
                f"\n超72小时轨迹未更新：（{track_stop_72_int}, {track_stop_72l}%）")
                # f"\n超120小时轨迹未更新：（{track_stop_120_int}, {track_stop_120l}%）")

    # track_stop_120_text = ""
    # if (track_stop_120_int > 0):
    #     text += f"\n-------超5天无轨迹补发-------"
    #     fs_text += f"\n-------超5天无轨迹补发-------"
    #     for key, value in track_stop_120.items():
    #         text += f"\n（单号：{key}, 快递单号：{value}）"
    #         fs_text += f"\n（单号：{key}, 快递单号：{value}）"
    #         track_stop_120_text += f"\n（单号：{key}, 快递单号：{value}）"

    print(text)
    delete_file(xlsx_path)

    sum_up_text = ""
    swl_flag = False
    qsl_flag = False
    bg = "#FFFFFF"

    # 上网率判断
    warning_levels = [
        (0, 30, "🧑‍🍳", "继续观察👀", "#FFFFFF"),
        (1, 30, "☁️注意", "未达30%！", "#F8F1D3"),
        (2, 70, "🌧️注意", "未达70%！", "#E3C49C"),
        (3, 97, "❄️异常", "未达97%！", "#F1C1BD"),
    ]

    actual_interval_time = interval_time - actual_interval
    # sum_up_text += f"\n间隔第{interval_time}{actual_interval}天"
    sum_up_text += f"\n间隔第{actual_interval_time}天"

    for days, threshold, icon, message, color in warning_levels:
        if actual_interval_time == days and swl < threshold:
            sum_up_text += f"\n{icon}：上网率为{swl}%，{message}"
            swl_flag = True
            if (actual_interval_time >= 2):
                bg = color
            break
    else:  # ✅ 只有 for 没有 break 时才会执行
        if (actual_interval_time >= 4):
            if (swl >= 99):
                sum_up_text += f"\n☀️上网率为{swl}%，优秀🌈"
            elif (swl >= 97 and swl < 99):
                sum_up_text += f"\n☀️上网率为{swl}%，达标✅"
            else:
                sum_up_text += f"\n⚡️异常：上网率为{swl}%，未达️97%"
                bg = "#F1C1BD"
                swl_flag = True
        else:
            sum_up_text += f"\n☀️上网率为{swl}%，达标✅"

    if (len(alert_intercepted_tracking_data) > 0):
        bg = "#FFF258"

    # if (track_stop_120_int > 0):
    #     bg = "#F54A45"

    if (len(unpaid_tracking_data) > 0):
        bg = "#A684F0"

    tat = get_token()
    # brief_sheet_value(tat, [fs_text], ck_time, gz_time, ClientConstants.dxm_tk_kj)
    # brief_sheet_bg(tat, ck_time, gz_time, ClientConstants.dxm_tk_kj, bg)

    khhz_sheet_value(tat, [
        f"{total_count}",
        f"（{no_track_count}, {wswl}%）",
        f"（{0}, {0}%）",
        f"（{delivered_count}, {delivered_countl}%）",
        f"（{unpaid_count}, {unpaid_countl}%）",
        f"（{track_stop_24_int}, {track_stop_24l}%）",
        f"（{track_stop_48_int}, {track_stop_48l}%）",
        f"（{track_stop_72_int}, {track_stop_72l}%）",
        # f"（{track_stop_120_int}, {track_stop_120l}%）",
    ], ck_time, ClientConstants.dxm_tk_kj)

    khhz_sheet_bg(tat, ck_time, ClientConstants.dxm_tk_kj, bg)

    result_fs_msg = f"客户：店小秘新引力TK_KJ\n"
    result_fs_msg += f"订单创建时间：{ck_time}\n"
    result_fs_msg += f"跟踪时间：{gz_time}\n"
    fs_msg_flag = False

    if len(current_day_unpaid_text) > 0:
        result_fs_msg += f"新增 {current_day_unpaid_len}单 unpaid: \n"
        result_fs_msg += current_day_unpaid_text
        fs_msg_flag = True

    # if track_stop_120_int > 0:
    #     result_fs_msg += f"超5天无轨迹需补发 {track_stop_120_int}单 : \n"
    #     result_fs_msg += track_stop_120_text
    #     fs_msg_flag = True

    if swl_flag:
        if no_track_count > 10:
            result_fs_msg += f"上网率异常: {swl}%\n"
            fs_msg_flag = True

    if fs_msg_flag:
        # print(result_fs_msg)
        fs_msg(FsUserID.WP_ID, result_fs_msg)
        # fs_msg(FsUserID.LW_ID, result_fs_msg)
        # fs_msg(FsUserID.LJ_ID, result_fs_msg)


if __name__ == '__main__':
    select = "请选择功能："
    select += "\n1：☀️写入阳单号"
    select += "\n2：📊轨迹分析️"
    select += "\n"
    select_input = input(select)
    if select_input == "1":
        sun_path = input("请输入阳单号文件的路径:")
        update_yd_number(sun_path, "/Users/zkp/Desktop/B&Y/轨迹统计/dxm_xyl_track/")
    elif select_input == "2":
        auto("/Users/zkp/Desktop/B&Y/轨迹统计/dxm_xyl_track")
    else:
        print("🈚️此项功能！")
