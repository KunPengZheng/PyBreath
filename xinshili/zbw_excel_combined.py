import os
import random
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def merge_and_sort_based_on_largest_header(input_folder, output_file, sort_column, ascending=True, hidden_columns=None):
    """
    合并文件夹中多个文件的相同列名数据，选择表头最多的文件作为模板，并按指定列排序。
    支持隐藏指定列。

    :param input_folder: 文件夹路径，包含多个 Excel 文件
    :param output_file: 输出的合并结果文件路径
    :param sort_column: 指定排序的列名
    :param ascending: 是否升序排序，True 为升序，False 为降序
    :param hidden_columns: 需要隐藏的列名列表
    """
    try:
        # 找到表头最多的文件
        largest_header_file = None
        largest_header = []
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(input_folder, file_name)
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]  # 获取第一行作为表头
                if len(headers) > len(largest_header):
                    largest_header = headers
                    largest_header_file = file_path
                wb.close()

        if not largest_header_file:
            raise FileNotFoundError("未找到任何 .xlsx 文件")

        print(f"表头最多的文件: {largest_header_file}")
        print(f"表头内容: {largest_header}")

        # 创建结果数据字典
        result_data = {header: [] for header in largest_header}

        # 遍历所有文件并合并数据
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(input_folder, file_name)
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]  # 获取第一行作为表头
                rows = list(ws.iter_rows(min_row=2, values_only=True))  # 获取数据行

                # 将行数据转换为字典形式
                for row in rows:
                    row_dict = {headers[idx]: row[idx] for idx in range(len(headers)) if headers[idx] in result_data}
                    for header in largest_header:
                        result_data[header].append(row_dict.get(header, None))  # 如果列不存在，则填充 None

                wb.close()

        # 转换为列表形式，以便排序
        merged_data = list(zip(*result_data.values()))
        headers = list(result_data.keys())

        # 检查排序列是否存在
        if sort_column not in headers:
            raise ValueError(f"排序列 '{sort_column}' 不存在，请检查列名是否正确！")

        # 按指定列排序
        sort_index = headers.index(sort_column)
        merged_data = sorted(merged_data, key=lambda x: (x[sort_index] is None, x[sort_index]), reverse=not ascending)

        # 写入合并结果到输出文件
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "合并结果"

        # 写入表头
        output_ws.append(headers)

        # 写入数据
        for row in merged_data:
            output_ws.append(row)

        # 提取 A 列内容
        column_a_values = [cell.value for cell in ws['A'][1:]]  # 排除表头
        unique_values = set(column_a_values)
        # 随机分配颜色
        color_mapping = {}
        for value in unique_values:
            if value is not None:
                # 生成随机颜色（RGB格式，确保是6位十六进制数）
                random_color = f"{random.randint(0, 255):02X}{random.randint(0, 255):02X}{random.randint(0, 255):02X}"
                color_mapping[value] = PatternFill(start_color=random_color, end_color=random_color, fill_type="solid")
        # 遍历 A 列并设置背景颜色
        for row_idx, cell in enumerate(ws['A'][1:], start=2):  # 从第二行开始
            if cell.value in color_mapping:
                color_fill = color_mapping[cell.value]
                cell.fill = color_fill

        # 自动调整列宽
        for col in output_ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # 列字母
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            output_ws.column_dimensions[col_letter].width = adjusted_width

        # 隐藏指定列
        if hidden_columns:
            for header in hidden_columns:
                if header in headers:
                    col_index = headers.index(header) + 1  # openpyxl 的列索引从 1 开始
                    col_letter = output_ws.cell(row=1, column=col_index).column_letter
                    output_ws.column_dimensions[col_letter].hidden = True

        # 保存结果文件
        output_wb.save(output_file)
        print(f"合并完成并按 '{sort_column}' 排序，结果已保存到: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


# 调用函数
input_folder = input("请输入需要合并的文件所在的目录路径：")  # 存放 Excel 文件的文件夹路径
output_file = input_folder + "/combined.xlsx"
sort_column = "实际发货时间"  # 替换为需要排序的列名
ascending_order = True  # True表示升序，False表示降序
hidden_columns = ["子订单号", "商品名称", "商品名称", "SKCID", "SPUID", "商品属性", "收货人姓名"
    , "收货人联系方式", "备用联系方式", "邮箱", "详细地址1", "详细地址2", "详细地址3"
    , "区县", "城市", "省份", "收货地址邮编", "国家", "要求最晚发货时间", "预计送达时间"]  # 替换为需要隐藏的列名

merge_and_sort_based_on_largest_header(input_folder, output_file, sort_column, ascending_order, hidden_columns)

