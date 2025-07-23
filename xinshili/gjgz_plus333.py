from datetime import datetime, date, timedelta
import os
import re
from openpyxl import load_workbook
import openpyxl
import pandas as pd
from collections import Counter, defaultdict, OrderedDict
from dataclasses import dataclass
import concurrent.futures
import time
import gc
from natsort import natsorted

from xinshili.fs_utils_plus import get_token, brief_sheet_value, detail_sheet_value, ClientConstants, detail_sheet_bg, \
    brief_sheet_bg, khhz_sheet_value, khhz_sheet_bg, FsUserID, fs_msg
from xinshili.pd_utils import remove_duplicates_by_column
from xinshili.usps_utils import track
from xinshili.utils import round2, getYmd, delete_file, is_us_weekend, get_weekday, get_american_holiday, \
    get_chinese_holiday, natural_key

"""
zbw轨迹跟踪分析
"""


@dataclass(frozen=True)
class RowName:
    Tracking_No = 'Tracking No./物流跟踪号'
    Courier = 'Courier/快递'
    CreationTime = "Creation time/创建时间"
    OutboundTime = "OutboundTime/出库时间"
    Warehouse = "Warehouse/仓库"
    Client = "Client/客户"
    CreationWaveTime = "Create wave time/生成波次时间"
    SKU = "SKU"
    ShippingService = "Shipping service/物流渠道"
    PossessionSfDate = "PossessionSfDate/揽收时间"
    LatestEventSfDate = "LatestEventSfDate/最新事件时间"
    LatestEventSfTime = "LatestEventSfTime/最新事件时间"
    LatestEventSfSite = "LatestEventSfSite/最新事件地点"
    SfDateInterval = "SfDateInterval/SF消息间隔"
    TrackTimeInterval = "TrackTimeInterval/跟踪时间间隔"
    TrackTimeIntervalState = "TrackTimeIntervalState/跟踪时间间隔状态"
    LastEventSfTime = "LastEventSfTime/上一条轨迹时间"
    Tacking_Time = "Tacking_Time/追踪时间"
    UnpaidDate = "UnpaidDate/unpaid记录时间"
    Recipient = "Recipient/收件人"
    Upload_Shipping_Label = "上传物流面单(Upload_Shipping_Label)"
    YD_Number = 'YD_Number/阳单号'
    YD_State = 'YD_State/阳单轨迹状态'

    Courier_File1 = 'Courier/快递_file1'
    SfDateInterval_File1 = 'SfDateInterval/SF消息间隔_file1'
    PossessionSfDate_File1 = 'PossessionSfDate/揽收时间_file1'
    LatestEventSfDate_File1 = 'LatestEventSfDate/最新事件时间_file1'
    Package1_Tracking = 'Package 1\nTracking No./物流跟踪号1'

    Order_Num = "订单号"
    Track_Num = "运单号"
    Platform_Num = "Platform Number/平台单号"
    Yang_Num = "阳单_运单号"
    Yin_Num = "阴单_运单号"
    Yang_Track_State = "阳单_物流状态"
    Yin_Track_State = "阴单_物流状态"
    Yang_Delivered_Time = "阳单_签收时间"
    Yin_Delivered_Time = "阴单_签收时间"
    YY_Delivered_Time = "阴阳单_签收间隔"
    Create_Time = "创建时间"

    Length = 'Length/长'
    Width = 'Width/宽'
    Height = 'Height/高'
    Unit = 'Unit/单位'

    Pay_Time = "付款时间"
    Ship_Time = "发货时间"
    Track_State = "轨迹状态"
    Analyse_State = "追踪时间"
    Last_Track_Time = "上一条轨迹时间"
    Latest_Track_Site = "最新轨迹位置"
    Latest_Track_Time = "最新轨迹时间"
    Interval_Time = "时间间隔"
    Process_Time = "处理状态"
    YD_Number2 = "阳单号"
    YD_State2 = "阳单轨迹状态"

    Total_Of_Product = "产品总数"
    Store_Account = "店铺账号"
    Store_Sales = "店铺销量"
    Store_Name = "店铺名称"
    Order_Sales = "订单数量"
    Sales_Update = "销量更新"
    Quantity = "数量"
    Inventory_Update = "库存更新"
    Sea_transportation = "海运在途"
    Air_transportation = "空运在途"
    Available_Quantity = "可用量"
    Available_Inventory = "Available Inventory/可用库存"


@dataclass(frozen=True)
class CourierStateMapKey:
    tracking_map = 'tracking_map'
    irregular_number_map = 'irregular_number_map'
    no_tracking_map = 'no_tracking_map'
    unpaid_map = "unpaid_map"
    not_yet_map = "not_yet_map"
    pre_ship_map = "pre_ship_map"
    delivered_map = "delivered_map"
    alert_map = "alert_map"
    possession_sf_date_map = "possession_sf_date_map"
    latest_event_sf_date_map = "latest_event_sf_date_map"
    sf_date_equality_map = "sf_date_equality_map"
    delivered_time_map = "delivered_time_map"
    latest_event_sf_time_map = "latest_event_sf_time_map"
    latest_event_sf_site_map = "latest_event_sf_site_map"


class CourierStateMapValue:
    irregular_no_tracking = 'irregular_no_tracking'
    not_yet = 'not_yet'
    pre_ship = "pre_ship"
    no_tracking = "no_tracking"  # 其它无轨迹
    no_track = "no_track"
    unpaid = "unpaid"
    delivered = "delivered"
    tracking = "tracking"
    acceptance_pending = "acceptance_pending"
    alert = "alert"


@dataclass(frozen=True)
class CellKey:
    Outbound_Time = "Outbound_Time"
    update_time = "update_time"
    warehouse_condition = "warehouse_condition"
    store_condition = "store_condition"
    sku_condition = "sku_condition"
    time_segment_condition = "time_segment_condition"
    sum_up = "sum_up"
    exception = "exception"
    shipping_service_condition = "shipping_service_condition"
    special_information = "special_information"
    wl = "wl"
    unpaid = "unpaid"


@dataclass(frozen=True)
class Pattern:
    no_track = r"not_yet|pre_ship|irregular_no_tracking|no_tracking"
    delivered = r"^delivered$"
    unpaid = r"^unpaid$"
    not_yet = r"^not_yet$"
    irregular_no_tracking = r"^irregular_no_tracking$"
    pre_ship = r"^pre_ship$"
    no_tracking = r"^no_tracking$"
    tracking = r"^tracking$"
    alert = r"^alert"


def find_irregular_tracking_numbers(filepath, column_name=RowName.Tracking_No):
    """
    查找不规则的快递单号（不是纯数字、不是9开头、或者包含下划线）
    :param filepath: Excel文件路径
    :param column_name: 需要检查的列名，默认为 'Tracking No./物流跟踪号'
    :return: 不规则快递单号字典
    """
    try:
        # 打开xlsx文件
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active  # 默认使用活动工作表

        # 获取 'Tracking No./物流跟踪号' 列索引
        tracking_no_col = next(
            (col for col in range(1, sheet.max_column + 1) if sheet.cell(row=1, column=col).value == column_name),
            None
        )

        if tracking_no_col is None:
            print(f"找不到 {column_name} 列")
            return {}

        # 存储不规则快递单号的字典
        irregular_number_map = {}

        # 遍历所有行，从第二行开始（跳过表头）
        for row in range(2, sheet.max_row + 1):
            tracking_no = str(sheet.cell(row=row, column=tracking_no_col).value).strip()  # 转换为字符串并去除前后空格
            # 判断是否为不合规单号
            if not tracking_no.isdigit() or not tracking_no.startswith(("92", "93", "94")) or "_" in tracking_no:
                irregular_number_map[tracking_no] = CourierStateMapValue.irregular_no_tracking

        # print("find_irregular_tracking_numbers 方法执行完成")
        return irregular_number_map

    except Exception as e:
        print(f"发生错误: {e}")
        return {}


def update_courier_status1(filepath, maps, wl=RowName.Tracking_No):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active  # 默认使用活动工作表

    data = pd.read_excel(filepath, dtype=str)
    # 获取 'Tracking No./物流跟踪号' 列和 'Courier/快递' 列的索引
    tracking_no_col = data.columns.get_loc(wl) + 1  # openpyxl索引从1开始
    courier_col = data.columns.get_loc(RowName.Courier) + 1  # openpyxl索引从1开始

    for tracking_no, status in maps.items():
        for row in range(2, sheet.max_row + 1):  # 从第二行开始（跳过表头）
            # 获取当前行的物流跟踪号
            current_tracking_no = sheet.cell(row=row, column=tracking_no_col).value
            # **处理 NoneType，转换为字符串 "None"**
            if current_tracking_no is None:
                current_tracking_no = "None"
            # 如果找到匹配的物流跟踪号，更新 Courier/快递 列
            if current_tracking_no == tracking_no:
                sheet.cell(row=row, column=courier_col, value=status)

    # 保存更新后的文件
    wb.save(filepath)


def extract_signature_receipt_time(text):
    # text = "Your item was delivered at the front door or porch at 11:03 am on May 2, 2025 in SANTA ANA, CA 92706."

    # 提取时间和日期部分
    match = re.search(r'at (\d{1,2}:\d{2} (?:am|pm)) on ([A-Za-z]+ \d{1,2}, \d{4})', text)
    if match:
        time_str = match.group(1)  # e.g. '11:03 am'
        date_str = match.group(2)  # e.g. 'May 2, 2025'

        # 合并并解析成 datetime 对象
        full_str = f"{date_str} {time_str}"
        dt = datetime.strptime(full_str, "%B %d, %Y %I:%M %p")

        # 转换为 24 小时格式
        formatted = dt.strftime("%Y-%m-%d %H:%M:%S")
        return formatted
    else:
        return ""


def extract_tracking_site_and_time11111(text, old_date: str = "", old_time: str = "", new_date: str = ""):
    """
    提取轨迹中的时间（24小时格式）和地址信息，并判断是否比已有时间更新。
    只有当 (new_date + 提取时间) > (old_date + old_time) 才返回新时间，否则返回 ""。

    :param text: 轨迹文本
    :param old_date: 原始记录的日期（YYYY-MM-DD）
    :param old_time: 原始记录的时间（HH:MM）
    :param new_date: 当前最新记录日期（YYYY-MM-DD）
    :return: (24小时格式时间字符串, 位置字符串)
    """
    time_24h = ""

    # 提取时间
    time_match = re.search(r'\bat\s+([\d:]+\s*[apAP][mM])', text)
    candidate_time = None
    if time_match:
        time_str = time_match.group(1).strip()
        try:
            time_obj = datetime.strptime(time_str, "%I:%M %p")
            candidate_time = time_obj.strftime("%H:%M")
        except ValueError:
            pass  # 保持 candidate_time 为 None

    # 判断是否要更新 time
    try:
        has_old_date = old_date and old_date.lower() not in ["", "nan"]  # 旧的日期存在且不是空或 nan
        has_old_time = old_time and old_time.lower() not in ["", "nan"]  # has_old_time：旧的时间存在且不是空或 nan
        has_new_date = new_date and new_date.lower() not in ["", "nan"]  # has_new_date：新的时间所对应的日期必须有效
        has_candidate_time = candidate_time is not None  # 是否成功提取出新时间（如“17:41”）

        if has_new_date and (has_old_date or has_old_time):  # 有新日期，并且之前的时间或日期有一个有效，就进入比较逻辑
            old_time_fixed = old_time if has_old_time else "00:00"  # 如果旧时间为空，则默认它为当天凌晨 00:00
            old_dt = datetime.strptime(f"{old_date} {old_time_fixed}", "%Y-%m-%d %H:%M")
            new_time_fixed = candidate_time if has_candidate_time else "00:00"  # 如果没有提取出新时间，就假设是 00:00，这样也能与旧时间比较。
            new_dt = datetime.strptime(f"{new_date} {new_time_fixed}", "%Y-%m-%d %H:%M")

            if new_dt >= old_dt:  # 只有当新时间整体比旧时间“更晚”才更新。
                # ✅ 如果有新时间，就用新时间（如 17:41）; ❌ 如果没有提取出时间（例如 “in transit…”），返回 ""
                time_24h = candidate_time if has_candidate_time else ""
            else:
                # 这个逻辑基本不会执行到
                time_24h = ""
        else:
            # 没有 old_time/old_date，可接受新时间
            time_24h = candidate_time if has_candidate_time else ""
    except Exception:
        time_24h = candidate_time if candidate_time else ""

    # 提取地点
    moving = "transit to the next facility"
    # 尝试匹配 "in xxx." 这种结构（例如 in NORCO, CA 92860.）
    in_match = re.search(r'\bin\s+([A-Z0-9 ,\-]+)[\.\n]?', text)
    # 尝试匹配 “our ... on” 这种结构
    our_on_match = re.search(r'\bour\s+(.*?)\s+on\b', text)

    location = ""
    if moving in text.lower():
        location = moving
    elif in_match:
        location = in_match.group(1).strip()
    elif our_on_match:
        location = our_on_match.group(1).strip()

    return time_24h, location


def extract_and_process_data(filepath: str, column_name: str, group_size: int, wl_name=RowName.Tracking_No,
                             request_interval: float = 30.0, ckjs_flag=False, dxm_xyl_yd_flag=False):
    data = pd.read_excel(filepath, dtype=str)

    if column_name not in data.columns:
        raise ValueError(f"列 '{column_name}' 不存在于 Excel 文件中")

    # 存储结果的 map（字典）
    results_map = {
        CourierStateMapKey.tracking_map: {},
        CourierStateMapKey.no_tracking_map: {},
        CourierStateMapKey.unpaid_map: {},
        CourierStateMapKey.not_yet_map: {},
        CourierStateMapKey.pre_ship_map: {},
        CourierStateMapKey.delivered_map: {},
        CourierStateMapKey.possession_sf_date_map: {},
        CourierStateMapKey.latest_event_sf_date_map: {},
        CourierStateMapKey.sf_date_equality_map: {},
        CourierStateMapKey.delivered_time_map: {},
        CourierStateMapKey.alert_map: {},
        CourierStateMapKey.latest_event_sf_time_map: {},
        CourierStateMapKey.latest_event_sf_site_map: {},
    }

    # 将无内容的单元格赋值""空字符串
    data[column_name] = data[column_name].fillna('')

    # 获取指定内容的数据
    if ckjs_flag:
        filtered_data = data[data[column_name].apply(
            lambda x: str(x).strip().lower() in [''])]
    else:
        filtered_data = data[data[column_name].apply(
            lambda x: str(x).strip().lower() in ['', CourierStateMapValue.not_yet,
                                                 CourierStateMapValue.pre_ship,
                                                 CourierStateMapValue.tracking,
                                                 CourierStateMapValue.no_tracking,
                                                 CourierStateMapValue.alert
                                                 ])]
    # 提取符合条件的 'Tracking No./物流跟踪号' 列数据，并剔除wl_name列中 NaN，去除字符串前后空格，排除只含空格或本身为空的字符串
    items = [x.strip() for x in filtered_data[wl_name].dropna().astype(str) if x.strip() != '']

    # 按组划分数据
    grouped_items = [items[i:i + group_size] for i in range(0, len(items), group_size)]

    # 顺序执行每组数据的处理
    for group in grouped_items:
        try:
            # 直接调用 track 函数，处理每一组数据
            track1 = track(group)
            print(f"处理第 {grouped_items.index(group) + 1} 组，共 {len(group)} 条数据")

            # 对返回的数据进行处理
            for package_id, info in track1['data'].items():
                # 判断错误类型并分类
                if info.get('err'):
                    if info.get('err_id') == '-2147219283':  # 无轨迹(Label Created, not yet in system)
                        results_map[CourierStateMapKey.not_yet_map][package_id] = CourierStateMapValue.not_yet
                    elif info.get('err_id') == 'pre-ship':  # 无轨迹(pre-ship)
                        results_map[CourierStateMapKey.pre_ship_map][package_id] = CourierStateMapValue.pre_ship
                    elif info.get('err_id') == '-2147219278':  # 无轨迹(unpaid)
                        # Tracking information will not be displayed because this package was shipped with counterfeit postage. Please contact the merchant or seller with any questions.
                        results_map[CourierStateMapKey.unpaid_map][package_id] = CourierStateMapValue.unpaid
                    else:
                        results_map[CourierStateMapKey.no_tracking_map][package_id] = CourierStateMapValue.no_tracking

                    results_map[CourierStateMapKey.possession_sf_date_map][package_id] = ""
                    results_map[CourierStateMapKey.latest_event_sf_date_map][package_id] = ""
                    results_map[CourierStateMapKey.sf_date_equality_map][package_id] = 0
                    results_map[CourierStateMapKey.latest_event_sf_time_map][package_id] = ""
                    results_map[CourierStateMapKey.latest_event_sf_site_map][package_id] = ""
                else:
                    statusLong = info.get('statusLong')
                    statusCategory = info.get('statusCategory')
                    latestEventSfDateTimeStr = info.get("latestEventSfDateTime")
                    possessionSfDateTimeStr = info.get("possessionSfDateTime")
                    statusShortStr = info.get("statusShort")

                    latestEventMatch = re.search(r"\d{4}-\d{2}-\d{2}", str(latestEventSfDateTimeStr))
                    possessionEventMatch = re.search(r"\d{4}-\d{2}-\d{2}", str(possessionSfDateTimeStr))

                    latestEventSfDateTimeGroup = latestEventMatch.group() if latestEventMatch else ""
                    possessionSfDateTimeGroup = possessionEventMatch.group() if possessionEventMatch else ""

                    if (latestEventSfDateTimeGroup == ""):
                        date1 = ""
                    else:
                        date1 = datetime.strptime(latestEventSfDateTimeGroup, "%Y-%m-%d")

                    if (possessionSfDateTimeGroup == ""):
                        date2 = ""
                    else:
                        date2 = datetime.strptime(possessionSfDateTimeGroup, "%Y-%m-%d")

                    if (latestEventSfDateTimeGroup == "" or possessionSfDateTimeGroup == ""):
                        days_diff = 0
                    else:
                        days_diff = (date1 - date2).days

                    # 表示不是
                    if (days_diff == 0 and statusShortStr != 'Shipment Received, Package Acceptance Pending'):
                        days_diff = 99

                    results_map[CourierStateMapKey.possession_sf_date_map][package_id] = possessionSfDateTimeGroup
                    results_map[CourierStateMapKey.latest_event_sf_date_map][package_id] = latestEventSfDateTimeGroup
                    results_map[CourierStateMapKey.sf_date_equality_map][package_id] = int(days_diff)
                    results_map[CourierStateMapKey.latest_event_sf_time_map][package_id] = ""
                    results_map[CourierStateMapKey.latest_event_sf_site_map][package_id] = ""

                    if "The package associated with this tracking number did not have proper postage applied and will not be delivered" in \
                            statusLong:
                        results_map[CourierStateMapKey.unpaid_map][package_id] = CourierStateMapValue.unpaid
                    elif "Delivered" in statusCategory or "Delivered to Agent" in statusCategory:
                        results_map[CourierStateMapKey.delivered_map][package_id] = CourierStateMapValue.delivered
                        receipt_time = extract_signature_receipt_time(statusLong)
                        results_map[CourierStateMapKey.delivered_time_map][package_id] = receipt_time
                    elif "Alert" in statusCategory:
                        results_map[CourierStateMapKey.alert_map][package_id] = CourierStateMapValue.alert

                        if dxm_xyl_yd_flag:
                            # 提取已有时间以供比较
                            existing_date = data.loc[data[wl_name] == package_id, RowName.LatestEventSfDate].values[
                                0] \
                                if RowName.LatestEventSfDate in data.columns else ""
                            existing_time = data.loc[data[wl_name] == package_id, RowName.LatestEventSfTime].values[
                                0] \
                                if RowName.LatestEventSfTime in data.columns else ""

                            time_24h, site_location = extract_tracking_site_and_time11111(statusLong,
                                                                                          str(existing_date),
                                                                                          str(existing_time),
                                                                                          latestEventSfDateTimeGroup)

                            results_map[CourierStateMapKey.latest_event_sf_time_map][package_id] = time_24h
                            results_map[CourierStateMapKey.latest_event_sf_site_map][package_id] = site_location
                    else:
                        results_map[CourierStateMapKey.tracking_map][package_id] = CourierStateMapValue.tracking

                        if dxm_xyl_yd_flag:
                            # 提取已有时间以供比较
                            existing_date = data.loc[data[wl_name] == package_id, RowName.LatestEventSfDate].values[
                                0] \
                                if RowName.LatestEventSfDate in data.columns else ""
                            existing_time = data.loc[data[wl_name] == package_id, RowName.LatestEventSfTime].values[
                                0] \
                                if RowName.LatestEventSfTime in data.columns else ""

                            time_24h, site_location = extract_tracking_site_and_time11111(statusLong,
                                                                                          str(existing_date),
                                                                                          str(existing_time),
                                                                                          latestEventSfDateTimeGroup)

                            results_map[CourierStateMapKey.latest_event_sf_time_map][package_id] = time_24h
                            results_map[CourierStateMapKey.latest_event_sf_site_map][package_id] = site_location

            # 等待指定的间隔时间，避免请求频率过高
            time.sleep(request_interval)

        except Exception as e:
            print(f"处理组 {grouped_items.index(group) + 1} 时发生错误: {e}")

    return results_map


def update_courier_status(filepath, maps_list, wl=RowName.Tracking_No, column_map=None):
    """
    批量更新多个状态，支持在更新"最新事件时间"前将原有轨迹时间备份到"上一条轨迹时间"
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    tracking_no_col = headers.index(wl) + 1

    column_indices = {key: headers.index(col_name) + 1 for key, col_name in column_map.items()}

    # 获取可能涉及的附加列索引
    try:
        unpaid_date_col_index = headers.index(RowName.UnpaidDate) + 1
    except ValueError:
        unpaid_date_col_index = None

    # 如果需要更新最新事件时间，确保相关列存在
    latest_date_col = RowName.LatestEventSfDate
    latest_time_col = RowName.LatestEventSfTime
    previous_event_time_col = RowName.LastEventSfTime

    try:
        latest_date_col_index = headers.index(latest_date_col) + 1
        latest_time_col_index = headers.index(latest_time_col) + 1
        previous_event_col_index = headers.index(previous_event_time_col) + 1
    except ValueError as ve:
        # 如果 column_map 包含 latest_date_col，但目标列不存在 → 提醒
        if latest_date_col in column_map.values():
            raise ValueError(f"需要更新 {latest_date_col}，但文件中缺失相关列: {ve}")

    strftime = datetime.now().strftime("%Y-%m-%d")

    # 提取所有 tracking_no 行号
    tracking_no_rows = {}
    for row in range(2, sheet.max_row + 1):
        tracking_no = sheet.cell(row=row, column=tracking_no_col).value
        if tracking_no:
            tracking_no_rows.setdefault(tracking_no, []).append(row)

    for state_map, col_name in column_map.items():
        col_index = column_indices[state_map]

        for tracking_no, status in maps_list.get(state_map, {}).items():
            if tracking_no in tracking_no_rows:
                for row_index in tracking_no_rows[tracking_no]:

                    # 特殊处理：若更新的是“LatestEventSfDate/最新事件时间”，先保存旧时间
                    if col_name == latest_date_col:
                        old_date = sheet.cell(row=row_index, column=latest_date_col_index).value
                        old_time = sheet.cell(row=row_index, column=latest_time_col_index).value

                        if old_date or old_time:
                            # 统一转为字符串并去除空白，过滤 nan
                            date_str = str(old_date).strip() if old_date not in [None, "nan", "NaT"] else ""
                            time_str = str(old_time).strip() if old_time not in [None, "nan", "NaT"] else ""
                            if date_str:
                                if time_str:
                                    combined = f"{date_str} {time_str}"
                                else:
                                    combined = f"{date_str} 00:00"
                                sheet.cell(row=row_index, column=previous_event_col_index, value=combined)

                    # 正常写入新状态
                    sheet.cell(row=row_index, column=col_index, value=status)

                    # 若是 unpaid_map 且 unpaid_date 列存在 → 写入日期
                    if state_map == CourierStateMapKey.unpaid_map and unpaid_date_col_index:
                        current_value = sheet.cell(row=row_index, column=unpaid_date_col_index).value
                        if current_value in [None, "", " "]:
                            sheet.cell(row=row_index, column=unpaid_date_col_index, value=strftime)

    wb.save(filepath)
    print("✅ update_courier_status 方法执行完成")


def count_pattern_state(file_path, column_name, patternStr):
    """
    统计指定列指定内容的数量
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        column_index = headers.index(column_name) + 1
        pattern = re.compile(patternStr, re.IGNORECASE)
        total_count = 0
        no_track_count = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            cell_value = row[column_index - 1]
            if cell_value is not None:
                total_count += 1
                if pattern.search(str(cell_value)):
                    no_track_count += 1
        return total_count, no_track_count
    except Exception as e:
        print(f"发生错误: {e}")
        return 0, 0


def count_pattern_and_tracking_with_sf_date(file_path, column_name, sfDateInterval_name, patterns):
    """
    统计指定列中多个正则表达式匹配内容的数量，并统计 "Courier/快递" 列内容为 'tracking' 且 "SfDateEquality" 列的内容为 0 的行数。

    :param file_path: Excel 文件路径
    :param column_name: 需要统计的列名
    :param patterns: 正则表达式字典，key 为模式名称，value 为模式字符串
    :return: 包含每个模式匹配计数和符合 'tracking' 且 'SfDateEquality' 为 0 的行数的字典
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # 检查列名是否存在
        if column_name not in headers or sfDateInterval_name not in headers:
            raise ValueError(f"Excel 文件中未找到 '{column_name}' 或 '{sfDateInterval_name}'列 找不到")

        # 获取列索引
        column_index = headers.index(column_name)
        sfDateInterval_index = headers.index(sfDateInterval_name)

        # 编译所有正则表达式
        compiled_patterns = {key: re.compile(pattern, re.IGNORECASE) for key, pattern in patterns.items()}

        # 初始化计数器
        count_dict = {**{key: 0 for key in patterns}, "sfDateInterval": 0}  # 显式初始化
        tracking_sf_date_equality_count = 0

        # 遍历每一行数据
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            courier_value = row[column_index]
            sfDateInterval_value = row[sfDateInterval_index]

            # 判断是否符合 'tracking' 且 'SfDateEquality' 为 0 或者 'acceptance_pending' 的条件
            if (courier_value == CourierStateMapValue.tracking and sfDateInterval_value == 0) or (
                    courier_value == CourierStateMapValue.acceptance_pending):
                tracking_sf_date_equality_count += 1

            # 判断正则表达式匹配
            if courier_value is not None:
                cell_value_str = str(courier_value)
                for key, pattern in compiled_patterns.items():
                    if pattern.search(cell_value_str):
                        count_dict[key] += 1

        # 将 'tracking' 且 'SfDateEquality' 为 0 的计数添加到字典中
        count_dict["sfDateInterval"] = tracking_sf_date_equality_count

        return count_dict

    except Exception as e:
        print(f"发生错误: {e}")
        # 返回一个包含所有模式及 'tracking_sf_date_equality' 键的字典
        return {key: 0 for key in patterns} | {"sfDateInterval": 0}  # 合并字典


def count_distribution_and_no_track(file_path, key_column, courier_column=RowName.Courier):
    """
    通用函数，统计指定列的分布情况及其对应 "无轨迹" 的数量。
    :param file_path: Excel 文件路径
    :param key_column: 需要统计的列名
    :param courier_column: 快递列名
    :return: 各值的总数和 "无轨迹" 数量的 Counter 对象
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if key_column not in headers or courier_column not in headers:
            raise ValueError(f"列名 '{key_column}' 或 '{courier_column}' 不存在！")
        key_index = headers.index(key_column) + 1
        courier_index = headers.index(courier_column) + 1
        pattern = re.compile(Pattern.no_track, re.IGNORECASE)
        key_counter = Counter()
        key_no_track_counter = Counter()
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            key_value = row[key_index - 1]
            courier_status = row[courier_index - 1]
            if key_value is not None:
                key_counter[key_value] += 1
                if courier_status is not None and pattern.search(str(courier_status)):
                    key_no_track_counter[key_value] += 1
        return key_counter, key_no_track_counter
    except Exception as e:
        print(f"发生错误: {e}")
        return Counter(), Counter()


def dimension_distribution(file_path, key_column, courier_column=RowName.Courier):
    """
    通用函数，统计指定列的分布情况及其对应 "无轨迹" 的数量。
    :param file_path: Excel 文件路径
    :param key_column: 需要统计的列名
    :param courier_column: 快递列名
    :return: 各值的总数和 "无轨迹" 数量的 Counter 对象（已排序）
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if key_column not in headers or courier_column not in headers:
            raise ValueError(f"列名 '{key_column}' 或 '{courier_column}' 不存在！")

        key_index = headers.index(key_column) + 1
        courier_index = headers.index(courier_column) + 1

        key_counter = Counter()
        key_no_track_counter = Counter()

        pattern = re.compile(Pattern.no_track, re.IGNORECASE)

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            key_value = row[key_index - 1]
            courier_status = row[courier_index - 1]

            if key_value is not None:
                key_counter[key_value] += 1
                if courier_status is not None and pattern.search(str(courier_status)):
                    key_no_track_counter[key_value] += 1

        # 排序逻辑
        if key_column == RowName.Warehouse:
            priority_order = ["美西", "美中", "美东"]

            def get_priority(k):
                for idx, keyword in enumerate(priority_order):
                    if keyword in str(k):
                        return idx
                return len(priority_order)

            key_counter = OrderedDict(sorted(key_counter.items(), key=lambda x: get_priority(x[0])))
            key_no_track_counter = OrderedDict(sorted(key_no_track_counter.items(), key=lambda x: get_priority(x[0])))

        elif key_column == RowName.Client:
            # 定义区域优先级
            region_priority = {
                "东谷": 1,
                "美西": 1,
                "美中": 2,
                "休斯顿": 2,
                "费城": 3,
                "美东": 3
            }

            def extract_region_and_suffix(k):
                # 提取区域
                region = ""
                for r in region_priority.keys():
                    if r in k:
                        region = r
                        break
                # 提取括号内的后缀，并忽略大小写
                match = re.search(r"\((?:[^()]*)([A-Za-z\-]+)\)", k)
                suffix = match.group(1).lower() if match else ""
                return region, suffix

            # 分组
            region_groups = {region: [] for region in region_priority.keys()}
            others = []

            for k in key_counter:
                region, suffix = extract_region_and_suffix(k)
                if region in region_groups:
                    region_groups[region].append((k, suffix))
                else:
                    others.append((k, ""))

            # 处理排序：按区域优先级，然后按后缀排序
            region_sorted_keys = []

            # 遍历区域优先级，进行排序
            for region in sorted(region_priority, key=lambda r: region_priority[r]):
                # 对同一区域内的条目按后缀进行排序
                sorted_by_suffix = sorted(region_groups[region], key=lambda x: x[1])
                region_sorted_keys.extend([k for k, _ in sorted_by_suffix])

            # 添加无法识别地区的
            region_sorted_keys.extend([k for k, _ in others])

            # 更新key_counter和key_no_track_counter
            key_counter = OrderedDict((k, key_counter[k]) for k in region_sorted_keys if k in key_counter)
            key_no_track_counter = OrderedDict(
                (k, key_no_track_counter[k]) for k in region_sorted_keys if k in key_no_track_counter)

        return key_counter, key_no_track_counter

    except Exception as e:
        print(f"发生错误: {e}")
        return Counter(), Counter()


def count_distribution_and_no_track3(file_path, key_column, courier_column=RowName.Courier,
                                     sf_date_column=RowName.SfDateInterval):
    """
    统计 key_column 分布及对应快递状态的数量（无轨迹、已送达、未支付、tracking 且 SF 为 0），并按自然顺序排序结果。
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        if key_column not in headers or courier_column not in headers or sf_date_column not in headers:
            raise ValueError(f"列名 '{key_column}'、'{courier_column}' 或 '{sf_date_column}' 不存在！")

        key_index = headers.index(key_column)
        courier_index = headers.index(courier_column)
        sf_date_index = headers.index(sf_date_column)

        pattern_no_track = re.compile(Pattern.no_track, re.IGNORECASE)
        pattern_no_tracking = re.compile(Pattern.no_tracking, re.IGNORECASE)
        pattern_pre_ship = re.compile(Pattern.pre_ship, re.IGNORECASE)
        pattern_not_yet = re.compile(Pattern.not_yet, re.IGNORECASE)
        pattern_delivered = re.compile(Pattern.delivered, re.IGNORECASE)
        pattern_unpaid = re.compile(Pattern.unpaid, re.IGNORECASE)

        key_counter = Counter()
        key_no_track_counter = Counter()
        key_no_tracking_counter = Counter()
        key_pre_ship_counter = Counter()
        key_not_yet_counter = Counter()
        key_delivered_counter = Counter()
        key_unpaid_counter = Counter()
        key_tracking_sf_zero_counter = Counter()

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            key_value = row[key_index]
            courier_status = row[courier_index]
            sf_date_value = row[sf_date_index]

            if key_value is not None:
                key_str = str(key_value)
                key_counter[key_str] += 1

                if courier_status is not None:
                    courier_str = str(courier_status)

                    if pattern_no_track.search(courier_str):
                        key_no_track_counter[key_str] += 1
                    if pattern_no_tracking.search(courier_str):
                        key_no_tracking_counter[key_str] += 1
                    if pattern_pre_ship.search(courier_str):
                        key_pre_ship_counter[key_str] += 1
                    if pattern_not_yet.search(courier_str):
                        key_not_yet_counter[key_str] += 1
                    if pattern_delivered.search(courier_str):
                        key_delivered_counter[key_str] += 1
                    if pattern_unpaid.search(courier_str):
                        key_unpaid_counter[key_str] += 1
                    if courier_str.lower() == CourierStateMapValue.tracking and sf_date_value == 0:
                        key_tracking_sf_zero_counter[key_str] += 1

        # 使用自然排序对所有 Counter 进行排序并转为 OrderedDict
        def sort_counter(counter):
            return OrderedDict((k, counter[k]) for k in natsorted(counter.keys()))

        return (
            sort_counter(key_counter),
            sort_counter(key_no_track_counter),
            sort_counter(key_no_tracking_counter),
            sort_counter(key_pre_ship_counter),
            sort_counter(key_not_yet_counter),
            sort_counter(key_delivered_counter),
            sort_counter(key_unpaid_counter),
            sort_counter(key_tracking_sf_zero_counter),
        )

    except Exception as e:
        print(f"发生错误: {e}")
        return Counter(), Counter(), Counter(), Counter(), Counter(), Counter(), Counter(), Counter()


def analyze_time_segments(file_path, data_map, time_column, courier_column, sf_date_column):
    """
    统计按 3 分钟为间隔的时间段数据，计算每个时间段的总数、无轨迹数，并找到最低上网率的时间段。
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if time_column not in headers or courier_column not in headers or sf_date_column not in headers:
            raise ValueError(f"列名 '{time_column}'、'{courier_column}' 或 '{sf_date_column}' 不存在！")

        # 获取列索引
        time_index, courier_index, sf_date_index = (
            headers.index(time_column), headers.index(courier_column), headers.index(sf_date_column)
        )

        pattern_no_track = re.compile(r"(no_tracking|pre_ship|not_yet)", re.IGNORECASE)

        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            order_time, courier_status, sf_date_value = row[time_index], row[courier_index], row[sf_date_index]
            if isinstance(order_time, str):
                try:
                    order_time = datetime.strptime(order_time, "%Y-%m-%d %H:%M:%S").replace(second=0)
                    data.append((order_time, courier_status, sf_date_value))
                except ValueError:
                    continue

        data.sort(key=lambda x: x[0])  # 按时间排序
        time_segments = defaultdict(list)

        if data:
            base_time = data[0][0]  # 作为时间段起点
            for order_time, courier_status, sf_date_value in data:
                if (order_time - base_time).total_seconds() <= 180:
                    time_segments[base_time].append((order_time, courier_status, sf_date_value))
                else:
                    base_time = order_time
                    time_segments[base_time].append((order_time, courier_status, sf_date_value))

        segment_statistics = {}
        lowest_segment, lowest_swl = "", 101  # 记录最低上网率的时间段

        for segment_start, entries in time_segments.items():
            total_count = len(entries)
            no_track_count = sum(
                1 for _, courier_status, sf_date_value in entries
                # if pattern_no_track.match(str(courier_status or "")) or (courier_status == "tracking" and sf_date_value == 0)
                if pattern_no_track.match(str(courier_status or ""))
            )
            segmentswl = round(100 - ((no_track_count / total_count) * 100), 2) if total_count else 0

            # segment_info = f"\n{segment_start.strftime('%y-%m-%d %H:%M')} - {(segment_start + timedelta(minutes=3)).strftime('%y-%m-%d %H:%M')}：（{total_count}, {no_track_count}, {segmentswl}%）"
            segment_info = f"\n{segment_start.strftime('%y-%m-%d %H:%M')} - {(segment_start + timedelta(minutes=3)).strftime('%y-%m-%d %H:%M')}：（{total_count}, {segmentswl}%）"
            segment_statistics[segment_start] = {
                "total_count": total_count,
                "no_track_count": no_track_count,
                "swl": segmentswl,
                "segment_info": segment_info
            }

            if segmentswl < lowest_swl:
                lowest_swl, lowest_segment = segmentswl, segment_info

        time_segment_text = "".join(
            stats["segment_info"] for stats in segment_statistics.values()
        )

        data_map[CellKey.time_segment_condition] = time_segment_text

        return time_segment_text, lowest_segment

    except Exception as e:
        print(f"发生错误: {e}")
        return "", ""


def process_tracking_no(file_path: str, row_name=RowName.Tracking_No):
    # 读取 Excel 文件
    data = pd.read_excel(file_path, dtype=str)  # 将数据全部读取为字符串类型

    # 确保 'Tracking No./物流跟踪号' 列存在
    if row_name not in data.columns:
        raise ValueError(f"文件中缺少 '{row_name}' 列")

    # 处理 'Tracking No./物流跟踪号' 列：去除空格并确保每个值是字符串，兼容 None 或空值
    data[row_name] = data[row_name].apply(
        lambda x: str(x).replace(" ", "") if x is not None and pd.notna(x) else "")

    # 保存处理后的数据
    data.to_excel(file_path, index=False, engine='openpyxl')

    return data


def check_and_add_courier_column(file_path):
    try:
        # 加载 Excel 文件
        data = pd.read_excel(file_path, dtype=str, engine='openpyxl')
        flag = False
        # 判断是否存在 '快递' 列
        if RowName.Courier not in data.columns:
            # 如果没有 '快递' 列，则在最后一列添加该列
            data[RowName.Courier] = ""  # 默认为空值，可以根据需求填充其他默认值
            flag = True
        if RowName.PossessionSfDate not in data.columns:
            data[RowName.PossessionSfDate] = ""
            flag = True
        if RowName.LatestEventSfDate not in data.columns:
            data[RowName.LatestEventSfDate] = ""
            flag = True
        if RowName.SfDateInterval not in data.columns:
            data[RowName.SfDateInterval] = ""
            flag = True
        if RowName.UnpaidDate not in data.columns:
            data[RowName.UnpaidDate] = ""
            flag = True
        if RowName.LatestEventSfTime not in data.columns:
            data[RowName.LatestEventSfTime] = ""
            flag = True
        if RowName.LatestEventSfSite not in data.columns:
            data[RowName.LatestEventSfSite] = ""
            flag = True
        if RowName.TrackTimeInterval not in data.columns:
            data[RowName.TrackTimeInterval] = ""
            flag = True
        if RowName.TrackTimeIntervalState not in data.columns:
            data[RowName.TrackTimeIntervalState] = ""
            flag = True
        if RowName.Tacking_Time not in data.columns:
            data[RowName.Tacking_Time] = ""
            flag = True
        if RowName.LastEventSfTime not in data.columns:
            data[RowName.LastEventSfTime] = ""
            flag = True
        if RowName.YD_Number not in data.columns:
            data[RowName.YD_Number] = ""
            flag = True
        if RowName.YD_State not in data.columns:
            data[RowName.YD_State] = ""
            flag = True
        # 保存修改后的文件
        data.to_excel(file_path, index=False, engine='openpyxl')
        return flag
    except Exception as e:
        print(f"发生错误: {e}")


def get_days_difference(file_path, column_name=RowName.CreationTime):
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active

        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        column_index = headers.index(column_name) + 1

        # 从第二行开始查找有效数据
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=column_index).value
            if not cell_value:
                continue

            # 如果是 datetime 类型，直接处理
            if isinstance(cell_value, datetime):
                return cell_value.strftime("%Y/%m/%d")
            else:
                try:
                    dt = datetime.strptime(str(cell_value), "%Y-%m-%d %H:%M:%S")
                    return dt.strftime("%Y/%m/%d")
                except ValueError:
                    continue  # 格式不符跳过

        raise ValueError("没有找到有效的时间格式数据！")

    except Exception as e:
        print(f"发生错误: {e}")
        return None


def get_unpaid_platform_tracking_map(file_path):
    """
    获取 'Courier/快递' 列为 'unpaid' 的记录，按照 'UnpaidDate/unpaid记录时间' 分组，
    返回每个分组的 platform_number 与其 tracking 信息组成的映射。

    :param file_path: Excel 文件路径
    :return: 分组数据字典：{ 分组时间（None 或 日期）: { platform_number: {tracking_number, kj} } }
    """
    # 读取 Excel 文件
    data = pd.read_excel(file_path)

    required_columns = [
        RowName.Courier, RowName.Platform_Num, RowName.Tracking_No,
        RowName.ShippingService, RowName.Recipient, RowName.UnpaidDate
    ]
    for col in required_columns:
        if col not in data.columns:
            raise ValueError(f"文件中缺少所需列：{col}")

    # 只处理 Courier 为 unpaid 的行
    unpaid_data = data[data[RowName.Courier] == CourierStateMapValue.unpaid].copy()

    # 将 UnpaidDate 列转换为 datetime 类型
    unpaid_data[RowName.UnpaidDate] = pd.to_datetime(
        unpaid_data[RowName.UnpaidDate], errors='coerce'
    )

    # 准备结果字典
    result = {}

    # 处理空日期的记录（NaT）
    null_date_group = unpaid_data[unpaid_data[RowName.UnpaidDate].isnull()]
    if not null_date_group.empty:
        group_dict = {}
        for _, row in null_date_group.iterrows():
            platform_number = row[RowName.Platform_Num]
            tracking_number = row[RowName.Tracking_No]
            shipping_service = row[RowName.ShippingService]
            recipient = row[RowName.Recipient]

            kj_ = (shipping_service == RowName.Upload_Shipping_Label and recipient in ['KJ', 'TK']) or \
                  (shipping_service != RowName.Upload_Shipping_Label)

            group_dict[platform_number] = {
                "tracking_number": tracking_number,
                "kj": kj_
            }
        result[None] = group_dict

    # 处理非空日期分组
    for group_time, group_rows in unpaid_data[unpaid_data[RowName.UnpaidDate].notnull()].groupby(
            RowName.UnpaidDate):
        group_dict = {}
        for _, row in group_rows.iterrows():
            platform_number = row[RowName.Platform_Num]
            tracking_number = row[RowName.Tracking_No]
            shipping_service = row[RowName.ShippingService]
            recipient = row[RowName.Recipient]

            kj_ = (shipping_service == RowName.Upload_Shipping_Label and recipient in ['KJ', 'TK']) or \
                  (shipping_service != RowName.Upload_Shipping_Label)

            group_dict[platform_number] = {
                "tracking_number": tracking_number,
                "kj": kj_
            }
        result[group_time] = group_dict

    return result


def get_shipment_received_numbers(filepath, gz_time):
    # 加载 Excel 文件
    workbook = load_workbook(filename=filepath, data_only=True)

    # 获取第一个工作表
    sheet = workbook.active

    # 读取表头（第一行），并构建列名到索引的映射
    header = [cell.value for cell in sheet[1]]

    # 确保所有必要的列都存在
    required_columns = [RowName.Courier, RowName.SfDateInterval, RowName.OutboundTime, RowName.Tracking_No]

    # 获取每个必要列的索引（列索引从 0 开始）
    column_indices = {}
    for col in required_columns:
        if col not in header:
            missing_cols = set(required_columns) - set(header)
            raise ValueError(f"Excel 文件缺少必要的列: {missing_cols}")
        column_indices[col] = header.index(col)

    # 存储符合条件的跟踪号
    tracking_numbers = []

    # 遍历数据行（从第二行开始）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        courier = row[column_indices[RowName.Courier]]
        sf_date_interval = row[column_indices[RowName.SfDateInterval]]
        outbound_time = row[column_indices[RowName.OutboundTime]]
        tracking_no = row[column_indices[RowName.Tracking_No]]

        # 筛选条件：Courier == 'tracking' 且 SfDateInterval == '0'
        if courier == CourierStateMapValue.tracking and sf_date_interval == 0:

            # try:
            # 解析并计算日期间隔
            parsed_outbound_time = datetime.strptime(outbound_time, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_outbound_time.strftime("%Y/%m/%d")
            interval_days = (datetime.strptime(gz_time, "%Y/%m/%d") -
                             datetime.strptime(formatted_date, "%Y/%m/%d")).days

            is_usweekend = is_us_weekend(formatted_date)
            actual_interval = 0
            if is_usweekend == 6:  # 6是中国周日，美国周六
                actual_interval = 2
            elif is_usweekend == 0:  # 0是中国周一，美国周日
                actual_interval = 1

            # 如果间隔为 2 天，添加到结果列表
            if (interval_days - actual_interval) >= 2 and tracking_no is not None:
                tracking_numbers.append(tracking_no)

        # except (ValueError, TypeError) as e:
        #     日期解析错误或数据错误时跳过
        # print(f"日期解析错误: {e}, 跟踪号: {tracking_no}")
        # continue

    # 返回唯一的跟踪号列表
    return list(set(tracking_numbers))


def get_filtered_count(filepath, gz_time, target_column, target_value):
    # 读取 Excel 文件
    df = pd.read_excel(filepath, dtype=str)

    # 确保需要的列存在
    required_columns = {RowName.Courier, RowName.SfDateInterval, RowName.OutboundTime, target_column}
    if not required_columns.issubset(df.columns):
        raise ValueError(f"Excel 文件缺少必要的列: {required_columns - set(df.columns)}")

    # 过滤指定列 `target_column` 的值为 `target_value`
    filtered_df = df[df[target_column] == target_value]

    # 进一步筛选条件：
    filtered_df = filtered_df[
        (filtered_df[RowName.Courier] == CourierStateMapValue.tracking) &
        (filtered_df[RowName.SfDateInterval] == "0")
        ]

    # 解析 "OutboundTime/出库时间" 并计算时间间隔
    def check_outbound_time(outbound_time):
        try:
            # interval_days = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(outbound_time,
            #                                                                             "%Y-%m-%d %H:%M:%S")).days

            parsed_outbound_time = datetime.strptime(outbound_time, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_outbound_time.strftime("%Y/%m/%d")
            # 跟踪时间 - 出库时间
            interval_days = (datetime.strptime(gz_time, "%Y/%m/%d") -
                             datetime.strptime(formatted_date, "%Y/%m/%d")).days

            is_usweekend = is_us_weekend(formatted_date)
            actual_interval = 0
            if is_usweekend == 6:  # 6是中国周日，美国周六
                actual_interval = 2
            elif is_usweekend == 0:  # 0是中国周一，美国周日
                actual_interval = 1

            return (interval_days - actual_interval) >= 2  # 计算时间间隔是否 >= 2
        except ValueError:
            return False  # 解析失败则跳过

    filtered_df = filtered_df[filtered_df[RowName.OutboundTime].apply(check_outbound_time)]

    # 返回符合条件的数据数量
    return len(filtered_df)


def convert_inch_to_cm(value_in_inch):
    """
    将英寸转换为厘米
    :param value_in_inch: 英寸数
    :return: 转换后的厘米数
    """
    return round2(value_in_inch * 2.54)


def get_in(file_path, sku_to_match):
    # 读取 Excel 文件
    data = pd.read_excel(file_path)

    # 确保相关列存在
    required_columns = [RowName.SKU, RowName.Length, RowName.Width, RowName.Height, RowName.Unit]
    if not all(col in data.columns for col in required_columns):
        raise ValueError(f"文件中缺少所需的列，请检查文件结构")

    # 查找匹配的第一个 SKU
    matched_row = data[data[RowName.SKU] == sku_to_match].iloc[0]  # 获取第一个匹配的行

    # 提取数据
    length = matched_row[RowName.Length]
    width = matched_row[RowName.Width]
    height = matched_row[RowName.Height]
    unit = matched_row[RowName.Unit]

    # 如果单位是英寸，进行转换
    if unit == 'in':
        length = convert_inch_to_cm(length)
        width = convert_inch_to_cm(width)
        height = convert_inch_to_cm(height)
        unit = 'cm'  # 转换后的单位为厘米

    # 返回一个字典，包含转换后的数据
    result = {
        RowName.SKU: sku_to_match,
        'Length': length,
        'Width': width,
        'Height': height,
        'Unit': unit
    }

    return result


def temu_count(file_path, sku_value,
               sku_column=RowName.SKU,
               shipping_service_column=RowName.Platform_Num,
               courier_column=RowName.Courier):
    # 读取 Excel 文件
    data = pd.read_excel(file_path, dtype=str)  # 确保数据按字符串读取，避免 NaN 影响筛选

    # 确保必要的列存在
    required_columns = {sku_column, shipping_service_column, courier_column}
    if not required_columns.issubset(data.columns):
        raise ValueError(f"文件中缺少必要的列: {required_columns - set(data.columns)}")

    # 初步筛选 SKU 列符合条件的数据
    filtered_data = data[(data[sku_column] == sku_value) & (data[shipping_service_column].str.startswith("PO-"))]

    # 编译正则模式
    compiled_pattern = re.compile(Pattern.no_track, re.IGNORECASE)

    # 进一步筛选 'Courier/快递' 列符合 Pattern.no_track 正则的数据
    matched_data = filtered_data[filtered_data[courier_column].str.match(compiled_pattern, na=False)]

    # 返回两个值：（初步筛选的数据数量，进一步匹配 'Courier/快递' 的订单数量）
    return len(filtered_data), len(matched_data)


def sku_kj_count(file_path, sku_value,
                 sku_column=RowName.SKU,
                 shipping_service_column=RowName.ShippingService,
                 recipient_column=RowName.Recipient,
                 courier_column=RowName.Courier):
    """
    计算符合条件的 SKU 订单总数，并返回匹配 'Courier/快递' 正则的数据数量
    :param file_path: Excel 文件路径
    :param sku_value: 要筛选的 SKU 值
    :param sku_column: SKU 列名
    :param shipping_service_column: 物流渠道列名
    :param recipient_column: 收件人列名
    :param courier_column: 快递列名
    :return: (符合 SKU 条件的订单数量, 进一步匹配 'Courier/快递' 的订单数量)
    """
    # 读取 Excel 文件
    data = pd.read_excel(file_path, dtype=str)  # 确保数据按字符串读取，避免 NaN 影响筛选

    # 确保必要的列存在
    required_columns = {sku_column, shipping_service_column, recipient_column, courier_column}
    if not required_columns.issubset(data.columns):
        raise ValueError(f"文件中缺少必要的列: {required_columns - set(data.columns)}")

    # 初步筛选 SKU 列符合条件的数据
    filtered_data = data[
        (data[sku_column] == sku_value) &
        (
                ((data[shipping_service_column] == RowName.Upload_Shipping_Label) &
                 ((data[recipient_column] == 'KJ') | (data[recipient_column] == 'TK'))) |
                (data[shipping_service_column] != RowName.Upload_Shipping_Label)
        )
        ]

    # 编译正则模式
    compiled_pattern = re.compile(Pattern.no_track, re.IGNORECASE)

    # 进一步筛选 'Courier/快递' 列符合 Pattern.no_track 正则的数据
    matched_data = filtered_data[filtered_data[courier_column].str.match(compiled_pattern, na=False)]

    # 返回两个值：（初步筛选的数据数量，进一步匹配 'Courier/快递' 的订单数量）
    return len(filtered_data), len(matched_data)


def kj_count(file_path, shipping_service_column=RowName.ShippingService, recipient_column=RowName.Recipient):
    # 读取 Excel 文件
    data = pd.read_excel(file_path)

    # 确保必要的列存在
    if shipping_service_column not in data.columns or recipient_column not in data.columns:
        raise ValueError(f"文件中缺少必要的列，请检查列名是否正确")

    # 条件1：'Shipping service/物流渠道' 不为 '上传物流面单(Upload_Shipping_Label)' 的行
    condition1 = (data[shipping_service_column] != RowName.Upload_Shipping_Label)

    # 条件2：'Shipping service/物流渠道' 为 '上传物流面单(Upload_Shipping_Label)' 且 'Recipient/收件人' 为 'KJ' 的行
    condition2 = (data[shipping_service_column] == RowName.Upload_Shipping_Label) & (
            (data[recipient_column] == 'KJ') | (data[recipient_column] == 'TK'))

    # 综合筛选符合任一条件的行
    kj_counts = data[condition1 | condition2]

    # 进一步筛选 'Courier/快递' 列符合 Pattern.no_track 正则的数据
    kj_no_track_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.no_track, re.IGNORECASE), na=False)]
    kj_tracking_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.tracking, re.IGNORECASE), na=False)]
    kj_delivered_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.delivered, re.IGNORECASE), na=False)]
    kj_pre_ship_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.pre_ship, re.IGNORECASE), na=False)]
    kj_not_yet_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.not_yet, re.IGNORECASE), na=False)]
    kj_unpaid_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.unpaid, re.IGNORECASE), na=False)]
    kj_irregular_number_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.irregular_no_tracking, re.IGNORECASE), na=False)]
    kj_no_tracking_count = kj_counts[
        kj_counts[RowName.Courier].str.match(re.compile(Pattern.no_tracking, re.IGNORECASE), na=False)]

    # 返回符合条件的行数
    return len(kj_counts), len(kj_no_track_count), len(kj_tracking_count), len(kj_delivered_count), \
        len(kj_pre_ship_count), len(kj_not_yet_count), len(kj_unpaid_count), len(kj_irregular_number_count), \
        len(kj_no_tracking_count)


def generate_distribution_report(xlsx_path, kj_column, gz_time, distribution, no_track_distribution, data_map,
                                 data_map_key):
    """
    通用的分布报告生成函数
    :param distribution: 订单分布字典
    :param no_track_distribution: 无轨迹分布字典
    :param data_map:
    :param data_map_key: 用于存储到 `data_map` 的 key（例如 `CellKey.warehouse_condition` 或 `CellKey.store_condition`）
    :return: 生成的分布报告文本
    """
    report_text = []
    report_text2 = []
    lowest_swl = 101  # 初始化为一个比 100 大的值，用于比较
    lowest_entity = ""  # 保存最低上网率的实体信息

    # 遍历分布数据
    for entity, count in distribution.items():
        no_track_count = no_track_distribution.get(entity, 0)
        swl = round2(100 - ((no_track_count / count) * 100))  # 计算上网率
        kjCount, kj_no_track_count = sku_kj_count(xlsx_path, entity, sku_column=kj_column)
        kjSwl = 0
        if (kjCount > 0):
            kjSwl = round2(100 - ((kj_no_track_count / kjCount) * 100))

        temuCount, temu_no_track_count = temu_count(xlsx_path, entity, sku_column=kj_column)
        temuSwl = 0
        if (temuCount > 0):
            temuSwl = round2(100 - ((temu_no_track_count / temuCount) * 100))

        shipment_received_interval2_count = get_filtered_count(xlsx_path, gz_time, kj_column, entity)

        # 使用 f-string 格式化输出文本
        report_text.append(
            f"\n{entity}：（{count}, {no_track_count}, {swl}%）,（{kjCount}, {kj_no_track_count}, {kjSwl}%）,（{temuCount}, {temu_no_track_count}, {temuSwl}%）,（{shipment_received_interval2_count}）")
        report_text2.append(f"\n{entity}：({count}, {swl}%)")

        # 判断是否是最低的上网率
        if swl < lowest_swl:
            lowest_swl = swl
            lowest_entity = f"\n{entity}：（{count}, {no_track_count}, {swl}%）,（{kjCount}, {kj_no_track_count}, {kjSwl}%）,（{temuCount}, {temu_no_track_count}, {temuSwl}%）,（{shipment_received_interval2_count}）"

    # 将结果存储到 data_map 中
    data_map[data_map_key] = "".join(report_text)  # 使用 join 合并字符串，减少内存消耗
    return "".join(report_text), lowest_entity, "".join(report_text2)


def generate_distribution_report2(distribution, no_track_distribution, sku_no_tracking_distribution,
                                  sku_pre_ship_distribution, sku_not_yet_distribution, sku_delivered_distribution,
                                  sku_unpaid_distribution, tracking_sf_zero_distribution,
                                  data_map, data_map_key, interval_time, xlsx_path, gz_time):
    """
    通用的分布报告生成函数，统计订单分布、无轨迹订单、计算上网率，并找出最低上网率的所有实体
    :param distribution: 订单分布字典
    :param no_track_distribution: 无轨迹分布字典
    :param data_map:
    :param data_map_key: 用于存储到 `data_map` 的 key（例如 `CellKey.warehouse_condition` 或 `CellKey.store_condition`）
    :return: 生成的分布报告文本, 最低上网率的所有实体信息, 精简版报告文本
    """
    report_text = ""
    report_text2 = ""
    lowest_swl = 101  # 初始化为比100大的值
    lowest_entities = {}  # 存储多个最低上网率的实体信息

    # 遍历分布数据
    for entity, count in distribution.items():
        no_track_count = no_track_distribution.get(entity, 0)
        no_tracking_count = sku_no_tracking_distribution.get(entity, 0)
        pre_ship_count = sku_pre_ship_distribution.get(entity, 0)
        not_yet_count = sku_not_yet_distribution.get(entity, 0)
        delivered_count = sku_delivered_distribution.get(entity, 0)
        unpaid_count = sku_unpaid_distribution.get(entity, 0)
        tracking_sf_zero_count = tracking_sf_zero_distribution.get(entity, 0)

        swl = round2(100 - ((int(no_track_count) / int(count)) * 100))  # 计算上网率
        in_data = get_in(xlsx_path, entity)
        length_ = in_data['Length']
        width_ = in_data['Width']
        height_ = in_data['Height']
        unit_ = in_data['Unit']
        sizes = f"{length_}*{width_}*{height_}*{unit_}"

        kjCount, kj_no_track_count = sku_kj_count(xlsx_path, entity)
        kjSwl = 0
        if (kjCount > 0):
            kjSwl = round2(100 - ((kj_no_track_count / kjCount) * 100))

        temuCount, temu_no_track_count = temu_count(xlsx_path, entity)
        temuSwl = 0
        if (temuCount > 0):
            temuSwl = round2(100 - ((temu_no_track_count / temuCount) * 100))

        shipment_received_interval2_count = get_filtered_count(xlsx_path, gz_time, RowName.SKU, entity)

        strs = ""
        # 生成报告内容
        if (swl != 100.0 or unpaid_count > 0 or shipment_received_interval2_count > 0):
            # strs = f"\n{entity}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{swl}%"
            strs = f"\n{entity}：（{count}, {no_track_count}, {swl}%）,（{no_tracking_count}, {pre_ship_count}, " \
                   f"{not_yet_count}, {delivered_count}, {unpaid_count}）,（{kjCount}, {kj_no_track_count}, {kjSwl}%）,（{temuCount}, {temu_no_track_count}, {temuSwl}%）,（{shipment_received_interval2_count}）"
            # f"{not_yet_count}, {delivered_count}, {unpaid_count}）,（{kjCount}, {kj_no_track_count}, {kjSwl}%）,（{shipment_received_interval2_count}, {sizes}）"
            strs2 = f"\n{entity}：({count},{swl}%)"
            report_text += strs
            report_text2 += strs2

        # 更新最低上网率的实体
        if swl < lowest_swl:
            lowest_swl = swl
            lowest_entities.clear()  # 清空数据
            lowest_entities[entity] = {"entity": entity, "count": count,
                                       "no_track_count": (no_track_count), "swl": swl,
                                       "strs": strs}
        elif swl == lowest_swl:
            lowest_entities[entity] = {"entity": entity, "count": count,
                                       "no_track_count": (no_track_count), "swl": swl,
                                       "strs": strs}

    resultList = []
    fs_lowest_sku_result_list = []
    for key, value in lowest_entities.items():
        count = value["count"]
        no_track_counts = value["no_track_count"]
        strss = value["strs"]

        if interval_time >= 3:
            resultList.append(strss)
        else:
            if no_track_counts >= 4:
                resultList.append(strss)

        # sku数量和未上网数量一致 且 数量>=4
        if count >= 4 and count == no_track_counts:
            fs_lowest_sku_result_list.append(value)

    data_map[data_map_key] = report_text  # 将结果存储到 data_map
    return report_text, resultList, report_text2, fs_lowest_sku_result_list


def filter_tracking_numbers(input_path, output_path):
    # 读取 Excel 文件
    df = pd.read_excel(input_path, dtype=str).fillna("")

    col = "Tracking No./物流跟踪号"

    if col not in df.columns:
        print(f"❌ 文件中缺少列: {col}")
        return

    # 保留以 92、93 或 94 开头的行
    df_filtered = df[df[col].str.startswith(("92", "93", "94"), na=False)]

    # 保存结果
    df_filtered.to_excel(output_path, index=False)
    # print(f"✅ 筛选后文件已保存到: {output_path}")


def go(analyse_obj, xlsx_path, api_flag):
    if analyse_obj is None:
        analyse_obj = input("请输跟踪对象（zbw/sanrio/xyl/mz_xsd/md_fc/mx_dg）：")

    if analyse_obj != ClientConstants.zbw \
            and analyse_obj != ClientConstants.sanrio \
            and analyse_obj != ClientConstants.xyl \
            and analyse_obj != ClientConstants.kaer \
            and analyse_obj != ClientConstants.mz_xsd \
            and analyse_obj != ClientConstants.md_fc \
            and analyse_obj != ClientConstants.mx_dg:
        raise ValueError(f"{analyse_obj} 未定义")

    if xlsx_path is None:
        xlsx_path = input("请输入文件的绝对路径：")

    # 针对zbw做的运单号过滤操作
    if analyse_obj == ClientConstants.zbw:
        output_file = os.path.splitext(xlsx_path)[0] + "_去重0.xlsx"
        all_total_count = remove_duplicates_by_column(xlsx_path, output_file, RowName.Tracking_No)  # 无筛选订单总数
        delete_file(output_file)

    process_tracking_no(xlsx_path)
    check_and_add_courier_column(xlsx_path)

    irregular_number_map = find_irregular_tracking_numbers(xlsx_path)
    irregular_number_list = []
    if irregular_number_map:
        irregular_number_list = list(irregular_number_map.keys())
        # print(f"存在无效的物流跟踪号：{irregular_number_list}")
        # update_courier_status(xlsx_path, {CourierStateMapKey.irregular_number_map: irregular_number_map})
        update_courier_status1(xlsx_path, irregular_number_map)

    if api_flag:
        results = extract_and_process_data(xlsx_path, RowName.Courier, 100)

        all_maps = {
            CourierStateMapKey.not_yet_map: results[CourierStateMapKey.not_yet_map],
            CourierStateMapKey.pre_ship_map: results[CourierStateMapKey.pre_ship_map],
            CourierStateMapKey.unpaid_map: results[CourierStateMapKey.unpaid_map],
            CourierStateMapKey.delivered_map: results[CourierStateMapKey.delivered_map],
            CourierStateMapKey.no_tracking_map: results[CourierStateMapKey.no_tracking_map],
            CourierStateMapKey.tracking_map: results[CourierStateMapKey.tracking_map],
            CourierStateMapKey.possession_sf_date_map: results[CourierStateMapKey.possession_sf_date_map],
            CourierStateMapKey.latest_event_sf_date_map: results[CourierStateMapKey.latest_event_sf_date_map],
            CourierStateMapKey.sf_date_equality_map: results[CourierStateMapKey.sf_date_equality_map],
            CourierStateMapKey.latest_event_sf_time_map: results[CourierStateMapKey.latest_event_sf_time_map],
            CourierStateMapKey.latest_event_sf_site_map: results[CourierStateMapKey.latest_event_sf_site_map],
            CourierStateMapKey.alert_map: results[CourierStateMapKey.alert_map],
        }

        column_mapping = {
            CourierStateMapKey.not_yet_map: RowName.Courier,
            CourierStateMapKey.pre_ship_map: RowName.Courier,
            CourierStateMapKey.unpaid_map: RowName.Courier,
            CourierStateMapKey.delivered_map: RowName.Courier,
            CourierStateMapKey.no_tracking_map: RowName.Courier,
            CourierStateMapKey.tracking_map: RowName.Courier,
            CourierStateMapKey.alert_map: RowName.Courier,
            CourierStateMapKey.possession_sf_date_map: RowName.PossessionSfDate,
            CourierStateMapKey.latest_event_sf_date_map: RowName.LatestEventSfDate,
            CourierStateMapKey.sf_date_equality_map: RowName.SfDateInterval,
            CourierStateMapKey.latest_event_sf_time_map: RowName.LatestEventSfTime,
            CourierStateMapKey.latest_event_sf_site_map: RowName.LatestEventSfSite,
        }

        update_courier_status(xlsx_path, all_maps, wl=RowName.Tracking_No, column_map=column_mapping)

    # 针对zbw做的运单号过滤操作
    if analyse_obj == ClientConstants.zbw:
        output_file = os.path.splitext(xlsx_path)[0] + "_去重1.xlsx"
        filter_tracking_numbers(xlsx_path, output_file)
        xlsx_path = output_file

    # 数据map
    data_map = {}

    text = ""

    # 文件的创建时间
    ck_time = get_days_difference(xlsx_path)
    # 获取跟踪日期（调用程序的当天）
    gz_time = getYmd()
    # 跟踪日期 - 创建时间
    interval_time = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(ck_time, "%Y/%m/%d")).days
    # 创建时间是否为美国周末
    is_usweekend = is_us_weekend(ck_time)
    date_obj = datetime.strptime(ck_time, "%Y/%m/%d").date()
    # 创建时间的前一天
    previous_day = date_obj - timedelta(days=1)

    Outbound_Time = ""
    Outbound_Time += ck_time
    Outbound_Time += f"\n{get_weekday(ck_time)}"
    us_holiday = get_american_holiday(previous_day)
    if us_holiday:
        Outbound_Time += f"\n美国节日: {us_holiday}"
    cn_holiday = get_chinese_holiday(date_obj)
    if cn_holiday:
        Outbound_Time += f"\n中国节日: {cn_holiday}"

    date_obj1 = datetime.strptime(gz_time, "%Y/%m/%d").date()
    previous_day1 = date_obj1 - timedelta(days=1)
    Update_Time = ""
    Update_Time += datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    Update_Time += f"\n{get_weekday(gz_time)}"
    us_holiday1 = get_american_holiday(previous_day1)
    if us_holiday1:
        Update_Time += f"\n美国节日: {us_holiday1}"
    cn_holiday1 = get_chinese_holiday(date_obj1)
    if cn_holiday1:
        Update_Time += f"\n中国节日: {cn_holiday1}"

    text += "\n----------------------时间----------------------"
    text += f"\n更新时间: {Update_Time}"
    text += f"\n创建日期：{Outbound_Time}"
    text += f"\n跟踪日期：{gz_time}"
    text += f"\n间隔时间：{interval_time}"
    data_map[CellKey.Outbound_Time] = Outbound_Time
    data_map[CellKey.update_time] = Update_Time

    text += "\n----------------------非usps物流跟踪号----------------------"
    irregular_number_text = ""
    if (len(irregular_number_list) > 0):
        irregular_number_text = "\n非usps物流跟踪号："
        for ele in irregular_number_list:
            irregular_number_text += f"\n"
            irregular_number_text += ele
    text += irregular_number_text

    text += "\n----------------------unpaid详情----------------------"
    unpaid_text = ""
    result_map = get_unpaid_platform_tracking_map(xlsx_path)
    current_day_unpaid_text = ""
    current_day_unpaid_len = 0
    now_strftime = datetime.now().strftime('%Y-%m-%d')
    if len(result_map) > 0:
        search_day = ""
        for group_time, records in result_map.items():
            if group_time is None:
                if len(records.items()) > 0:
                    unpaid_text += f"\n🕒"
            else:
                dt = pd.to_datetime(group_time)
                search_day = dt.strftime('%Y-%m-%d')
                unpaid_text += f"\n🕒{search_day}"

            for platform_number, info in records.items():
                # print(f"\n平台单号：{platform_number} , 物流跟踪号：{info['tracking_number']}, 是否kj: {info['kj']}")
                number_ = f"\n{info['tracking_number']}"
                unpaid_text += number_
                if search_day == now_strftime:
                    current_day_unpaid_text += number_
                    current_day_unpaid_len += 1

            unpaid_text += "\n"

        # for key, value in result_map.items():
        #     value1 = value["tracking_number"]
        #     value2 = value["kj"]
        #     # if (analyse_obj == ClientConstants.sanrio):
        #     #     unpaid_text += f"\n平台单号：{key}, 物流跟踪号：{value1}, 是否kj：{value2}"
        #     # else:
        #     #     unpaid_text += f"\n物流跟踪号：{value1}, 是否kj：{value2}"
        #     # unpaid_text += f"\n物流跟踪号：{value1}, 是否kj：{value2}"
        #     unpaid_text += f"\n{value1}"
    text += unpaid_text

    text += "\n----------------------shipment_received详情----------------------"
    shipment_received_text = ""
    shipment_received_interval2_list = get_shipment_received_numbers(xlsx_path, gz_time)
    change_shipment_received_count = len(shipment_received_interval2_list)
    if (change_shipment_received_count > 0):
        shipment_received_text = "\nshipment_received物流跟踪号："
        for ele in shipment_received_interval2_list:
            shipment_received_text += f"\n"
            shipment_received_text += ele
    text += shipment_received_text

    data_map[CellKey.special_information] = irregular_number_text + unpaid_text + shipment_received_text

    text += "\n----------------------SKU分布----------------------"
    sku_distribution, sku_no_track_distribution, sku_no_tracking_distribution, sku_pre_ship_distribution, \
        sku_not_yet_distribution, sku_delivered_distribution, sku_unpaid_distribution, key_tracking_sf_zero_counter = count_distribution_and_no_track3(
        xlsx_path, key_column=RowName.SKU)
    sku_text, lowest_sku, sku_text2, fs_lowest_sku_result_list = generate_distribution_report2(
        sku_distribution, sku_no_track_distribution, sku_no_tracking_distribution, sku_pre_ship_distribution,
        sku_not_yet_distribution, sku_delivered_distribution, sku_unpaid_distribution, key_tracking_sf_zero_counter,
        data_map, CellKey.sku_condition,
        interval_time, xlsx_path, gz_time
    )
    text += sku_text

    output_file = os.path.splitext(xlsx_path)[0] + "_去重.xlsx"
    # 同一单会有多个sku，多个sku会生成多行数据，分析sku的时候不能去重，其它的需要去重
    total_count = remove_duplicates_by_column(xlsx_path, output_file, RowName.Tracking_No)

    patterns = {
        CourierStateMapValue.no_track: Pattern.no_track,
        CourierStateMapValue.delivered: Pattern.delivered,
        CourierStateMapValue.unpaid: Pattern.unpaid,
        CourierStateMapValue.not_yet: Pattern.not_yet,
        CourierStateMapValue.pre_ship: Pattern.pre_ship,
        CourierStateMapValue.irregular_no_tracking: Pattern.irregular_no_tracking,
        CourierStateMapValue.no_tracking: Pattern.no_tracking,
        CourierStateMapValue.tracking: Pattern.tracking
    }

    count_dict = count_pattern_and_tracking_with_sf_date(output_file, RowName.Courier, RowName.SfDateInterval, patterns)

    no_track_count = count_dict[CourierStateMapValue.no_track]
    delivered_count = count_dict[CourierStateMapValue.delivered]
    unpaid_count = count_dict[CourierStateMapValue.unpaid]
    not_yet_count = count_dict[CourierStateMapValue.not_yet]
    pre_ship_count = count_dict[CourierStateMapValue.pre_ship]
    irregular_no_tracking_count = count_dict[CourierStateMapValue.irregular_no_tracking]
    no_tracking_count = count_dict[CourierStateMapValue.no_tracking]
    tracking_count = count_dict[CourierStateMapValue.tracking]
    tracking_zero_count = count_dict["sfDateInterval"]

    # 先进行一次计算，并缓存结果
    total_count_int = int(total_count)
    no_track_count_int = int(no_track_count)
    track_count_int = total_count_int - no_track_count_int

    tracking_zero_count_int = int(tracking_zero_count)
    delivered_count_int = int(delivered_count)
    unpaid_count_int = int(unpaid_count)
    not_yet_count_int = int(not_yet_count)
    pre_ship_count_int = int(pre_ship_count)
    irregular_no_tracking_count_int = int(irregular_no_tracking_count)
    no_tracking_count_int = int(no_tracking_count)
    tracking_count_int = int(tracking_count)
    # real_no_track_count = no_track_count_int + tracking_zero_count_int  # 真正的未上网数
    # real_track_count = total_count_int - no_track_count  # 真正的上网数
    # real_tracking_count = tracking_count_int - tracking_zero_count_int

    # 计算百分比
    swl = round2(100 - ((no_track_count_int) / total_count_int * 100))
    wswl = round2(100 - swl)
    qsl = round2((delivered_count_int / total_count_int) * 100)
    unpaidl = round2((unpaid_count_int / total_count_int) * 100)
    not_yetl = round2((not_yet_count_int / total_count_int) * 100)
    pre_shipl = round2((pre_ship_count_int / total_count_int) * 100)
    irregular_no_trackingl = round2((irregular_no_tracking_count_int / total_count_int) * 100)
    no_tracking_countl = round2((no_tracking_count_int / total_count_int) * 100)
    tracking_countl = round2((tracking_count_int / total_count_int) * 100)
    tracking_zero_countl = round2((tracking_zero_count_int / total_count_int) * 100)
    change_shipment_received_countl = round2((change_shipment_received_count / total_count_int) * 100)

    kj_counts, kj_no_track_count, kj_tracking_count, kj_delivered_count, kj_pre_ship_count, kj_not_yet_count, \
        kj_unpaid_count, kj_irregular_number_count, kj_no_tracking_count = kj_count(output_file)

    kj_swl = 0
    kj_wswl = 0
    kj_track_count = kj_counts - kj_no_track_count
    if (track_count_int > 0):
        kj_swl = round2((kj_track_count / track_count_int) * 100)
    if (no_track_count_int > 0):
        kj_wswl = round2((kj_no_track_count / no_track_count_int) * 100)

    kj_tracking_swl = 0
    kj_delivered_swl = 0
    kj_pre_ship_swl = 0
    kj_not_yet_swl = 0
    kj_unpaid_swl = 0
    kj_irregular_number_swl = 0
    kj_other_no_tracking_swl = 0

    if (tracking_count_int > 0):
        kj_tracking_swl = round2((kj_tracking_count / tracking_count_int) * 100)
    if (delivered_count_int > 0):
        kj_delivered_swl = round2((kj_delivered_count / delivered_count_int) * 100)
    if (pre_ship_count_int > 0):
        kj_pre_ship_swl = round2((kj_pre_ship_count / pre_ship_count_int) * 100)
    if (not_yet_count_int > 0):
        kj_not_yet_swl = round2((kj_not_yet_count / not_yet_count_int) * 100)
    if (unpaid_count_int > 0):
        kj_unpaid_swl = round2((kj_unpaid_count / unpaid_count_int) * 100)
    if (irregular_no_tracking_count_int > 0):
        kj_irregular_number_swl = round2((kj_irregular_number_count / irregular_no_tracking_count_int) * 100)
    if (no_tracking_count_int > 0):
        kj_other_no_tracking_swl = round2((kj_no_tracking_count / no_tracking_count_int) * 100)

    zongshu = ""
    # 针对zbw做的运单号过滤操作
    if analyse_obj == ClientConstants.zbw:
        zongshu = f"\n订单总数：{total_count_int}【{all_total_count}】"
    else:
        zongshu = f"\n订单总数：{total_count_int}"

    # 构建字符串
    wl = (
        f"{zongshu}"
        f"\nKJ订单总数：{kj_counts}"
        f"\n"
        f"\n上网：（{track_count_int}, {swl}%）,（{kj_track_count}, {kj_swl}%）"
        f"\n未上网：（{no_track_count_int}, {wswl}%）,（{kj_no_track_count}, {kj_wswl}%）"
        f"\n"

        f"\n上网状态细分："
        f"\ntracking：（{tracking_count_int}, {tracking_countl}%）,（{kj_tracking_count}, {kj_tracking_swl}%）"
        f"\ndelivered：（{delivered_count_int}, {qsl}%）,（{kj_delivered_count}, {kj_delivered_swl}%）"
        f"\n"

        f"\n未上网状态细分："
        f"\npre_ship：（{pre_ship_count_int}, {pre_shipl}%）,（{kj_pre_ship_count}, {kj_pre_ship_swl}%）"
        f"\nnot_yet：（{not_yet_count_int}, {not_yetl}%）,（{kj_not_yet_count}, {kj_not_yet_swl}%）"
        f"\nunpaid：（{unpaid_count_int}, {unpaidl}%）,（{kj_unpaid_count}, {kj_unpaid_swl}%）"
        f"\nirregular_number：（{irregular_no_tracking_count_int}, {irregular_no_trackingl}%）,（{kj_irregular_number_count}, {kj_irregular_number_swl}%）"
        f"\nother_no_tracking：（{no_tracking_count_int}, {no_tracking_countl}%）,（{kj_no_tracking_count}, {kj_other_no_tracking_swl}%）"
        f"\n"

        # 接口无法获取过往物流状态，所以无法得出该订单是否使用了提货单。但是如果当天物流状态为shipment_received且(更新时间 - 出库时间 >= 2天)则一定为提货单且该订单物流异常
        f"\n提货单异常："
        # f"\n🔔提示：（满足'更新时间 - 出库时间 >= 2天' 且 流状态停留在shipment_received状态）"
        # f"\nshipment_received：（{tracking_zero_count_int}, {tracking_zero_countl}%）"
        f"\nstay_shipment_received：（{change_shipment_received_count}, {change_shipment_received_countl}%）"

        # f"\n{irregular_number_text + unpaid_text + shipment_received_text}"
        # f"\n{irregular_number_text + unpaid_text}"
    )
    data_map[CellKey.wl] = wl
    data_map[CellKey.unpaid] = unpaid_text

    text += "\n----------------------轨迹概览----------------------"
    text += wl

    # 避免重复读取 Excel 文件
    warehouse_count, warehouse_no_track_count = dimension_distribution(output_file, key_column=RowName.Warehouse)
    warehouse_text, lowest_warehouse, warehouse_text2 = generate_distribution_report(output_file,
                                                                                     RowName.Warehouse,
                                                                                     gz_time,
                                                                                     warehouse_count,
                                                                                     warehouse_no_track_count,
                                                                                     data_map,
                                                                                     CellKey.warehouse_condition)
    text += "\n----------------------仓库分布----------------------" + warehouse_text

    store_count, store_no_track_count = dimension_distribution(output_file, key_column=RowName.Client)
    store_text, lowest_store, store_text2 = generate_distribution_report(output_file, RowName.Client, gz_time,
                                                                         store_count, store_no_track_count,
                                                                         data_map, CellKey.store_condition)
    text += "\n----------------------店铺分布----------------------" + store_text

    shipping_service_count, shipping_service_no_track_count = dimension_distribution(output_file,
                                                                                     key_column=RowName.ShippingService)
    shipping_service_text, lowest_shipping_service, shipping_service_text2 = generate_distribution_report(
        output_file, RowName.ShippingService, gz_time,
        shipping_service_count,
        shipping_service_no_track_count,
        data_map,
        CellKey.shipping_service_condition)
    text += "\n----------------------物流渠道分布----------------------" + shipping_service_text

    time_segment_text, lowest_segment = analyze_time_segments(output_file, data_map,
                                                              time_column=RowName.CreationWaveTime,
                                                              courier_column=RowName.Courier,
                                                              sf_date_column=RowName.SfDateInterval)

    text += "\n----------------------时间段分布----------------------" + time_segment_text

    lowest_txt = "\n"
    if lowest_sku:
        lowest_txt += "\n最低上网率的 SKU：" + "".join(lowest_sku)
    lowest_txt += f"""
    最低上网率的 仓库：{lowest_warehouse}
    最低上网率的 商店：{lowest_store}
    最低上网率的 时间段：{lowest_segment}
    最低上网率的 物流渠道：{lowest_shipping_service}
    """

    sum_up_text = ""
    actual_interval = 0
    if is_usweekend == 6:  # 6是中国周日，美国周六
        sum_up_text = "美国时间：周六（和中国相差13-16个小时）\n"
        actual_interval = 2
    elif is_usweekend == 0:  # 0是中国周一，美国周日
        sum_up_text = "美国时间：周日（相差13-16个小时）\n"
        actual_interval = 1

    # if (len(irregular_number_list) > 0):
    #     sum_up_text += f"存在不规则单号：{irregular_number_list}"
    #     sum_up_text += f"\n"

    swl_flag = False
    qsl_flag = False
    bg = "#FFFFFF"

    # 上网率判断
    warning_levels = [
        (0, 30, "🧑‍🍳", "继续观察👀", "#FFFFFF"),
        (1, 30, "☁️注意", "未达30%！", "#F8F1D3"),
        (2, 70, "🌧️注意", "未达70%！", "#E3C49C"),
        (3, 97, "❄️异常", "未达97%！", "#F1C1BD"),
    ]

    # 跟踪时间 - 创建时间 - （创建时间美国周末）
    actual_interval_time = interval_time - actual_interval
    # sum_up_text += f"\n间隔第{interval_time}{actual_interval}天"
    sum_up_text += f"\n间隔第{actual_interval_time}天"

    for days, threshold, icon, message, color in warning_levels:
        if actual_interval_time == days and swl < threshold:
            sum_up_text += f"\n{icon}：上网率为{swl}%，{message}"
            if actual_interval_time >= 2:
                swl_flag = True
                bg = color
            break
    else:  # ✅ 只有 for 没有 break 时才会执行
        if actual_interval_time >= 4:
            if swl >= 99:
                sum_up_text += f"\n☀️上网率为{swl}%，优秀🌈"
            elif 97 <= swl < 99:
                sum_up_text += f"\n☀️上网率为{swl}%，达标✅"
            else:
                sum_up_text += f"\n⚡️异常：上网率为{swl}%，未达️97%"
                bg = "#F1C1BD"
                swl_flag = True
        else:
            sum_up_text += f"\n☀️上网率为{swl}%，达标✅"

    # 签收率判断
    qsl_warnings = [
        (0, 0, 1, "🧑‍🍳", "继续观察👀"),
        (1, 3, 1, "☁️注意", "未达1%"),
        (3, 5, 20, "🌧️注意", "未达20%"),
        (5, 7, 50, "⛈️注意", "未达50%"),
        (7, 9, 80, "❄️注意", "未达80%"),
    ]

    for start, end, threshold, icon, message in qsl_warnings:
        if start <= actual_interval_time <= end and qsl <= threshold:
            sum_up_text += f"\n{icon}：签收率为{qsl}%，{message}"
            # qsl_flag = True
            break
    else:
        if actual_interval_time >= 10:
            if qsl >= 95:
                sum_up_text += f"\n☀️签收率为{qsl}%，优秀！🌈"
            elif qsl >= 90 and qsl < 95:
                sum_up_text += f"\n☀️签收率为{qsl}%，达标！✅"
            else:
                sum_up_text += f"\n⚡️异常：️签收率为{qsl}%，未达️90%"
                # qsl_flag = True
        else:
            sum_up_text += f"\n☀️签收率为{qsl}%，达标！✅"

    text += "\n----------------------总结&建议----------------------"
    if swl < 100:
        sum_up_text += lowest_txt
    text += f"\n{sum_up_text}"
    data_map[CellKey.sum_up] = sum_up_text

    # 异常状态记录
    exception_text = ""
    bgFlag = False

    if qsl_flag:
        exception_text += "\n✍️签收率异常（签收率目前无法量化，只为提醒⏰）"
        bgFlag = True
    if change_shipment_received_count >= 10:
        bg = "#B3D600"
        exception_text += "\n📒提货单未更新轨迹（>=2天） 异常"
        bgFlag = True
    if swl_flag:
        exception_text += "\n📦上网率异常"
        bgFlag = True
    if unpaid_count_int > 0:
        bg = "#A684F0"
        exception_text += "\n💰unpaid 异常"
        bgFlag = True
    data_map[CellKey.exception] = exception_text

    # 删除去重文件
    delete_file(output_file)
    # 针对zbw做的运单号过滤操作
    if analyse_obj == ClientConstants.zbw:
        delete_file(xlsx_path)
    print(text)

    # 写入飞书在线文档
    tat = get_token()
    if analyse_obj == ClientConstants.zbw or \
            analyse_obj == ClientConstants.sanrio or \
            analyse_obj == ClientConstants.xyl or \
            analyse_obj == ClientConstants.kaer:

        zongshu1 = ""
        # 针对zbw做的运单号过滤操作
        if analyse_obj == ClientConstants.zbw:
            zongshu1 = f"{total_count_int}【{all_total_count}】"
        else:
            zongshu1 = f"{total_count_int}"

        khhz_sheet_value(tat, [
            f"{zongshu1}",
            f"（{no_track_count_int}, {wswl}%）",
            f"（{change_shipment_received_count}, {change_shipment_received_countl}%）",
            f"（{delivered_count_int}, {qsl}%）",
            f"（{unpaid_count_int}, {unpaidl}%）",
        ], ck_time, analyse_obj)

        khhz_sheet_bg(tat, ck_time, analyse_obj, bg)

        zongshu2 = ""
        # 针对zbw做的运单号过滤操作
        if analyse_obj == ClientConstants.zbw:
            zongshu2 = f"上网：({total_count},{swl}%)【{all_total_count}】"
        else:
            zongshu2 = f"上网：({total_count},{swl}%)"

        lists = f"{zongshu2}"
        lists += f"\n提货单未上网：({change_shipment_received_count},{change_shipment_received_countl}%)"
        lists += f"\n{warehouse_text2}"
        brief_sheet_value(tat, [lists], ck_time, gz_time, analyse_obj)
        # if (bgFlag):
        brief_sheet_bg(tat, ck_time, gz_time, analyse_obj, bg)
    else:
        lists = f"({total_count},{swl}%)"
        brief_sheet_value(tat, [lists], ck_time, gz_time, analyse_obj)
        # if (bgFlag):
        brief_sheet_bg(tat, ck_time, gz_time, analyse_obj, bg)

    if analyse_obj == ClientConstants.mz_xsd or \
            analyse_obj == ClientConstants.mx_dg or \
            analyse_obj == ClientConstants.md_fc:
        detail_sheet_value(tat, [
            data_map[CellKey.Outbound_Time],
            data_map[CellKey.update_time],
            data_map[CellKey.wl],
            data_map[CellKey.store_condition],
            data_map[CellKey.time_segment_condition],
            data_map[CellKey.shipping_service_condition],
            data_map[CellKey.sum_up],
            data_map[CellKey.exception],
        ], ck_time, analyse_obj)

        # if (bgFlag):
        detail_sheet_bg(tat, ck_time, analyse_obj, bg)
    else:
        detail_sheet_value(tat, [
            data_map[CellKey.Outbound_Time],
            data_map[CellKey.update_time],
            data_map[CellKey.wl],
            data_map[CellKey.warehouse_condition],
            data_map[CellKey.store_condition],
            data_map[CellKey.time_segment_condition],
            data_map[CellKey.shipping_service_condition],
            data_map[CellKey.sku_condition],
            data_map[CellKey.sum_up],
            data_map[CellKey.exception],
            data_map[CellKey.unpaid],
        ], ck_time, analyse_obj)

        # if (bgFlag):
        detail_sheet_bg(tat, ck_time, analyse_obj, bg)

    result_fs_msg = f"客户：{analyse_obj}\n"
    result_fs_msg += f"订单创建时间：{ck_time}\n"
    result_fs_msg += f"跟踪时间：{gz_time}\n"
    lj_fs_msg = ""
    fs_msg_flag = False
    lj_msg_flag = False

    if len(current_day_unpaid_text) > 0:
        result_fs_msg += f"新增 {current_day_unpaid_len}单 unpaid: \n"
        result_fs_msg += current_day_unpaid_text
        lj_fs_msg = result_fs_msg
        fs_msg_flag = True
        lj_msg_flag = True

    if change_shipment_received_count >= 10:
        result_fs_msg += f"{change_shipment_received_count}单 提货单未更新轨迹（>=2天）\n"
        fs_msg_flag = True

    if swl_flag:
        swl_text = "⚠️未上网数和应发数量一致：\n"
        swl_text_flag = False
        for lowest_sku in fs_lowest_sku_result_list:
            count = lowest_sku["count"]
            no_track_counts = lowest_sku["no_track_count"]
            entity = lowest_sku["entity"]
            swl_text += f"({entity}：{count})\n"
            swl_text_flag = True

        if swl <= 90 or swl_text_flag:
            result_fs_msg += f"上网率异常: {swl}%\n"
            result_fs_msg += swl_text
            fs_msg_flag = True

    if fs_msg_flag:
        # print(result_fs_msg)
        fs_msg(FsUserID.WP_ID, result_fs_msg)
        fs_msg(FsUserID.LW_ID, result_fs_msg)

    if lj_msg_flag:
        fs_msg(FsUserID.LJ_ID, lj_fs_msg)


def call2():
    automatic("/Users/zkp/Desktop/B&Y/轨迹统计/zbw/", ClientConstants.zbw, False, False, False)
    automatic("/Users/zkp/Desktop/B&Y/轨迹统计/sanrio/", ClientConstants.sanrio, False, False, False)
    automatic("/Users/zkp/Desktop/B&Y/轨迹统计/xyl/", ClientConstants.xyl, False, True, False)
    automatic("/Users/zkp/Desktop/B&Y/轨迹统计/kaer/", ClientConstants.kaer, False, True, False)


def is_time_difference_exceed(start_time_str, end_time_str):
    try:
        time_format = "%Y-%m-%d"
        start_time = datetime.strptime(start_time_str, time_format)
        end_time = datetime.strptime(end_time_str, time_format)
        difference = abs((end_time - start_time).days)
        return difference
    except ValueError as e:
        print(f"❌ 时间格式错误: {e}")
        return False


def automatic(root_dir, analyse_obj, ignore=False, analyse_obj_ignore=False, api_flag=True):
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    is_morning = (datetime.now().hour) < 12
    gz_time = datetime.strptime(getYmd(), "%Y/%m/%d")
    # print(current_year, current_month, current_day)
    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = re.search(r"创建时间(\d+)_", file)
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= 15:
                        print(f"正在处理文件: {xlsx_path}")

                        if ignore:
                            go(analyse_obj, xlsx_path, api_flag)
                            continue

                        ck_time = get_days_difference(xlsx_path)
                        interval_time = (gz_time - datetime.strptime(ck_time, "%Y/%m/%d")).days

                        if interval_time == 1 and is_morning:
                            go(analyse_obj, xlsx_path, api_flag)
                        else:
                            if is_morning:
                                continue

                            if analyse_obj_ignore:
                                go(analyse_obj, xlsx_path, api_flag)
                                continue

                            if check_and_add_courier_column(xlsx_path):
                                go(analyse_obj, xlsx_path, api_flag)
                            else:
                                shipment_received_interval2_list = get_shipment_received_numbers(xlsx_path, getYmd())
                                change_shipment_received_count = len(shipment_received_interval2_list)

                                output_file = os.path.splitext(xlsx_path)[0] + "_去重.xlsx"
                                filter_tracking_numbers(xlsx_path, output_file)

                                # 同一单会有多个sku，多个sku会生成多行数据，分析sku的时候不能去重，其它的需要去重
                                total_count = remove_duplicates_by_column(output_file, output_file, RowName.Tracking_No)

                                patterns = {
                                    CourierStateMapValue.no_track: Pattern.no_track,
                                    CourierStateMapValue.delivered: Pattern.delivered,
                                    CourierStateMapValue.unpaid: Pattern.unpaid,
                                    CourierStateMapValue.not_yet: Pattern.not_yet,
                                    CourierStateMapValue.pre_ship: Pattern.pre_ship,
                                    CourierStateMapValue.irregular_no_tracking: Pattern.irregular_no_tracking,
                                    CourierStateMapValue.no_tracking: Pattern.no_tracking,
                                    CourierStateMapValue.tracking: Pattern.tracking
                                }

                                count_dict = count_pattern_and_tracking_with_sf_date(output_file, RowName.Courier,
                                                                                     RowName.SfDateInterval, patterns)

                                no_track_count = count_dict[CourierStateMapValue.no_track]
                                delivered_count = count_dict[CourierStateMapValue.delivered]
                                unpaid_count = count_dict[CourierStateMapValue.unpaid]
                                not_yet_count = count_dict[CourierStateMapValue.not_yet]
                                pre_ship_count = count_dict[CourierStateMapValue.pre_ship]
                                irregular_no_tracking_count = count_dict[CourierStateMapValue.irregular_no_tracking]
                                no_tracking_count = count_dict[CourierStateMapValue.no_tracking]
                                tracking_count = count_dict[CourierStateMapValue.tracking]
                                tracking_zero_count = count_dict["sfDateInterval"]

                                # 先进行一次计算，并缓存结果
                                total_count_int = int(total_count)
                                no_track_count_int = int(no_track_count)
                                track_count_int = total_count_int - no_track_count_int

                                tracking_zero_count_int = int(tracking_zero_count)
                                delivered_count_int = int(delivered_count)
                                unpaid_count_int = int(unpaid_count)
                                not_yet_count_int = int(not_yet_count)
                                pre_ship_count_int = int(pre_ship_count)
                                irregular_no_tracking_count_int = int(irregular_no_tracking_count)
                                no_tracking_count_int = int(no_tracking_count)
                                tracking_count_int = int(tracking_count)
                                # real_no_track_count = no_track_count_int + tracking_zero_count_int  # 真正的未上网数
                                # real_track_count = total_count_int - no_track_count  # 真正的上网数
                                # real_tracking_count = tracking_count_int - tracking_zero_count_int

                                if tracking_zero_count_int == 0 and delivered_count_int == 0 and unpaid_count_int == 0 \
                                        and not_yet_count_int == 0 and pre_ship_count_int == 0 and no_tracking_count_int == 0 \
                                        and tracking_count_int == 0:
                                    delete_file(output_file)
                                    go(analyse_obj, xlsx_path, api_flag)
                                    continue

                                # 计算百分比
                                swl = round2(100 - ((no_track_count_int) / total_count_int * 100))
                                wswl = round2(100 - swl)
                                qsl = round2((delivered_count_int / total_count_int) * 100)
                                unpaidl = round2((unpaid_count_int / total_count_int) * 100)
                                not_yetl = round2((not_yet_count_int / total_count_int) * 100)
                                pre_shipl = round2((pre_ship_count_int / total_count_int) * 100)
                                irregular_no_trackingl = round2(
                                    (irregular_no_tracking_count_int / total_count_int) * 100)
                                no_tracking_countl = round2((no_tracking_count_int / total_count_int) * 100)
                                tracking_countl = round2((tracking_count_int / total_count_int) * 100)
                                tracking_zero_countl = round2((tracking_zero_count_int / total_count_int) * 100)
                                change_shipment_received_countl = round2(
                                    (change_shipment_received_count / total_count_int) * 100)

                                delete_file(output_file)

                                if swl < 99 or change_shipment_received_count >= 10 or unpaid_count > 0 or \
                                        (exceed >= 14 and qsl < 98):
                                    go(analyse_obj, xlsx_path, api_flag)


if __name__ == '__main__':
    call2()
