from openpyxl import load_workbook
import xlwings as xw
import re  # 用于提取括号中的数值

try:
    # 打开文件1和文件2
    try:
        wb1 = load_workbook("/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF待发货 订单-2024-12-24-19_09.xlsx")  # 表1文件
    except FileNotFoundError:
        raise FileNotFoundError("未找到文件 file1.xlsx，请检查文件路径。")

    try:
        # wb2 = load_workbook("/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF供应商对账表1223-111.xlsx")  # 表2文件
        wb2 = load_workbook("/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF供应商对账表1223-111.xlsx")  # 表2文件
    except FileNotFoundError:
        raise FileNotFoundError("未找到文件 file2.xlsx，请检查文件路径。")

    try:
        wb4 = load_workbook("/Users/zkp/Desktop/B&Y/CZFF供应商对账/CZFF产品核对表1114.xlsx")
    except FileNotFoundError:
        raise FileNotFoundError("未找到文件 file4.xlsx，请检查文件路径。")

    # 获取活动表
    try:
        sheet1 = wb1.active
        sheet2 = wb2.active
        sheet4 = wb4.active
    except Exception as e:
        raise Exception(f"获取活动表时发生错误: {e}")

    # 动态查找表1的列
    paid_time_col = None
    order_id_col = None
    quantity_col = None
    seller_sku_col = None
    try:
        for col in sheet1.iter_cols(1, sheet1.max_column):
            if col[0].value == "Paid Time":  # 假设标题在第1行
                paid_time_col = col[0].column  # 获取列号
            elif col[0].value == "Order ID":  # 假设标题在第1行
                order_id_col = col[0].column  # 获取列号
            elif col[0].value == "Quantity":  # 假设标题在第1行
                quantity_col = col[0].column  # 获取列号
            elif col[0].value == "Seller SKU":  # 假设标题在第1行
                seller_sku_col = col[0].column  # 获取列号
            if paid_time_col and order_id_col and quantity_col and seller_sku_col:
                break
    except Exception as e:
        raise Exception(f"查找表1中的列时发生错误: {e}")

    if paid_time_col is None:
        raise ValueError("表1中未找到 'Paid Time' 列，请确认列名是否正确。")
    if order_id_col is None:
        raise ValueError("表1中未找到 'Order ID' 列，请确认列名是否正确。")
    if quantity_col is None:
        raise ValueError("表1中未找到 'Quantity' 列，请确认列名是否正确。")
    if seller_sku_col is None:
        raise ValueError("表1中未找到 'Seller SKU' 列，请确认列名是否正确。")

    # 动态查找表2的列
    date_col = None
    order_number_col = None
    quantity_dest_col = None
    freight_col = None
    sku_col = None
    exchange_rate_col = None
    try:
        for col in sheet2.iter_cols(1, sheet2.max_column):
            if col[0].value == "日期":  # 假设标题在第1行
                date_col = col[0].column
            elif col[0].value == "订单单号":  # 假设标题在第1行
                order_number_col = col[0].column
            elif col[0].value == "数量":  # 假设标题在第1行
                quantity_dest_col = col[0].column
            elif col[0].value == "运费":  # 假设标题在第1行
                freight_col = col[0].column
            elif col[0].value == "款号":  # 假设标题在第1行
                sku_col = col[0].column
            elif col[0].value == "汇率":  # 假设标题在第1行
                exchange_rate_col = col[0].column
            if date_col and order_number_col and quantity_dest_col and freight_col and sku_col and exchange_rate_col:
                break
    except Exception as e:
        raise Exception(f"查找表2中的列时发生错误: {e}")

    if date_col is None:
        raise ValueError("表2中未找到 '日期' 列，请确认列名是否正确。")
    if order_number_col is None:
        raise ValueError("表2中未找到 '订单单号' 列，请确认列名是否正确。")
    if quantity_dest_col is None:
        raise ValueError("表2中未找到 '数量' 列，请确认列名是否正确。")
    if freight_col is None:
        raise ValueError("表2中未找到 '运费' 列，请确认列名是否正确。")
    if sku_col is None:
        raise ValueError("表2中未找到 '款号' 列，请确认列名是否正确。")
    if exchange_rate_col is None:
        raise ValueError("表2中未找到 '汇率' 列，请确认列名是否正确。")

    # 遍历表1并复制数据到表2（从表2的第2行开始）
    try:
        target_row = 2  # 表2从第2行开始写入
        for row in range(3, sheet1.max_row + 1):  # 跳过表1的第2行
            # 从表1获取数据
            paid_time_value = sheet1.cell(row, paid_time_col).value
            order_id_value = sheet1.cell(row, order_id_col).value
            quantity_value = sheet1.cell(row, quantity_col).value
            seller_sku_value = sheet1.cell(row, seller_sku_col).value

            # 写入表2的各列
            sheet2.cell(target_row, date_col).value = paid_time_value
            sheet2.cell(target_row, order_number_col).value = order_id_value
            sheet2.cell(target_row, quantity_dest_col).value = quantity_value
            sheet2.cell(target_row, sku_col).value = seller_sku_value

            # 接口获取到到汇率写入表2的汇率列
            sheet2.cell(target_row, freight_col).value = f"=2.99+IF(E{target_row}=1,0,(E{target_row}-1)*0.5)"

            # 写入公式到运费列
            sheet2.cell(target_row, exchange_rate_col).value = 7.31

            target_row += 1  # 表2写入下一行
    except Exception as e:
        raise Exception(f"写入数据时发生错误: {e}")

    # 动态查找表2的“款号”和“采购价”列
    sku_col_2 = None
    purchase_price_col_2 = None
    result_price_rmb_col = None

    try:
        for col in sheet2.iter_cols(1, sheet2.max_column):
            if col[0].value == "款号":  # 表2的“款号”列
                sku_col_2 = col[0].column
            elif col[0].value == "采购价":  # 表2的“采购价”列
                purchase_price_col_2 = col[0].column
            elif col[0].value == "采购价（RMB）":  # 假设标题在第1行
                result_price_rmb_col = col[0].column
            if sku_col_2 and purchase_price_col_2 and result_price_rmb_col:
                break
    except Exception as e:
        raise Exception(f"查找表2中的列时发生错误: {e}")

    if sku_col_2 is None:
        raise ValueError("表2中未找到 '款号' 列，请确认列名是否正确。")
    if purchase_price_col_2 is None:
        raise ValueError("表2中未找到 '采购价' 列，请确认列名是否正确。")
    if result_price_rmb_col is None:
        raise ValueError("表2中未找到 '采购价（RMB）' 列，请确认列名是否正确。")

    # 表4的列位置直接使用
    sku_col_4 = 8  # 表4的 H 列对应的列号
    cost_col_4 = 5  # 表4的 E 列对应的列号

    # 遍历表2的款号列，匹配表4的款号列，并提取括号中的值
    try:
        for row_2 in range(2, sheet2.max_row + 1):  # 从表2的第2行开始
            sku_value_2 = sheet2.cell(row_2, sku_col_2).value
            if sku_value_2:
                matched_cost = None
                for row_4 in range(2, sheet4.max_row + 1):  # 从表4的第2行开始
                    sku_value_4 = sheet4.cell(row_4, sku_col_4).value
                    if sku_value_4 == sku_value_2:  # 匹配款号
                        cost_value_4 = sheet4.cell(row_4, cost_col_4).value
                        # print(f"dsdsdsd:{cost_value_4}")
                        # 提取括号中的数值
                        match = re.search(r"（([\d.]+)", cost_value_4)
                        # print(f"dsdsdsd1111:{match}")
                        if match:
                            matched_cost = float(match.group(1))  # 转为浮点数
                            # print(f"dsdsdsd22222:{matched_cost}")
                        break

                # 将匹配到的成本价写入表2的采购价列
                sheet2.cell(row_2, purchase_price_col_2).value = matched_cost

                purchase_price_value = sheet2.cell(row_2, purchase_price_col_2).value
                quantity_value = sheet2.cell(row_2, quantity_dest_col).value
                freight_value = sheet2.cell(row_2, freight_col).value
                exchange_rate_value = sheet2.cell(row_2, exchange_rate_col).value

                print(f"dsdsdsds: {purchase_price_value},{quantity_value},{freight_value},{exchange_rate_value}")

                # 公式
                formula = f"(({purchase_price_value} * {quantity_value}) + {freight_value}) * {exchange_rate_value}"
                # 将公式写入采购价（RMB）列
                sheet2.cell(row_2, result_price_rmb_col).value = f"={formula}"

    except Exception as e:
        raise Exception(f"匹配和写入数据时发生错误: {e}")

    # 保存修改后的表2
    try:
        wb2.save("/Users/zkp/Desktop/B&Y/CZFF供应商对账/3333.xlsx")
        print("文件更新成功，保存为 3333.xlsx")
    except Exception as e:
        raise Exception(f"保存文件时发生错误: {e}")

except Exception as error:
    print(f"程序运行时发生错误: {error}")
