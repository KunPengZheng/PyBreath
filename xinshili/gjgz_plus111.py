from datetime import datetime, date, timedelta
import os
import re
from openpyxl import load_workbook
import openpyxl
import pandas as pd
from collections import Counter, defaultdict
from dataclasses import dataclass
import concurrent.futures
import time

from xinshili.fs_utils_plus import get_token, brief_sheet_value, detail_sheet_value, ClientConstants
from xinshili.pd_utils import remove_duplicates_by_column
from xinshili.usps_utils import track
from xinshili.utils import round2, getYmd, delete_file, is_us_weekend

"""
zbw轨迹跟踪分析
"""


@dataclass(frozen=True)
class RowName:
    Tracking_No = 'Tracking No./物流跟踪号'
    Courier = 'Courier/快递'
    OutboundTime = "OutboundTime/出库时间"
    Warehouse = "Warehouse/仓库"
    Client = "Client/客户"
    CreationWaveTime = "Create wave time/生成波次时间"
    SKU = "SKU"
    ShippingService = "Shipping service/物流渠道"


@dataclass(frozen=True)
class CourierStateMapKey:
    tracking_map = 'tracking_map'
    no_tracking_map = 'no_tracking_map'
    unpaid_map = "unpaid_map"
    not_yet_map = "not_yet_map"
    pre_ship_map = "pre_ship_map"
    delivered_map = "delivered_map"


class CourierStateMapValue:
    irregular_no_tracking = 'irregular_no_tracking'
    not_yet = 'not_yet'
    pre_ship = "pre_ship"
    no_tracking = "no_tracking"
    unpaid = "unpaid"
    delivered = "delivered"
    tracking = "tracking"


@dataclass(frozen=True)
class CellKey:
    update_time = "update_time"
    order_count = "order_count"
    no_track_number = "no_track_number"
    track_percent = "track_percent"
    delivered_counts = "delivered_counts"
    delivered_percent = "delivered_percent"
    no_track_percent = "no_track_percent"
    warehouse_condition = "warehouse_condition"
    store_condition = "store_condition"
    sku_condition = "sku_condition"
    time_segment_condition = "time_segment_condition"
    sum_up = "sum_up"
    exception = "exception"
    shipping_service_condition = "shipping_service_condition"
    unpaid_count = "unpaid_count"


@dataclass(frozen=True)
class Pattern:
    no_track = r"not_yet|pre_ship|irregular_no_tracking|no_tracking"
    delivered = r"delivered"
    unpaid = r"unpaid"


def find_irregular_tracking_numbers(filepath, column_name=RowName.Tracking_No):
    """
    查找不规则的快递单号（不是纯数字或者不是9开头）
    :param filepath: Excel文件路径
    :return: 不规则快递单号字典
    """
    try:
        # 打开xlsx文件
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active  # 默认使用活动工作表

        # 获取 'Tracking No./物流跟踪号' 列索引
        tracking_no_col = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == column_name:
                tracking_no_col = col
                break

        if tracking_no_col is None:
            print(f"找不到 {column_name} 列")
            return {}

        # 存储不规则快递单号的字典
        irregular_number_map = {}

        # 遍历所有行，从第二行开始（跳过表头）
        for row in range(2, sheet.max_row + 1):
            tracking_no = str(sheet.cell(row=row, column=tracking_no_col).value)  # 转换为字符串
            # 判断是否是纯数字并且以9开头
            if not tracking_no.isdigit() or not tracking_no.startswith('9'):
                irregular_number_map[tracking_no] = CourierStateMapValue.irregular_no_tracking

        return irregular_number_map

    except Exception as e:
        print(f"发生错误: {e}")
        return {}


def update_courier_status(filepath, maps, wl=RowName.Tracking_No):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active  # 默认使用活动工作表

    data = pd.read_excel(filepath)
    # 获取 'Tracking No./物流跟踪号' 列和 'Courier/快递' 列的索引
    tracking_no_col = data.columns.get_loc(wl) + 1  # openpyxl索引从1开始
    courier_col = data.columns.get_loc(RowName.Courier) + 1  # openpyxl索引从1开始

    for tracking_no, status in maps.items():
        for row in range(2, sheet.max_row + 1):  # 从第二行开始（跳过表头）
            # 获取当前行的物流跟踪号
            current_tracking_no = sheet.cell(row=row, column=tracking_no_col).value
            # 如果找到匹配的物流跟踪号，更新 Courier/快递 列
            if current_tracking_no == tracking_no:
                sheet.cell(row=row, column=courier_col, value=status)

    # 保存更新后的文件
    wb.save(filepath)


def extract_and_process_data(filepath: str, column_name: str, group_size: int, wl_name=RowName.Tracking_No,
                             request_interval: float = 5.0):
    data = pd.read_excel(filepath)

    if column_name not in data.columns:
        raise ValueError(f"列 '{column_name}' 不存在于 Excel 文件中")

    # 存储结果的 map（字典）
    results_map = {
        CourierStateMapKey.tracking_map: {},
        CourierStateMapKey.no_tracking_map: {},
        CourierStateMapKey.unpaid_map: {},
        CourierStateMapKey.not_yet_map: {},
        CourierStateMapKey.pre_ship_map: {},
        CourierStateMapKey.delivered_map: {},
    }

    # 将无内容的单元格赋值""空字符串
    data[column_name] = data[column_name].fillna('')

    # 获取指定内容的数据
    filtered_data = data[data[column_name].apply(
        lambda x: str(x).strip().lower() in ['', CourierStateMapValue.not_yet,
                                             CourierStateMapValue.pre_ship,
                                             CourierStateMapValue.tracking,
                                             CourierStateMapValue.no_tracking])]

    # 提取符合条件的 'Tracking No./物流跟踪号' 列数据
    items = filtered_data[wl_name].tolist()

    # 按组划分数据
    grouped_items = [items[i:i + group_size] for i in range(0, len(items), group_size)]

    # 使用线程池来并发请求每组数据
    with concurrent.futures.ThreadPoolExecutor() as executor:
        # 提交任务
        futures = {executor.submit(track, group): group for group in grouped_items}

        # 遍历所有的任务（每组数据的请求）
        for future in concurrent.futures.as_completed(futures):
            group = futures[future]  # 获取当前完成的任务的 group 数据
            try:
                track1 = future.result()  # 获取请求结果
                print(f"处理第 {grouped_items.index(group) + 1} 组，共 {len(group)} 条数据")

                # 对返回的数据进行处理
                for package_id, info in track1['data'].items():
                    # 判断错误类型并分类
                    if info.get('err'):
                        if info.get('err_id') == '-2147219283':  # 无轨迹(Label Created, not yet in system)
                            results_map[CourierStateMapKey.not_yet_map][package_id] = CourierStateMapValue.not_yet
                        elif info.get('err_id') == 'pre-ship':  # 无轨迹(pre-ship)
                            results_map[CourierStateMapKey.pre_ship_map][package_id] = CourierStateMapValue.pre_ship
                        else:
                            results_map[CourierStateMapKey.no_tracking_map][
                                package_id] = CourierStateMapValue.no_tracking
                    else:
                        if "The package associated with this tracking number did not have proper postage applied and will not be delivered" in \
                                info.get('statusLong'):
                            results_map[CourierStateMapKey.unpaid_map][package_id] = CourierStateMapValue.unpaid
                        elif "Delivered" in info.get('statusCategory'):
                            results_map[CourierStateMapKey.delivered_map][package_id] = CourierStateMapValue.delivered
                        elif "Delivered to Agent" in info.get('statusCategory'):
                            results_map[CourierStateMapKey.delivered_map][package_id] = CourierStateMapValue.delivered
                        else:
                            results_map[CourierStateMapKey.tracking_map][package_id] = CourierStateMapValue.tracking

                # 等待指定的间隔时间，避免请求频率过高
                time.sleep(request_interval)

            except Exception as e:
                print(f"处理组 {grouped_items.index(group) + 1} 时发生错误: {e}")

    return results_map


def count_pattern_state(file_path, column_name, patternStr):
    """
    统计指定列指定内容的数量
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        column_index = headers.index(column_name) + 1
        pattern = re.compile(patternStr, re.IGNORECASE)
        total_count = 0
        no_track_count = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            cell_value = row[column_index - 1]
            if cell_value is not None:
                total_count += 1
                if pattern.search(str(cell_value)):
                    no_track_count += 1
        return total_count, no_track_count
    except Exception as e:
        print(f"发生错误: {e}")
        return 0, 0


def count_distribution_and_no_track(file_path, key_column, courier_column=RowName.Courier):
    """
    通用函数，统计指定列的分布情况及其对应 "无轨迹" 的数量。
    :param file_path: Excel 文件路径
    :param key_column: 需要统计的列名
    :param courier_column: 快递列名
    :return: 各值的总数和 "无轨迹" 数量的 Counter 对象
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if key_column not in headers or courier_column not in headers:
            raise ValueError(f"列名 '{key_column}' 或 '{courier_column}' 不存在！")
        key_index = headers.index(key_column) + 1
        courier_index = headers.index(courier_column) + 1
        pattern = re.compile(Pattern.no_track, re.IGNORECASE)
        key_counter = Counter()
        key_no_track_counter = Counter()
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            key_value = row[key_index - 1]
            courier_status = row[courier_index - 1]
            if key_value is not None:
                key_counter[key_value] += 1
                if courier_status is not None and pattern.search(str(courier_status)):
                    key_no_track_counter[key_value] += 1
        return key_counter, key_no_track_counter
    except Exception as e:
        print(f"发生错误: {e}")
        return Counter(), Counter()


def analyze_time_segments(file_path, time_column, courier_column):
    """
    按时间段（每3分钟为一段，忽略秒进行判断）统计总数和 "无轨迹" 的数量。
    输出时包括秒显示。
    """
    try:
        # 加载 Excel 文件
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if time_column not in headers or courier_column not in headers:
            raise ValueError(f"列名 '{time_column}' 或 '{courier_column}' 不存在！")

        # 获取列索引
        time_index = headers.index(time_column) + 1
        courier_index = headers.index(courier_column) + 1

        # 正则表达式匹配 "无轨迹"
        pattern = re.compile(Pattern.no_track, re.IGNORECASE)

        # 读取并解析数据
        data = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            order_time = row[time_index - 1]
            courier_status = row[courier_index - 1]
            if order_time is not None and isinstance(order_time, str):
                try:
                    # 解析时间格式为 "2025-01-22 23:11:43"
                    order_time = datetime.strptime(order_time, "%Y-%m-%d %H:%M:%S")
                    order_time_without_seconds = order_time.replace(second=0)
                    data.append((order_time, order_time_without_seconds, courier_status))
                except ValueError:
                    continue

        # 按时间段归类
        data.sort(key=lambda x: x[1])  # 按无秒时间排序
        time_segments = defaultdict(list)
        if data:
            base_time = data[0][1]  # 使用无秒时间作为基准
            current_segment = []
            for full_time, order_time_without_seconds, courier_status in data:
                if (order_time_without_seconds - base_time).total_seconds() <= 180:  # 3分钟内
                    current_segment.append((full_time, courier_status))
                else:
                    time_segments[base_time].extend(current_segment)
                    base_time = order_time_without_seconds
                    current_segment = [(full_time, courier_status)]
            if current_segment:
                time_segments[base_time].extend(current_segment)

        # 统计每个时间段的总数和无轨迹数量
        segment_statistics = {}
        for segment_start, entries in time_segments.items():
            total_count = len(entries)
            no_track_count = sum(
                1 for _, courier_status in entries if courier_status is not None and pattern.match(str(courier_status)))
            segment_statistics[segment_start] = {
                "total_count": total_count,
                "no_track_count": no_track_count,
                "entries": entries,
            }

        return segment_statistics

    except Exception as e:
        print(f"发生错误: {e}")
        return {}


def check_and_add_courier_column(file_path, courier_column=RowName.Courier):
    """
    检查 Excel 文件是否存在 '快递' 列，如果没有，则在最后一列添加该列。

    :param file_path: Excel 文件路径
    :param courier_column: 快递列名，默认为 'Courier/快递'
    :return: None
    """
    try:
        # 加载 Excel 文件
        data = pd.read_excel(file_path, engine='openpyxl')
        # 判断是否存在 '快递' 列
        if courier_column not in data.columns:
            # 如果没有 '快递' 列，则在最后一列添加该列
            data[courier_column] = ""  # 默认为空值，可以根据需求填充其他默认值
            # 保存修改后的文件
            data.to_excel(file_path, index=False, engine='openpyxl')
        #     print(f"列 '{courier_column}' 已添加到文件中，并保存。")
        # else:
        #     print(f"列 '{courier_column}' 已存在，无需添加。")
    except Exception as e:
        print(f"发生错误: {e}")


def get_days_difference(file_path, column_name=RowName.OutboundTime):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        # 获取列索引
        column_index = headers.index(column_name) + 1
        # 获取第一条数据
        first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
        if not first_row_value:
            raise ValueError(f"'{column_name}' 列的第一条数据为空！")
        # 解析日期
        outbound_time = datetime.strptime(first_row_value, "%Y-%m-%d %H:%M:%S")
        # 格式化为 "%Y/%m/%d" 格式
        formatted_date = outbound_time.strftime("%Y/%m/%d")
        return formatted_date
    except Exception as e:
        print(f"发生错误: {e}")
        return None

def generate_distribution_report(distribution, no_track_distribution, data_map, data_map_key):
    """
    通用的分布报告生成函数
    :param distribution: 订单分布字典
    :param no_track_distribution: 无轨迹分布字典
    :param data_map:
    :param data_map_key: 用于存储到 `data_map` 的 key（例如 `CellKey.warehouse_condition` 或 `CellKey.store_condition`）
    :return: 生成的分布报告文本
    """
    report_text = ""
    report_text2 = ""
    lowest_swl = 101  # 初始化为一个比 100 大的值，用于比较
    lowest_entity = ""  # 保存最低上网率的实体信息

    # 遍历分布数据
    for entity, count in distribution.items():
        no_track_count = no_track_distribution.get(entity, 0)
        swl = round2(100 - ((int(no_track_count) / int(count)) * 100))
        strs = f"\n{entity}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{swl}%"
        strs2 = f"\n{entity}：({count},{swl}%)"
        report_text += strs
        report_text2 += strs2

        # 判断是否是最低的上网率
        if swl < lowest_swl:
            lowest_swl = swl
            lowest_entity = strs

    data_map[data_map_key] = report_text  # 将结果存储到 data_map 中
    return report_text, lowest_entity, report_text2


def generate_distribution_report2(distribution, no_track_distribution, data_map, data_map_key, interval_time):
    """
    通用的分布报告生成函数，统计订单分布、无轨迹订单、计算上网率，并找出最低上网率的所有实体
    :param distribution: 订单分布字典
    :param no_track_distribution: 无轨迹分布字典
    :param data_map:
    :param data_map_key: 用于存储到 `data_map` 的 key（例如 `CellKey.warehouse_condition` 或 `CellKey.store_condition`）
    :return: 生成的分布报告文本, 最低上网率的所有实体信息, 精简版报告文本
    """
    report_text = ""
    report_text2 = ""
    lowest_swl = 101  # 初始化为比100大的值
    lowest_entities = {}  # 存储多个最低上网率的实体信息

    # 遍历分布数据
    for entity, count in distribution.items():
        no_track_count = no_track_distribution.get(entity, 0)
        swl = round2(100 - ((int(no_track_count) / int(count)) * 100))  # 计算上网率

        # 生成报告内容
        strs = f"\n{entity}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{swl}%"
        strs2 = f"\n{entity}：({count},{swl}%)"
        report_text += strs
        report_text2 += strs2

        # 更新最低上网率的实体
        if swl < lowest_swl:
            lowest_swl = swl
            lowest_entities.clear()  # 清空数据
            lowest_entities[entity] = {"entity": entity, "count": count, "no_track_count": no_track_count, "swl": swl,
                                       "strs": strs}
        elif swl == lowest_swl:
            lowest_entities[entity] = {"entity": entity, "count": count, "no_track_count": no_track_count, "swl": swl,
                                       "strs": strs}

    resultList = []
    for key, value in lowest_entities.items():
        no_track_counts = value["no_track_count"]
        strss = value["strs"]
        if (interval_time >= 3):
            resultList.append(strss)
        else:
            if (no_track_counts >= 4):
                resultList.append(strss)

    data_map[data_map_key] = report_text  # 将结果存储到 data_map
    return report_text, resultList, report_text2


def go(analyse_obj, xlsx_path):
    if analyse_obj is None:
        analyse_obj = input("请输跟踪对象（zbw/sanrio/xyl/mz_xsd/md_fc/mx_dg）：")

    if analyse_obj != ClientConstants.zbw \
            and analyse_obj != ClientConstants.sanrio \
            and analyse_obj != ClientConstants.xyl \
            and analyse_obj != ClientConstants.mz_xsd \
            and analyse_obj != ClientConstants.md_fc \
            and analyse_obj != ClientConstants.mx_dg:
        raise ValueError(f"{analyse_obj} 未定义")

    if xlsx_path is None:
        xlsx_path = input("请输入文件的绝对路径：")

    check_and_add_courier_column(xlsx_path)

    irregular_number_map = find_irregular_tracking_numbers(xlsx_path)
    irregular_number_list = []
    if (len(irregular_number_map) > 0):
        irregular_number_list = list(irregular_number_map.keys())
        print(f"存在无效的物流跟踪号：{irregular_number_list}")
        update_courier_status(xlsx_path, irregular_number_map)

    results = extract_and_process_data(xlsx_path, RowName.Courier, 100)

    # no_tracking_count = len(results[CourierStateMapKey.not_yet_results]) + len(
    #     results[CourierStateMapKey.pre_ship_results]) + len(results[CourierStateMapKey.no_tracking_results])
    # tracking_count = len(results[CourierStateMapKey.unpaid_results]) + len(
    #     results[CourierStateMapKey.delivered_results]) + len(results[CourierStateMapKey.tracking_results])
    # print(f"没有轨迹数： {no_tracking_count} 条，有轨迹数： {tracking_count} 条")
    # print(f"\nunpaid数： {len(results[CourierStateMapKey.unpaid_results])} 条")
    # print(f"\nnot_yet数： {len(results[CourierStateMapKey.not_yet_results])} 条")
    # print(f"\npre_ship数： {len(results[CourierStateMapKey.pre_ship_results])} 条")
    # print(f"\ndelivered数： {len(results[CourierStateMapKey.delivered_results])} 条")

    update_courier_status(xlsx_path, results[CourierStateMapKey.not_yet_map])
    update_courier_status(xlsx_path, results[CourierStateMapKey.pre_ship_map])
    update_courier_status(xlsx_path, results[CourierStateMapKey.unpaid_map])
    update_courier_status(xlsx_path, results[CourierStateMapKey.delivered_map])
    update_courier_status(xlsx_path, results[CourierStateMapKey.no_tracking_map])
    update_courier_status(xlsx_path, results[CourierStateMapKey.tracking_map])

    # 数据map
    data_map = {}

    text = ""

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ck_time = get_days_difference(xlsx_path)
    gz_time = getYmd()
    interval_time = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(ck_time, "%Y/%m/%d")).days
    is_usweekend = is_us_weekend(ck_time)
    text += "\n----------------------时间----------------------"
    text += f"\n更新时间: {current_time}"
    text += f"\n出库日期：{ck_time}"
    text += f"\n跟踪日期：{gz_time}"
    text += f"\n间隔时间：{interval_time}"
    data_map[CellKey.update_time] = current_time

    text += "\n----------------------SKU分布----------------------"
    sku_distribution, sku_no_track_distribution = count_distribution_and_no_track(xlsx_path, key_column=RowName.SKU)
    sku_text, lowest_sku, sku_text2 = generate_distribution_report2(
        sku_distribution, sku_no_track_distribution, data_map, CellKey.sku_condition, interval_time
    )
    text += sku_text

    output_file = os.path.splitext(xlsx_path)[0] + "_去重.xlsx"
    # 同一单会有多个sku，多个sku会生成多行数据，分析sku的时候不能去重，其它的需要去重
    remove_duplicates_by_column(xlsx_path, output_file, RowName.Tracking_No)

    total_count, no_track_count = count_pattern_state(output_file, RowName.Courier, Pattern.no_track)
    total_count2, delivered_count = count_pattern_state(output_file, RowName.Courier, Pattern.delivered)
    total_count3, unpaid_count = count_pattern_state(output_file, RowName.Courier, Pattern.unpaid)
    swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))
    wswl = round2(100 - swl)
    qsl = round2((int(delivered_count) / int(total_count)) * 100)
    text += "\n----------------------概览----------------------"
    text += f"\n订单总数：{total_count}"
    text += f"\n邮资未付数：{unpaid_count}"
    text += f"\n签收数：{delivered_count}"
    text += f"\n签收率：{qsl}%"
    text += f"\n未上网数：{no_track_count}"
    text += f"\n上网率：{swl}%"
    text += f"\n未上网率：{wswl}%"
    data_map[CellKey.order_count] = total_count
    data_map[CellKey.unpaid_count] = unpaid_count
    data_map[CellKey.delivered_counts] = delivered_count
    data_map[CellKey.delivered_percent] = qsl
    data_map[CellKey.no_track_number] = no_track_count
    data_map[CellKey.track_percent] = swl
    data_map[CellKey.no_track_percent] = wswl

    text += "\n----------------------仓库分布----------------------"
    warehouse_distribution, warehouse_no_track = count_distribution_and_no_track(
        output_file, key_column=RowName.Warehouse)
    warehouse_text, lowest_warehouse, warehouse_text2 = generate_distribution_report(
        warehouse_distribution, warehouse_no_track, data_map, CellKey.warehouse_condition
    )
    text += warehouse_text

    text += "\n----------------------店铺分布----------------------"
    store_distribution, store_no_track_distribution = count_distribution_and_no_track(
        output_file, key_column=RowName.Client)
    store_text, lowest_store, store_text2 = generate_distribution_report(
        store_distribution, store_no_track_distribution, data_map, CellKey.store_condition
    )
    text += store_text

    text += "\n----------------------物流渠道分布----------------------"
    shipping_service_distribution, shipping_service_no_track_distribution = count_distribution_and_no_track(
        output_file, key_column=RowName.ShippingService)
    shipping_service_text, lowest_shipping_service, shipping_service_text2 = generate_distribution_report(
        shipping_service_distribution, shipping_service_no_track_distribution, data_map,
        CellKey.shipping_service_condition
    )
    text += shipping_service_text

    text += "\n----------------------时间段分布----------------------"
    time_segment_analysis = analyze_time_segments(
        output_file, time_column=RowName.CreationWaveTime, courier_column=RowName.Courier)
    time_segment_text = ""
    lowest_segment = ""  # 保存上网率最低的时间段
    lowest_swl = 101  # 初始化为比 100 大的值
    for segment_start, stats in time_segment_analysis.items():
        segment_end = segment_start + timedelta(minutes=3)
        total_count_temp = stats["total_count"]
        no_track_count = stats["no_track_count"]
        segmentswl = round2(100 - ((int(no_track_count) / int(total_count_temp)) * 100))
        strs = f"\n{segment_start.strftime('%y-%m-%d %H:%M')} - {segment_end.strftime('%y-%m-%d %H:%M')}： 订单总数：{total_count_temp}；无轨迹数：{no_track_count}；上网率：{segmentswl}%"
        text += strs
        time_segment_text += strs
        # 判断是否是最低的上网率
        if segmentswl < lowest_swl:
            lowest_swl = segmentswl
            lowest_segment = strs
    data_map[CellKey.time_segment_condition] = time_segment_text

    lowest_txt = ""
    lowest_txt += f"\n"
    if (len(lowest_sku) > 0):
        lowest_txt += f"\n最低上网率的 SKU："
        for item in lowest_sku:
            lowest_txt += item
    lowest_txt += f"\n最低上网率的 仓库：{lowest_warehouse}"
    lowest_txt += f"\n最低上网率的 商店：{lowest_store}"
    lowest_txt += f"\n最低上网率的 时间段：{lowest_segment}"
    lowest_txt += f"\n最低上网率的 物流渠道：{lowest_shipping_service}"

    sum_up_text = ""

    actual_interval = ""
    if (is_usweekend == 6):  # 6是中国周日，美国周六
        sum_up_text += f"美国时间：周六（和中国相差13-16个小时）"
        sum_up_text += f"\n"
        actual_interval = "（-2）"
    elif (is_usweekend == 0):  # 0是中国周一，美国周日
        sum_up_text += f"美国时间：周日（相差13-16个小时）"
        sum_up_text += f"\n"
        actual_interval = "（-1）"
    else:
        actual_interval = ""

    if (len(irregular_number_list) > 0):
        sum_up_text += f"存在不规则单号：{irregular_number_list}"
        sum_up_text += f"\n"

    swl_flag = False
    qsl_flag = False

    # 如果三天后的上网率没有99%以上，那么就严重有问题；隔天应该要 》= 三分之一，隔两天应该要有》=75
    if (interval_time == 1):
        if (swl < 30):
            sum_up_text += f"\n☁️注意：间隔第{interval_time}{actual_interval}天，上网率为{swl}%，未达30%，建议跟进！"
            swl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，上网率为{swl}%，上网率优秀"
    elif (interval_time == 2):
        if (swl < 70):
            sum_up_text += f"\n🌧️异常：间隔第{interval_time}{actual_interval}天，上网率为{swl}%，未达75%，建议分析数据尝试定位问题！"
            swl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，上网率为{swl}%，上网率优秀"
    else:  # 间隔时间 >= 3天
        if (swl < 97):
            sum_up_text += f"\n❄️⛈️🌀⚠️🚨警报：间隔第{interval_time}{actual_interval}天，上网率为{swl}%，未达97%，分析数据反馈问题！"
            swl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，上网率为{swl}%，上网率优秀"

    # 要持续监控一个星期才行，从出库开始计算，三天内没有签收的不正常，五天内签收没达到50%也不正常，7天内没到90也不正常
    if (interval_time >= 1 and interval_time <= 3):
        if (interval_time >= 2 and qsl == 0):
            sum_up_text += f"\n🚨警报：间隔第{interval_time}{actual_interval}天，签收率为0%，异常状态！"
            qsl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，继续跟进！"
    elif (interval_time > 3 and interval_time <= 5):
        if (qsl <= 30):
            sum_up_text += f"\n🚨警报：间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，异常状态！"
            qsl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，继续跟进！"
    elif (interval_time > 5 and interval_time <= 7):
        if (qsl <= 70):
            sum_up_text += f"\n🚨警报：间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，异常状态！"
            qsl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，继续跟进！"
    elif (interval_time > 7 and interval_time <= 9):
        if (qsl <= 90):
            sum_up_text += f"\n🚨警报：间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，异常状态！"
            qsl_flag = True
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，继续跟进！"
    else:
        if (qsl >= 98):
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，签收率优秀！"
        else:
            sum_up_text += f"\n☀️间隔第{interval_time}{actual_interval}天，签收率为{qsl}%，继续跟进！"

    text += "\n----------------------总结&建议----------------------"
    if (swl < 100):
        sum_up_text += lowest_txt
    text += f"\n{sum_up_text}"
    data_map[CellKey.sum_up] = sum_up_text

    if (swl_flag or qsl_flag):
        data_map[CellKey.exception] = "异常"
    else:
        data_map[CellKey.exception] = ""

    # 删除去重文件
    delete_file(output_file)
    # 数据打印
    # print(data_map)
    print(text)

    # 写入飞书在线文档
    tat = get_token()
    if analyse_obj == ClientConstants.zbw or analyse_obj == ClientConstants.sanrio or analyse_obj == ClientConstants.xyl:
        lists = f"({total_count},{swl}%)"
        lists += f"\n{warehouse_text2}"
        brief_sheet_value(tat, [lists], ck_time, gz_time, analyse_obj)
    else:
        lists = f"({total_count},{swl}%)"
        brief_sheet_value(tat, [lists], ck_time, gz_time, analyse_obj)

    if analyse_obj == ClientConstants.mz_xsd or \
            analyse_obj == ClientConstants.mx_dg or \
            analyse_obj == ClientConstants.md_fc:
        detail_sheet_value(tat, [
            data_map[CellKey.update_time],
            data_map[CellKey.order_count],
            data_map[CellKey.unpaid_count],
            data_map[CellKey.delivered_counts],
            data_map[CellKey.delivered_percent],
            data_map[CellKey.no_track_number],
            data_map[CellKey.track_percent],
            data_map[CellKey.no_track_percent],
            data_map[CellKey.warehouse_condition],
            data_map[CellKey.shipping_service_condition],
            data_map[CellKey.store_condition],
            # data_map[CellKey.sku_condition],
            data_map[CellKey.time_segment_condition],
            data_map[CellKey.sum_up],
            data_map[CellKey.exception],
        ], ck_time, analyse_obj)
    else:
        detail_sheet_value(tat, [
            data_map[CellKey.update_time],
            data_map[CellKey.order_count],
            data_map[CellKey.unpaid_count],
            data_map[CellKey.delivered_counts],
            data_map[CellKey.delivered_percent],
            data_map[CellKey.no_track_number],
            data_map[CellKey.track_percent],
            data_map[CellKey.no_track_percent],
            data_map[CellKey.warehouse_condition],
            data_map[CellKey.shipping_service_condition],
            data_map[CellKey.store_condition],
            data_map[CellKey.sku_condition],
            data_map[CellKey.time_segment_condition],
            data_map[CellKey.sum_up],
            data_map[CellKey.exception],
        ], ck_time, analyse_obj)


def automatic(dir_path, analyse_obj):
    for root, dirs, files in os.walk(dir_path):
        """
        root: 当前文件夹路径
        dirs: 当前文件夹下的子文件夹列表
        files: 当前文件夹下的文件列表
        """
        # print(f"当前文件夹: {root}")
        # print(f"子文件夹: {dirs}")
        # print(f"文件: {files}")
        # print("--------")
        pattern = r"^出库时间\d+_\d+\.xlsx$"  # 正则表达式
        for ele in files:
            if re.match(pattern, ele):
                xlsx_path = f"{root}/{ele}"
                print(f"匹配的文件: {xlsx_path}")
                try:
                    total_count, no_track_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.no_track)
                    total_count2, delivered_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.delivered)

                    if total_count == 0:
                        swl = 0  # 处理 total_count 为 0 的情况
                    else:
                        swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))

                    if total_count == 0:
                        qsl = 0  # 处理 total_count 为 0 的情况
                    else:
                        qsl = round2((int(delivered_count) / int(total_count)) * 100)

                    if swl < 99 or qsl < 98:
                        go(analyse_obj, xlsx_path)
                except ZeroDivisionError:
                    print(f"警告：{xlsx_path} 的 total_count 为 0，跳过计算。")
                    go(analyse_obj, xlsx_path)  # 仍然执行 go 但避免除零错误
                except Exception as e:
                    print(f"处理 {xlsx_path} 时发生错误: {e}")
                    go(analyse_obj, xlsx_path)


if __name__ == '__main__':
    # 手动
    go(None, None)
    # 自动
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/zbw", ClientConstants.zbw)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/zbw/2025.1", ClientConstants.zbw)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/zbw/2025.2", ClientConstants.zbw)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/sanrio", ClientConstants.sanrio)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/sanrio/2025.1", ClientConstants.sanrio)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/sanrio/2025.2", ClientConstants.sanrio)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/xyl", ClientConstants.xyl)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/xyl/2025.2", ClientConstants.xyl)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/mzxsd", ClientConstants.mz_xsd)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/mxdg", ClientConstants.mx_dg)
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/mdfc", ClientConstants.md_fc)
