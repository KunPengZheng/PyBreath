from datetime import datetime
import pandas as pd
import shutil
import os
from collections import defaultdict

from xinshili import utils
from xinshili.fs_utils_plus import insert_col_row, FsConstants, get_token, ClientMapConstants, ClientConstants, \
    MapFields, fs_col_to_index, value_range, values_batch_update
from xinshili.gjgz_plus333 import RowName
from openpyxl import load_workbook


def update_sales_data(file1_path, output_path):
    # 读取文件1和复制后的文件2
    df1 = pd.read_excel(file1_path).fillna("")
    df2 = pd.read_excel(output_path, sheet_name=None)

    # 👉 获取下单时间第一条有效数据并格式化为 "7月2日"
    order_time_str = df1.get("下单时间", "").astype(str).str.strip().replace("nan", "").values
    order_time = next((t for t in order_time_str if t), None)
    formatted_date = ""
    if order_time:
        try:
            dt = datetime.strptime(order_time, "%Y-%m-%d %H:%M:%S")
            formatted_date = f"{dt.month}月{dt.day}日"
            print(f"📅 首个下单时间对应日期为：{formatted_date}")
        except Exception as e:
            print(f"⚠️ 日期格式错误：{e}")

    # 产品总数要能转为数字以便汇总
    if RowName.Total_Of_Product in df1.columns:
        df1[RowName.Total_Of_Product] = pd.to_numeric(df1[RowName.Total_Of_Product], errors="coerce").fillna(0)

    # 店铺订单数（订单号去重）
    store_count_map = (
        df1.drop_duplicates(subset=[RowName.Store_Account, RowName.Order_Num])
        [RowName.Store_Account]
        .value_counts()
        .to_dict()
    )
    # 店铺订单数（订单号没有去重）
    # store_count_map = df1[RowName.Store_Account].value_counts().to_dict()
    sku_sum_map = df1.groupby(RowName.SKU)[RowName.Total_Of_Product].sum().to_dict()

    # 规则1: "31302413112" -> 乘 2，加到 "3130241311"
    target = "3130241311"
    target1 = "3130242931"

    key1 = "31302413112"
    if key1 in sku_sum_map:
        sku_sum_map[target] = sku_sum_map.get(target, 0) + (sku_sum_map.get(key1, 0) * 2)

    # 规则2: "3130242931+3130241311" -> 加到 "3130241311" 和 加到 "3130242931"
    key2 = "3130242931+3130241311"
    if key2 in sku_sum_map:
        sku_sum_map[target] = sku_sum_map.get(target, 0) + sku_sum_map.get(key2, 0)
        sku_sum_map[target1] = sku_sum_map.get(target1, 0) + sku_sum_map.get(key2, 0)

    updated_sheets = {}

    # --- 店铺销量表 ---
    if RowName.Store_Sales in df2:
        sheet = df2[RowName.Store_Sales].copy()
        if RowName.Store_Name in sheet.columns and RowName.Order_Sales in sheet.columns:
            sheet[RowName.Order_Sales] = sheet[RowName.Store_Name].apply(
                lambda x: store_count_map.get(str(x).strip(), 0))
        updated_sheets[RowName.Store_Sales] = sheet

    # --- 销量更新表 ---
    if RowName.Sales_Update in df2:
        sheet = df2[RowName.Sales_Update].copy()
        if RowName.SKU in sheet.columns and RowName.Quantity in sheet.columns:
            sheet[RowName.Quantity] = sheet[RowName.SKU].apply(lambda x: sku_sum_map.get(str(x).strip(), 0))
        updated_sheets[RowName.Sales_Update] = sheet

    # --- 库存更新表（直接复制） ---
    if RowName.Inventory_Update in df2:
        updated_sheets[RowName.Inventory_Update] = df2[RowName.Inventory_Update].copy()

    # 写入更新后的 sheet
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df_sheet in updated_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✅ 已更新并保存到：{output_path}")
    return formatted_date


def merge_excels_in_folder(folder_path, output_path):
    utils.delete_file(output_path)

    all_files = [f for f in os.listdir(folder_path) if f.endswith((".xlsx", ".xls"))]
    combined_df = pd.DataFrame()
    header_saved = False  # 只保留第一个文件的列头

    for filename in all_files:
        file_path = os.path.join(folder_path, filename)
        try:
            # 始终使用 header=0 保证列头是从文件中第一行读的
            df = pd.read_excel(file_path, header=0)

            if not header_saved:
                combined_df = pd.concat([combined_df, df], ignore_index=True)
                header_saved = True
            else:
                combined_df = pd.concat([combined_df, df], ignore_index=True)  # 直接合并，无需 iloc[1:]

            print(f"✅ 已处理文件：{filename}")
        except Exception as e:
            print(f"❌ 处理文件 {filename} 出错：{e}")

    # 保存合并结果
    combined_df.to_excel(output_path, index=False)
    print(f"\n✅ 合并完成，保存路径：{output_path}")


def update_available_inventory(file1_path, file2_path, output_path):
    # 读取文件1
    df1 = pd.read_excel(file1_path)
    df1.fillna('', inplace=True)

    # 构建 SKU → 可用库存总和 map
    sku_inventory_map = defaultdict(float)
    for _, row in df1.iterrows():
        sku = row.get(RowName.SKU, "").strip()
        inv = row.get(RowName.Available_Inventory, "")
        try:
            inv_val = float(inv)
        except:
            inv_val = 0
        if sku:
            sku_inventory_map[sku] += inv_val

    # 读取文件2所有 Sheet
    df2_sheets = pd.read_excel(file2_path, sheet_name=None)

    # 修改库存更新 Sheet
    if RowName.Inventory_Update in df2_sheets:
        df_kc = df2_sheets[RowName.Inventory_Update].copy()
        df_kc.fillna('', inplace=True)
        if RowName.SKU in df_kc.columns and RowName.Available_Quantity in df_kc.columns:
            df_kc[RowName.Available_Quantity] = df_kc[RowName.SKU].apply(
                lambda x: sku_inventory_map.get(str(x).strip(), 0))
        df2_sheets[RowName.Inventory_Update] = df_kc
    else:
        print("⚠️ 文件2中未找到 sheet：库存更新，已跳过更新")

    # 写入所有 sheet（保留原有的）
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df_sheet in df2_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✅ 已更新“库存更新”并保存至：{output_path}")


def update_total_inventory(file1_path, file2_path, output_path):
    # 读取文件1，填充空值
    df1 = pd.read_excel(file1_path)

    # 将库存数值列转换为 float
    df1["Available Inventory/可用库存"] = pd.to_numeric(df1.get("Available Inventory/可用库存", 0),
                                                        errors="coerce").fillna(0)
    df1["In-transit inventory/在途库存"] = pd.to_numeric(df1.get("In-transit inventory/在途库存", 0),
                                                         errors="coerce").fillna(0)

    # 构建 SKU → 可用库存/在途库存 映射表
    available_map = defaultdict(float)
    transit_map = defaultdict(float)

    for _, row in df1.iterrows():
        sku = row.get("SKU", "").strip()
        if sku:
            available_map[sku] += row["Available Inventory/可用库存"]
            transit_map[sku] += row["In-transit inventory/在途库存"]

    # 读取文件2的所有 sheet
    df2_sheets = pd.read_excel(file2_path, sheet_name=None)

    # 处理“库存更新”sheet
    sheet_name_target = "库存更新"
    if sheet_name_target in df2_sheets:
        df_kc = df2_sheets[sheet_name_target].copy()
        df_kc.fillna('', inplace=True)

        if "SKU" in df_kc.columns:
            df_kc["库存(全部)"] = df_kc["SKU"].apply(lambda x: available_map.get(str(x).strip(), 0))
            df_kc["在途（全部）"] = df_kc["SKU"].apply(lambda x: transit_map.get(str(x).strip(), 0))
        else:
            print("⚠️ '库存更新' 表中缺少 SKU 列，未更新")

        # 更新 sheet 回 sheet dict
        df2_sheets[sheet_name_target] = df_kc
    else:
        print("⚠️ 文件2中未找到 sheet：库存更新，已跳过更新")

    # 写入所有 sheet 到 output_path
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet, df_sheet in df2_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet, index=False)

    print(f"✅ 已更新“库存更新”并保存至：{output_path}")


def collect_sku_values(file, sheet_name, sku_name="SKU", count_name="个数"):
    wb = load_workbook(file, data_only=True)
    ws = wb[sheet_name]

    result_map = {}

    # === 第一步：找到列号（通过第2行表头匹配） ===
    header_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    header_index = {name: idx + 1 for idx, name in enumerate(header_row) if name}

    if sku_name not in header_index or count_name not in header_index:
        raise ValueError(f"表头中没有找到指定列: {sku_name}, {count_name}")

    sku_col = header_index[sku_name]
    count_col = header_index[count_name]

    # === 第二步：遍历数据行（第3行开始） ===
    for row in ws.iter_rows(min_row=3):
        # 用第一列判断整行背景色
        cell = row[0]
        fill = cell.fill
        bg_color = None
        if fill and fill.fgColor and fill.fgColor.type == "rgb":
            bg_color = fill.fgColor.rgb

        # 跳过指定颜色的行
        if bg_color in ("FF34C724", "FFFBBFBC"):
            continue

        sku = str(row[sku_col - 1].value)
        count = row[count_col - 1].value

        if sku and isinstance(count, (int, float)):
            result_map[sku] = result_map.get(sku, 0) + count

    return result_map


def update_shipping_inventory(src_file_path, xlsx_file_path, output_path):
    if src_file_path.endswith("dszs_inventory.csv"):
        # 读取 CSV 文件（文件1）
        df1 = pd.read_csv(src_file_path).fillna("")

        # 转换为数值，便于求和
        df1[RowName.Sea_transportation] = pd.to_numeric(df1.get(RowName.Sea_transportation, 0), errors="coerce").fillna(
            0)
        df1[RowName.Air_transportation] = pd.to_numeric(df1.get(RowName.Air_transportation, 0), errors="coerce").fillna(
            0)

        # 构建 SKU → 海/空 运在途数量 Map
        sea_map = defaultdict(float)
        air_map = defaultdict(float)

        for _, row in df1.iterrows():
            sku = row.get(RowName.SKU, "").strip()
            if sku:
                sea_map[sku] += row[RowName.Sea_transportation]
                air_map[sku] += row[RowName.Air_transportation]
    else:
        air_map = collect_sku_values(src_file_path, "空运头程计划US", sku_name="SKU", count_name="个数")
        sea_map = collect_sku_values(src_file_path, "海运头程计划US", sku_name="SKU", count_name="个数")

    # 读取 Excel 所有工作表
    xls = pd.read_excel(xlsx_file_path, sheet_name=None)
    updated_sheets = {}

    for sheet_name, df_sheet in xls.items():
        df_sheet = df_sheet.fillna("")
        if sheet_name == RowName.Inventory_Update and RowName.SKU in df_sheet.columns:
            # 更新库存更新表的海运在途和空运在途
            df_sheet[RowName.Sea_transportation] = df_sheet[RowName.SKU].apply(lambda x: sea_map.get(str(x).strip(), 0))
            df_sheet[RowName.Air_transportation] = df_sheet[RowName.SKU].apply(lambda x: air_map.get(str(x).strip(), 0))
        updated_sheets[sheet_name] = df_sheet

    # 写入到新文件，保留所有 sheet
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df_sheet in updated_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"✅ 所有工作表已写入，库存更新已处理并保存至：{output_path}")


# def get_range_column_data(file_path, sheet_name, column_name, start_row, end_row):
#     """
#     获取 Excel 指定 sheet 的某一列数据（包含列头），支持行区间选择。
#
#     :param file_path: Excel 文件路径
#     :param sheet_name: 工作表名称
#     :param column_name: 要读取的列名
#     :param start_row: 起始行号（从 0 开始，0 表示包含列头）
#     :param end_row: 结束行号（包含，None 表示到末尾）
#     :return: 列表形式返回数据
#     """
#     try:
#         df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=object).fillna("")
#
#         if column_name not in df.columns:
#             raise ValueError(f"❌ 指定列名 '{column_name}' 不存在于工作表 '{sheet_name}' 中。")
#
#         # 列数据（包含列头）
#         data_with_header = [column_name] + df[column_name].tolist()
#
#         # 切片区间（包含列头），行号从 0 开始，包含 end_row
#         start_idx = max(0, start_row - 1)
#         end_idx = end_row
#
#         return data_with_header[start_idx:end_idx]
#
#     except Exception as e:
#         print(f"⚠️ 获取失败: {e}")
#         return []


def get_range_column_data(file_path, sheet_name, column_name, start_row, end_row):
    """
    获取 Excel 指定 sheet 的某一列数据（包含列头），支持行区间选择。
    返回二维数组，每个元素是一个子数组，每个子数组只有一个值。

    :param file_path: Excel 文件路径
    :param sheet_name: 工作表名称
    :param column_name: 要读取的列名
    :param start_row: 起始行号（从 0 开始，0 表示包含列头）
    :param end_row: 结束行号（包含，None 表示到末尾）
    :return: 列表形式返回二维数据
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=object)

        if column_name not in df.columns:
            raise ValueError(f"❌ 指定列名 '{column_name}' 不存在于工作表 '{sheet_name}' 中。")

        # 列数据（包含列头）
        data_with_header = [column_name] + df[column_name].tolist()

        # 切片区间
        start_idx = max(0, start_row - 1)
        end_idx = end_row

        result = data_with_header[start_idx:end_idx]
        return [[item] for item in result]

    except Exception as e:
        print(f"⚠️ 获取失败: {e}")
        return []


def get_range_column_row_data(file_path, sheet_name, start_row, end_row, start_col, end_col):
    """
    获取 Excel 指定 sheet 中指定行列区域的数据，返回二维数组形式。

    :param file_path: Excel 文件路径
    :param sheet_name: 表名（Sheet 名）
    :param start_row: 起始行号（从 0 开始，包含）
    :param end_row: 结束行号（从 0 开始，包含）
    :param start_col: 起始列号（从 0 开始，包含）
    :param end_col: 结束列号（从 0 开始，包含）
    :return: 二维数组（嵌套列表）
    """
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df.fillna('', inplace=True)

    # 将 Excel 行号（从 1 开始）转换为 Pandas 行索引（从 0 开始）
    pandas_start_row = start_row - 1
    pandas_end_row = end_row - 1

    # 将列名（如 "C"）转换为列索引（从 0 开始）
    start_col_num = fs_col_to_index(start_col) - 1
    end_col_num = fs_col_to_index(end_col) - 1

    region = df.iloc[pandas_start_row:pandas_end_row + 1, start_col_num:end_col_num + 1]
    return region.values.tolist()


def xyl_fs(formatted_date, template_copy_path):
    xyl_sku_zjhz_start_index = 2
    xyl_sku_zjhz_end_index = 124

    xyl_sku_ejyxhz_start_index = 126
    xyl_sku_ejyxhz_end_index = 239

    # 获取数据数组
    xyl_sku_zjhz_arr = get_range_column_data(template_copy_path, RowName.Sales_Update, RowName.Quantity,
                                             xyl_sku_zjhz_start_index, xyl_sku_zjhz_end_index)
    # 数组数组头部插入日期数据
    xyl_sku_zjhz_arr.insert(0, [formatted_date])
    # 数组数组头部插入日期数据
    xyl_sku_zjhz_arr.append([{"type": "formula", "text": f"=SUM(L2:L{len(xyl_sku_zjhz_arr) - 1})"}])

    xyl_sku_ejyxhz_arr = get_range_column_data(template_copy_path, RowName.Sales_Update, RowName.Quantity,
                                               xyl_sku_ejyxhz_start_index, xyl_sku_ejyxhz_end_index)
    xyl_sku_ejyxhz_arr.insert(0, [formatted_date])
    xyl_sku_ejyxhz_arr.append([{"type": "formula", "text": f"=SUM(L2:L{len(xyl_sku_ejyxhz_arr) - 1})"}])

    xyl_store_arr = get_range_column_data(template_copy_path, RowName.Store_Sales, RowName.Order_Sales, 2, 51)
    xyl_store_arr.insert(0, [formatted_date])
    xyl_store_arr.append([{"type": "formula", "text": f"=SUM(E2:E{len(xyl_store_arr)})"}])

    print(len(xyl_sku_zjhz_arr), xyl_sku_zjhz_arr)
    print(len(xyl_sku_ejyxhz_arr), xyl_sku_ejyxhz_arr)
    print(len(xyl_store_arr), xyl_store_arr)

    xyl_sku_inventory_zjhz_arr = get_range_column_row_data(template_copy_path, RowName.Inventory_Update,
                                                           xyl_sku_zjhz_start_index, xyl_sku_zjhz_end_index, "C", "E")
    xyl_sku_inventory_ejyxhz_arr = get_range_column_row_data(template_copy_path, RowName.Inventory_Update,
                                                             xyl_sku_ejyxhz_start_index, xyl_sku_ejyxhz_end_index, "C",
                                                             "E")
    print(len(xyl_sku_inventory_zjhz_arr), xyl_sku_inventory_zjhz_arr)
    print(len(xyl_sku_inventory_ejyxhz_arr), xyl_sku_inventory_ejyxhz_arr)

    token = get_token()
    startIndex = fs_col_to_index("K")
    endIndex = fs_col_to_index("L")

    insert_col_row(token, FsConstants.xyl_sales_repertory_token,
                   ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_sku_zjhz],
                   startIndex, endIndex, FsConstants.COLUMNS, FsConstants.AFTER)
    insert_col_row(token, FsConstants.xyl_sales_repertory_token,
                   ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_sku_ejyxhz],
                   startIndex, endIndex, FsConstants.COLUMNS, FsConstants.AFTER)
    insert_col_row(token, FsConstants.xyl_sales_repertory_token,
                   ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_store],
                   fs_col_to_index("D"), fs_col_to_index("E"), FsConstants.COLUMNS, FsConstants.AFTER)

    values_batch_update(token, FsConstants.xyl_sales_repertory_token,
                        post_data={
                            "valueRanges": [
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_sku_zjhz]}!L1:L{len(xyl_sku_zjhz_arr)}",
                                    "values": xyl_sku_zjhz_arr
                                },
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_sku_zjhz]}!D2:F{len(xyl_sku_inventory_zjhz_arr) + 1}",
                                    "values": xyl_sku_inventory_zjhz_arr
                                }
                            ]
                        })

    values_batch_update(token, FsConstants.xyl_sales_repertory_token,
                        post_data={
                            "valueRanges": [
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_sku_ejyxhz]}!L1:L{len(xyl_sku_ejyxhz_arr)}",
                                    "values": xyl_sku_ejyxhz_arr
                                },
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_sku_ejyxhz]}!D2:F{len(xyl_sku_inventory_ejyxhz_arr) + 1}",
                                    "values": xyl_sku_inventory_ejyxhz_arr
                                }
                            ]
                        })

    value_range(token, FsConstants.xyl_sales_repertory_token,
                ClientMapConstants[ClientConstants.xyl_sales_repertory][MapFields.xyl_store],
                f"E1:E{len(xyl_store_arr) + 1}", xyl_store_arr)


def sanrio_fs(formatted_date, template_copy_path):
    sanrio_sku_wjshypl_arr = get_range_column_data(template_copy_path, RowName.Sales_Update, RowName.Quantity, 2, 55)
    sanrio_sku_wjshypl_arr.insert(0, [formatted_date])
    sanrio_sku_wjshypl_arr.append([{"type": "formula", "text": f"=SUM(H2:H55)"}])

    sanrio_sku_wjl_arr = get_range_column_data(template_copy_path, RowName.Sales_Update, RowName.Quantity, 57, 193)
    sanrio_sku_wjl_arr.insert(0, [formatted_date])
    sanrio_sku_wjl_arr.append([{"type": "formula", "text": f"=SUM(H2:H124)"}])

    sanrio_sku_snb_arr = get_range_column_data(template_copy_path, RowName.Sales_Update, RowName.Quantity, 195, 262)
    sanrio_sku_snb_arr.insert(0, [formatted_date])
    sanrio_sku_snb_arr.append([{"type": "formula", "text": f"=SUM(H2:H69)"}])

    sanrio_sku_ph_arr = get_range_column_data(template_copy_path, RowName.Sales_Update, RowName.Quantity, 264, 320)
    sanrio_sku_ph_arr.insert(0, [formatted_date])
    sanrio_sku_ph_arr.append([{"type": "formula", "text": f"=SUM(H2:H58)"}])

    sanrio_sku_dpxl_arr = get_range_column_data(template_copy_path, RowName.Store_Sales, RowName.Order_Sales, 2, 12)
    sanrio_sku_dpxl_arr.insert(0, [formatted_date])
    sanrio_sku_dpxl_arr.append([{"type": "formula", "text": f"=SUM(D2:D12)"}])

    sanrio_sku_inventory_wjshypl_arr = get_range_column_row_data(template_copy_path, RowName.Inventory_Update, 2, 55,
                                                                 "C", "D")
    sanrio_sku_inventory_wjl_arr = get_range_column_row_data(template_copy_path, RowName.Inventory_Update, 57, 193,
                                                             "C", "D")
    sanrio_sku_inventory_snb_arr = get_range_column_row_data(template_copy_path, RowName.Inventory_Update, 195, 262,
                                                             "C", "D")
    sanrio_sku_inventory_ph_arr = get_range_column_row_data(template_copy_path, RowName.Inventory_Update, 264, 320,
                                                            "C", "D")

    token = get_token()
    startIndex = fs_col_to_index("G")
    endIndex = fs_col_to_index("H")

    insert_col_row(token, FsConstants.sanrio_sales_repertory_token,
                   ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_wjshypl],
                   startIndex, endIndex, FsConstants.COLUMNS, FsConstants.AFTER)
    insert_col_row(token, FsConstants.sanrio_sales_repertory_token,
                   ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_wjl],
                   startIndex, endIndex, FsConstants.COLUMNS, FsConstants.AFTER)
    insert_col_row(token, FsConstants.sanrio_sales_repertory_token,
                   ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_snb],
                   startIndex, endIndex, FsConstants.COLUMNS, FsConstants.AFTER)
    insert_col_row(token, FsConstants.sanrio_sales_repertory_token,
                   ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_ph],
                   startIndex, endIndex, FsConstants.COLUMNS, FsConstants.AFTER)
    insert_col_row(token, FsConstants.sanrio_sales_repertory_token,
                   ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_dpxl],
                   fs_col_to_index("C"), fs_col_to_index("D"), FsConstants.COLUMNS, FsConstants.AFTER)

    values_batch_update(token, FsConstants.sanrio_sales_repertory_token,
                        post_data={
                            "valueRanges": [
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_wjshypl]}!H1:H{len(sanrio_sku_wjshypl_arr)}",
                                    "values": sanrio_sku_wjshypl_arr
                                },
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_wjshypl]}!C2:D{len(sanrio_sku_inventory_wjshypl_arr) + 1}",
                                    "values": sanrio_sku_inventory_wjshypl_arr
                                }
                            ]
                        })

    values_batch_update(token, FsConstants.sanrio_sales_repertory_token,
                        post_data={
                            "valueRanges": [
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_wjl]}!H1:H{len(sanrio_sku_wjl_arr)}",
                                    "values": sanrio_sku_wjl_arr
                                },
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_wjl]}!C2:D{len(sanrio_sku_inventory_wjl_arr) + 1}",
                                    "values": sanrio_sku_inventory_wjl_arr
                                }
                            ]
                        })

    values_batch_update(token, FsConstants.sanrio_sales_repertory_token,
                        post_data={
                            "valueRanges": [
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_snb]}!H1:H{len(sanrio_sku_snb_arr)}",
                                    "values": sanrio_sku_snb_arr
                                },
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_snb]}!C2:D{len(sanrio_sku_inventory_snb_arr) + 1}",
                                    "values": sanrio_sku_inventory_snb_arr
                                }
                            ]
                        })

    values_batch_update(token, FsConstants.sanrio_sales_repertory_token,
                        post_data={
                            "valueRanges": [
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_ph]}!H1:H{len(sanrio_sku_ph_arr)}",
                                    "values": sanrio_sku_ph_arr
                                },
                                {
                                    "range": f"{ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_ph]}!C2:D{len(sanrio_sku_inventory_ph_arr) + 1}",
                                    "values": sanrio_sku_inventory_ph_arr
                                }
                            ]
                        })

    value_range(token, FsConstants.sanrio_sales_repertory_token,
                ClientMapConstants[ClientConstants.sanrio_sales_repertory][MapFields.sanrio_sku_dpxl],
                f"D1:D{len(sanrio_sku_dpxl_arr)}", sanrio_sku_dpxl_arr)


def update_skus(file_a, file_b,
                sku_column="SKU",
                name_column="Product Name/产品名称",
                sheet_update="销量更新",
                sheet_exclude="不需要统计的SKU",
                sheet_stock="库存更新"):
    # 读取 A 文件
    data_a = pd.read_excel(file_a, dtype=str)

    # 读取 B 文件的指定 sheet
    data_b_update = pd.read_excel(file_b, sheet_name=sheet_update, dtype=str)
    data_b_exclude = pd.read_excel(file_b, sheet_name=sheet_exclude, dtype=str)

    # 检查列
    if sku_column not in data_a.columns or name_column not in data_a.columns:
        raise ValueError(f"A 文件缺少 {sku_column} 或 {name_column}")
    if "SKU" not in data_b_update.columns or "名称" not in data_b_update.columns or "数量" not in data_b_update.columns:
        raise ValueError(f"B 文件的 [{sheet_update}] 表缺少 'SKU'、'名称' 或 '数量' 列")

    # 构建 dict {SKU: 产品名称}
    sku_map = dict(zip(data_a[sku_column].dropna().str.strip(),
                       data_a[name_column].dropna().str.strip()))

    # 取 B 文件“不需要统计的SKU”表的 A 列
    excluded_skus = data_b_exclude.iloc[:, 0].dropna().astype(str).str.strip().tolist()

    # 取 B 文件“销量更新”表的 SKU 列
    existing_skus = data_b_update["SKU"].dropna().astype(str).str.strip().tolist()

    # 过滤
    filtered_map = {sku: pname for sku, pname in sku_map.items()
                    if sku not in excluded_skus and sku not in existing_skus}

    # 打开 B 文件
    book = load_workbook(file_b)

    # ---------------- 处理销量更新 ----------------
    sheet = book[sheet_update]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        col_sku = header.index("SKU") + 1
        col_name = header.index("名称") + 1
        col_qty = header.index("数量") + 1
    except ValueError:
        raise ValueError(f"在 {sheet_update} 表头中未找到 'SKU'、'名称' 或 '数量' 列")

    for sku, pname in filtered_map.items():
        new_row_idx = sheet.max_row + 1
        sheet.cell(row=new_row_idx, column=col_sku, value=sku)
        sheet.cell(row=new_row_idx, column=col_name, value=pname)
        sheet.cell(row=new_row_idx, column=col_qty,
                   value=f'=IFERROR(VLOOKUP(B{new_row_idx},E:F,2,0),0)')

    # ---------------- 处理库存更新 ----------------
    if sheet_stock not in book.sheetnames:
        raise ValueError(f"B 文件缺少表 {sheet_stock}")
    stock_sheet = book[sheet_stock]

    for sku, pname in filtered_map.items():
        new_row_idx = stock_sheet.max_row + 1
        stock_sheet.cell(row=new_row_idx, column=1, value=pname)  # A 列 名称
        stock_sheet.cell(row=new_row_idx, column=2, value=sku)  # B 列 SKU
        stock_sheet.cell(row=new_row_idx, column=3,
                         value=f"=SUMIF(N:N,B{new_row_idx},S:S)")  # C 列
        stock_sheet.cell(row=new_row_idx, column=4,
                         value=f"=IFERROR(VLOOKUP(B{new_row_idx},G:L,3,0),0)")  # D 列
        stock_sheet.cell(row=new_row_idx, column=5,
                         value=f"=IFERROR(VLOOKUP(B{new_row_idx},G:L,4,0),0)")  # E 列

    # 保存文件
    book.save(file_b)

    if filtered_map:
        print(f"✅ 已追加 {len(filtered_map)} 个 SKU 到 {sheet_update} 和 {sheet_stock}")
        for sku, pname in filtered_map.items():
            print(f"SKU: {sku} | 名称: {pname}")
    else:
        print("ℹ️ 没有需要追加的 SKU。")


def call(analyse_obj):
    root_path = "/Users/zkp/Desktop/B&Y/dxm/sku_store_statistics"
    analyse_dir_path = f"{root_path}/{analyse_obj}"
    dxm_order_path = f"{analyse_dir_path}/dxm_order.xlsx"
    template_path = f"{root_path}/{analyse_obj}运营统计.xlsx"
    template_copy_path = f"{analyse_dir_path}/{analyse_obj}运营统计_copy.xlsx"
    oms_store_dir = f"{analyse_dir_path}/oms_store"
    oms_store_merger_path = f"{oms_store_dir}/oms_store_merger.xlsx"
    dszs_inventory_path = f"{analyse_dir_path}/dszs_inventory.csv"
    jhb_inventory_path = f"{analyse_dir_path}/头程计划跟踪表&产品信息.xlsx"

    # 合并oms库存文件
    merge_excels_in_folder(oms_store_dir, oms_store_merger_path)

    # 更新sku
    update_skus(oms_store_merger_path, template_path)

    # 创建输出文件路径（复制一份文件2）
    template_copy_path = os.path.join(template_path, template_copy_path)
    shutil.copy(template_path, template_copy_path)

    # 更新店铺和sku的销量
    formatted_date = update_sales_data(dxm_order_path, template_copy_path)

    if analyse_obj == ClientConstants.xyl:
        # 更新库存
        update_available_inventory(oms_store_merger_path, template_copy_path, template_copy_path)
        # 更新海运空运
        # update_shipping_inventory(dszs_inventory_path, template_copy_path, template_copy_path)
        update_shipping_inventory(jhb_inventory_path, template_copy_path, template_copy_path)
        # 数据写入飞书表格
        xyl_fs(formatted_date, template_copy_path)
    elif ClientConstants.sanrio:
        update_total_inventory(oms_store_merger_path, template_copy_path, template_copy_path)
        sanrio_fs(formatted_date, template_copy_path)


if __name__ == '__main__':
    call(ClientConstants.sanrio)
    call(ClientConstants.xyl)
