from datetime import datetime, timedelta
import os
import re
from openpyxl import load_workbook
import pandas as pd
from collections import Counter, defaultdict
import requests
import json

from xinshili.fs_utils import get_token, detail_sheet_value


def count_no_track(file_path, column_name="快递"):
    """统计 '快递' 列中所有行数和内容为 '无轨迹' 的数量"""
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        column_index = headers.index(column_name) + 1
        pattern = re.compile(r"^\s*无轨迹\s*$", re.IGNORECASE)
        total_count = 0
        no_track_count = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            cell_value = row[column_index - 1]
            if cell_value is not None:
                total_count += 1
                if pattern.match(str(cell_value)):
                    no_track_count += 1
        return total_count, no_track_count
    except Exception as e:
        print(f"发生错误: {e}")
        return 0, 0


def count_distribution_and_no_track(file_path, key_column, courier_column="快递"):
    """
    通用函数，统计指定列的分布情况及其对应 "无轨迹" 的数量。
    :param file_path: Excel 文件路径
    :param key_column: 需要统计的列名
    :param courier_column: 快递列名
    :return: 各值的总数和 "无轨迹" 数量的 Counter 对象
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if key_column not in headers or courier_column not in headers:
            raise ValueError(f"列名 '{key_column}' 或 '{courier_column}' 不存在！")
        key_index = headers.index(key_column) + 1
        courier_index = headers.index(courier_column) + 1
        pattern = re.compile(r"^\s*无轨迹\s*$", re.IGNORECASE)
        key_counter = Counter()
        key_no_track_counter = Counter()
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            key_value = row[key_index - 1]
            courier_status = row[courier_index - 1]
            if key_value is not None:
                key_counter[key_value] += 1
                if courier_status is not None and pattern.match(str(courier_status)):
                    key_no_track_counter[key_value] += 1
        return key_counter, key_no_track_counter
    except Exception as e:
        print(f"发生错误: {e}")
        return Counter(), Counter()


def analyze_time_segments(file_path, time_column="订购时间", courier_column="快递"):
    """
    按时间段（每3分钟为一段，忽略秒进行判断）统计总数和 "无轨迹" 的数量。
    输出时包括秒显示。
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if time_column not in headers or courier_column not in headers:
            raise ValueError(f"列名 '{time_column}' 或 '{courier_column}' 不存在！")

        time_index = headers.index(time_column) + 1
        courier_index = headers.index(courier_column) + 1
        pattern = re.compile(r"^\s*无轨迹\s*$", re.IGNORECASE)

        # 读取并解析数据
        data = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            order_time = row[time_index - 1]
            courier_status = row[courier_index - 1]
            if order_time is not None and isinstance(order_time, str):
                try:
                    order_time = datetime.strptime(order_time, "%m-%d %H:%M:%S")
                    order_time_without_seconds = order_time.replace(second=0)
                    data.append((order_time, order_time_without_seconds, courier_status))
                except ValueError:
                    continue

        # 按时间段归类
        data.sort(key=lambda x: x[1])  # 按无秒时间排序
        time_segments = defaultdict(list)
        if data:
            base_time = data[0][1]  # 使用无秒时间作为基准
            current_segment = []
            for full_time, order_time_without_seconds, courier_status in data:
                if (order_time_without_seconds - base_time).total_seconds() <= 180:  # 3分钟内
                    current_segment.append((full_time, courier_status))
                else:
                    time_segments[base_time].extend(current_segment)
                    base_time = order_time_without_seconds
                    current_segment = [(full_time, courier_status)]
            if current_segment:
                time_segments[base_time].extend(current_segment)

        # 统计每个时间段的总数和无轨迹数量
        segment_statistics = {}
        for segment_start, entries in time_segments.items():
            total_count = len(entries)
            no_track_count = sum(
                1 for _, courier_status in entries if courier_status is not None and pattern.match(str(courier_status)))
            segment_statistics[segment_start] = {
                "total_count": total_count,
                "no_track_count": no_track_count,
                "entries": entries,
            }

        return segment_statistics

    except Exception as e:
        print(f"发生错误: {e}")
        return {}


def handle_file(input_file):
    file_extension = os.path.splitext(input_file)[1].lower()
    file_dir = os.path.dirname(input_file)
    file_name = os.path.splitext(os.path.basename(input_file))[0]
    if file_extension == '.csv':
        xlsx_file_path = os.path.join(file_dir, f"{file_name}.xlsx")
        try:
            data = pd.read_csv(input_file, encoding='utf-8')
            data.to_excel(xlsx_file_path, index=False)
            print(f"已将 CSV 文件转换为 XLSX 文件：{xlsx_file_path}")
            os.remove(input_file)
            print(f"已删除原始 CSV 文件：{input_file}")
            return xlsx_file_path
        except Exception as e:
            print(f"处理 CSV 文件时发生错误：{e}")
            return None
    else:
        print("文件不是 CSV 格式，执行其他逻辑")
        return input_file


def extract_number_from_filepath(filepath):
    """
    从文件路径中提取文件名中 '出库时间' 和 '_' 之间的数字。

    参数:
    - filepath: str，文件路径

    返回:
    - 提取到的数字（字符串形式），若未找到，返回 None
    """
    # 获取文件名（去掉路径部分）
    filename = os.path.basename(filepath)

    # 使用正则匹配 '出库时间' 和 '_' 之间的内容
    match = re.search(r"出库时间(\d+)_", filename)
    if match:
        return match.group(1)
    return None


update_time = "update_time"
order_count = "order_count"
no_track_number = "no_track_number"
track_percent = "track_percent"
no_track_percent = "no_track_percent"
warehouse_condition = "warehouse_condition"
store_condition = "store_condition"
sku_condition = "sku_condition"
time_segment_condition = "time_segment_condition"
time_segment_condition = "time_segment_condition"
sum_up = "sum_up"

input_file = input("请输入文件的绝对路径：")
xlsx_path = handle_file(input_file)
# 出库时间
ck_time = extract_number_from_filepath(xlsx_path)
# 获取今天的日期
today = datetime.today()
# 获取今天是几号
day_of_month = today.day
# 数据map
data_map = {}

text = ""
current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
text += "\n----------------------时间----------------------"
text += f"\n更新时间: {current_time}"
data_map[update_time] = current_time

interval_time = int(day_of_month) - int(ck_time)
text += f"\n出库日期：{ck_time}"
text += f"\n跟踪日期：{day_of_month}"
text += f"\n间隔时间：{interval_time}"
# print(f"出库日期：{ck_time}，跟踪日期：{day_of_month}，间隔时间：{time}")

total_count, no_track_count = count_no_track(xlsx_path, column_name="快递")
swl = round(100 - ((int(no_track_count) / int(total_count)) * 100))
# print(f"总条数（除列头）：{total_count}，内容为 '无轨迹' 的总数：{no_track_count}，上网率为：{swl}%")
text += "\n----------------------概览----------------------"
text += f"\n订单总数：{total_count}"
text += f"\n未上网数：{no_track_count}"
text += f"\n上网率：{swl}%"
text += f"\n未上网率：{100 - swl}%"

data_map[order_count] = total_count
data_map[no_track_number] = no_track_count
data_map[track_percent] = swl
data_map[no_track_percent] = 100 - swl

text += "\n----------------------仓库分布----------------------"
warehouse_distribution, warehouse_no_track = count_distribution_and_no_track(
    xlsx_path, key_column="发货仓库", courier_column="快递"
)
# print("\n发货仓库分布情况：")
warehouse_text = ""
lowest_swl = 101  # 初始化为比 100 大的值
lowest_warehouse = ""  # 保存最低上网率的仓库信息
for warehouse, count in warehouse_distribution.items():
    no_track_count = warehouse_no_track[warehouse]
    warehouseswl = round(100 - ((int(no_track_count) / int(count)) * 100))
    # print(f"{warehouse}: 总数 {count} 条，其中 '无轨迹' {no_track_count} 条，上网率为：{warehouseswl}%")
    text += f"\n{warehouse}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{warehouseswl}%"
    warehouse_text += f"\n{warehouse}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{warehouseswl}%"
    # 判断是否是最低的上网率
    if warehouseswl < lowest_swl:
        lowest_swl = warehouseswl
        lowest_warehouse = f"{warehouse}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{warehouseswl}%"
data_map[warehouse_condition] = warehouse_text

text += "\n----------------------店铺分布----------------------"
store_distribution, store_no_track_distribution = count_distribution_and_no_track(
    xlsx_path, key_column="店铺", courier_column="快递"
)
# print("\n店铺分布及对应的 '无轨迹' 情况：")
store_text = ""
lowest_store = ""
lowest_swl = 101  # 初始化为一个比 100 大的值，用于比较
for store, count in store_distribution.items():
    no_track_count = store_no_track_distribution[store]
    storeswl = round(100 - ((int(no_track_count) / int(count)) * 100))
    # print(f"{store}: 总数 {count} 条，其中 '无轨迹' {no_track_count} 条，上网率为：{storeswl}%")
    text += f"\n{store}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{storeswl}%"
    store_text += f"\n{store}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{storeswl}%"
    # 判断是否是最低的上网率
    if storeswl < lowest_swl:
        lowest_swl = storeswl
        lowest_store = f"{store}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{storeswl}%"
data_map[store_condition] = store_text

text += "\n----------------------sku分布----------------------"
sku_distribution, sku_no_track_distribution = count_distribution_and_no_track(
    xlsx_path, key_column="sku", courier_column="快递"
)
# print("\nSKU 分布及对应的 '无轨迹' 情况：")
sku_text = ""
lowest_sku = ""
lowest_swl = 101  # 初始化为比 100 大的值
for sku, count in sku_distribution.items():
    no_track_count = sku_no_track_distribution[sku]
    skuswl = round(100 - ((int(no_track_count) / int(count)) * 100))
    # print(f"{sku}: 总数 {count} 条，其中 '无轨迹' {no_track_count} 条，上网率为：{skuswl}%")
    text += f"\n{sku}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{skuswl}%"
    sku_text += f"\n{sku}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{skuswl}%"
    # 判断是否是最低的上网率
    if skuswl < lowest_swl:
        lowest_swl = skuswl
        lowest_sku = f"{sku}： 订单总数：{count}；无轨迹数：{no_track_count}；上网率：{skuswl}%"
# 将 sku_text 保存到 data_map
data_map[sku_condition] = sku_text

# 分析时间段
text += "\n----------------------时间段分布----------------------"
time_segment_analysis = analyze_time_segments(xlsx_path, time_column="订购时间", courier_column="快递")
# print("\n按时间段统计结果：")
time_segment_text = ""
lowest_segment = ""  # 保存上网率最低的时间段
lowest_swl = 101  # 初始化为比 100 大的值
for segment_start, stats in time_segment_analysis.items():
    segment_end = segment_start + timedelta(minutes=3)
    total_count = stats["total_count"]
    no_track_count = stats["no_track_count"]
    segmentswl = round(100 - ((int(no_track_count) / int(total_count)) * 100))
    # print(f"时间段 {segment_start.strftime('%m-%d %H:%M:%S')} - {segment_end.strftime('%m-%d %H:%M:%S')}:")
    # print(f"  总数: {total_count} 条, 其中 '无轨迹': {no_track_count} 条，上网率为：{segmentswl}%")
    text += f"\n{segment_start.strftime('%m-%d %H:%M')} - {segment_end.strftime('%m-%d %H:%M')}： 订单总数：{total_count}；无轨迹数：{no_track_count}；上网率：{segmentswl}%"
    time_segment_text += f"\n{segment_start.strftime('%m-%d %H:%M')} - {segment_end.strftime('%m-%d %H:%M')}： 订单总数：{total_count}；无轨迹数：{no_track_count}；上网率：{segmentswl}%"
    # 判断是否是最低的上网率
    if segmentswl < lowest_swl:
        lowest_swl = segmentswl
        lowest_segment = f"{segment_start.strftime('%m-%d %H:%M')} - {segment_end.strftime('%m-%d %H:%M')}： 订单总数：{total_count}；无轨迹数：{no_track_count}；上网率：{segmentswl}%"
data_map[time_segment_condition] = time_segment_text

lowest_txt = ""
lowest_txt += f"\n最低上网率的 仓库：{lowest_warehouse}"
lowest_txt += f"\n最低上网率的 SKU：{lowest_sku}"
lowest_txt += f"\n最低上网率的 商店：{lowest_store}"
lowest_txt += f"\n最低上网率的 时间段：{lowest_segment}"

sum_up_text = ""
# 如果三天后的上网率没有99%以上，那么就严重有问题；隔天应该要 》= 三分之一，隔两天应该要有》=75
if (interval_time == 1):
    if (swl < 30):
        sum_up_text += "☁️注意：间隔第1天，上网率未达30%，建议跟进！"
        sum_up_text += lowest_txt
    else:
        if (swl >= 50):
            sum_up_text += "☀️间隔第1天，上网率优秀"
        else:
            sum_up_text += "☀️间隔第1天，上网率良好"
elif (interval_time == 2):
    if (swl < 70):
        sum_up_text += "🌧️异常：间隔第2天，上网率未达75%，建议分析数据尝试定位问题！"
        sum_up_text += lowest_txt
    else:
        if (swl >= 85):
            sum_up_text += "☀️间隔第2天，上网率优秀"
        else:
            sum_up_text += "☀️间隔第2天，上网率良好"
else:
    if (swl < 95):
        sum_up_text += f"❄️⛈️🌀⚠️🚨警报：间隔第{interval_time}天，上网率未达95%，异常，定位问题后联系仓库反馈问题！"
        sum_up_text += lowest_txt
    else:
        if (swl >= 99):
            sum_up_text += f"☀️间隔第{interval_time}天，上网率优秀"
        else:
            sum_up_text += f"☀️间隔第{interval_time}天，上网率良好"

data_map[sum_up] = sum_up_text
text += "\n----------------------总结&建议----------------------"
text += f"\n{sum_up_text}"

# 数据打印
# print(data_map)
print(text)

# 写入飞书在线文档
tat = get_token()
detail_sheet_value(tat, [
    data_map[update_time],
    data_map[order_count],
    data_map[no_track_number],
    data_map[track_percent],
    data_map[no_track_percent],
    data_map[warehouse_condition],
    data_map[store_condition],
    data_map[sku_condition],
    data_map[time_segment_condition],
    data_map[sum_up],
], ck_time)
