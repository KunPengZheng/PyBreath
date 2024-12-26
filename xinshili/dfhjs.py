import openpyxl
import xlwings as xw
import os
import utils


def copy(source_file, target_file, output_file, columns_to_replace):
    # 读取表（source）和表（target）
    source_wb = openpyxl.load_workbook(source_file)
    target_wb = openpyxl.load_workbook(target_file, data_only=False)  # 保留公式

    source_ws = source_wb.active
    target_ws = target_wb.active

    # 获取表头（去除头尾空格）
    source_headers = [header.strip() for header in next(source_ws.iter_rows(max_row=1, values_only=True))]
    target_headers = [header.strip() for header in next(target_ws.iter_rows(max_row=1, values_only=True))]

    # 找到需要替换的列的索引
    source_column_indices = {col: source_headers.index(col) + 1 for col in columns_to_replace if col in source_headers}
    target_column_indices = {col: target_headers.index(col) + 1 for col in columns_to_replace if col in target_headers}

    # 检查匹配情况
    if not source_column_indices or not target_column_indices:
        raise ValueError("Some columns to replace are missing in source or target files.")

    # 读取表1的数据
    source_data = list(source_ws.iter_rows(min_row=2, max_row=source_ws.max_row, values_only=True))

    # 调整表2的行数以与表1匹配
    current_target_rows = target_ws.max_row - 1  # 减去表头行
    source_rows = len(source_data)

    if current_target_rows > source_rows:
        # 表2行数多，删除多余行
        target_ws.delete_rows(source_rows + 2, current_target_rows - source_rows)
    elif current_target_rows < source_rows:
        # 表2行数少，新增行
        for _ in range(source_rows - current_target_rows):
            target_ws.append([None] * target_ws.max_column)

    # 替换数据
    for row_idx, source_row in enumerate(source_data, start=2):  # 从第2行开始写入
        for col in columns_to_replace:
            if col in source_column_indices and col in target_column_indices:
                source_value = source_row[source_column_indices[col] - 1]
                target_cell = target_ws.cell(row=row_idx, column=target_column_indices[col])
                if target_cell.data_type != "f":  # 如果不是公式，替换值
                    target_cell.value = source_value

    # 保存结果
    target_wb.save(output_file)
    print(f"数据替换完成，结果保存为 {output_file}")
    return source_ws.max_row  # 返回原表的总行数


def tip_format(file_path, exchange_rate):
    app = xw.App(visible=True)  # 显示 Excel
    workbook = app.books.open(file_path)  # 加载 Excel 文件
    sheet = workbook.sheets[0]  # 使用第一个工作表

    # 获取 H 列最后一行，并计算总和
    last_row = sheet.range('H' + str(sheet.cells.last_cell.row)).end('up').row
    h_values = sheet.range(f'H2:H{last_row}').value
    h_total = sum(value for value in h_values if isinstance(value, (int, float)))

    # 设置字体颜色为红色
    red_color = (255, 0, 0)  # RGB 值表示红色

    # 第一步：写入 "共计" 和 H 列总和
    summary_row = last_row + 1  # n+1 行
    sheet.range(f'G{summary_row}').value = "共计"
    sheet.range(f'G{summary_row}').color = red_color  # 设置字体颜色为红色
    sheet.range(f'H{summary_row}').value = h_total
    sheet.range(f'H{summary_row}').color = red_color  # 设置字体颜色为红色

    # 第二步：写入 "汇率" 并等待用户输入
    exchange_rate_row = summary_row + 1  # n+2 行
    sheet.range(f'G{exchange_rate_row}').value = "汇率"
    sheet.range(f'G{exchange_rate_row}').color = red_color  # 设置字体颜色为红色
    sheet.range(f'H{exchange_rate_row}').value = exchange_rate  # 空白单元格，等待用户输入
    sheet.range(f'H{exchange_rate_row}').color = red_color  # 设置字体颜色为红色

    sheet.range(f'A{exchange_rate_row}').value = "注：单个发货操作费1.00美元"
    sheet.range(f'A{exchange_rate_row}').color = red_color  # 设置字体颜色为红色

    # 第三步：写入 "结算" 和公式
    settlement_row = exchange_rate_row + 1  # n+3 行
    sheet.range(f'G{settlement_row}').value = "结算"
    sheet.range(f'G{settlement_row}').color = red_color  # 设置字体颜色为红色
    sheet.range(f'H{settlement_row}').formula = f"=H{summary_row}*H{exchange_rate_row}"  # 设置公式
    sheet.range(f'H{settlement_row}').color = red_color  # 设置字体颜色为红色

    sheet.range(f'A{settlement_row}').value = "多个：另加打包袋0.3美元一个，产品多一个多加0.2美元一个"
    sheet.range(f'A{settlement_row}').color = red_color  # 设置字体颜色为红色

    # 自动调整所有单元格的列宽和行高
    sheet.used_range.columns.autofit()  # 自动调整所有列宽
    sheet.used_range.rows.autofit()  # 自动调整所有行高

    # 设置所有单元格内容居中
    used_range = sheet.used_range
    used_range.api.HorizontalAlignment = -4108  # xlCenter，水平居中
    used_range.api.VerticalAlignment = -4108  # xlCenter，垂直居中

    # 保存并关闭
    workbook.save()
    workbook.close()
    app.quit()

    print("格式化 和 底部总和公式 插入成功！")


# 源表文件路径
source_file = input("请输入源表文件的绝对路径：")
# exchange_rate_input = input("请输入（美对中）汇率（四舍五入 保留两位数，再+0.01）：")
# exchange_rate = float(exchange_rate_input)
exchange_rate = round(float(utils.get_usd_to_cny_rate()), 2) + 0.01

# 检查路径是否有效
if os.path.isabs(source_file) and os.path.isfile(source_file):
    file_name_without_extension = os.path.splitext(os.path.basename(source_file))[0]
    # last_four_chars = file_name_without_extension[-4:]  # 获取文件名最后四个字符
    target_file = "/Users/zkp/Desktop/B&Y/代发货结算/代发货结算表模版.xlsx"  # 目标表文件路径
    output_file = "/Users/zkp/Desktop/B&Y/代发货结算/" + file_name_without_extension + "_temp.xlsx"  # 输出文件路径

    # 替换的列名（注意名称要与表格列头匹配，去除空格）
    columns_to_replace = ["客户订单号", "派送单号", "RefNo2", "运费(USD)", "旺季附加费(USD)"]
    # 需要-1，第一行为表名
    total_rows = copy(source_file, target_file, output_file, columns_to_replace) - 1

    tip_format(output_file, exchange_rate)

    old_file_path = output_file
    new_file_path = "/Users/zkp/Desktop/B&Y/代发货结算/" + "代发货结算表" + str(
        total_rows) + "单" + utils.get_yd() + ".xlsx"

    try:
        # 修改文件名称
        os.rename(old_file_path, new_file_path)
        print(f"文件已重命名为：{new_file_path}")
    except FileNotFoundError:
        print("文件未找到，请检查路径！")
    except PermissionError:
        print("没有权限修改文件名！")
    except Exception as e:
        print(f"发生错误：{e}")

else:
    print("输入的路径无效或文件不存在，请检查并重新输入！")
