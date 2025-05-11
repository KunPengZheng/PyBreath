import re
import os
from pathlib import Path
from PIL import Image
from openpyxl import load_workbook


def crop_to_square(image_path, output_path, target_size=800):
    img = Image.open(image_path)
    w, h = img.size
    if w >= h:
        left = (w - h) // 2
        upper = 0
        right = left + h
        lower = h
    else:
        left = 0
        upper = (h - w) // 2
        right = w
        lower = upper + w

    square_img = img.crop((left, upper, right, lower))
    square_img = square_img.resize((target_size, target_size), Image.LANCZOS)
    square_img.save(output_path)
    print(f"Cropped and resized: {output_path}")


def process_folder(input_dir, output_dir, target_size=800):
    input_path = Path(input_dir)
    output_path = Path(output_dir)
    # 确保输出目录存在
    output_path.mkdir(parents=True, exist_ok=True)

    for img_file in input_path.iterdir():
        if img_file.suffix.lower() in ('.jpg', '.jpeg', '.png', '.bmp', '.gif'):
            out_file = output_path / f"{img_file.stem}{img_file.suffix}"
            crop_to_square(img_file, out_file, target_size)


def check_cell_value(value):
    # 去除前后空格并统一大小写
    result = value.strip()

    # 定义正则表达式
    pattern = r"^[A-Z]\d{2,4}$"

    # 使用正则匹配或以 'i am in bag' 开头（忽略大小写）
    if re.match(pattern, result) or result.lower().startswith("i am in bag"):
        return True  # 匹配成功
    else:
        return False  # 不匹配


def append_data_based_on_condition(file1, file2):
    wb1 = load_workbook(file1, data_only=True)
    ws1 = wb1.active

    wb2 = load_workbook(file2)
    ws2 = wb2.active

    # 判断 file1 A1 单元格是否满足条件
    cell_value = ws1.cell(row=1, column=1).value
    is_condition_met = check_cell_value(cell_value)

    # 根据条件选择源列
    source_cols = (2, 3) if is_condition_met else (1, 2)

    # 提取数据
    data_to_append = []
    for row in ws1.iter_rows(min_row=1, values_only=True):
        if row[source_cols[0] - 1] is None and row[source_cols[1] - 1] is None:
            continue
        data_to_append.append((row[source_cols[0] - 1], row[source_cols[1] - 1]))

    # ✅ 找 file2 中 B/C 列第一个空行的行号（从1开始）
    def find_first_empty_row(ws, col):
        for row in range(1, ws.max_row + 2):  # +2 是为了包含完全空的末尾行
            if ws.cell(row=row, column=col).value is None:
                return row
        return ws.max_row + 1

    start_row = min(
        find_first_empty_row(ws2, 2),  # B列
        find_first_empty_row(ws2, 3)  # C列
    )

    # 写入 file2 的 B/C 列
    for idx, (val1, val2) in enumerate(data_to_append):
        ws2.cell(row=start_row + idx, column=2).value = val1
        ws2.cell(row=start_row + idx, column=3).value = val2

    wb2.save(file2)
    print(f"✅ 已将数据从 {file1} 追加到 {file2} 的 B/C 列，从第 {start_row} 行开始。")


def natural_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]


def append_image_paths_to_excel(image_folder, excel_path, column_num):
    # 获取所有图片文件路径（自然排序）
    image_files = [f for f in os.listdir(image_folder) if
                   f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp'))]
    image_files.sort(key=natural_key)
    absolute_paths = [os.path.abspath(os.path.join(image_folder, f)) for f in image_files]

    # 打开 Excel 文件
    wb = load_workbook(excel_path)
    ws = wb.active

    # 找到 D 列第一个空白单元格的行号。column_num为4表示D列
    row = 1
    while ws.cell(row=row, column=column_num).value is not None:
        row += 1
    start_row = row

    # 写入路径到 D 列
    for idx, path in enumerate(absolute_paths):
        ws.cell(row=start_row + idx, column=column_num).value = path

    wb.save(excel_path)
    print(f"✅ 已将 {len(absolute_paths)} 条图片路径追加到 {excel_path} 的 {column_num} 列中（从第 {start_row} 行开始）。")


def append_columns_from_excel(source_path, target_path):
    # 加载 source 和 target 工作簿
    wb_source = load_workbook(source_path, data_only=True)
    ws_source = wb_source.active

    wb_target = load_workbook(target_path)
    ws_target = wb_target.active

    # 获取 source 可用的最大列和最大行
    source_max_col = ws_source.max_column
    source_max_row = ws_source.max_row

    # 获取 target 当前已有数据的最大列数
    target_max_col = ws_target.max_column

    # 确保 target 中每行的 list 是对齐的
    for row in ws_target.iter_rows(min_row=1, max_row=ws_target.max_row):
        while len(row) < target_max_col:
            ws_target.cell(row=row[0].row, column=len(row) + 1).value = None

    # 开始将 source 中的每列数据复制到 target 中，从 target_max_col + 1 开始
    for col_offset in range(source_max_col):
        for row_idx in range(1, source_max_row + 1):
            value = ws_source.cell(row=row_idx, column=col_offset + 1).value
            target_col = target_max_col + col_offset + 1
            ws_target.cell(row=row_idx, column=target_col).value = value

    # 保存结果
    wb_target.save(target_path)
    print(f"✅ 已将 {source_max_col} 列从 {source_path} 追加到 {target_path} 的列尾。")


def fill_blank_in_column_a(xlsx_path, content):
    wb = load_workbook(xlsx_path)
    ws = wb.active

    filled_count = 0

    # 遍历 A 列的所有单元格，直到没有行数据为止
    max_row = ws.max_row

    for row in range(1, max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value is None or str(cell.value).strip() == "":
            cell.value = content
            filled_count += 1

    wb.save(xlsx_path)
    print(f"✅ 已将内容 '{content}' 填入 {filled_count} 个空白单元格中。")


if __name__ == "__main__":
    # 生成800*800的1:1图片
    pic_dir = "/Users/zkp/Documents/fz20250403黑色女装N301-N450/N301-N450"

    pic800_dir = pic_dir + "_800/"
    # process_folder(pic_dir, pic800_dir)

    result_xlsx_path = "/Users/zkp/Documents/data.xlsx"

    # 复制标题
    # append_data_based_on_condition("/Users/zkp/Documents/fz20250403黑色女装N301-N450/0402黑色女装标题N301-N450.xlsx",
    #                                result_xlsx_path)

    # 复制800*800图片的绝对路径
    # append_image_paths_to_excel(pic800_dir, result_xlsx_path, 4)

    # 复制原图（3:4）图片的绝对路径
    # append_image_paths_to_excel(pic_dir, result_xlsx_path, 5)

    # 复制SKU
    # append_columns_from_excel("/Users/zkp/Documents/fz20250403黑色女装N301-N450/0402黑色女装货号N301-N450.xlsx",
    #                           result_xlsx_path)

    # 填充模版ID
    fill_blank_in_column_a(result_xlsx_path, "123456789")
