import openpyxl
from openpyxl.styles import PatternFill

def mark_matching_cells_in_file2(file1_path, file2_path):
    # 打开文件1和文件2
    wb1 = openpyxl.load_workbook(file1_path)
    sheet1 = wb1.active  # 文件1的工作表

    wb2 = openpyxl.load_workbook(file2_path)
    sheet2 = wb2.active  # 文件2的工作表

    # 获取文件1和文件2的表头
    header1 = [cell.value for cell in sheet1[1]]  # 读取文件1的表头
    header2 = [cell.value for cell in sheet2[1]]  # 读取文件2的表头

    # 确定列的索引
    try:
        tracking_no_col_idx = header1.index("Tracking No./物流跟踪号") + 1  # 文件1的物流跟踪号列
        waybill_no_col_idx = header2.index("运单号") + 1  # 文件2的运单号列
    except ValueError:
        print("未找到对应的列，请检查表头名称是否正确")
        return

    # 获取文件1的 Tracking No./物流跟踪号 列数据（用于匹配）
    tracking_numbers = [sheet1.cell(row=row, column=tracking_no_col_idx).value for row in range(2, sheet1.max_row + 1)]

    # 红色背景填充
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 遍历文件2的运单号列并与文件1的Tracking No./物流跟踪号列进行匹配
    for row in range(2, sheet2.max_row + 1):
        waybill_no = sheet2.cell(row=row, column=waybill_no_col_idx).value
        if waybill_no in tracking_numbers:
            # 如果匹配，设置文件2中的该单元格背景颜色为红色
            sheet2.cell(row=row, column=waybill_no_col_idx).fill = red_fill

    # 保存修改后的文件2
    wb2.save(file2_path)
    print(f"匹配完成，文件已保存至: {file2_path}")

# 示例调用
file1_path = '/Users/zkp/Documents/副本ParcelOutbound_20250214155734.xlsx'  # 请替换为实际文件1路径
file2_path = '/Users/zkp/Documents/副本新引力海外仓1月.xlsx'  # 请替换为实际文件2路径

mark_matching_cells_in_file2(file1_path, file2_path)
