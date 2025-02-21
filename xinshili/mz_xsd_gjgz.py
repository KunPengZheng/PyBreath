import openpyxl
import pandas as pd

from xinshili.gjgz_plus111 import check_and_add_courier_column, extract_and_process_data, RowName, CourierStateMapKey, \
    update_courier_status
from xinshili.pd_utils import remove_duplicates_by_column


def delete_rows_based_on_conditions(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active  # 默认使用活动工作表

    data = pd.read_excel(filepath)
    # 获取 'Shipping service/物流渠道' 和 'Recipient/收件人' 列的索引
    shipping_service_col = data.columns.get_loc('Shipping service/物流渠道') + 1  # openpyxl索引从1开始
    recipient_col = data.columns.get_loc('Recipient/收件人') + 1  # openpyxl索引从1开始

    # 记录要删除的行
    rows_to_delete = []

    # 遍历所有行（从第二行开始，跳过表头）
    for row in range(2, sheet.max_row + 1):
        shipping_service_value = sheet.cell(row=row, column=shipping_service_col).value
        recipient_value = sheet.cell(row=row, column=recipient_col).value

        # 如果满足删除条件，则标记该行
        if shipping_service_value == '上传物流面单(Upload_Shipping_Label)' and recipient_value != 'KJ':
            rows_to_delete.append(row)

    # 删除标记的行（从最后一行开始删除，避免索引错误）
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    # 保存更新后的文件
    wb.save(filepath)
    print(f"文件已更新：{filepath}")


def delete_rows_based_on_conditions222(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active  # 默认读取活动工作表

    # 查找 SKU 列的列号（假设表头在第一行）
    sku_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == 'SKU':
            sku_col = col
            break

    # 确保找到了 SKU 列
    if sku_col is None:
        raise ValueError("找不到 'SKU' 列")

    # 遍历工作表的每一行，跳过表头（假设表头是第一行）
    rows_to_delete = []
    for row in range(2, sheet.max_row + 1):  # 从第二行开始，跳过表头
        sku_value = sheet.cell(row=row, column=sku_col).value
        if sku_value != 'HS11168WE':  # 如果 SKU 不等于 'HS11168WE'
            rows_to_delete.append(row)

    # 删除不符合条件的行
    for row in reversed(rows_to_delete):  # 从后往前删除，避免删除影响到行号
        sheet.delete_rows(row)

    # 保存修改后的文件
    wb.save(filepath)


def delete_rows_not_tracking(filepath):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active  # 默认使用活动工作表

    data = pd.read_excel(filepath)
    # 获取 'Courier/快递' 列的索引
    courier_col = data.columns.get_loc('Courier/快递') + 1  # openpyxl索引从1开始

    # 记录要删除的行
    rows_to_delete = []

    # 遍历所有行（从第二行开始，跳过表头）
    for row in range(2, sheet.max_row + 1):
        courier_value = sheet.cell(row=row, column=courier_col).value

        # 如果 'Courier/快递' 列的内容不为 'tracking'，则标记删除
        if courier_value != 'tracking':
            rows_to_delete.append(row)

    # 删除标记的行（从最后一行开始删除，避免索引错误）
    for row in reversed(rows_to_delete):
        sheet.delete_rows(row)

    # 保存更新后的文件
    wb.save(filepath)
    print(f"文件已更新：{filepath}")


xlsx_path = input("请输入文件的绝对路径：")
remove_duplicates_by_column(xlsx_path, xlsx_path, 'Tracking No./物流跟踪号')
# delete_rows_based_on_conditions(xlsx_path)
delete_rows_based_on_conditions222(xlsx_path)
check_and_add_courier_column(xlsx_path)
results = extract_and_process_data(xlsx_path, RowName.Courier, 100)

update_courier_status(xlsx_path, results[CourierStateMapKey.not_yet_map])
update_courier_status(xlsx_path, results[CourierStateMapKey.pre_ship_map])
update_courier_status(xlsx_path, results[CourierStateMapKey.unpaid_map])
update_courier_status(xlsx_path, results[CourierStateMapKey.delivered_map])
update_courier_status(xlsx_path, results[CourierStateMapKey.no_tracking_map])
update_courier_status(xlsx_path, results[CourierStateMapKey.tracking_map])

# delete_rows_not_tracking(xlsx_path)

# no_tracking_count = len(results[CourierStateMapKey.not_yet_results]) + len(
#     results[CourierStateMapKey.pre_ship_results]) + len(results[CourierStateMapKey.no_tracking_results])
# tracking_count = len(results[CourierStateMapKey.unpaid_results]) + len(
#     results[CourierStateMapKey.delivered_results]) + len(results[CourierStateMapKey.tracking_results])
# print(f"没有轨迹数： {no_tracking_count} 条，有轨迹数： {tracking_count} 条")
# print(f"\nunpaid数： {len(results[CourierStateMapKey.unpaid_results])} 条")
# print(f"\nnot_yet数： {len(results[CourierStateMapKey.not_yet_results])} 条")
# print(f"\npre_ship数： {len(results[CourierStateMapKey.pre_ship_results])} 条")
# print(f"\ndelivered数： {len(results[CourierStateMapKey.delivered_results])} 条")
