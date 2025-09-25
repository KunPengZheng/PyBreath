import os
import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def insert_blank_row(filepath: str, row: int):
    wb = load_workbook(filepath)
    ws = wb.active

    # 检查指定行是否全为空（None 或 空字符串）
    def is_row_empty(worksheet, row_idx):
        for cell in worksheet[row_idx]:
            if cell.value not in (None, ""):
                return False
        return True

    if is_row_empty(ws, row):
        print(f"⚠️ {filepath} 的第 {row} 行已是空白行，跳过插入")
    else:
        ws.insert_rows(row)
        wb.save(filepath)
        print(f"✅ 已在 {filepath} 的第 {row} 行插入空白行")

    wb.close()


def get_color_palette():
    """生成一组颜色（30种）"""
    return [
        # 基础亮色
        "FFFF00",  # 黄色
        "00FF00",  # 绿色
        "00FFFF",  # 青色
        "FF00FF",  # 紫色
        "FF9900",  # 橙色
        "9999FF",  # 蓝紫
        "FF6666",  # 浅红
        "66CC99",  # 青绿
        "CC99FF",  # 淡紫
        "CCCC00",  # 橄榄

        # 额外扩展色
        "FFCC00",  # 金黄
        "0099FF",  # 天蓝
        "FF3399",  # 桃红
        "33CC33",  # 草绿
        "9966FF",  # 深紫
        "FF9933",  # 杏橙
        "33CCCC",  # 湖蓝
        "CC3333",  # 红棕
        "6699FF",  # 浅蓝
        "99CC33",  # 青柠

        # 柔和淡色
        "FFCCCC",  # 粉红
        "CCFFCC",  # 浅绿
        "CCCCFF",  # 浅蓝紫
        "FFFFCC",  # 浅黄
        "FFCCFF",  # 浅粉紫
        "CCFFFF",  # 浅青
        "FFE5CC",  # 米橙
        "E6CCFF",  # 淡紫
        "CCE5FF",  # 淡天蓝
        "E6FFE6",  # 淡绿
    ]


def unify_orders_inplace(folder: str):
    """
    修改后直接保存回原文件（覆盖）
    """
    files = [f for f in os.listdir(folder) if f.endswith(".xlsx") and not f.startswith("~$")]
    files.sort()

    all_data = []
    files_data = []

    # 读取所有文件
    for i, fname in enumerate(files):
        fpath = os.path.join(folder, fname)
        df = pd.read_excel(fpath, dtype=str).fillna("")
        if not {"标签", "系统单号", "平台单号"}.issubset(df.columns):
            print(f"⚠️ {fname} 缺少必要列，跳过")
            continue

        files_data.append({"filename": fname, "path": fpath, "df": df})

        for idx in df.index:
            all_data.append({
                "file_idx": i,
                "row_idx": idx,
                "标签": df.at[idx, "标签"],
                "sysno": df.at[idx, "系统单号"].strip(),
                "platform": df.at[idx, "平台单号"].strip()
            })

    if not all_data:
        print("⚠️ 没有找到可处理的数据")
        return

    # ---------------- 合并订单逻辑 ---------------- #
    for file_meta in files_data:
        df = file_meta["df"]
        merged_mask = df["标签"].str.contains("合并订单", na=False)
        for sys_id, group in df[merged_mask].groupby("系统单号"):
            if len(group) > 1:
                first_value = group["平台单号"].iloc[0]
                df.loc[group.index, "平台单号"] = first_value

    # ---------------- 拆分订单逻辑（跨文件） ---------------- #
    split_groups = defaultdict(list)
    for r in all_data:
        if "拆分订单" in r["标签"]:
            split_groups[(r["sysno"], r["platform"])].append(r)

    for (sysno, platform), group in split_groups.items():
        if len(group) <= 1:
            continue
        group_sorted = sorted(group, key=lambda x: (x["file_idx"], x["row_idx"]))
        for i, rec in enumerate(group_sorted[1:], start=1):
            file_meta = files_data[rec["file_idx"]]
            df = file_meta["df"]
            old_val = df.at[rec["row_idx"], "平台单号"]
            new_val = f"{old_val}-{i}"
            df.at[rec["row_idx"], "平台单号"] = new_val
            print(f"✏️ 修改 {file_meta['filename']} 第 {rec['row_idx'] + 2} 行: {old_val} → {new_val}")

    # ---------------- 保存并标色（覆盖原文件） ---------------- #
    palette = get_color_palette()
    color_map = {}
    color_idx = 0

    for meta in files_data:
        df = meta["df"]
        output_path = meta["path"]  # ⚠️ 覆盖原文件
        df.to_excel(output_path, index=False)

        wb = load_workbook(output_path)
        ws = wb.active

        # 合并订单 + 拆分订单
        mask = df["标签"].str.contains("合并订单|拆分订单", na=False)
        for (tag, sys_id), group in df[mask].groupby(["标签", "系统单号"]):
            if (tag, sys_id) not in color_map:
                color_map[(tag, sys_id)] = PatternFill(
                    start_color=palette[color_idx % len(palette)],
                    end_color=palette[color_idx % len(palette)],
                    fill_type="solid"
                )
                color_idx += 1

            fill = color_map[(tag, sys_id)]
            for idx in group.index:
                for cell in ws[idx + 2]:
                    cell.fill = fill

        wb.save(output_path)
        wb.close()
        print(f"✅ 已覆盖保存: {output_path}")


if __name__ == '__main__':
    folder_path = "/Users/zkp/Downloads/未命名文件夹 2/"
    unify_orders_inplace(folder_path)
