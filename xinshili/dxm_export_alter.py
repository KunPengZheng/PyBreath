import pandas as pd
import re
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from xinshili import utils


def extract_spec_ids(file_path):
    df = pd.read_excel(file_path, dtype=str).fillna("")

    spec_ids = {}

    for size in ["XXL", "XL", "XXS", "XS"]:
        # 找到第一个匹配行
        row = df[df["变种属性值二"] == size].head(1)

        if not row.empty:
            sku_attr_str = row.iloc[0]["SKU属性"]
            try:
                # 尝试将字符串解析为 JSON
                sku_attr_list = json.loads(sku_attr_str)
                for item in sku_attr_list:
                    if item.get("specName") == size:
                        spec_ids[size] = str(item.get("specId"))
                        break
            except json.JSONDecodeError as e:
                print(f"❌ JSON 解析失败：{e}")
        else:
            print(f"⚠️ 未找到“变种属性值二”为 {size} 的行")

    return spec_ids


def process_excel(file_path, output_path, ids):
    # 读取 Excel 数据
    df = pd.read_excel(file_path, dtype=str).fillna("")

    # 尺码映射
    size_map = {
        "S": "1",
        "M": "2",
        "L": "3",
        "XL": "4",
        "XXL": "5"
    }

    highlight_rows = set()

    # 遍历每一行处理 XXS/XS 替换
    for idx, row in df.iterrows():
        row_str = " ".join(row.values)

        # 替换 XXS → XXL，10002 → 8002
        if "XXS" in row_str:
            df.loc[idx] = df.loc[idx].apply(
                lambda x: x.replace("XXS", "XXL") if isinstance(x, str) else x
            )
            df.loc[idx] = df.loc[idx].apply(
                lambda x: x.replace(ids["XXS"], ids["XXL"]) if isinstance(x, str) else x
            )
            highlight_rows.add(idx)

        # 替换 XS → XL，12001 → 12003
        if "XS" in row_str:
            df.loc[idx] = df.loc[idx].apply(
                lambda x: x.replace("XS", "XL") if isinstance(x, str) else x
            )
            df.loc[idx] = df.loc[idx].apply(
                lambda x: x.replace(ids["XS"], ids["XL"]) if isinstance(x, str) else x
            )
            highlight_rows.add(idx)

    # 变种属性值二与 SKU货号 调整
    if "变种属性值二" in df.columns and "SKU货号" in df.columns:
        for idx, row in df.iterrows():
            size = row["变种属性值二"].strip()
            if size in size_map:
                sku = row["SKU货号"]
                if "-" in sku:
                    prefix, suffix = sku.split("-", 1)
                    # 使用正则替换第一个数字
                    import re
                    # 查找 - 后面第一个数字并替换
                    new_suffix = re.sub(r'^\d', size_map[size], suffix, count=1)
                    new_sku = f"{prefix}-{new_suffix}"
                    df.at[idx, "SKU货号"] = new_sku

    # 保存结果
    df.to_excel(output_path, index=False)

    # 加背景色（绿色）到指定行
    wb = load_workbook(output_path)
    ws = wb.active
    green_fill = PatternFill(start_color="C0D79B", end_color="C0D79B", fill_type="solid")

    for row_idx in highlight_rows:
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx + 2, column=col_idx).fill = green_fill  # Excel中数据从第2行开始

    wb.save(output_path)

    print(f"✅ 文件已处理并保存至：{output_path}")
    utils.open_dir(utils.get_file_dir(output_path))


if __name__ == '__main__':
    source_file = input("请输入源表文件的绝对路径：")
    output_path = f"/Users/zkp/Desktop/B&Y/dxm/export/{utils.get_filename_with_extension(source_file)}"
    ids = extract_spec_ids(source_file)
    process_excel(source_file, output_path, ids)
