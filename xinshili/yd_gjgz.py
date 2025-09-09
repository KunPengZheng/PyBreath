from datetime import datetime
import os
import re
import glob
import openpyxl
from openpyxl import load_workbook
import pandas as pd
from dataclasses import dataclass
from collections import defaultdict
import platform
import subprocess


@dataclass(frozen=True)
class ClientConstants:
    zbw = "zbw"
    sanrio = "sanrio"
    xyl = "xyl"
    kaer = "kaer"
    mz_xsd = "mz_xsd"
    mx_dg = "mx_dg"
    md_fc = "md_fc"
    md_flld = "md_flld"
    cksj = "cksj"
    cjsj = "cjsj"
    khhz = "khhz"
    ckoms = "ckoms"
    yy = "yy"
    xyl_sales_repertory = "xyl_sales_repertory"
    sanrio_sales_repertory = "sanrio_sales_repertory"
    dxm_xyl_yd = "dxm_xyl_yd"


@dataclass(frozen=True)
class RowName:
    Tracking_No = 'Tracking No./物流跟踪号'
    Courier = 'Courier/快递'
    CreationTime = "Creation time/创建时间"
    OutboundTime = "OutboundTime/出库时间"
    Warehouse = "Warehouse/仓库"
    Client = "Client/客户"
    CreationWaveTime = "Create wave time/生成波次时间"
    SKU = "SKU"
    ShippingService = "Shipping service/物流渠道"
    PossessionSfDate = "PossessionSfDate/揽收时间"
    LatestEventSfDate = "LatestEventSfDate/最新事件时间"
    LatestEventSfTime = "LatestEventSfTime/最新事件时间"
    LatestEventSfSite = "LatestEventSfSite/最新事件地点"
    SfDateInterval = "SfDateInterval/SF消息间隔"
    TrackTimeInterval = "TrackTimeInterval/跟踪时间间隔"
    TrackTimeIntervalState = "TrackTimeIntervalState/跟踪时间间隔状态"
    LastEventSfTime = "LastEventSfTime/上一条轨迹时间"
    Tacking_Time = "Tacking_Time/追踪时间"
    UnpaidDate = "UnpaidDate/unpaid记录时间"
    Recipient = "Recipient/收件人"
    Upload_Shipping_Label = "上传物流面单(Upload_Shipping_Label)"
    YD_Number = 'YD_Number/阳单号'
    YD_State = 'YD_State/阳单轨迹状态'

    Courier_File1 = 'Courier/快递_file1'
    SfDateInterval_File1 = 'SfDateInterval/SF消息间隔_file1'
    PossessionSfDate_File1 = 'PossessionSfDate/揽收时间_file1'
    LatestEventSfDate_File1 = 'LatestEventSfDate/最新事件时间_file1'
    Package1_Tracking = 'Package 1\nTracking No./物流跟踪号1'

    Order_Num = "订单号"
    Track_Num = "运单号"
    Platform_Num = "Platform Number/平台单号"
    Yang_Num = "阳单_运单号"
    Yin_Num = "阴单_运单号"
    Yang_Track_State = "阳单_物流状态"
    Yin_Track_State = "阴单_物流状态"
    Yang_Delivered_Time = "阳单_签收时间"
    Yin_Delivered_Time = "阴单_签收时间"
    YY_Delivered_Time = "阴阳单_签收间隔"
    Create_Time = "创建时间"

    Length = 'Length/长'
    Width = 'Width/宽'
    Height = 'Height/高'
    Unit = 'Unit/单位'

    Pay_Time = "付款时间"
    Ship_Time = "发货时间"
    Track_State = "轨迹状态"
    Analyse_State = "追踪时间"
    Last_Track_Time = "上一条轨迹时间"
    Latest_Track_Site = "最新轨迹位置"
    Latest_Track_Time = "最新轨迹时间"
    Interval_Time = "时间间隔"
    Process_Time = "处理状态"
    YD_Number2 = "阳单号"
    YD_State2 = "阳单轨迹状态"

    Total_Of_Product = "产品总数"
    Store_Account = "店铺账号"
    Store_Sales = "店铺销量"
    Store_Name = "店铺名称"
    Order_Sales = "订单数量"
    Sales_Update = "销量更新"
    Quantity = "数量"
    Inventory_Update = "库存更新"
    Sea_transportation = "海运在途"
    Air_transportation = "空运在途"
    Available_Quantity = "可用量"
    Available_Inventory = "Available Inventory/可用库存"


@dataclass(frozen=True)
class CourierStateMapKey:
    tracking_map = 'tracking_map'
    irregular_number_map = 'irregular_number_map'
    no_tracking_map = 'no_tracking_map'
    unpaid_map = "unpaid_map"
    not_yet_map = "not_yet_map"
    pre_ship_map = "pre_ship_map"
    delivered_map = "delivered_map"
    alert_map = "alert_map"
    possession_sf_date_map = "possession_sf_date_map"
    latest_event_sf_date_map = "latest_event_sf_date_map"
    sf_date_equality_map = "sf_date_equality_map"
    delivered_time_map = "delivered_time_map"
    latest_event_sf_time_map = "latest_event_sf_time_map"
    latest_event_sf_site_map = "latest_event_sf_site_map"


class CourierStateMapValue:
    irregular_no_tracking = 'irregular_no_tracking'
    not_yet = 'not_yet'
    pre_ship = "pre_ship"
    no_tracking = "no_tracking"
    no_track = "no_track"
    unpaid = "unpaid"
    delivered = "delivered"
    tracking = "tracking"
    alert = "alert"


@dataclass(frozen=True)
class CellKey:
    Outbound_Time = "Outbound_Time"
    update_time = "update_time"
    warehouse_condition = "warehouse_condition"
    store_condition = "store_condition"
    sku_condition = "sku_condition"
    time_segment_condition = "time_segment_condition"
    sum_up = "sum_up"
    exception = "exception"
    shipping_service_condition = "shipping_service_condition"
    special_information = "special_information"
    wl = "wl"
    unpaid = "unpaid"


@dataclass(frozen=True)
class Pattern:
    no_track = r"not_yet|pre_ship|irregular_no_tracking|no_tracking"
    delivered = r"^delivered$"
    unpaid = r"^unpaid$"
    not_yet = r"^not_yet$"
    irregular_no_tracking = r"^irregular_no_tracking$"
    pre_ship = r"^pre_ship$"
    no_tracking = r"^no_tracking$"
    tracking = r"^tracking$"
    alert = r"^alert"


def getYmd():
    # 获取今天的日期
    today = datetime.today()
    # 格式化为 "%Y/%m/%d" 格式
    formatted_today = today.strftime("%Y/%m/%d")
    # print(formatted_today)
    return formatted_today


# 自然排序的辅助函数
def natural_key(s):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r'(\d+)', s)]


def is_time_difference_exceed(start_time_str, end_time_str):
    try:
        time_format = "%Y-%m-%d"
        start_time = datetime.strptime(start_time_str, time_format)
        end_time = datetime.strptime(end_time_str, time_format)
        difference = abs((end_time - start_time).days)
        return difference
    except ValueError as e:
        print(f"❌ 时间格式错误: {e}")
        return False


def get_days_difference(file_path, column_name="Creation time/创建时间"):
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        # 获取列索引
        column_index = headers.index(column_name) + 1
        # 获取第一条数据
        first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
        if not first_row_value:
            raise ValueError(f"'{column_name}' 列的第一条数据为空！")
        # 解析日期
        outbound_time = datetime.strptime(first_row_value, "%Y-%m-%d %H:%M:%S")
        # 格式化为 "%Y/%m/%d" 格式
        formatted_date = outbound_time.strftime("%Y/%m/%d")
        return formatted_date
    except Exception as e:
        print(f"发生错误: {e}")
        return None


def get_days_difference_flld(file_path, column_name="打单时间"):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        # 获取列索引
        column_index = headers.index(column_name) + 1
        # 获取第一条数据
        first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
        if not first_row_value:
            raise ValueError(f"'{column_name}' 列的第一条数据为空！")

        # 获取当前年份
        current_year = datetime.now().year

        # 检查日期类型，如果是字符串并且没有年份，补充当前年份
        if isinstance(first_row_value, str):
            # 如果是字符串并且没有年份，补充当前年份
            if len(first_row_value.split('-')) == 2:  # 格式为 'MM-DD' 或 'DD-MM'
                first_row_value = f"{current_year}-{first_row_value}"

            # 尝试解析日期,支持带秒或不带秒的时间格式
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                try:
                    outbound_time = datetime.strptime(first_row_value.strip(), fmt)
                    break
                except ValueError:
                    continue
            else:
                raise ValueError(f"❌ 无法识别时间格式: {first_row_value}")
        elif isinstance(first_row_value, datetime):
            # 如果是已经是 datetime 类型，直接处理
            outbound_time = first_row_value
        else:
            raise ValueError(f"无法解析日期: {first_row_value}")

        # 格式化为 "%Y/%m/%d" 格式
        formatted_date = outbound_time.strftime("%Y/%m/%d")
        return formatted_date
    except Exception as e:
        print(f"发生错误: {e}")
        return None


def get_days_difference_tkkj(file_path, column_name="发货时间"):
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        # 获取列索引
        column_index = headers.index(column_name) + 1
        # 获取第一条数据
        first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
        if not first_row_value:
            raise ValueError(f"'{column_name}' 列的第一条数据为空！")
        # 解析日期
        outbound_time = datetime.strptime(first_row_value, "%Y-%m-%d")
        # 格式化为 "%Y/%m/%d" 格式
        formatted_date = outbound_time.strftime("%Y/%m/%d")
        return formatted_date
    except Exception as e:
        print(f"发生错误: {e}")
        return None


def check_and_add_courier_column(file_path):
    try:
        # 加载 Excel 文件
        data = pd.read_excel(file_path, dtype=str, engine='openpyxl')
        flag = False
        # 判断是否存在 '快递' 列
        if RowName.Courier not in data.columns:
            # 如果没有 '快递' 列，则在最后一列添加该列
            data[RowName.Courier] = ""  # 默认为空值，可以根据需求填充其他默认值
            flag = True
        if RowName.PossessionSfDate not in data.columns:
            data[RowName.PossessionSfDate] = ""
            flag = True
        if RowName.LatestEventSfDate not in data.columns:
            data[RowName.LatestEventSfDate] = ""
            flag = True
        if RowName.SfDateInterval not in data.columns:
            data[RowName.SfDateInterval] = ""
            flag = True
        if RowName.UnpaidDate not in data.columns:
            data[RowName.UnpaidDate] = ""
            flag = True
        if RowName.LatestEventSfTime not in data.columns:
            data[RowName.LatestEventSfTime] = ""
            flag = True
        if RowName.LatestEventSfSite not in data.columns:
            data[RowName.LatestEventSfSite] = ""
            flag = True
        if RowName.TrackTimeInterval not in data.columns:
            data[RowName.TrackTimeInterval] = ""
            flag = True
        if RowName.TrackTimeIntervalState not in data.columns:
            data[RowName.TrackTimeIntervalState] = ""
            flag = True
        if RowName.Tacking_Time not in data.columns:
            data[RowName.Tacking_Time] = ""
            flag = True
        if RowName.LastEventSfTime not in data.columns:
            data[RowName.LastEventSfTime] = ""
            flag = True
        if RowName.YD_Number not in data.columns:
            data[RowName.YD_Number] = ""
            flag = True
        if RowName.YD_State not in data.columns:
            data[RowName.YD_State] = ""
            flag = True
        # 保存修改后的文件
        data.to_excel(file_path, index=False, engine='openpyxl')
        return flag
    except Exception as e:
        print(f"发生错误: {e}")


def is_us_weekend(date_str):
    """
    中国和美国的时差相差：13-16 个钟头，目前日期的单位最小是日期，没有到小时，所以这里我们默认和美国相差一天，
    也就是中国时间周日为美国的周六，中国时间周一为美国的周日
    """
    # 解析字符串为 datetime 对象
    date_obj = datetime.strptime(date_str, "%Y/%m/%d")

    # 判断是否为 周日（6）或者周一 (0)，即为美国的周六和周日
    return date_obj.weekday()


def get_shipment_received_numbers(filepath, gz_time):
    # 加载 Excel 文件
    workbook = load_workbook(filename=filepath, data_only=True)

    # 获取第一个工作表
    sheet = workbook.active

    # 读取表头（第一行），并构建列名到索引的映射
    header = [cell.value for cell in sheet[1]]

    # 确保所有必要的列都存在
    required_columns = [RowName.Courier, RowName.SfDateInterval, RowName.OutboundTime, RowName.Tracking_No]

    # 获取每个必要列的索引（列索引从 0 开始）
    column_indices = {}
    for col in required_columns:
        if col not in header:
            missing_cols = set(required_columns) - set(header)
            raise ValueError(f"Excel 文件缺少必要的列: {missing_cols}")
        column_indices[col] = header.index(col)

    # 存储符合条件的跟踪号
    tracking_numbers = []

    # 遍历数据行（从第二行开始）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        courier = row[column_indices[RowName.Courier]]
        sf_date_interval = row[column_indices[RowName.SfDateInterval]]
        outbound_time = row[column_indices[RowName.OutboundTime]]
        tracking_no = row[column_indices[RowName.Tracking_No]]

        # 筛选条件：Courier == 'tracking' 且 SfDateInterval == '0'
        if courier == CourierStateMapValue.tracking and sf_date_interval == 0:

            # try:
            # 解析并计算日期间隔
            parsed_outbound_time = datetime.strptime(outbound_time, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_outbound_time.strftime("%Y/%m/%d")
            interval_days = (datetime.strptime(gz_time, "%Y/%m/%d") -
                             datetime.strptime(formatted_date, "%Y/%m/%d")).days

            is_usweekend = is_us_weekend(formatted_date)
            actual_interval = 0
            if is_usweekend == 6:  # 6是中国周日，美国周六
                actual_interval = 2
            elif is_usweekend == 0:  # 0是中国周一，美国周日
                actual_interval = 1

            # 如果间隔为 2 天，添加到结果列表
            if (interval_days - actual_interval) >= 2 and tracking_no is not None:
                tracking_numbers.append(tracking_no)

        # except (ValueError, TypeError) as e:
        #     日期解析错误或数据错误时跳过
        # print(f"日期解析错误: {e}, 跟踪号: {tracking_no}")
        # continue

    # 返回唯一的跟踪号列表
    return list(set(tracking_numbers))


def filter_tracking_numbers(input_path, output_path):
    # 读取 Excel 文件
    df = pd.read_excel(input_path, dtype=str).fillna("")

    col = "Tracking No./物流跟踪号"

    if col not in df.columns:
        print(f"❌ 文件中缺少列: {col}")
        return

    # 保留以 92、93 或 94 开头的行
    df_filtered = df[df[col].str.startswith(("92", "93", "94"), na=False)]

    # 保存结果
    df_filtered.to_excel(output_path, index=False)
    # print(f"✅ 筛选后文件已保存到: {output_path}")


def remove_duplicates_by_column(input_file, output_file, column_name):
    """
    去重：删除指定列中重复的行，仅保留第一条，并覆盖源文件，同时返回去重后的行数。

    参数：
    - input_file: str，输入文件路径
    - column_name: str，要检查重复的列名
    """
    try:
        # 读取 Excel 文件
        df = pd.read_excel(input_file, dtype=str)

        # 检查列名是否存在
        if column_name not in df.columns:
            raise ValueError(f"列 '{column_name}' 不存在于输入文件中！")

        # 删除指定列的重复项，仅保留第一条
        df_deduplicated = df.drop_duplicates(subset=[column_name], keep='first')

        # 将去重后的数据保存到输出文件
        df_deduplicated.to_excel(output_file, index=False)

        # 返回去重后的行数
        return len(df_deduplicated)

    except Exception as e:
        print(f"处理文件时发生错误：{e}")
        return 0  # 如果出现错误，返回 0 行数


def count_pattern_and_tracking_with_sf_date(file_path, column_name, sfDateInterval_name, patterns):
    """
    统计指定列中多个正则表达式匹配内容的数量，并统计 "Courier/快递" 列内容为 'tracking' 且 "SfDateEquality" 列的内容为 0 的行数。

    :param file_path: Excel 文件路径
    :param column_name: 需要统计的列名
    :param patterns: 正则表达式字典，key 为模式名称，value 为模式字符串
    :return: 包含每个模式匹配计数和符合 'tracking' 且 'SfDateEquality' 为 0 的行数的字典
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # 检查列名是否存在
        if column_name not in headers or sfDateInterval_name not in headers:
            raise ValueError(f"Excel 文件中未找到 '{column_name}' 或 '{sfDateInterval_name}'列 找不到")

        # 获取列索引
        column_index = headers.index(column_name)
        sfDateInterval_index = headers.index(sfDateInterval_name)

        # 编译所有正则表达式
        compiled_patterns = {key: re.compile(pattern, re.IGNORECASE) for key, pattern in patterns.items()}

        # 初始化计数器
        count_dict = {**{key: 0 for key in patterns}, "sfDateInterval": 0}  # 显式初始化
        tracking_sf_date_equality_count = 0

        # 遍历每一行数据
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            courier_value = row[column_index]
            sfDateInterval_value = row[sfDateInterval_index]

            # 判断是否符合 'tracking' 且 'SfDateEquality' 为 0 的条件
            if courier_value == CourierStateMapValue.tracking and sfDateInterval_value == 0:
                tracking_sf_date_equality_count += 1

            # 判断正则表达式匹配
            if courier_value is not None:
                cell_value_str = str(courier_value)
                for key, pattern in compiled_patterns.items():
                    if pattern.search(cell_value_str):
                        count_dict[key] += 1

        # 将 'tracking' 且 'SfDateEquality' 为 0 的计数添加到字典中
        count_dict["sfDateInterval"] = tracking_sf_date_equality_count

        return count_dict

    except Exception as e:
        print(f"发生错误: {e}")
        # 返回一个包含所有模式及 'tracking_sf_date_equality' 键的字典
        return {key: 0 for key in patterns} | {"sfDateInterval": 0}  # 合并字典


def round2(nums):
    """
    四舍五入，保留两位数
    """
    return round(nums, 2)


def delete_file(file_path):
    """
    删除指定的文件
    :param file_path: 要删除的文件的绝对路径
    """
    try:
        if os.path.exists(file_path):  # 检查文件是否存在
            os.remove(file_path)  # 删除文件
            # print(f"文件 {file_path} 已成功删除")
        else:
            print(f"文件 {file_path} 不存在")
    except Exception as e:
        print(f"删除文件时发生错误: {e}")


def find_irregular_tracking_numbers(filepath, column_name=RowName.Tracking_No):
    """
    查找不规则的快递单号（不是纯数字、不是9开头、或者包含下划线）
    :param filepath: Excel文件路径
    :param column_name: 需要检查的列名，默认为 'Tracking No./物流跟踪号'
    :return: 不规则快递单号字典
    """
    try:
        # 打开xlsx文件
        wb = openpyxl.load_workbook(filepath, data_only=True)
        sheet = wb.active  # 默认使用活动工作表

        # 获取 'Tracking No./物流跟踪号' 列索引
        tracking_no_col = next(
            (col for col in range(1, sheet.max_column + 1) if sheet.cell(row=1, column=col).value == column_name),
            None
        )

        if tracking_no_col is None:
            print(f"找不到 {column_name} 列")
            return {}

        # 存储不规则快递单号的字典
        irregular_number_map = {}

        # 遍历所有行，从第二行开始（跳过表头）
        for row in range(2, sheet.max_row + 1):
            tracking_no = str(sheet.cell(row=row, column=tracking_no_col).value).strip()  # 转换为字符串并去除前后空格
            # 判断是否为不合规单号
            if not tracking_no.isdigit() or not tracking_no.startswith(("92", "93", "94")) or "_" in tracking_no:
                irregular_number_map[tracking_no] = CourierStateMapValue.irregular_no_tracking

        # print("find_irregular_tracking_numbers 方法执行完成")
        return irregular_number_map

    except Exception as e:
        print(f"发生错误: {e}")
        return {}


def update_courier_status1(filepath, maps, wl=RowName.Tracking_No):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active  # 默认使用活动工作表

    data = pd.read_excel(filepath, dtype=str)
    # 获取 'Tracking No./物流跟踪号' 列和 'Courier/快递' 列的索引
    tracking_no_col = data.columns.get_loc(wl) + 1  # openpyxl索引从1开始
    courier_col = data.columns.get_loc(RowName.Courier) + 1  # openpyxl索引从1开始

    for tracking_no, status in maps.items():
        for row in range(2, sheet.max_row + 1):  # 从第二行开始（跳过表头）
            # 获取当前行的物流跟踪号
            current_tracking_no = sheet.cell(row=row, column=tracking_no_col).value
            # **处理 NoneType，转换为字符串 "None"**
            if current_tracking_no is None:
                current_tracking_no = "None"
            # 如果找到匹配的物流跟踪号，更新 Courier/快递 列
            if current_tracking_no == tracking_no:
                sheet.cell(row=row, column=courier_col, value=status)

    # 保存更新后的文件
    wb.save(filepath)


def process_tracking_no(file_path: str, row_name=RowName.Tracking_No):
    # 读取 Excel 文件
    data = pd.read_excel(file_path, dtype=str)  # 将数据全部读取为字符串类型

    # 确保 'Tracking No./物流跟踪号' 列存在
    if row_name not in data.columns:
        raise ValueError(f"文件中缺少 '{row_name}' 列")

    # 处理 'Tracking No./物流跟踪号' 列：去除空格并确保每个值是字符串，兼容 None 或空值
    data[row_name] = data[row_name].apply(
        lambda x: str(x).replace(" ", "") if x is not None and pd.notna(x) else "")

    # 保存处理后的数据
    data.to_excel(file_path, index=False, engine='openpyxl')

    return data


def extract_and_process_data(filepath: str, column_name: str, wl_name=RowName.Tracking_No, ckjs_flag=False):
    data = pd.read_excel(filepath, dtype=str)

    if column_name not in data.columns:
        raise ValueError(f"列 '{column_name}' 不存在于 Excel 文件中")

    # 将无内容的单元格赋值""空字符串
    data[column_name] = data[column_name].fillna('')

    # 获取指定内容的数据
    if ckjs_flag:
        filtered_data = data[data[column_name].apply(
            lambda x: str(x).strip().lower() in [''])]
    else:
        filtered_data = data[data[column_name].apply(
            lambda x: str(x).strip().lower() in ['',
                                                 CourierStateMapValue.not_yet,
                                                 CourierStateMapValue.pre_ship,
                                                 CourierStateMapValue.tracking,
                                                 CourierStateMapValue.no_tracking
                                                 ]
                      or str(x).strip().lower().startswith(CourierStateMapValue.alert)
        )]
    # 提取符合条件的 'Tracking No./物流跟踪号' 列数据，并剔除wl_name列中 NaN，去除字符串前后空格，排除只含空格或本身为空的字符串
    items = [x.strip() for x in filtered_data[wl_name].dropna().astype(str) if x.strip() != '']

    return items


def get_computer_model():
    system = platform.system()

    try:
        if system == "Windows":
            result = subprocess.check_output(["wmic", "computersystem", "get", "model"], shell=True)
            return result.decode().split("\n")[1].strip()

        elif system == "Darwin":  # macOS
            result = subprocess.check_output(["system_profiler", "SPHardwareDataType"])
            for line in result.decode().splitlines():
                if "Model Identifier" in line:
                    return line.split(":")[1].strip()

        elif system == "Linux":
            try:
                # 尝试读取 product_name 文件
                with open("/sys/devices/virtual/dmi/id/product_name", "r") as f:
                    return f.read().strip()
            except FileNotFoundError:
                # 使用 dmidecode（需要 root 权限）
                result = subprocess.check_output(["sudo", "dmidecode", "-s", "system-product-name"])
                return result.decode().strip()

        else:
            return f"Unsupported OS: {system}"

    except Exception as e:
        return f"Error retrieving model: {e}"


def automatic(analyse_obj, ignore, analyse_obj_ignore):
    root_dir = ""
    if "MacBookPro" in get_computer_model():
        if analyse_obj == "zbw":
            root_dir = "/Users/zkp/Desktop/B&Y/轨迹统计/zbw/"
        elif analyse_obj == "sanrio":
            root_dir = "/Users/zkp/Desktop/B&Y/轨迹统计/sanrio/"
        elif analyse_obj == "xyl":
            root_dir = "/Users/zkp/Desktop/B&Y/轨迹统计/xyl/"
        elif analyse_obj == "kaer":
            root_dir = "/Users/zkp/Desktop/B&Y/轨迹统计/kaer/"
        else:
            root_dir = ""
    else:
        if analyse_obj == "zbw":
            root_dir = "/Volumes/B&Y/轨迹统计/zbw/"
        elif analyse_obj == "sanrio":
            root_dir = "/Volumes/B&Y/轨迹统计/sanrio/"
        elif analyse_obj == "xyl":
            root_dir = "/Volumes/B&Y/轨迹统计/xyl/"
        elif analyse_obj == "kaer":
            root_dir = "/Volumes/B&Y/轨迹统计/kaer/"
        else:
            root_dir = ""

    usps_arr = []
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    is_morning = (datetime.now().hour) < 12
    gz_time = datetime.strptime(getYmd(), "%Y/%m/%d")
    # print(current_year, current_month, current_day)
    if analyse_obj == ClientConstants.xyl:
        track_day = 10
    else:
        track_day = 5
    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = re.search(r"创建时间(\d+)_", file)
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= track_day:
                        print(f"正在处理文件: {xlsx_path}")

                        if ignore:
                            # go(analyse_obj, xlsx_path)
                            usps_arr.append(xlsx_path)
                            continue

                        ck_time = get_days_difference(xlsx_path)
                        interval_time = (gz_time - datetime.strptime(ck_time, "%Y/%m/%d")).days

                        if interval_time == 1 and is_morning:
                            # go(analyse_obj, xlsx_path)
                            usps_arr.append(xlsx_path)
                        else:
                            if is_morning:
                                continue

                            if (analyse_obj_ignore):
                                # go(analyse_obj, xlsx_path)
                                usps_arr.append(xlsx_path)
                                continue

                            if (check_and_add_courier_column(xlsx_path)):
                                # go(analyse_obj, xlsx_path)
                                usps_arr.append(xlsx_path)
                            else:
                                shipment_received_interval2_list = get_shipment_received_numbers(xlsx_path, getYmd())
                                change_shipment_received_count = len(shipment_received_interval2_list)

                                output_file = os.path.splitext(xlsx_path)[0] + "_去重.xlsx"
                                filter_tracking_numbers(xlsx_path, output_file)

                                # 同一单会有多个sku，多个sku会生成多行数据，分析sku的时候不能去重，其它的需要去重
                                total_count = remove_duplicates_by_column(output_file, output_file, RowName.Tracking_No)

                                patterns = {
                                    CourierStateMapValue.no_track: Pattern.no_track,
                                    CourierStateMapValue.delivered: Pattern.delivered,
                                    CourierStateMapValue.unpaid: Pattern.unpaid,
                                    CourierStateMapValue.not_yet: Pattern.not_yet,
                                    CourierStateMapValue.pre_ship: Pattern.pre_ship,
                                    CourierStateMapValue.irregular_no_tracking: Pattern.irregular_no_tracking,
                                    CourierStateMapValue.no_tracking: Pattern.no_tracking,
                                    CourierStateMapValue.tracking: Pattern.tracking
                                }

                                count_dict = count_pattern_and_tracking_with_sf_date(output_file, RowName.Courier,
                                                                                     RowName.SfDateInterval, patterns)

                                no_track_count = count_dict[CourierStateMapValue.no_track]
                                delivered_count = count_dict[CourierStateMapValue.delivered]
                                unpaid_count = count_dict[CourierStateMapValue.unpaid]
                                not_yet_count = count_dict[CourierStateMapValue.not_yet]
                                pre_ship_count = count_dict[CourierStateMapValue.pre_ship]
                                irregular_no_tracking_count = count_dict[CourierStateMapValue.irregular_no_tracking]
                                no_tracking_count = count_dict[CourierStateMapValue.no_tracking]
                                tracking_count = count_dict[CourierStateMapValue.tracking]
                                tracking_zero_count = count_dict["sfDateInterval"]

                                # 先进行一次计算，并缓存结果
                                total_count_int = int(total_count)
                                no_track_count_int = int(no_track_count)
                                track_count_int = total_count_int - no_track_count_int
                                tracking_zero_count_int = int(tracking_zero_count)
                                delivered_count_int = int(delivered_count)
                                unpaid_count_int = int(unpaid_count)
                                not_yet_count_int = int(not_yet_count)
                                pre_ship_count_int = int(pre_ship_count)
                                irregular_no_tracking_count_int = int(irregular_no_tracking_count)
                                no_tracking_count_int = int(no_tracking_count)
                                tracking_count_int = int(tracking_count)
                                # real_no_track_count = no_track_count_int + tracking_zero_count_int  # 真正的未上网数
                                # real_track_count = total_count_int - no_track_count  # 真正的上网数
                                # real_tracking_count = tracking_count_int - tracking_zero_count_int
                                if tracking_zero_count_int == 0 and delivered_count_int == 0 and unpaid_count_int == 0 \
                                        and not_yet_count_int == 0 and pre_ship_count_int == 0 and no_tracking_count_int == 0 \
                                        and tracking_count_int == 0:
                                    delete_file(output_file)
                                    usps_arr.append(xlsx_path)
                                    continue

                                # 计算百分比
                                swl = round2(100 - ((no_track_count_int) / total_count_int * 100))
                                wswl = round2(100 - swl)
                                qsl = round2((delivered_count_int / total_count_int) * 100)
                                unpaidl = round2((unpaid_count_int / total_count_int) * 100)
                                not_yetl = round2((not_yet_count_int / total_count_int) * 100)
                                pre_shipl = round2((pre_ship_count_int / total_count_int) * 100)
                                irregular_no_trackingl = round2(
                                    (irregular_no_tracking_count_int / total_count_int) * 100)
                                no_tracking_countl = round2((no_tracking_count_int / total_count_int) * 100)
                                tracking_countl = round2((tracking_count_int / total_count_int) * 100)
                                tracking_zero_countl = round2((tracking_zero_count_int / total_count_int) * 100)
                                change_shipment_received_countl = round2(
                                    (change_shipment_received_count / total_count_int) * 100)

                                delete_file(output_file)

                                print(
                                    f"swl:{swl}, change_shipment_received_count:{change_shipment_received_count}, unpaid_count:{unpaid_count}, exceed:{exceed}, qsl:{qsl},")
                                if swl < 99 or change_shipment_received_count >= 10 or unpaid_count > 0 or \
                                        (exceed >= 14 and qsl < 98):
                                    # go(analyse_obj, xlsx_path)
                                    usps_arr.append(xlsx_path)

    return usps_arr


def filter_track_num(xlsx_path):
    process_tracking_no(xlsx_path)
    check_and_add_courier_column(xlsx_path)

    irregular_number_map = find_irregular_tracking_numbers(xlsx_path)
    irregular_number_list = []
    if irregular_number_map:
        irregular_number_list = list(irregular_number_map.keys())
        update_courier_status1(xlsx_path, irregular_number_map)

    return split_and_join_by_35_to_map(extract_and_process_data(xlsx_path, RowName.Courier))


def update_tracking_info(input_file, output_file, tracking_dict, analyse_obj):
    df = pd.read_excel(input_file, dtype=str)

    needed_columns = [
        RowName.Courier, RowName.LatestEventSfDate, RowName.LatestEventSfTime, RowName.UnpaidDate,
        RowName.LatestEventSfSite, RowName.PossessionSfDate, RowName.SfDateInterval
    ]
    for col in needed_columns:
        if col not in df.columns:
            df[col] = ""

    time_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    strftime = datetime.now().strftime("%Y-%m-%d")

    for idx, row in df.iterrows():
        if analyse_obj == "flld":
            tracking_no = str(row.get("快递单号", "")).strip()
        elif analyse_obj == "dxm_xyl_yd":
            tracking_no = str(row.get("运单号", "")).strip()
        else:
            tracking_no = str(row.get(RowName.Tracking_No, "")).strip()
        if tracking_no in tracking_dict:
            data = tracking_dict[tracking_no]

            courier_state = str(data.get("Courier", "") or "")
            df.at[idx, RowName.Courier] = courier_state
            df.at[idx, RowName.LatestEventSfDate] = str(data.get("LastEventDate", "") or "")
            df.at[idx, RowName.LatestEventSfTime] = str(data.get("LastEventTime", "") or "")
            df.at[idx, RowName.LatestEventSfSite] = str(data.get("LastEventSite", "") or "")
            df.at[idx, RowName.PossessionSfDate] = str(data.get("PossessionLastDate", "") or "")

            if courier_state == "unpaid":
                df.at[idx, RowName.UnpaidDate] = strftime

            if analyse_obj == "flld" or analyse_obj == "zbw" or analyse_obj == "sanrio" or analyse_obj == "xyl":
                df.at[idx, RowName.Tacking_Time] = time_str

    df.to_excel(output_file, index=False, engine='openpyxl')


############################flld############################
############################flld############################
############################flld############################

def delete_files(file_paths):
    """
    遍历文件路径集合并删除文件。

    :param file_paths: 可迭代对象，如 list、set，包含完整文件路径字符串
    """
    for path in file_paths:
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"✅ 已删除: {path}")
            else:
                print(f"⚠️ 文件不存在: {path}")
        except Exception as e:
            print(f"❌ 删除失败: {path}，原因: {e}")


def merge_csv_files_to_excel(csv_files, output_dir):
    """
    合并多个 CSV 文件，只保留第一个文件的列头，并保存为 Excel 文件。
    同时提取“打单时间”首条值的月份格式（如2025.5）和日期（如05-24），打印有效行数。
    """
    if not csv_files:
        print("❌ 输入文件列表为空")
        return

    combined_df = pd.DataFrame()

    for i, file in enumerate(csv_files):
        try:
            df = pd.read_csv(file, dtype=str).fillna("")
            if df.empty:
                print(f"⚠️ 文件为空，已跳过: {file}")
                continue

            if i == 0:
                combined_df = df
            else:
                if df.columns.equals(combined_df.columns):
                    combined_df = pd.concat([combined_df, df], ignore_index=True)
                else:
                    print(f"⚠️ 列不匹配，跳过文件: {file}")
        except Exception as e:
            print(f"❌ 读取失败: {file}，原因: {e}")

    if combined_df.empty:
        print("❌ 无有效数据，未生成 Excel 文件")
        return

    # 有效行：去除全空行
    valid_rows = combined_df[combined_df.apply(lambda row: row.astype(str).str.strip().any(), axis=1)]
    valid_count = len(valid_rows)

    # 提取“打单时间”第一条数据并处理
    if "打单时间" in valid_rows.columns:
        first_time_str = valid_rows["打单时间"].dropna().astype(str).str.strip().iloc[0]
        try:
            # 解析格式，如 05-24 14:54
            dt = datetime.strptime(first_time_str, "%m-%d %H:%M")
            # 添加年份，拼接为 “2025.5”
            month_info = f"2025.{dt.month}"
            # 获取日期部分
            date_only = str(dt.day)
            print(f"📅 打单时间首条记录：{first_time_str}")
            print(f"📅 转换后的月份格式：{month_info}")
            print(f"📅 提取的日期：{date_only}")
        except Exception as e:
            print(f"⚠️ 时间格式解析失败: {first_time_str}，错误: {e}")
    else:
        print("⚠️ 未找到“打单时间”列，跳过时间处理")

    print(f"✅ 有效数据行数（不含表头）：{valid_count}")

    # ✅ 创建输出目录（如果不存在）
    output_dir_month = output_dir + month_info + "/"
    create_dir = os.path.dirname(output_dir + month_info + "/")
    os.makedirs(create_dir, exist_ok=True)

    output_path = output_dir_month + f"打单时间{date_only}_{valid_count}.xlsx"

    combined_df.to_excel(output_path, index=False)
    print(f"✅ 合并完成，保存至: {output_path}")


def copy_new_file(input_path, output_path):
    """
    复制生成新文件
    """
    # 读取原始 Excel 文件
    df = pd.read_excel(input_path, dtype=str)
    # 将数据写入新的 Excel 文件
    df.to_excel(output_path, index=False)


def detect_duplicate_prefix_suffix(root_path, dir_path):
    prefix_suffix_map = defaultdict(list)

    for filename in os.listdir(dir_path):
        if not filename.lower().endswith('.csv'):
            continue  # 只处理 .csv 文件

        # 提取 '打单时间X' 前缀
        match = re.match(r"(打单时间\d+)_\d+\.csv", filename)
        if match:
            prefix = match.group(1)  # 如 "打单时间1"
            suffix = ".csv"
            full_path = os.path.join(dir_path, filename)
            prefix_suffix_map[(prefix, suffix)].append(full_path)

    for (prefix, suffix), files in prefix_suffix_map.items():
        if len(files) >= 1:
            print(f"📁 找到同组文件（前缀: {prefix}, 后缀: {suffix}）共 {len(files)} 个:")
            for f in files:
                print(f"   - {f}")
            merge_csv_files_to_excel(files, root_path)
            delete_files(files)


def count_pattern_state(file_path, column_name, patternStr):
    """
    统计指定列指定内容的数量
    """
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        column_index = headers.index(column_name) + 1
        pattern = re.compile(patternStr, re.IGNORECASE)
        total_count = 0
        no_track_count = 0
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            cell_value = row[column_index - 1]
            if cell_value is not None:
                total_count += 1
                if pattern.search(str(cell_value)):
                    no_track_count += 1
        return total_count, no_track_count
    except Exception as e:
        print(f"发生错误: {e}")
        return 0, 0


def tkkj_auto():
    if "MacBookPro" in get_computer_model():
        root_dir = "/Users/zkp/Desktop/B&Y/轨迹统计/dxm_xyl_track/"
    else:
        root_dir = "/Volumes/B&Y/轨迹统计/dxm_xyl_track/"
    usps_arr = []
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    track_day = 10
    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = re.search(r"发货时间(\d+)_", file)
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= track_day:
                        print(f"正在处理文件: {xlsx_path}")
                        usps_arr.append(xlsx_path)
    return usps_arr


def flld_automatic(analyse_obj, ignore, analyse_obj_ignore):
    root_dir = ""
    if "MacBookPro" in get_computer_model():
        if analyse_obj == "flld":
            root_dir = "/Users/zkp/Desktop/B&Y/轨迹统计/flld/"
        else:
            root_dir = ""
    else:
        if analyse_obj == "flld":
            root_dir = "/Volumes/B&Y/轨迹统计/flld/"
        else:
            root_dir = ""

    usps_arr = []
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    is_morning = (datetime.now().hour) < 12
    gz_time = getYmd()
    track_day = 10

    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            if analyse_obj == "flld":
                detect_duplicate_prefix_suffix(root_dir, sun_dir_path)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = None
                if analyse_obj == "flld":
                    match = re.search(r"打单时间(\d+)_", file)
                else:
                    match = None
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= track_day:
                        print(f"正在处理文件: {xlsx_path}")

                        if ignore:
                            # go(xlsx_path)
                            usps_arr.append(xlsx_path)
                            continue

                        if analyse_obj == "flld":
                            ck_time = get_days_difference_flld(xlsx_path)
                        else:
                            ck_time = None
                        interval_time = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(ck_time,
                                                                                                    "%Y/%m/%d")).days

                        if interval_time == 1 and is_morning:
                            # go(xlsx_path)
                            usps_arr.append(xlsx_path)
                        else:
                            if is_morning:
                                continue

                            if (analyse_obj_ignore):
                                # go(xlsx_path)
                                usps_arr.append(xlsx_path)
                                continue

                            if (check_and_add_courier_column(xlsx_path)):
                                # go(xlsx_path)
                                usps_arr.append(xlsx_path)
                            else:

                                output_file = os.path.splitext(xlsx_path)[0] + "_复制.xlsx"
                                copy_new_file(xlsx_path, output_file)

                                total_count, no_track_count = count_pattern_state(output_file, RowName.Courier,
                                                                                  Pattern.no_track)
                                track_count = total_count - no_track_count
                                total_count2, delivered_count = count_pattern_state(output_file, RowName.Courier,
                                                                                    Pattern.delivered)
                                total_count3, unpaid_count = count_pattern_state(output_file, RowName.Courier,
                                                                                 Pattern.unpaid)
                                total_count4, not_yet_count = count_pattern_state(output_file, RowName.Courier,
                                                                                  r"not_yet")
                                total_count5, pre_ship_count = count_pattern_state(output_file, RowName.Courier,
                                                                                   r"pre_ship")
                                total_count6, alert_count = count_pattern_state(output_file, RowName.Courier,
                                                                                Pattern.alert)

                                if delivered_count == 0 and unpaid_count == 0 \
                                        and not_yet_count == 0 and pre_ship_count == 0 and alert_count == 0 \
                                        and track_count == 0:
                                    delete_file(output_file)
                                    usps_arr.append(xlsx_path)
                                    continue

                                swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))
                                unpaid_countl = round2((int(unpaid_count) / int(total_count)) * 100)
                                delivered_countl = round2((int(delivered_count) / int(total_count)) * 100)

                                delete_file(output_file)

                                if swl < 99 or unpaid_count > 0 or (exceed >= 14 and delivered_countl < 98):
                                    # go(xlsx_path)
                                    usps_arr.append(xlsx_path)

    return usps_arr


def convert_csv_to_xlsx(csv_file, xlsx_file):
    """
    将 CSV 文件转换为 XLSX 文件格式。

    :param csv_file: 输入的 CSV 文件路径
    :param xlsx_file: 输出的 XLSX 文件路径
    """
    try:
        # 读取 CSV 文件
        data = pd.read_csv(csv_file)

        # 将数据写入 XLSX 文件
        data.to_excel(xlsx_file, index=False, engine="openpyxl")

        print(f"文件已成功转换为 XLSX 格式: {xlsx_file}")
    except Exception as e:
        print(f"转换过程中发生错误: {e}")


def extract_path_before_csv(file_path):
    # 判断文件路径是否以 .csv 结尾
    if file_path.endswith('.csv'):
        # 提取 .csv 前的所有字符
        base_path = file_path.rsplit('.csv', 1)[0]
        xlsx_ = base_path + ".xlsx"
        convert_csv_to_xlsx(file_path, xlsx_)
        delete_file(file_path)
        return xlsx_
    else:
        return file_path


def str_strip(filepath: str, column_name: str):
    data = pd.read_excel(filepath, dtype=str)
    data[column_name] = data[column_name].str.replace('\t', '', regex=False).str.strip()
    data.to_excel(filepath, index=False)


def flld_filter_track_num(analyse_obj, input_path):
    if analyse_obj == "flld":
        xlsx_path = extract_path_before_csv(input_path)
        column_name = "快递单号"
    else:
        xlsx_path = input_path
        column_name = "运单号"
    str_strip(xlsx_path, column_name)
    check_and_add_courier_column(xlsx_path)

    irregular_number_map = find_irregular_tracking_numbers(xlsx_path, column_name)
    irregular_number_list = []
    if irregular_number_map:
        irregular_number_list = list(irregular_number_map.keys())
        update_courier_status1(xlsx_path, irregular_number_map, column_name)

    return split_and_join_by_35_to_map(extract_and_process_data(xlsx_path, RowName.Courier, column_name))


def split_and_join_by_35_to_map(original_list):
    num35_map = {}
    for i in range(0, len(original_list), 35):
        chunk = original_list[i:i + 35]
        joined = ''.join(str(item) + '%2C' for item in chunk)
        num35_map[joined] = len(chunk) + 1
    return num35_map


def extract_datetime_info(text):
    date_str = ""
    time_str = ""

    # 匹配完整的日期 + 时间（支持 "at" 或 "," 连接）
    datetime_match = re.search(
        r'([A-Za-z]+ \d{1,2}, \d{4})[, ]+(?:at )?(\d{1,2}:\d{2}) ?(am|pm)',
        text, re.IGNORECASE
    )

    if datetime_match:
        try:
            full_str = f"{datetime_match.group(1)} {datetime_match.group(2)} {datetime_match.group(3)}"
            dt = datetime.strptime(full_str, "%B %d, %Y %I:%M %p")
            date_str = dt.strftime("%Y-%m-%d")
            time_str = dt.strftime("%H:%M")
        except Exception as e:
            print(f"⚠️ 解析失败: {e}")
    else:
        # 只提取日期
        date_match = re.search(r'([A-Za-z]+ \d{1,2}, \d{4})', text)
        if date_match:
            try:
                dt = datetime.strptime(date_match.group(1), "%B %d, %Y")
                date_str = dt.strftime("%Y-%m-%d")
            except:
                pass

        # 只提取时间
        time_match = re.search(r'(\d{1,2}:\d{2}) ?(am|pm)', text, re.IGNORECASE)
        if time_match:
            try:
                t = datetime.strptime(time_match.group(0).lower(), "%I:%M %p")
                time_str = t.strftime("%H:%M")
            except:
                pass

    return date_str, time_str


def extract_site_info(text):
    # 提取地点
    moving = "transit to the next facility"
    # 尝试匹配 "in xxx." 这种结构（例如 in NORCO, CA 92860.）
    in_match = re.search(r'\bin\s+([A-Z0-9 ,\-]+)[\.\n]?', text)
    # 尝试匹配 “our ... on” 这种结构
    our_on_match = re.search(r'\bour\s+(.*?)\s+on\b', text)

    location = ""
    if moving in text.lower():
        location = moving
    elif in_match:
        location = in_match.group(1).strip()
    elif our_on_match:
        location = our_on_match.group(1).strip()

    return location


def extract_info(result_map, excel_result_map):
    for key, value in result_map.items():
        latest_event = value.get("LatestEvent", "")
        possession_last_event = value.get("PossessionLastEvent", "")
        possession_first_event = value.get("PossessionFirstEvent", "")
        possession_second_event = value.get("PossessionSecondEvent", "")
        possession_newest_time_event = value.get("PossessionNewestTimeEvent", "")

        objs = {}
        if "Alert" in possession_first_event:
            if "Unpaid postage" in possession_last_event \
                    or "Delayed for postage assessment, held awaiting payment" in possession_second_event:
                objs["Courier"] = "unpaid"
            else:
                if "Vacant" in possession_second_event:
                    objs["Courier"] = "alert_vacant"
                elif "Awaiting Delivery" in possession_second_event:
                    objs["Courier"] = "alert_Awaiting Delivery"
                elif "Sent to Mail Recovery Center" in possession_second_event:
                    objs["Courier"] = "alert_Sent to Mail Recovery Center"
                elif "Contact Customer Care" in possession_second_event:
                    objs["Courier"] = "alert_Contact Customer Care"
                elif "No Access to Delivery Location" in possession_second_event:
                    objs["Courier"] = "alert_No Access to Delivery Location"
                else:
                    objs["Courier"] = "alert"
        elif "Delivery Attempt" in possession_first_event:
            objs["Courier"] = "alert_Delivery Attempt: Action Needed"
        elif "Delivered" in possession_first_event:
            if "Delivered, Individual Picked Up at Post Office" in possession_first_event:
                objs["Courier"] = "delivered"
            else:
                if "Your item was delivered" in latest_event or "Your item has been delivered" in latest_event:
                    objs["Courier"] = "delivered"
                else:
                    objs["Courier"] = "tracking"
        elif "Pre-Shipment" in possession_first_event:
            objs["Courier"] = "pre_ship"
        else:
            if "pending" in latest_event:
                objs["Courier"] = "acceptance_pending"
            elif "not yet" in latest_event:
                objs["Courier"] = "not_yet"
            else:
                objs["Courier"] = "tracking"

        date_str, time_str = extract_datetime_info(latest_event)
        location = extract_site_info(latest_event)
        if date_str == "" and time_str == "" and location == "":  # 如果左侧最新轨迹信息没有日期
            date_str2, time_str2 = extract_datetime_info(possession_second_event)  # 获取轨迹历史消息中的第二条消息获取
            location2 = extract_site_info(possession_second_event)
            objs["LastEventDate"] = date_str2
            objs["LastEventTime"] = time_str2
            objs["LastEventSite"] = location2
        else:
            objs["LastEventDate"] = date_str
            objs["LastEventTime"] = time_str
            objs["LastEventSite"] = location

        if not (isinstance(objs["LastEventDate"], str) and objs["LastEventDate"].strip() != ""):
            newest_dates, newest_times = extract_first_date_time_node(possession_newest_time_event)
            objs["LastEventDate"] = newest_dates
            objs["LastEventTime"] = newest_times

        if not (isinstance(objs["LastEventSite"], str) and objs["LastEventSite"].strip() != ""):
            objs["LastEventSite"] = extract_first_datetime_site(possession_newest_time_event)

        if objs["Courier"] == "alert_Delivery Attempt: Action Needed":
            newest_dates, newest_times = extract_first_date_time_node(possession_newest_time_event, "last")
            objs["LastEventDate"] = newest_dates
            objs["LastEventTime"] = newest_times

        date_str1, time_str1 = extract_datetime_info(possession_last_event)
        objs["PossessionLastDate"] = date_str1
        objs["PossessionLastTime"] = time_str1

        excel_result_map[key] = objs


def get_first_datetime_element(arr):
    # 匹配 "August 20, 2025" 或 "August 19, 2025, 5:31 am"
    pattern = re.compile(r"([A-Za-z]+\s+\d{1,2},\s+\d{4}(?:,\s*\d{1,2}:\d{2}\s*(?:am|pm))?)")

    for ele in arr:
        match = pattern.search(ele)
        if match:
            return ele.strip()  # 返回第一个包含日期的完整元素
    return None


def extract_first_date_time_node(text, mode="first"):
    """
    提取文本中的日期和时间
    :param text: 输入文本
    :param mode: "first" 表示提取第一个匹配日期; "last" 表示提取最后一个 \n\n 后的日期
    :return: (日期, 时间) -> ("%Y-%m-%d", "%H:%M")
    """
    # 匹配类似 "August 19, 2025" 或 "August 19, 2025, 5:31 am"
    pattern = re.compile(
        r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{1,2},\s+\d{4}(?:,\s*\d{1,2}:\d{2}\s*(?:am|pm))?",
        re.I)

    raw_date = None

    if mode == "first":
        # 取第一个匹配
        match = pattern.search(text)
        if match:
            raw_date = match.group(0).strip()

    elif mode == "last":
        # 取最后一个 \n\n 后的时间
        if "\n\n" in text:
            last_part = text.split("\n\n")[-1].strip()
            match = pattern.search(last_part)
            if match:
                raw_date = match.group(0).strip()
        else:
            # 如果没有 \n\n，就退化为第一个匹配
            match = pattern.search(text)
            if match:
                raw_date = match.group(0).strip()

    if raw_date:
        try:
            # 先尝试 "月 日, 年, 时:分 am/pm"
            dt = datetime.strptime(raw_date, "%B %d, %Y, %I:%M %p")
            return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")
        except ValueError:
            # 再尝试 "月 日, 年"
            dt = datetime.strptime(raw_date, "%B %d, %Y")
            return dt.strftime("%Y-%m-%d"), ""

    return "", ""


def extract_first_datetime_site(text):
    # 按两个换行符切分
    parts = text.split("\n\n")
    if len(parts) >= 3:
        return parts[-2].strip()  # 倒数第二个就是时间前的内容
    return ""


def get_specified_node_info(progress_sub_ele_arr):
    possession_first_event = ""
    possession_second_event = ""
    possession_last_event = ""
    possession_newest_time_event = ""

    if progress_sub_ele_arr:
        # 获取第一个元素
        first_ele = progress_sub_ele_arr[0]
        possession_first_event = first_ele
        if first_ele == "Alert" or first_ele == "Delivery Attempt: Action Needed":
            second_elem = progress_sub_ele_arr[1]
            possession_second_event = second_elem

            # 获取最后一个元素
        last_elem = progress_sub_ele_arr[-1]
        if last_elem == "See All Tracking History":
            if len(progress_sub_ele_arr) >= 2:
                raw_text = progress_sub_ele_arr[-2]
                cleaned = re.sub(r'[\n\t]+', '', raw_text).strip()
                possession_last_event = cleaned
        else:
            possession_last_event = last_elem

        possession_newest_time_event = get_first_datetime_element(progress_sub_ele_arr)

    return {"possession_first_event": possession_first_event,
            "possession_second_event": possession_second_event,
            "possession_last_event": possession_last_event,
            "possession_newest_time_event": possession_newest_time_event}


def find_existing_same_prefix_files(new_file_path):
    """查找与 new_file_path 同目录且符合相同文件条件的文件"""
    dir_path = os.path.dirname(new_file_path)
    new_file = os.path.basename(new_file_path)

    if "_" not in new_file or "." not in new_file:
        return []

    new_prefix = new_file.split("_")[0]  # 第一个 _ 之前
    new_suffix = new_file.split(".")[-1]  # 最后 . 之后

    existing_files = []
    for f in os.listdir(dir_path):
        f_prefix = f.split("_")[0]
        f_suffix = f.split(".")[-1]
        if f_prefix == new_prefix and f_suffix == new_suffix:
            existing_files.append(os.path.join(dir_path, f))
    return existing_files


def split_excel_by_date_and_unique_count(
        input_file,
        time_column,
        unique_column,
        track_column,
        file_prefix,
        output_dir
):
    # 读取 Excel 文件
    df = pd.read_excel(input_file, dtype=str)

    if time_column not in df.columns:
        raise ValueError(f"Excel 中没有找到列: {time_column}")
    if unique_column not in df.columns:
        raise ValueError(f"Excel 中没有找到列: {unique_column}")

    # ✅ 针对 flld 文件夹：自定义时间解析逻辑
    if output_dir == "/Users/zkp/Desktop/B&Y/轨迹统计/flld/":
        current_year = datetime.now().year

        def parse_custom_time(x):
            if isinstance(x, str):
                if len(x.split("-")[0]) == 2:  # 没有年份，补上
                    x = f"{current_year}-{x}"
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                    try:
                        return datetime.strptime(x.strip(), fmt)
                    except ValueError:
                        continue
                return pd.NaT
            elif isinstance(x, datetime):
                return x
            else:
                return pd.NaT

        df[time_column] = df[time_column].apply(parse_custom_time)
    else:
        df[time_column] = pd.to_datetime(df[time_column], errors="coerce")

    # 删除无法解析的时间
    df = df.dropna(subset=[time_column])

    # 提取年月日
    df["date_only"] = df[time_column].dt.strftime("%Y-%m-%d")

    # 需要赋值的列
    merge_columns = [
        "Courier/快递",
        "PossessionSfDate/揽收时间",
        "LatestEventSfDate/最新事件时间",
        "SfDateInterval/SF消息间隔",
        "UnpaidDate/unpaid记录时间",
        "LatestEventSfTime/最新事件时间",
        "LatestEventSfSite/最新事件地点",
        "TrackTimeInterval/跟踪时间间隔",
        "TrackTimeIntervalState/跟踪时间间隔状态",
        "Tacking_Time/追踪时间",
        "LastEventSfTime/上一条轨迹时间",
        "YD_Number/阳单号",
        "YD_State/阳单轨迹状态"
    ]

    # 按日期分组
    for dates, group in df.groupby("date_only"):
        dt = datetime.strptime(dates, "%Y-%m-%d")
        year, month, day = dt.year, dt.month, dt.day

        # 去重计数
        unique_count = group[unique_column].nunique()

        # 输出目录
        sub_output_dir = os.path.join(output_dir, f"{year}.{month}")
        if not os.path.exists(sub_output_dir):
            os.makedirs(sub_output_dir)

        # 新文件路径
        new_file = os.path.join(sub_output_dir, f"{file_prefix}{day}_{unique_count}.xlsx")

        # === 查找相同前缀的旧文件 ===
        existing_files = find_existing_same_prefix_files(new_file)

        if existing_files:
            for old_file in existing_files:
                df_old = pd.read_excel(old_file)

                if track_column in df_old.columns and track_column in group.columns:
                    # 用 map 填充空值，避免 merge 导致重复行
                    for col in merge_columns:
                        if col not in group.columns:
                            group[col] = pd.NA
                        if col in df_old.columns:
                            mapping = df_old.set_index(track_column)[col].to_dict()
                            mask = group[col].isna()
                            group.loc[mask, col] = group.loc[mask, track_column].map(mapping)

                os.remove(old_file)  # 删除旧文件

        # 保存文件
        group.drop(columns=["date_only"]).to_excel(new_file, index=False)

    # ✅ 删除输入文件
    if os.path.exists(input_file):
        os.remove(input_file)


def merge_csvs_to_excel(folder_path, output_file):
    """
    将指定文件夹中的所有 CSV 文件合并为一个 Excel 文件
    只保留第一个文件的表头，合并完成后删除原 CSV 文件
    """
    # 找到文件夹中所有 CSV 文件
    csv_files = glob.glob(os.path.join(folder_path, "*.csv"))
    if not csv_files:
        print("❌ 没有找到 CSV 文件")
        return

    combined_df = pd.DataFrame()

    for i, file in enumerate(csv_files):
        try:
            df = pd.read_csv(file, dtype=str).fillna("")
            if df.empty:
                print(f"⚠️ 文件为空，跳过: {file}")
                continue

            if i == 0:
                combined_df = df
            else:
                if df.columns.equals(combined_df.columns):
                    combined_df = pd.concat([combined_df, df], ignore_index=True)
                else:
                    print(f"⚠️ 列不匹配，跳过文件: {file}")
        except Exception as e:
            print(f"❌ 读取失败 {file}，原因: {e}")

    if combined_df.empty:
        print("❌ 没有有效数据")
        return

    # 保存为 Excel
    combined_df.to_excel(output_file, index=False)
    print(f"✅ 合并完成，保存到: {output_file}")

    # 删除原始 CSV 文件
    for file in csv_files:
        try:
            os.remove(file)
            print(f"🗑️ 已删除: {file}")
        except Exception as e:
            print(f"⚠️ 删除失败 {file}，原因: {e}")
