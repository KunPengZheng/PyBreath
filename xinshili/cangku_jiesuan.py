import os
import re
import pandas as pd
from xinshili.gjgz_plus333 import check_and_add_courier_column, RowName, CourierStateMapValue, \
    find_irregular_tracking_numbers, update_courier_status1, extract_and_process_data, CourierStateMapKey, \
    update_courier_status
from xinshili.pd_utils import remove_duplicates_by_column
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from xinshili.utils import ensure_directory_exists


def update_courier(file1_path, file2_path, output_path):
    """
    根据两个 Excel 文件中的物流跟踪号进行匹配，
    如果匹配成功，则将文件1中对应行的“Courier/快递”列内容和“SfDateInterval/SF消息间隔”列内容
    复制到文件2中对应行的“Courier/快递”列和“SfDateInterval/SF消息间隔”列，
    并将更新后的结果保存到 output_path 中。

    参数：
      file1_path: str，文件1路径，必须包含“Tracking No./物流跟踪号”和“Courier/快递”列
      file2_path: str，文件2路径，必须包含“Package 1 Tracking No./物流跟踪号1”和“Courier/快递”列
      output_path: str，保存更新后文件的路径
    """
    # 读取 Excel 文件
    df1 = pd.read_excel(file1_path)
    df2 = pd.read_excel(file2_path)

    # 自动选择 df2 中的匹配列（left_on）
    if RowName.Tracking_No in df2.columns:
        left_on_col = RowName.Tracking_No
    elif RowName.Package1_Tracking in df2.columns:
        left_on_col = RowName.Package1_Tracking
    else:
        raise ValueError(
            f"❌ 文件2中未找到可用的跟踪号列（如 '{RowName.Package1_Tracking}' 或 '{RowName.Tracking_No}'）")

    # 提取文件1中需要的列：用于匹配的物流跟踪号、快递信息和消息间隔
    df1_subset = df1[[RowName.Tracking_No, RowName.Courier, RowName.SfDateInterval]]

    # 检查文件1是否包含其他需要的列（可选的列）
    optional_columns = [RowName.PossessionSfDate, RowName.LatestEventSfDate]
    for col in optional_columns:
        if col in df1.columns:
            df1_subset[col] = df1[col]

    # 合并两个 DataFrame：以文件2为主表，通过文件2的“Package 1 Tracking No./物流跟踪号1”与文件1的“Tracking No./物流跟踪号”匹配
    merged_df = pd.merge(
        df2,
        df1_subset,
        left_on=left_on_col,
        right_on=RowName.Tracking_No,
        how="left",
        suffixes=("", "_file1")
    )

    # 将合并后文件1中的“Courier/快递”和“SfDateInterval/SF消息间隔”列数据赋值到文件2对应的列中
    merged_df[RowName.Courier] = merged_df[RowName.Courier_File1]
    merged_df[RowName.SfDateInterval] = merged_df[RowName.SfDateInterval_File1]

    # 处理可选列：检查是否存在并复制值
    if RowName.PossessionSfDate_File1 in merged_df.columns:
        merged_df[RowName.PossessionSfDate] = merged_df[RowName.PossessionSfDate_File1]

    if RowName.LatestEventSfDate_File1 in merged_df.columns:
        merged_df[RowName.LatestEventSfDate] = merged_df[RowName.LatestEventSfDate_File1]

    # 删除合并过程中产生的临时列
    merged_df.drop(columns=[RowName.Courier_File1,
                            RowName.SfDateInterval_File1,
                            RowName.PossessionSfDate_File1,
                            RowName.LatestEventSfDate_File1],
                   inplace=True)

    # 将更新后的 DataFrame 保存为新的 Excel 文件
    merged_df.to_excel(output_path, index=False)

    print(f"更新后的文件已保存为 {output_path}")
    return left_on_col


def me(save_paths, target_dir, pattern_flag=True):
    lists = []
    pattern = r"^(出库时间|创建时间)\d+_\d+\.xlsx$"

    # 遍历目录，匹配符合条件的文件
    for root, dirs, files in os.walk(target_dir):
        for ele in files:
            if (pattern_flag):
                if re.match(pattern, ele):
                    xlsx_path = os.path.join(root, ele)  # 规范路径
                    try:
                        df = pd.read_excel(xlsx_path)  # 读取 Excel 文件
                        lists.append(df)  # 存入 DataFrame
                    except Exception as e:
                        print(f"错误: 无法读取文件 {xlsx_path}，错误信息: {e}")
            else:
                # 仅处理 .xlsx 或 .xls 文件
                if not ele.lower().endswith(('.xlsx', '.xls')):
                    continue  # 跳过非 Excel 文件

                xlsx_path = os.path.join(root, ele)  # 规范路径
                try:
                    df = pd.read_excel(xlsx_path)  # 读取 Excel 文件
                    lists.append(df)  # 存入 DataFrame 列表
                except Exception as e:
                    print(f"错误: 无法读取文件 {xlsx_path}，错误信息: {e}")

    # 确保 lists 不是空的
    if lists:
        merged_df = pd.concat(lists, ignore_index=True)  # 合并 DataFrame
        merged_df.to_excel(save_paths, index=False)  # 保存合并后的文件
        print(f"合并完成，结果保存为 {save_paths}")
    else:
        print("错误: 没有找到可合并的 Excel 文件")


def highlight_courier_column(file_path, solid_color="C0D79B"):
    wb = load_workbook(file_path)
    ws = wb.active  # 获取当前活动的工作表

    red_fill = PatternFill(start_color=solid_color, end_color=solid_color, fill_type="solid")

    # 找到 RowName.Courier 列和 "SfDateInterval/SF消息间隔" 列的索引
    courier_col_index = None
    sf_date_interval_col_index = None

    # 遍历第一行（列头）找到相关列的索引
    for col_num, col_cell in enumerate(ws[1], 1):
        if col_cell.value == RowName.Courier:
            courier_col_index = col_num
        elif col_cell.value == RowName.SfDateInterval:
            sf_date_interval_col_index = col_num

    if courier_col_index is None:
        print(f"错误: 未找到 {RowName.Courier} 列")
        return
    if sf_date_interval_col_index is None:
        print(f"警告: 未找到 {RowName.SfDateInterval} 列，跳过该列的检查")
        # 如果没有该列，我们只根据 'Courier/快递' 列进行标记
        sf_date_interval_col_index = -1  # 设置为 -1，表示不使用这个列

    # 遍历列中的每一行，如果满足条件则设置红色背景
    for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过列头）
        courier_cell = ws.cell(row=row, column=courier_col_index)

        # 如果存在 "SfDateInterval/SF消息间隔" 列，并且需要进行比较
        if sf_date_interval_col_index != -1:
            sf_date_interval_cell = ws.cell(row=row, column=sf_date_interval_col_index)
            # 检查 Courier/快递 为特定值并且 SfDateInterval/SF消息间隔 为 0
            if courier_cell.value in [CourierStateMapValue.pre_ship,
                                      CourierStateMapValue.not_yet,
                                      CourierStateMapValue.no_tracking] or \
                    (courier_cell.value == CourierStateMapValue.tracking and sf_date_interval_cell.value == 0):
                for col in range(1, ws.max_column + 1):  # 让整行变红
                    ws.cell(row=row, column=col).fill = red_fill
        else:
            # 如果没有 "SfDateInterval/SF消息间隔" 列，只根据 "Courier/快递" 列进行判断
            if courier_cell.value in [CourierStateMapValue.pre_ship,
                                      CourierStateMapValue.not_yet,
                                      CourierStateMapValue.no_tracking]:
                for col in range(1, ws.max_column + 1):  # 让整行变红
                    ws.cell(row=row, column=col).fill = red_fill

    # 保存 Excel 文件
    wb.save(file_path)
    print(f"标记完成，已更新 {file_path}")


def handle(lists, month_start, month_end, merge_exit=False):
    zbw_gjgz_path = "/Users/zkp/Desktop/B&Y/轨迹统计/zbw/2025."
    sanrio_gjgz_path = "/Users/zkp/Desktop/B&Y/轨迹统计/sanrio/2025."
    xyl_gjgz_path = "/Users/zkp/Desktop/B&Y/轨迹统计/xyl/2025."

    zbw_start_month_dir = f"{zbw_gjgz_path}{month_start}"
    sanrio_start_month_dir = f"{sanrio_gjgz_path}{month_start}"
    xyl_start_month_dir = f"{xyl_gjgz_path}{month_start}"

    zbw_end_month_dir = f"{zbw_gjgz_path}{month_end}"
    sanrio_end_month_dir = f"{sanrio_gjgz_path}{month_end}"
    xyl_end_month_dir = f"{xyl_gjgz_path}{month_end}"

    merge_dir = f"/Users/zkp/Desktop/B&Y/仓库结算/merge_{month_start}_{month_end}/"
    ensure_directory_exists(merge_dir)
    merge_file = f"{merge_dir}merge_{month_start}_{month_end}.xlsx"

    zbw_merge = "zbw_merge_"
    sanrio_merge = "sanrio_merge_"
    xyl_merge = "xyl_merge_"

    if not merge_exit:
        # 客户各自的数据合并
        me(f"{merge_dir}{zbw_merge}{month_start}.xlsx", zbw_start_month_dir)
        me(f"{merge_dir}{sanrio_merge}{month_start}.xlsx", sanrio_start_month_dir)
        me(f"{merge_dir}{xyl_merge}{month_start}.xlsx", xyl_start_month_dir)

        me(f"{merge_dir}{zbw_merge}{month_end}.xlsx", zbw_end_month_dir)
        me(f"{merge_dir}{sanrio_merge}{month_end}.xlsx", sanrio_end_month_dir)
        me(f"{merge_dir}{xyl_merge}{month_end}.xlsx", xyl_end_month_dir)

        # 合并多个客户
        me(merge_file, merge_dir, False)

    # 去重
    remove_duplicates_me_file = os.path.splitext(merge_file)[0] + "_remove_duplicates.xlsx"
    if not merge_exit:
        remove_duplicates_by_column(merge_file, remove_duplicates_me_file, RowName.Tracking_No)

    # 遍历每个待更新的文件
    for item in lists:
        # 补充缺少的列名
        check_and_add_courier_column(item)

        # 根据指定列匹配，匹配成功后将对应列的内容复制过来
        wl_names = update_courier(remove_duplicates_me_file, item, item)

        irregular_number_map = find_irregular_tracking_numbers(item)
        if irregular_number_map:
            update_courier_status1(item, irregular_number_map)

        results = extract_and_process_data(item, RowName.Courier, 100, wl_name=wl_names, ckjs_flag=True)

        all_maps = {
            CourierStateMapKey.not_yet_map: results[CourierStateMapKey.not_yet_map],
            CourierStateMapKey.pre_ship_map: results[CourierStateMapKey.pre_ship_map],
            CourierStateMapKey.unpaid_map: results[CourierStateMapKey.unpaid_map],
            CourierStateMapKey.delivered_map: results[CourierStateMapKey.delivered_map],
            CourierStateMapKey.no_tracking_map: results[CourierStateMapKey.no_tracking_map],
            CourierStateMapKey.tracking_map: results[CourierStateMapKey.tracking_map],
            CourierStateMapKey.possession_sf_date_map: results[CourierStateMapKey.possession_sf_date_map],
            CourierStateMapKey.latest_event_sf_date_map: results[CourierStateMapKey.latest_event_sf_date_map],
            CourierStateMapKey.sf_date_equality_map: results[CourierStateMapKey.sf_date_equality_map],
        }

        column_mapping = {
            CourierStateMapKey.not_yet_map: RowName.Courier,
            CourierStateMapKey.pre_ship_map: RowName.Courier,
            CourierStateMapKey.unpaid_map: RowName.Courier,
            CourierStateMapKey.delivered_map: RowName.Courier,
            CourierStateMapKey.no_tracking_map: RowName.Courier,
            CourierStateMapKey.tracking_map: RowName.Courier,
            CourierStateMapKey.possession_sf_date_map: RowName.PossessionSfDate,
            CourierStateMapKey.latest_event_sf_date_map: RowName.LatestEventSfDate,
            CourierStateMapKey.sf_date_equality_map: RowName.SfDateInterval,
        }

        update_courier_status(item, all_maps, wl=wl_names, column_map=column_mapping)

        # 根据指定条件标记背景颜色
        highlight_courier_column(item)

    print("✅ 所有文件处理完毕。")


if __name__ == '__main__':
    handle(
        lists=[
            "/Users/zkp/Desktop/B&Y/仓库结算/merge_4_5/副本5月其他客户结算.xlsx",
            "/Users/zkp/Desktop/B&Y/仓库结算/merge_4_5/副本5月自发货结算.xlsx",
        ],
        month_start="4",
        month_end="5",
        merge_exit=False
    )
