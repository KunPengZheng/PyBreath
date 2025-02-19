from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os


def load_excel_file(file_path):
    """
    加载 Excel 文件
    """
    try:
        return load_workbook(file_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"未找到文件: {file_path}")
    except Exception as e:
        raise Exception(f"加载文件 {file_path} 时发生错误: {e}")


def get_active_sheet(workbook):
    """
    获取活动表
    """
    try:
        return workbook.active
    except Exception as e:
        raise Exception(f"获取活动表时发生错误: {e}")


def find_column(sheet, column_name):
    """
    动态查找指定列名的列号。

    :param sheet: openpyxl 的 Worksheet 对象，表示 Excel 工作表。
    :param column_name: 需要查找的单一列名（字符串）。
    :return: 对应的列号（整数）。
    :raises: 如果未找到指定列名，抛出 ValueError。

    eg: find_column(sheet,"价格")
    """
    try:
        # 遍历工作表的每一列
        for col in sheet.iter_cols(1, sheet.max_column):
            # 获取当前列的第一行值（表头）
            header = col[0].value

            # 如果表头匹配目标列名，返回对应的列号
            if header == column_name:
                return col[0].column

        # 如果未找到目标列名，抛出异常
        raise ValueError(f"未找到列名: {column_name}")

    except Exception as e:
        # 捕获异常，并抛出自定义错误信息
        raise Exception(f"查找列时发生错误: {e}")


def find_columns(sheet, column_names):
    """
    动态查找多个列名的列号。

    :param sheet: openpyxl 的 Worksheet 对象，表示 Excel 工作表。
    :param column_names: 需要查找的列名列表（list）。
    :return: 包含列名和对应列号的字典 {列名: 列号}。
    :raises: 当某些列名未找到时，抛出 ValueError。
    """
    column_map = {}
    missing_columns = []

    for column_name in column_names:
        try:
            # 调用单列查找函数
            column_map[column_name] = find_column(sheet, column_name)
        except ValueError:
            # 如果列名未找到，记录到缺失列表
            missing_columns.append(column_name)

    # 如果有未找到的列名，抛出异常
    if missing_columns:
        raise ValueError(f"未找到以下列名: {missing_columns}")

    return column_map


def get_cell_value(sheet, row, column):
    """
    通过 openpyxl 获取工作表 sheet 中第 row 行的 column 列单元格的值
    eg: get_columns_value_by_row(sheet1,1,"A")
    """
    return sheet.cell(row, column).value


def get_merged_cell_value(sheet, row, col):
    """
    获取合并单元格的内容，如果单元格是合并区域中的一部分，获取合并区域的左上角单元格的值。
    :param sheet: 工作表
    :param row: 行号
    :param col: 列号
    :return: 单元格的值
    """
    # 遍历所有合并单元格的区域
    for merged_cell in sheet.merged_cells.ranges:
        # 检查当前单元格是否在合并单元格的范围内
        if merged_cell.min_row <= row <= merged_cell.max_row and merged_cell.min_col <= col <= merged_cell.max_col:
            # 如果该单元格在合并区域内，则返回该合并区域的左上角单元格的值
            top_left_cell = sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col)
            return top_left_cell.value

    # 如果没有合并单元格，直接返回单元格的值
    return sheet.cell(row=row, column=col).value


def set_cell_value(sheet, row, column, value):
    """
    通过 openpyxl 设置工作表 sheet 中第 row 行的 column 列单元格的值
    eg: set_columns_value_by_row(sheet1,1,"A","黑色")
    """
    sheet.cell(row, column).value = value


def generate_column_name():
    """
    生成从 A 列到 BZ 列的常量，例如 column_A = "A"
    """
    for i in range(1, 79):  # 1 到 78 对应 A 到 BZ
        column_letter = get_column_letter(i)
        print(f"column_{column_letter} = \"{column_letter}\"")


def generate_column_value():
    """
    生成 A 列到 BZ 列对应的数字常量，例如 column_A = "1"
    """
    columns = {}
    for i in range(1, 79):  # 1 到 78 对应 A 到 BZ
        column_letter = get_column_letter(i)  # 获取列名
        columns[column_letter] = i

    # 打印常量
    for col, num in columns.items():
        print(f"column_{col} = {num}")


def get_max_row(sheet):
    # 获取最大行号
    return sheet.max_row


def match_sign(input_file1, input_file2, output_file, file1_column_name, file2_column_name):
    """
    将input_file1的指定列的内容 和 将input_file2的指定列的内容 进行匹配，匹配到了 则将input_file2指定列的内容 标记为红色
    :param input_file1: 文件1的路径
    :param input_file2: 文件2的路径
    :param output_file: 输出文件的路径
    :param file1_column_name: 文件1指定的列名
    :param file2_column_name: 文件2指定的列名
    """
    wb1 = load_excel_file(input_file1)
    wb2 = load_excel_file(input_file2)

    sheet1 = wb1.active
    sheet2 = wb2.active

    # 获取文件1和文件2的表头
    header1 = [cell.value for cell in sheet1[1]]  # 读取文件1的表头
    header2 = [cell.value for cell in sheet2[1]]  # 读取文件2的表头

    # 找到对应列索引
    try:
        col_index_1 = header1.index(file1_column_name) + 1
        col_index_2 = header2.index(file2_column_name) + 1
    except ValueError:
        print("未找到对应的列，请检查表头名称是否正确")
        exit()

    # 读取文件1的指定列数据，存入集合以加快查询
    tracking_numbers = set(
        sheet1.cell(row, col_index_1).value for row in range(2, sheet1.max_row + 1) if
        sheet1.cell(row, col_index_1).value
    )

    # 颜色填充样式（红色）
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 遍历文件2的指定列，匹配数据
    for row in range(2, sheet2.max_row + 1):  # 从第2行开始（跳过表头）
        cell = sheet2.cell(row, col_index_2)
        if cell.value and cell.value in tracking_numbers:  # 如果匹配成功
            cell.fill = red_fill  # **正确方式：使用 PatternFill 设置背景颜色**

    # 保存修改后的文件
    wb2.save(output_file)
    print("匹配完成，结果已保存到", output_file)


def mark_by_row_content(filepath, src_column_name, dst_column_name, src_search_list):
    """
      根据文件指定列的指定内容，将指定列的对应行的背景颜色设置为黄色
     :param input_file1: 文件1的路径
     :param input_file2: 文件2的路径
     :param output_file: 输出文件的路径
     :param file1_column_name: 文件1指定的列名
     :param file2_column_name: 文件2指定的列名
    """

    # 打开 Excel 文件
    wb = load_workbook(filepath)
    sheet = wb.active  # 默认选择活动工作表

    # 获取文件的表头
    header = [cell.value for cell in sheet[1]]  # 读取表头

    # 获取 src_column_name 和 dst_column_name 的列索引
    try:
        courier_col_idx = header.index(src_column_name) + 1
        tracking_no_col_idx = header.index(dst_column_name) + 1
    except ValueError:
        print(f"未找到{src_column_name}或{dst_column_name}的列，请检查表头名称是否正确")
        return

    # 定义黄色背景填充
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 遍历每一行数据
    for row in range(2, sheet.max_row + 1):  # 从第2行开始，跳过表头
        courier_value = sheet.cell(row=row, column=courier_col_idx).value
        if courier_value in src_search_list:
            tracking_cell = sheet.cell(row=row, column=tracking_no_col_idx)
            tracking_cell.fill = yellow_fill  # 设置背景颜色为黄色

    # 保存修改后的文件
    wb.save(filepath)
    print(f"修改完成，已将背景颜色标记为黄色并保存到 {filepath}")


def merge_xlsx_files(file_paths: list, output_path: str):
    """
    将多个 Excel 文件合并为一个新的 Excel 文件。
    :param file_paths: 要合并的文件路径列表
    :param output_path: 合并后的输出文件路径
    """
    # 创建一个新的工作簿
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = "合并数据"  # 新工作簿中的第一个工作表

    for file_path in file_paths:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            print(f"文件 {file_path} 不存在，跳过")
            continue

        # 加载现有的 Excel 文件
        wb_old = load_workbook(file_path)
        sheet_old = wb_old.active  # 默认读取第一个工作表

        # 获取原工作表的所有数据
        for row in sheet_old.iter_rows(values_only=True):
            ws_new.append(row)  # 将每一行数据添加到新的工作簿

    # 保存合并后的工作簿
    wb_new.save(output_path)
    print(f"所有文件已合并，结果保存为: {output_path}")
