import os
from openpyxl import load_workbook, Workbook


def merge_based_on_largest_header(input_folder, output_file):
    """
    合并文件夹中多个文件的相同列名数据，选择表头最多的文件作为模板。

    :param input_folder: 文件夹路径，包含多个 Excel 文件
    :param output_file: 输出的合并结果文件路径
    """
    try:
        # 找到表头最多的文件
        largest_header_file = None
        largest_header = []
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(input_folder, file_name)
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]  # 获取第一行作为表头
                if len(headers) > len(largest_header):
                    largest_header = headers
                    largest_header_file = file_path
                wb.close()

        if not largest_header_file:
            raise FileNotFoundError("未找到任何 .xlsx 文件")

        print(f"表头最多的文件: {largest_header_file}")
        print(f"表头内容: {largest_header}")

        # 创建结果数据字典
        result_data = {header: [] for header in largest_header}

        # 遍历所有文件并合并数据
        for file_name in os.listdir(input_folder):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(input_folder, file_name)
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]  # 获取第一行作为表头
                rows = list(ws.iter_rows(min_row=2, values_only=True))  # 获取数据行

                # 将行数据转换为字典形式
                for row in rows:
                    row_dict = {headers[idx]: row[idx] for idx in range(len(headers)) if headers[idx] in result_data}
                    for header in largest_header:
                        result_data[header].append(row_dict.get(header, None))  # 如果列不存在，则填充 None

                wb.close()

        # 写入合并结果到输出文件
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "合并结果"

        # 写入表头
        output_ws.append(list(result_data.keys()))

        # 写入数据
        for row in zip(*result_data.values()):
            output_ws.append(row)

        # 自动调整列宽
        for col in output_ws.columns:
            max_length = 0
            col_letter = col[0].column_letter  # 列字母
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            output_ws.column_dimensions[col_letter].width = adjusted_width

        # 保存结果文件
        output_wb.save(output_file)
        print(f"合并完成，结果已保存到: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")

# 示例使用
# input_folder = input("请输入需要合并的文件所在的目录路径：")  # 存放 Excel 文件的文件夹路径
# output_file = input_folder + "/combined.xlsx"
# merge_based_on_largest_header(input_folder, output_file)
