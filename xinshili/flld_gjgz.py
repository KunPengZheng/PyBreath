import os
import re
from datetime import datetime, timedelta
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook

from xinshili.fs_utils_plus import get_token, brief_sheet_value, ClientConstants, brief_sheet_bg, khhz_sheet_value, \
    khhz_sheet_bg, fs_msg, FsUserID
from xinshili.gjgz_plus333 import RowName, check_and_add_courier_column, extract_and_process_data_flld, \
    update_courier_status_flld, count_pattern_state, CourierStateMapKey, Pattern, is_time_difference_exceed, \
    extract_and_process_data, update_courier_status
from xinshili.pd_utils import copy_new_file
from xinshili.utils import convert_csv_to_xlsx, delete_file, getYmd, round2, is_us_weekend, natural_key


def extract_path_before_csv(file_path):
    # 判断文件路径是否以 .csv 结尾
    if file_path.endswith('.csv'):
        # 提取 .csv 前的所有字符
        base_path = file_path.rsplit('.csv', 1)[0]
        xlsx_ = base_path + ".xlsx"
        convert_csv_to_xlsx(file_path, xlsx_)
        delete_file(file_path)
        return xlsx_
    else:
        return file_path


def str_strip(filepath: str, column_name: str):
    data = pd.read_excel(filepath)
    data[column_name] = data[column_name].str.replace('\t', '', regex=False).str.strip()
    data.to_excel(filepath, index=False)


def get_alert_intercepted_data(file_path, courier_column='Courier/快递', waybill_column='单号',
                               tracking_column='快递单号', key_value='unpaid'):
    # 读取Excel文件
    data = pd.read_excel(file_path)

    # 确保必要的列存在
    if courier_column not in data.columns or waybill_column not in data.columns or tracking_column not in data.columns:
        raise ValueError(f"文件中缺少必要的列，请检查列名是否正确")

    # 筛选出 Courier/快递 列内容为 'unpaid' 的数据
    unpaid_data = data[data[courier_column].str.strip().str.lower() == key_value]

    # 使用 map 存储结果，单号列作为 key，快递单号列作为 value
    result_map = dict(zip(unpaid_data[waybill_column], unpaid_data[tracking_column]))

    return result_map


def get_unpaid_tracking_data(file_path,
                             courier_column='Courier/快递',
                             waybill_column='单号',
                             tracking_column='快递单号',
                             date_column='UnpaidDate/unpaid记录时间',
                             key_value='unpaid'):
    # 读取 Excel 文件
    data = pd.read_excel(file_path)

    # 确保必要的列存在
    for col in [courier_column, waybill_column, tracking_column, date_column]:
        if col not in data.columns:
            raise ValueError(f"文件中缺少必要的列：{col}")

    # 筛选出 Courier/快递 列内容为 'unpaid' 的数据，并复制以避免警告
    unpaid_data = data[data[courier_column].astype(str).str.strip().str.lower() == key_value].copy()

    # 处理日期列，填充空值为 'EMPTY'
    unpaid_data.loc[:, date_column] = unpaid_data[date_column].fillna('EMPTY')

    # 格式化日期列为字符串
    unpaid_data.loc[:, date_column] = unpaid_data[date_column].apply(
        lambda x: x.strftime('%Y-%m-%d') if isinstance(x, pd.Timestamp) else str(x)
    )

    # 根据日期列分组，并构建结果字典
    grouped_result = {}
    for group_key, group_df in unpaid_data.groupby(date_column):
        # 每组数据生成一个 map：{单号: 快递单号}
        group_map = dict(zip(group_df[waybill_column], group_df[tracking_column]))
        grouped_result[group_key] = group_map

    return grouped_result


def get_days_difference(file_path, column_name="打单时间"):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        # 获取列索引
        column_index = headers.index(column_name) + 1
        # 获取第一条数据
        first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
        if not first_row_value:
            raise ValueError(f"'{column_name}' 列的第一条数据为空！")

        # 获取当前年份
        current_year = datetime.now().year

        # 检查日期类型，如果是字符串并且没有年份，补充当前年份
        if isinstance(first_row_value, str):
            # 如果是字符串并且没有年份，补充当前年份
            if len(first_row_value.split('-')) == 2:  # 格式为 'MM-DD' 或 'DD-MM'
                first_row_value = f"{current_year}-{first_row_value}"
            # 尝试解析日期
            outbound_time = datetime.strptime(first_row_value, "%Y-%m-%d %H:%M")
        elif isinstance(first_row_value, datetime):
            # 如果是已经是 datetime 类型，直接处理
            outbound_time = first_row_value
        else:
            raise ValueError(f"无法解析日期: {first_row_value}")

        # 格式化为 "%Y/%m/%d" 格式
        formatted_date = outbound_time.strftime("%Y/%m/%d")
        return formatted_date
    except Exception as e:
        print(f"发生错误: {e}")
        return None


def merge_csv_files_to_excel(csv_files, output_dir):
    """
    合并多个 CSV 文件，只保留第一个文件的列头，并保存为 Excel 文件。
    同时提取“打单时间”首条值的月份格式（如2025.5）和日期（如05-24），打印有效行数。
    """
    if not csv_files:
        print("❌ 输入文件列表为空")
        return

    combined_df = pd.DataFrame()

    for i, file in enumerate(csv_files):
        try:
            df = pd.read_csv(file, dtype=str).fillna("")
            if df.empty:
                print(f"⚠️ 文件为空，已跳过: {file}")
                continue

            if i == 0:
                combined_df = df
            else:
                if df.columns.equals(combined_df.columns):
                    combined_df = pd.concat([combined_df, df], ignore_index=True)
                else:
                    print(f"⚠️ 列不匹配，跳过文件: {file}")
        except Exception as e:
            print(f"❌ 读取失败: {file}，原因: {e}")

    if combined_df.empty:
        print("❌ 无有效数据，未生成 Excel 文件")
        return

    # 有效行：去除全空行
    valid_rows = combined_df[combined_df.apply(lambda row: row.astype(str).str.strip().any(), axis=1)]
    valid_count = len(valid_rows)

    # 提取“打单时间”第一条数据并处理
    if "打单时间" in valid_rows.columns:
        first_time_str = valid_rows["打单时间"].dropna().astype(str).str.strip().iloc[0]
        try:
            # 解析格式，如 05-24 14:54
            dt = datetime.strptime(first_time_str, "%m-%d %H:%M")
            # 添加年份，拼接为 “2025.5”
            month_info = f"2025.{dt.month}"
            # 获取日期部分
            date_only = str(dt.day)
            print(f"📅 打单时间首条记录：{first_time_str}")
            print(f"📅 转换后的月份格式：{month_info}")
            print(f"📅 提取的日期：{date_only}")
        except Exception as e:
            print(f"⚠️ 时间格式解析失败: {first_time_str}，错误: {e}")
    else:
        print("⚠️ 未找到“打单时间”列，跳过时间处理")

    print(f"✅ 有效数据行数（不含表头）：{valid_count}")

    # ✅ 创建输出目录（如果不存在）
    output_dir_month = output_dir + month_info + "/"
    create_dir = os.path.dirname(output_dir + month_info + "/")
    os.makedirs(create_dir, exist_ok=True)

    output_path = output_dir_month + f"打单时间{date_only}_{valid_count}.xlsx"

    combined_df.to_excel(output_path, index=False)
    print(f"✅ 合并完成，保存至: {output_path}")


def delete_files(file_paths):
    """
    遍历文件路径集合并删除文件。

    :param file_paths: 可迭代对象，如 list、set，包含完整文件路径字符串
    """
    for path in file_paths:
        try:
            if os.path.exists(path):
                os.remove(path)
                print(f"✅ 已删除: {path}")
            else:
                print(f"⚠️ 文件不存在: {path}")
        except Exception as e:
            print(f"❌ 删除失败: {path}，原因: {e}")


def go(input_path):
    if input_path is None:
        input_path = input("请输入文件的绝对路径：")
    xlsx_path = extract_path_before_csv(input_path)
    str_strip(xlsx_path, "快递单号")
    check_and_add_courier_column(xlsx_path)

    results = extract_and_process_data(xlsx_path, RowName.Courier, 100, "快递单号")

    all_maps = {
        CourierStateMapKey.not_yet_map: results[CourierStateMapKey.not_yet_map],
        CourierStateMapKey.pre_ship_map: results[CourierStateMapKey.pre_ship_map],
        CourierStateMapKey.unpaid_map: results[CourierStateMapKey.unpaid_map],
        CourierStateMapKey.delivered_map: results[CourierStateMapKey.delivered_map],
        CourierStateMapKey.no_tracking_map: results[CourierStateMapKey.no_tracking_map],
        CourierStateMapKey.tracking_map: results[CourierStateMapKey.tracking_map],
        CourierStateMapKey.possession_sf_date_map: results[CourierStateMapKey.possession_sf_date_map],
        CourierStateMapKey.latest_event_sf_date_map: results[CourierStateMapKey.latest_event_sf_date_map],
        CourierStateMapKey.sf_date_equality_map: results[CourierStateMapKey.sf_date_equality_map],
    }

    column_mapping = {
        CourierStateMapKey.not_yet_map: RowName.Courier,
        CourierStateMapKey.pre_ship_map: RowName.Courier,
        CourierStateMapKey.unpaid_map: RowName.Courier,
        CourierStateMapKey.delivered_map: RowName.Courier,
        CourierStateMapKey.no_tracking_map: RowName.Courier,
        CourierStateMapKey.tracking_map: RowName.Courier,
        CourierStateMapKey.possession_sf_date_map: RowName.PossessionSfDate,
        CourierStateMapKey.latest_event_sf_date_map: RowName.LatestEventSfDate,
        CourierStateMapKey.sf_date_equality_map: RowName.SfDateInterval,
    }

    update_courier_status(xlsx_path, all_maps, wl="快递单号", column_map=column_mapping)

    total_count, no_track_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.no_track)
    track_count = total_count - no_track_count
    total_count2, delivered_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.delivered)
    total_count3, unpaid_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.unpaid)
    total_count4, not_yet_count = count_pattern_state(xlsx_path, RowName.Courier, r"not_yet")
    total_count5, pre_ship_count = count_pattern_state(xlsx_path, RowName.Courier, r"pre_ship")
    total_count6, alert_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.alert)

    text = ""
    fs_text = ""
    ck_time = get_days_difference(xlsx_path)
    gz_time = getYmd()
    interval_time = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(ck_time, "%Y/%m/%d")).days
    is_usweekend = is_us_weekend(ck_time)
    date_obj = datetime.strptime(ck_time, "%Y/%m/%d").date()
    previous_day = date_obj - timedelta(days=1)
    actual_interval = 0
    if is_usweekend == 6:  # 6是中国周日，美国周六
        actual_interval = interval_time - 2
    elif is_usweekend == 0:  # 0是中国周一，美国周日
        actual_interval = interval_time - 1

    text += f"打单日期：{ck_time}"
    text += f"\n跟踪日期：{gz_time}"

    text += f"\n订单总数：{total_count}"
    fs_text += f"\n订单总数：{total_count}"

    swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))
    wswl = round2(100 - swl)
    text += f"\n上网：（{track_count}, {swl}%）"
    fs_text += f"\n上网：（{track_count}, {swl}%）"

    text += f"\n未上网：（{no_track_count}, {wswl}%）"
    fs_text += f"\n未上网：（{no_track_count}, {wswl}%）"

    not_yet_countl = round2((int(not_yet_count) / int(total_count)) * 100)
    text += f"\nnot_yet：（{not_yet_count}, {not_yet_countl}%）"
    fs_text += f"\nnot_yet：（{not_yet_count}, {not_yet_countl}%）"

    pre_ship_countl = round2((int(pre_ship_count) / int(total_count)) * 100)
    text += f"\npre_ship：（{pre_ship_count}, {pre_ship_countl}%）"
    fs_text += f"\npre_ship：（{pre_ship_count}, {pre_ship_countl}%）"

    delivered_countl = round2((int(delivered_count) / int(total_count)) * 100)
    text += f"\ndelivered：（{delivered_count}, {delivered_countl}%）"
    fs_text += f"\ndelivered：（{delivered_count}, {delivered_countl}%）"

    unpaid_countl = round2((int(unpaid_count) / int(total_count)) * 100)
    text += f"\nunpaid：（{unpaid_count}, {unpaid_countl}%）"
    fs_text += f"\nunpaid：（{unpaid_count}, {unpaid_countl}%）"

    alert_countl = round2((int(alert_count) / int(total_count)) * 100)
    text += f"\nalert：（{alert_count}, {alert_countl}%）"
    fs_text += f"\nalert：（{alert_count}, {alert_countl}%）"

    unpaid_tracking_data = get_unpaid_tracking_data(xlsx_path)
    current_day_unpaid_text = ""
    current_day_unpaid_len = 0
    now_strftime = datetime.now().strftime('%Y-%m-%d')
    if (len(unpaid_tracking_data) > 0):
        text += f"\n-------unpaid详情-------"
        fs_text += f"\n-------unpaid详情-------"
        for times, value in unpaid_tracking_data.items():
            result = f"\n 🕒：{times}"
            for key, val in value.items():
                result += f"\n（单号：{key}, 快递单号：{val}）"
                if times == now_strftime:
                    current_day_unpaid_text += f"{val}\n"
                    current_day_unpaid_len += 1
            result += "\n"
            text += result
            fs_text += result

    alert_intercepted_tracking_data = get_alert_intercepted_data(xlsx_path, key_value="alert")
    if (len(alert_intercepted_tracking_data) > 0):
        text += f"\n-------alert详情-------"
        fs_text += f"\n-------alert详情-------"
        for key, value in alert_intercepted_tracking_data.items():
            text += f"\n（单号：{key}, 快递单号：{value}）"
            fs_text += f"\n（单号：{key}, 快递单号：{value}）"

    print(text)

    sum_up_text = ""
    swl_flag = False
    qsl_flag = False
    bg = "#FFFFFF"

    # 上网率判断
    warning_levels = [
        (0, 30, "🧑‍🍳", "继续观察👀", "#FFFFFF"),
        (1, 30, "☁️注意", "未达30%！", "#F8F1D3"),
        (2, 70, "🌧️注意", "未达70%！", "#E3C49C"),
        (3, 97, "❄️异常", "未达97%！", "#F1C1BD"),
    ]

    actual_interval_time = interval_time - actual_interval
    # sum_up_text += f"\n间隔第{interval_time}{actual_interval}天"
    sum_up_text += f"\n间隔第{actual_interval_time}天"

    for days, threshold, icon, message, color in warning_levels:
        if actual_interval_time == days and swl < threshold:
            sum_up_text += f"\n{icon}：上网率为{swl}%，{message}"
            swl_flag = True
            if (actual_interval_time >= 2):
                bg = color
            break
    else:  # ✅ 只有 for 没有 break 时才会执行
        if (actual_interval_time >= 4):
            if (swl >= 99):
                sum_up_text += f"\n☀️上网率为{swl}%，优秀🌈"
            elif (swl >= 97 and swl < 99):
                sum_up_text += f"\n☀️上网率为{swl}%，达标✅"
            else:
                sum_up_text += f"\n⚡️异常：上网率为{swl}%，未达️97%"
                bg = "#F1C1BD"
                swl_flag = True
        else:
            sum_up_text += f"\n☀️上网率为{swl}%，达标✅"

    if (len(alert_intercepted_tracking_data) > 0):
        bg = "#FFF258"

    if (len(unpaid_tracking_data) > 0):
        bg = "#A684F0"

    tat = get_token()
    brief_sheet_value(tat, [fs_text], ck_time, gz_time, ClientConstants.md_flld)
    brief_sheet_bg(tat, ck_time, gz_time, ClientConstants.md_flld, bg)

    khhz_sheet_value(tat, [
        f"{total_count}",
        f"（{no_track_count}, {wswl}%）",
        f"（{0}, {0}%）",
        f"（{delivered_count}, {delivered_countl}%）",
        f"（{unpaid_count}, {unpaid_countl}%）",
    ], ck_time, ClientConstants.md_flld)

    khhz_sheet_bg(tat, ck_time, ClientConstants.md_flld, bg)

    result_fs_msg = f"客户：佛罗里达\n"
    result_fs_msg += f"订单创建时间：{ck_time}\n"
    result_fs_msg += f"跟踪时间：{gz_time}\n"
    fs_msg_flag = False

    if len(current_day_unpaid_text) > 0:
        result_fs_msg += f"新增 {current_day_unpaid_len}单 unpaid: \n"
        result_fs_msg += current_day_unpaid_text
        fs_msg_flag = True

    if swl_flag:
        result_fs_msg += f"上网率异常: {swl}%\n"
        fs_msg_flag = True

    if fs_msg_flag:
        # print(result_fs_msg)
        fs_msg(FsUserID.WP_ID, result_fs_msg)
        fs_msg(FsUserID.LW_ID, result_fs_msg)


def detect_duplicate_prefix_suffix(dir_path):
    prefix_suffix_map = defaultdict(list)

    for filename in os.listdir(dir_path):
        if not filename.lower().endswith('.csv'):
            continue  # 只处理 .csv 文件

        # 提取 '打单时间X' 前缀
        match = re.match(r"(打单时间\d+)_\d+\.csv", filename)
        if match:
            prefix = match.group(1)  # 如 "打单时间1"
            suffix = ".csv"
            full_path = os.path.join(dir_path, filename)
            prefix_suffix_map[(prefix, suffix)].append(full_path)

    for (prefix, suffix), files in prefix_suffix_map.items():
        if len(files) >= 1:
            print(f"📁 找到同组文件（前缀: {prefix}, 后缀: {suffix}）共 {len(files)} 个:")
            for f in files:
                print(f"   - {f}")
            merge_csv_files_to_excel(files, '/Users/zkp/Desktop/B&Y/轨迹统计/flld/')
            delete_files(files)


def automatic(root_dir, ignore=False, analyse_obj_ignore=False):
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    is_morning = (datetime.now().hour) < 12
    gz_time = getYmd()
    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            detect_duplicate_prefix_suffix(sun_dir_path)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = re.search(r"打单时间(\d+)_", file)
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= 15:
                        print(f"正在处理文件: {xlsx_path}")

                        if ignore:
                            go(xlsx_path)
                            continue

                        ck_time = get_days_difference(xlsx_path)
                        interval_time = (datetime.strptime(gz_time, "%Y/%m/%d") - datetime.strptime(ck_time,
                                                                                                    "%Y/%m/%d")).days

                        if interval_time == 1 and is_morning:
                            go(xlsx_path)
                        else:
                            if is_morning:
                                continue

                            if (analyse_obj_ignore):
                                go(xlsx_path)
                                continue

                            if (check_and_add_courier_column(xlsx_path)):
                                go(xlsx_path)
                            else:

                                output_file = os.path.splitext(xlsx_path)[0] + "_复制.xlsx"
                                copy_new_file(xlsx_path, output_file)

                                total_count, no_track_count = count_pattern_state(output_file, RowName.Courier,
                                                                                  Pattern.no_track)
                                track_count = total_count - no_track_count
                                total_count2, delivered_count = count_pattern_state(output_file, RowName.Courier,
                                                                                    Pattern.delivered)
                                total_count3, unpaid_count = count_pattern_state(output_file, RowName.Courier,
                                                                                 Pattern.unpaid)
                                total_count4, not_yet_count = count_pattern_state(output_file, RowName.Courier,
                                                                                  r"not_yet")
                                total_count5, pre_ship_count = count_pattern_state(output_file, RowName.Courier,
                                                                                   r"pre_ship")
                                total_count6, alert_count = count_pattern_state(output_file, RowName.Courier,
                                                                                Pattern.alert)
                                swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))
                                unpaid_countl = round2((int(unpaid_count) / int(total_count)) * 100)
                                delivered_countl = round2((int(delivered_count) / int(total_count)) * 100)

                                delete_file(output_file)

                                if swl < 99 or unpaid_count > 0 or (exceed >= 14 and delivered_countl < 98):
                                    go(xlsx_path)


def call():
    automatic("/Users/zkp/Desktop/B&Y/轨迹统计/flld/", False, True)


if __name__ == '__main__':
    call()
