from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from xinshili.utils import get_filename_without_extension, dirname


def aggregate_and_save(input_file, output_file):
    """
    对相同 SKU和客户 的出库数量 进行统计，并对客户列排序后保存到新的文件。
    :param input_file: 输入文件路径
    :param output_file: 输出文件路径
    """
    try:
        # 加载输入文件
        wb = load_workbook(input_file)
        sheet = wb.active

        # 获取列索引
        data = sheet.iter_rows(min_row=2, values_only=True)  # 跳过表头
        header = [cell.value for cell in sheet[1]]  # 获取表头
        ar_index = header.index("SKU")
        az_index = header.index("Outbound Qty/出库数量")
        b_index = header.index("Client/客户")

        # 汇总数据到字典
        result_dict = {}
        for row in data:
            ar_value = row[ar_index]
            az_value = row[az_index]
            b_value = row[b_index]

            if ar_value is not None and b_value is not None:
                key = (b_value, ar_value)  # 客户为第一关键字，SKU 为第二关键字
                result_dict[key] = result_dict.get(key, 0) + (az_value if isinstance(az_value, (int, float)) else 0)

        # 对结果按客户列排序（其次按 SKU 排序）
        sorted_results = sorted(result_dict.items(), key=lambda x: x[0])

        # 创建新的工作簿
        new_wb = Workbook()
        new_sheet = new_wb.active
        new_sheet.title = "汇总结果"

        # 写入表头
        new_sheet.append(["Client/客户", "SKU", "Outbound Qty/出库数量"])

        # 写入汇总结果
        for (b, ar), total in sorted_results:
            new_sheet.append([b, ar, total])

        # 调整所有单元格宽度和对齐方式
        for column_cells in new_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = max_length + 2  # 增加一些宽度间隔
            new_sheet.column_dimensions[column_letter].width = adjusted_width

        # 保存到新文件
        new_wb.save(output_file)
        print(f"汇总结果已保存到文件: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


if __name__ == "__main__":
    input_file = input("请输入源表文件的绝对路径：")
    output_file = dirname(input_file) + "/" + get_filename_without_extension(input_file) + "_统计.xlsx"
    aggregate_and_save(input_file, output_file)
