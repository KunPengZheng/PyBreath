import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import glob
from datetime import datetime, timedelta
from xinshili.flld_gjgz import extract_path_before_csv
from xinshili.utils import get_filename_without_extension, get_file_ext, delete_file

# 固定列（不要改）
fixed_columns = [
    "代码", "名称", "浮标",
    "2024-03(股东人数)", "2024-06(股东人数)",
    "2024-09(股东人数)", "2024-12(股东人数)",
    "2025-03(股东人数)", "2025-06(股东人数)"
]


def remove_st_rows_excel(input_file, output_file=None):
    # 读取 Excel
    df = pd.read_excel(input_file, dtype=object)

    # 删除 "名称" 列中包含 "ST" 或 "st" 的行
    df = df[~df["名称"].str.contains("ST|st", na=False)]

    # 输出文件路径
    if output_file is None:
        output_file = input_file.replace(".xlsx", "_clean.xlsx")

    # 保存为 Excel
    df.to_excel(output_file, index=False)

    print(f"✅ 已处理完成，结果保存到: {output_file}")


def merge_excels_with_dynamic_columns(file1, file2, output_file):
    # 固定列（顺序必须固定）
    fixed_columns = [
        "代码", "名称", "浮标",
        "2024-03(股东人数)", "2024-06(股东人数)",
        "2024-09(股东人数)", "2024-12(股东人数)",
        "2025-03(股东人数)", "2025-06(股东人数)"
    ]

    # 读取两个 Excel
    df1 = pd.read_excel(file1, dtype=object)
    df2 = pd.read_excel(file2, dtype=object)

    # 收集所有列名（固定列 + 动态列）
    all_columns = list(dict.fromkeys(fixed_columns + list(df1.columns) + list(df2.columns)))

    # 确保所有 DataFrame 都有这些列，缺失填 0
    for df in [df1, df2]:
        for col in all_columns:
            if col not in df.columns:
                df[col] = 0

    # 统一列顺序（固定列在前，动态列按字母/时间顺序排列）
    dynamic_columns = [c for c in all_columns if c not in fixed_columns]
    dynamic_columns = sorted(dynamic_columns)  # 按时间排序（如果列名规范化了）
    final_columns = fixed_columns + dynamic_columns

    # 合并
    merged_df = pd.concat([df1[final_columns], df2[final_columns]], ignore_index=True)

    # 保存结果
    merged_df.to_excel(output_file, index=False)
    print(f"✅ 合并完成，保存到: {output_file}")


def remove_zero_cells_and_shift_left(input_file, output_file):
    """
    规则：
    - 固定列（不会被修改/删除）:
        "代码", "名称",
        "2024-03(股东人数)", "2024-06(股东人数)",
        "2024-09(股东人数)", "2024-12(股东人数)",
        "2025-03(股东人数)"
    - 对于其它（动态）列：如果单元格值为 0（数字 0 或字符串 "0"），
      则删除该单元格，使右侧单元格左移一格（仅本行操作）。
    - 最终把所有动态列的表头清空（保留固定列表头）。
    """

    # 读取表头以确定列顺序（只读取表头也可以，但这里读全部以防万一）
    df = pd.read_excel(input_file)
    all_columns = list(df.columns)

    # 动态列 = 全部列中不属于固定列的
    dynamic_columns = [c for c in all_columns if c not in fixed_columns]
    if not dynamic_columns:
        print("⚠️ 没有检测到动态列，未做任何修改。")
        # 直接拷贝保存原文件或退出
        if output_file != input_file:
            df.to_excel(output_file, index=False)
            print(f"已将原文件复制为 {output_file}")
        return

    # 打开 Excel 工作簿
    wb = load_workbook(input_file)
    ws = wb.active

    # 列名 -> 列号 映射（openpyxl 列号从1开始）
    col_index_map = {c: i + 1 for i, c in enumerate(all_columns)}
    dynamic_start = min(col_index_map[c] for c in dynamic_columns)
    dynamic_end = max(col_index_map[c] for c in dynamic_columns)

    max_row = ws.max_row

    # 遍历每一行，从第2行开始（第1行为表头）
    for row in range(2, max_row + 1):
        col = dynamic_start
        # 对当前行，在动态列范围内循环
        while col <= dynamic_end:
            cell = ws.cell(row=row, column=col)
            val = cell.value

            # 判断是否等于 0 （支持 int/float 或字符串 '0'，忽略 None/空字符串）
            is_zero = False
            if val is None:
                is_zero = False
            elif isinstance(val, (int, float)):
                is_zero = (val == 0)
            elif isinstance(val, str):
                is_zero = (val.strip() == "0")
            else:
                is_zero = False

            if is_zero:
                # 将该单元格右边所有动态列值左移一格
                for c in range(col, dynamic_end):
                    ws.cell(row=row, column=c).value = ws.cell(row=row, column=c + 1).value
                # 最右侧动态列位置置空
                ws.cell(row=row, column=dynamic_end).value = None
                # 注意：不要增加 col，这样会再次检查当前位置（因为已经左移了一个新值到当前位置）
                # 动态_end 不变（我们不删除列，只是左移数据）
            else:
                col += 1

    # 最后清空动态列的表头（保持固定列标题不变）
    for c in range(dynamic_start, dynamic_end + 1):
        ws.cell(row=1, column=c).value = ""

    # 保存
    # 如果输出目录不存在则创建
    out_dir = os.path.dirname(os.path.abspath(output_file))
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    wb.save(output_file)
    print(f"✅ 已处理并保存到: {output_file}")


def calculate_diff_ratio(input_file, output_file):
    wb = load_workbook(input_file)
    ws = wb.active

    max_row = ws.max_row

    # 动态获取最大数据列（假设前两列是“代码”“名称”，从第3列开始是数值列）
    max_col = 0
    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=3):
        for idx, cell in enumerate(row, start=3):
            if isinstance(cell.value, (int, float)):
                max_col = max(max_col, idx)

    # 计算结果列从最大数据列向右跳 3~6 列
    p_col_index = max_col + 3
    y_col_index = max_col + 4
    z_col_index = max_col + 5
    aa_col_index = max_col + 6

    # 表头
    ws.cell(row=1, column=p_col_index).value = "2025-03 vs 2024-12 比率"
    ws.cell(row=1, column=y_col_index).value = "2025-06 vs 2025-03 比率"
    ws.cell(row=1, column=z_col_index).value = "倒数第一 vs 2025-03 比率"
    ws.cell(row=1, column=aa_col_index).value = "倒数第一 vs 倒数第二 比率"

    # 找到关键列索引
    header_row = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    col_2025_06 = header_row.get("2025-06(股东人数)")
    col_2025_03 = header_row.get("2025-03(股东人数)")
    col_2024_12 = header_row.get("2024-12(股东人数)")

    for row in range(2, max_row + 1):
        values = []
        for col in range(3, max_col + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, (int, float)):
                values.append(val)

        # P 列计算（2025-03 vs 2024-12）
        if col_2024_12 and col_2025_03:
            a_val = ws.cell(row=row, column=col_2025_03).value
            b_val = ws.cell(row=row, column=col_2024_12).value
            ws.cell(row=row, column=p_col_index).value = ((a_val - b_val) / b_val) if b_val else 0

        # Q 列计算（2025-06 vs 2025-03）
        if col_2025_06 and col_2025_03:
            a_val = ws.cell(row=row, column=col_2025_06).value
            b_val = ws.cell(row=row, column=col_2025_03).value
            ws.cell(row=row, column=y_col_index).value = ((a_val - b_val) / b_val) if b_val else 0

        # R 列计算（倒数第一 vs 2025-03）
        if col_2025_03 and values:
            b_val = ws.cell(row=row, column=col_2025_03).value
            last_val = values[-1]
            ws.cell(row=row, column=z_col_index).value = ((last_val - b_val) / b_val) if b_val else 0

        # S 列计算（倒数第一 vs 倒数第二）
        if len(values) >= 2:
            last_val = values[-1]
            second_last_val = values[-2]
            ws.cell(row=row, column=aa_col_index).value = (
                    (last_val - second_last_val) / second_last_val) if second_last_val else 0
        else:
            ws.cell(row=row, column=aa_col_index).value = 0

    wb.save(output_file)
    print(f"✅ 已完成，结果已保存到 {output_file}")


def compare_and_mark(file_a, file_b, output_file):
    wb_a = load_workbook(file_a)
    ws_a = wb_a.active

    wb_b = load_workbook(file_b)
    ws_b = wb_b.active

    # 获取列名所在行，这里假设是第2行
    header_a = [cell.value for cell in ws_a[1]]
    header_b = [cell.value for cell in ws_b[1]]

    # 找到“代码”列索引
    code_col_a = header_a.index("代码") + 1
    code_col_b = header_b.index("代码") + 1

    # 黄色填充样式
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 构建 B 文件 code->行对象的映射
    b_code_map = {}
    for row in ws_b.iter_rows(min_row=3):
        code = row[code_col_b - 1].value
        if code:
            b_code_map[code] = row

    # 遍历 A 文件每一行
    for row_a in ws_a.iter_rows(min_row=3):
        code = row_a[code_col_a - 1].value
        if not code or code not in b_code_map:
            continue  # 不匹配的行保持默认背景

        row_b = b_code_map[code]

        # 计算动态有效列数量（没有列名的列）
        dyn_cols_a = [cell for idx, cell in enumerate(row_a) if
                      idx >= 2 and (header_a[idx] is None or header_a[idx] == "") and cell.value not in (None, "")]
        dyn_cols_b = [cell for idx, cell in enumerate(row_b) if
                      idx >= 2 and (header_b[idx] is None or header_b[idx] == "") and cell.value not in (None, "")]

        # 如果 A 文件对应行动态列数量 > B 文件对应行数量，则标黄
        if len(dyn_cols_a) > len(dyn_cols_b):
            for cell in row_a:
                cell.fill = yellow_fill
        # 否则保持默认背景，网格线自动显示

    wb_a.save(output_file)
    print(f"✅ 已完成，结果已保存到 {output_file}")


def extract_path_before_csv(file_path):
    # 判断文件路径是否以 .csv 结尾
    if file_path.endswith('.csv'):
        # 提取 .csv 前的所有字符
        base_path = file_path.rsplit('.csv', 1)[0]
        xlsx_ = base_path + ".xlsx"
        convert_csv_to_xlsx(file_path, xlsx_)
        delete_file(file_path)
        return xlsx_
    else:
        return file_path


def convert_csv_to_xlsx(csv_file, xlsx_file):
    """
    将 CSV 文件转换为 XLSX 文件格式。

    :param csv_file: 输入的 CSV 文件路径
    :param xlsx_file: 输出的 XLSX 文件路径
    """
    try:
        # 读取 CSV 文件
        data = pd.read_csv(csv_file, encoding="gbk", dtype={"代码": str})

        # 将数据写入 XLSX 文件
        data.to_excel(xlsx_file, index=False)

        print(f"文件已成功转换为 XLSX 格式: {xlsx_file}")
    except Exception as e:
        print(f"转换过程中发生错误: {e}")


if __name__ == "__main__":
    days = 1
    gdrs_folder_path = "/Users/zkp/Desktop/B&Y/gdrs"
    csv_files = glob.glob(os.path.join(gdrs_folder_path, "*.csv"))
    arr = []
    for i, file in enumerate(csv_files):
        xlsx = extract_path_before_csv(file)
        path_without_ext = os.path.splitext(xlsx)[0]
        file_extension = get_file_ext(xlsx)
        result_file = path_without_ext + "_copy" + file_extension
        remove_st_rows_excel(xlsx, xlsx)
        arr.append(xlsx)

    time_str = datetime.now().strftime("%Y-%m-%d")
    yesterday = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
    merged = gdrs_folder_path + f"/gdrs_merged_{time_str}.xlsx"
    yesterday_merged = gdrs_folder_path + f"/gdrs_merged_{yesterday}.xlsx"
    if len(arr) == 2:
        merge_excels_with_dynamic_columns(arr[0], arr[1], merged)
        remove_zero_cells_and_shift_left(merged, merged)
        calculate_diff_ratio(merged, merged)
        compare_and_mark(merged, yesterday_merged, merged)
        delete_file(arr[0])
        delete_file(arr[1])