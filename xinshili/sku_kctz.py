import csv
import os
from collections import defaultdict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from xinshili.utils import get_filename_without_extension, dirname


def extract_and_count_sku_from_csv_to_xlsx(input_file, output_file, column_name, plus_or_minus):
    """
    从 CSV 文件提取 SKU 列中的内容，以 `*` 为分隔符统计每个SKU的数量
    并按 SKU 列排序后输出为另一个 XLSX 文件。

    :param input_file: 输入 CSV 文件路径
    :param output_file: 输出统计结果的 XLSX 文件路径
    :param column_name: 要统计的列名，一般是 "SKU"
    :param plus_or_minus: 表示增加或减少的标志
    """
    try:
        # 初始化统计字典
        sku_counts = defaultdict(int)

        # 打开 CSV 文件并读取
        with open(input_file, mode="r", encoding="utf-8") as infile:
            reader = csv.DictReader(infile)
            if column_name not in reader.fieldnames:
                raise ValueError(f"列名 '{column_name}' 不存在于文件中")

            # 遍历文件行
            for row in reader:
                sku_data = row[column_name]
                if not sku_data:
                    continue

                # 分割并统计 SKU
                for part in sku_data.split():
                    if "*" in part:
                        sku, count = part.split("*")
                        sku_counts[sku] += int(count)

        # 对统计结果按 SKU 排序
        sorted_sku_counts = sorted(sku_counts.items())

        # 创建新的 XLSX 文件并写入统计结果
        wb = Workbook()
        ws = wb.active
        ws.title = "SKU统计"

        # 写入表头
        ws.append(["SKU", "数量", "增加/减少"])

        # 写入数据
        for sku, count in sorted_sku_counts:
            ws.append([sku, count, plus_or_minus])

        # 调整单元格宽度并居中
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = max_length + 2  # 增加适当的宽度间隔
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        wb.save(output_file)
        print(f"统计完成，结果已保存到: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


def extract_and_count_sku_from_xlsx_to_xlsx(input_file, output_file, column_name, plus_or_minus):
    """
    从 XLSX 文件提取 SKU 列中的内容，以 `*` 为分隔符统计每个SKU的数量
    并按 SKU 列排序后输出为另一个 XLSX 文件。

    :param input_file: 输入 XLSX 文件路径
    :param sheet_name: 要读取的工作表名称
    :param column_name: 要统计的列名，一般是 "SKU"
    :param plus_or_minus: 表示增加或减少的标志
    """
    try:
        # 加载输入 XLSX 文件
        wb = load_workbook(input_file, data_only=True)
        ws = wb.active

        # 查找列名对应的列索引
        header_row = [cell.value for cell in ws[1]]
        if column_name not in header_row:
            raise ValueError(f"列名 '{column_name}' 不存在于工作表中")
        column_index = header_row.index(column_name) + 1

        # 初始化统计字典
        sku_counts = defaultdict(int)

        # 遍历数据行
        for row in ws.iter_rows(min_row=2, values_only=True):
            sku_data = row[column_index - 1]
            if not sku_data:
                continue

            # 分割并统计 SKU
            for part in str(sku_data).split():
                if "*" in part:
                    sku, count = part.split("*")
                    sku_counts[sku] += int(count)

        # 对统计结果按 SKU 排序
        sorted_sku_counts = sorted(sku_counts.items())

        # 创建新的 XLSX 文件并写入统计结果
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.title = "SKU统计"

        # 写入表头
        output_ws.append(["SKU", "数量", "增加/减少"])

        # 写入数据
        for sku, count in sorted_sku_counts:
            output_ws.append([sku, count, plus_or_minus])

        # 调整单元格宽度并居中
        for column_cells in output_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            adjusted_width = max_length + 2  # 增加适当的宽度间隔
            output_ws.column_dimensions[column_letter].width = adjusted_width

        # 保存文件
        output_wb.save(output_file)
        print(f"统计完成，结果已保存到: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


if __name__ == "__main__":
    """
    ⚠️注意：不区分客户或者仓库，如果需要区分 客户或者仓库。
    """

    input_file = input("请输入源表文件的绝对路径：")
    output_file = dirname(input_file) + "/" + get_filename_without_extension(input_file) + "_统计.xlsx"
    column_name = "SKU"  # 列头名
    plus_or_minus = "-"  # 是增加还是减少

    file_extension = os.path.splitext(input_file)[-1].lower()
    if file_extension == ".csv":
        print("检测到 CSV 文件，开始处理...")
        extract_and_count_sku_from_csv_to_xlsx(input_file, output_file, column_name, plus_or_minus)
    elif file_extension == ".xlsx":
        print("检测到 XLSX 文件，开始处理...")
        extract_and_count_sku_from_xlsx_to_xlsx(input_file, output_file, column_name, plus_or_minus)
    else:
        raise ValueError("不支持的文件格式，仅支持 .csv 和 .xlsx")
