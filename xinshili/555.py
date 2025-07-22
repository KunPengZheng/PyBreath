from datetime import datetime
import os
import re
from openpyxl import load_workbook
import pandas as pd
from dataclasses import dataclass


@dataclass(frozen=True)
class ClientConstants:
    zbw = "zbw"
    sanrio = "sanrio"
    xyl = "xyl"
    kaer = "kaer"
    mz_xsd = "mz_xsd"
    mx_dg = "mx_dg"
    md_fc = "md_fc"
    md_flld = "md_flld"
    cksj = "cksj"
    cjsj = "cjsj"
    khhz = "khhz"
    ckoms = "ckoms"
    yy = "yy"
    xyl_sales_repertory = "xyl_sales_repertory"
    sanrio_sales_repertory = "sanrio_sales_repertory"
    dxm_xyl_yd = "dxm_xyl_yd"


@dataclass(frozen=True)
class RowName:
    Tracking_No = 'Tracking No./物流跟踪号'
    Courier = 'Courier/快递'
    CreationTime = "Creation time/创建时间"
    OutboundTime = "OutboundTime/出库时间"
    Warehouse = "Warehouse/仓库"
    Client = "Client/客户"
    CreationWaveTime = "Create wave time/生成波次时间"
    SKU = "SKU"
    ShippingService = "Shipping service/物流渠道"
    PossessionSfDate = "PossessionSfDate/揽收时间"
    LatestEventSfDate = "LatestEventSfDate/最新事件时间"
    LatestEventSfTime = "LatestEventSfTime/最新事件时间"
    LatestEventSfSite = "LatestEventSfSite/最新事件地点"
    SfDateInterval = "SfDateInterval/SF消息间隔"
    TrackTimeInterval = "TrackTimeInterval/跟踪时间间隔"
    TrackTimeIntervalState = "TrackTimeIntervalState/跟踪时间间隔状态"
    LastEventSfTime = "LastEventSfTime/上一条轨迹时间"
    Tacking_Time = "Tacking_Time/追踪时间"
    UnpaidDate = "UnpaidDate/unpaid记录时间"
    Recipient = "Recipient/收件人"
    Upload_Shipping_Label = "上传物流面单(Upload_Shipping_Label)"
    YD_Number = 'YD_Number/阳单号'
    YD_State = 'YD_State/阳单轨迹状态'

    Courier_File1 = 'Courier/快递_file1'
    SfDateInterval_File1 = 'SfDateInterval/SF消息间隔_file1'
    PossessionSfDate_File1 = 'PossessionSfDate/揽收时间_file1'
    LatestEventSfDate_File1 = 'LatestEventSfDate/最新事件时间_file1'
    Package1_Tracking = 'Package 1\nTracking No./物流跟踪号1'

    Order_Num = "订单号"
    Track_Num = "运单号"
    Platform_Num = "Platform Number/平台单号"
    Yang_Num = "阳单_运单号"
    Yin_Num = "阴单_运单号"
    Yang_Track_State = "阳单_物流状态"
    Yin_Track_State = "阴单_物流状态"
    Yang_Delivered_Time = "阳单_签收时间"
    Yin_Delivered_Time = "阴单_签收时间"
    YY_Delivered_Time = "阴阳单_签收间隔"
    Create_Time = "创建时间"

    Length = 'Length/长'
    Width = 'Width/宽'
    Height = 'Height/高'
    Unit = 'Unit/单位'

    Pay_Time = "付款时间"
    Ship_Time = "发货时间"
    Track_State = "轨迹状态"
    Analyse_State = "追踪时间"
    Last_Track_Time = "上一条轨迹时间"
    Latest_Track_Site = "最新轨迹位置"
    Latest_Track_Time = "最新轨迹时间"
    Interval_Time = "时间间隔"
    Process_Time = "处理状态"
    YD_Number2 = "阳单号"
    YD_State2 = "阳单轨迹状态"

    Total_Of_Product = "产品总数"
    Store_Account = "店铺账号"
    Store_Sales = "店铺销量"
    Store_Name = "店铺名称"
    Order_Sales = "订单数量"
    Sales_Update = "销量更新"
    Quantity = "数量"
    Inventory_Update = "库存更新"
    Sea_transportation = "海运在途"
    Air_transportation = "空运在途"
    Available_Quantity = "可用量"
    Available_Inventory = "Available Inventory/可用库存"


@dataclass(frozen=True)
class CourierStateMapKey:
    tracking_map = 'tracking_map'
    irregular_number_map = 'irregular_number_map'
    no_tracking_map = 'no_tracking_map'
    unpaid_map = "unpaid_map"
    not_yet_map = "not_yet_map"
    pre_ship_map = "pre_ship_map"
    delivered_map = "delivered_map"
    alert_map = "alert_map"
    possession_sf_date_map = "possession_sf_date_map"
    latest_event_sf_date_map = "latest_event_sf_date_map"
    sf_date_equality_map = "sf_date_equality_map"
    delivered_time_map = "delivered_time_map"
    latest_event_sf_time_map = "latest_event_sf_time_map"
    latest_event_sf_site_map = "latest_event_sf_site_map"


class CourierStateMapValue:
    irregular_no_tracking = 'irregular_no_tracking'
    not_yet = 'not_yet'
    pre_ship = "pre_ship"
    no_tracking = "no_tracking"
    no_track = "no_track"
    unpaid = "unpaid"
    delivered = "delivered"
    tracking = "tracking"
    acceptance_pending = "acceptance_pending"
    alert = "alert"


@dataclass(frozen=True)
class CellKey:
    Outbound_Time = "Outbound_Time"
    update_time = "update_time"
    warehouse_condition = "warehouse_condition"
    store_condition = "store_condition"
    sku_condition = "sku_condition"
    time_segment_condition = "time_segment_condition"
    sum_up = "sum_up"
    exception = "exception"
    shipping_service_condition = "shipping_service_condition"
    special_information = "special_information"
    wl = "wl"
    unpaid = "unpaid"


@dataclass(frozen=True)
class Pattern:
    no_track = r"not_yet|pre_ship|irregular_no_tracking|no_tracking"
    delivered = r"^delivered$"
    unpaid = r"^unpaid$"
    not_yet = r"^not_yet$"
    irregular_no_tracking = r"^irregular_no_tracking$"
    pre_ship = r"^pre_ship$"
    no_tracking = r"^no_tracking$"
    tracking = r"^tracking$"
    alert = r"^alert"


def getYmd():
    # 获取今天的日期
    today = datetime.today()
    # 格式化为 "%Y/%m/%d" 格式
    formatted_today = today.strftime("%Y/%m/%d")
    # print(formatted_today)
    return formatted_today


# 自然排序的辅助函数
def natural_key(s):
    return [int(text) if text.isdigit() else text.lower()
            for text in re.split(r'(\d+)', s)]


def is_time_difference_exceed(start_time_str, end_time_str):
    try:
        time_format = "%Y-%m-%d"
        start_time = datetime.strptime(start_time_str, time_format)
        end_time = datetime.strptime(end_time_str, time_format)
        difference = abs((end_time - start_time).days)
        return difference
    except ValueError as e:
        print(f"❌ 时间格式错误: {e}")
        return False


def get_days_difference(file_path, column_name="Creation time/创建时间"):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active
    # 获取表头
    headers = [cell.value for cell in sheet[1]]
    if column_name not in headers:
        raise ValueError(f"列名 '{column_name}' 不存在！")
    # 获取列索引
    column_index = headers.index(column_name) + 1
    # 获取第一条数据
    first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
    if not first_row_value:
        raise ValueError(f"'{column_name}' 列的第一条数据为空！")
    # 解析日期
    outbound_time = datetime.strptime(first_row_value, "%Y-%m-%d %H:%M:%S")
    # 格式化为 "%Y/%m/%d" 格式
    formatted_date = outbound_time.strftime("%Y/%m/%d")
    return formatted_date


def check_and_add_courier_column(file_path):
    try:
        # 加载 Excel 文件
        data = pd.read_excel(file_path, dtype=str, engine='openpyxl')
        flag = False
        # 判断是否存在 '快递' 列
        if RowName.Courier not in data.columns:
            # 如果没有 '快递' 列，则在最后一列添加该列
            data[RowName.Courier] = ""  # 默认为空值，可以根据需求填充其他默认值
            flag = True
        if RowName.PossessionSfDate not in data.columns:
            data[RowName.PossessionSfDate] = ""
            flag = True
        if RowName.LatestEventSfDate not in data.columns:
            data[RowName.LatestEventSfDate] = ""
            flag = True
        if RowName.SfDateInterval not in data.columns:
            data[RowName.SfDateInterval] = ""
            flag = True
        if RowName.UnpaidDate not in data.columns:
            data[RowName.UnpaidDate] = ""
            flag = True
        if RowName.LatestEventSfTime not in data.columns:
            data[RowName.LatestEventSfTime] = ""
            flag = True
        if RowName.LatestEventSfSite not in data.columns:
            data[RowName.LatestEventSfSite] = ""
            flag = True
        if RowName.TrackTimeInterval not in data.columns:
            data[RowName.TrackTimeInterval] = ""
            flag = True
        if RowName.TrackTimeIntervalState not in data.columns:
            data[RowName.TrackTimeIntervalState] = ""
            flag = True
        if RowName.Tacking_Time not in data.columns:
            data[RowName.Tacking_Time] = ""
            flag = True
        if RowName.LastEventSfTime not in data.columns:
            data[RowName.LastEventSfTime] = ""
            flag = True
        if RowName.YD_Number not in data.columns:
            data[RowName.YD_Number] = ""
            flag = True
        if RowName.YD_State not in data.columns:
            data[RowName.YD_State] = ""
            flag = True
        # 保存修改后的文件
        data.to_excel(file_path, index=False, engine='openpyxl')
        return flag
    except Exception as e:
        print(f"发生错误: {e}")


def is_us_weekend(date_str):
    """
    中国和美国的时差相差：13-16 个钟头，目前日期的单位最小是日期，没有到小时，所以这里我们默认和美国相差一天，
    也就是中国时间周日为美国的周六，中国时间周一为美国的周日
    """
    # 解析字符串为 datetime 对象
    date_obj = datetime.strptime(date_str, "%Y/%m/%d")

    # 判断是否为 周日（6）或者周一 (0)，即为美国的周六和周日
    return date_obj.weekday()


def get_shipment_received_numbers(filepath, gz_time):
    # 加载 Excel 文件
    workbook = load_workbook(filename=filepath, data_only=True)

    # 获取第一个工作表
    sheet = workbook.active

    # 读取表头（第一行），并构建列名到索引的映射
    header = [cell.value for cell in sheet[1]]

    # 确保所有必要的列都存在
    required_columns = [RowName.Courier, RowName.SfDateInterval, RowName.OutboundTime, RowName.Tracking_No]

    # 获取每个必要列的索引（列索引从 0 开始）
    column_indices = {}
    for col in required_columns:
        if col not in header:
            missing_cols = set(required_columns) - set(header)
            raise ValueError(f"Excel 文件缺少必要的列: {missing_cols}")
        column_indices[col] = header.index(col)

    # 存储符合条件的跟踪号
    tracking_numbers = []

    # 遍历数据行（从第二行开始）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        courier = row[column_indices[RowName.Courier]]
        sf_date_interval = row[column_indices[RowName.SfDateInterval]]
        outbound_time = row[column_indices[RowName.OutboundTime]]
        tracking_no = row[column_indices[RowName.Tracking_No]]

        # 筛选条件：Courier == 'tracking' 且 SfDateInterval == '0'
        if courier == CourierStateMapValue.tracking and sf_date_interval == 0:

            # try:
            # 解析并计算日期间隔
            parsed_outbound_time = datetime.strptime(outbound_time, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_outbound_time.strftime("%Y/%m/%d")
            interval_days = (datetime.strptime(gz_time, "%Y/%m/%d") -
                             datetime.strptime(formatted_date, "%Y/%m/%d")).days

            is_usweekend = is_us_weekend(formatted_date)
            actual_interval = 0
            if is_usweekend == 6:  # 6是中国周日，美国周六
                actual_interval = 2
            elif is_usweekend == 0:  # 0是中国周一，美国周日
                actual_interval = 1

            # 如果间隔为 2 天，添加到结果列表
            if (interval_days - actual_interval) >= 2 and tracking_no is not None:
                tracking_numbers.append(tracking_no)

        # except (ValueError, TypeError) as e:
        #     日期解析错误或数据错误时跳过
        # print(f"日期解析错误: {e}, 跟踪号: {tracking_no}")
        # continue

    # 返回唯一的跟踪号列表
    return list(set(tracking_numbers))


def filter_tracking_numbers(input_path, output_path):
    # 读取 Excel 文件
    df = pd.read_excel(input_path, dtype=str).fillna("")

    col = "Tracking No./物流跟踪号"

    if col not in df.columns:
        print(f"❌ 文件中缺少列: {col}")
        return

    # 保留以 92、93 或 94 开头的行
    df_filtered = df[df[col].str.startswith(("92", "93", "94"), na=False)]

    # 保存结果
    df_filtered.to_excel(output_path, index=False)
    # print(f"✅ 筛选后文件已保存到: {output_path}")


def remove_duplicates_by_column(input_file, output_file, column_name):
    """
    去重：删除指定列中重复的行，仅保留第一条，并覆盖源文件，同时返回去重后的行数。

    参数：
    - input_file: str，输入文件路径
    - column_name: str，要检查重复的列名
    """
    try:
        # 读取 Excel 文件
        df = pd.read_excel(input_file)

        # 检查列名是否存在
        if column_name not in df.columns:
            raise ValueError(f"列 '{column_name}' 不存在于输入文件中！")

        # 删除指定列的重复项，仅保留第一条
        df_deduplicated = df.drop_duplicates(subset=[column_name], keep='first')

        # 将去重后的数据保存到输出文件
        df_deduplicated.to_excel(output_file, index=False)

        # 返回去重后的行数
        return len(df_deduplicated)

    except Exception as e:
        print(f"处理文件时发生错误：{e}")
        return 0  # 如果出现错误，返回 0 行数


def count_pattern_and_tracking_with_sf_date(file_path, column_name, sfDateInterval_name, patterns):
    """
    统计指定列中多个正则表达式匹配内容的数量，并统计 "Courier/快递" 列内容为 'tracking' 且 "SfDateEquality" 列的内容为 0 的行数。

    :param file_path: Excel 文件路径
    :param column_name: 需要统计的列名
    :param patterns: 正则表达式字典，key 为模式名称，value 为模式字符串
    :return: 包含每个模式匹配计数和符合 'tracking' 且 'SfDateEquality' 为 0 的行数的字典
    """
    try:
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]

        # 检查列名是否存在
        if column_name not in headers or sfDateInterval_name not in headers:
            raise ValueError(f"Excel 文件中未找到 '{column_name}' 或 '{sfDateInterval_name}'列 找不到")

        # 获取列索引
        column_index = headers.index(column_name)
        sfDateInterval_index = headers.index(sfDateInterval_name)

        # 编译所有正则表达式
        compiled_patterns = {key: re.compile(pattern, re.IGNORECASE) for key, pattern in patterns.items()}

        # 初始化计数器
        count_dict = {**{key: 0 for key in patterns}, "sfDateInterval": 0}  # 显式初始化
        tracking_sf_date_equality_count = 0

        # 遍历每一行数据
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            courier_value = row[column_index]
            sfDateInterval_value = row[sfDateInterval_index]

            # 判断是否符合 'tracking' 且 'SfDateEquality' 为 0 或者 'acceptance_pending' 的条件
            if (courier_value == CourierStateMapValue.tracking and sfDateInterval_value == 0) or (
                    courier_value == CourierStateMapValue.acceptance_pending):
                tracking_sf_date_equality_count += 1

            # 判断正则表达式匹配
            if courier_value is not None:
                cell_value_str = str(courier_value)
                for key, pattern in compiled_patterns.items():
                    if pattern.search(cell_value_str):
                        count_dict[key] += 1

        # 将 'tracking' 且 'SfDateEquality' 为 0 的计数添加到字典中
        count_dict["sfDateInterval"] = tracking_sf_date_equality_count

        return count_dict

    except Exception as e:
        print(f"发生错误: {e}")
        # 返回一个包含所有模式及 'tracking_sf_date_equality' 键的字典
        return {key: 0 for key in patterns} | {"sfDateInterval": 0}  # 合并字典


def round2(nums):
    """
    四舍五入，保留两位数
    """
    return round(nums, 2)


def delete_file(file_path):
    """
    删除指定的文件
    :param file_path: 要删除的文件的绝对路径
    """
    try:
        if os.path.exists(file_path):  # 检查文件是否存在
            os.remove(file_path)  # 删除文件
            # print(f"文件 {file_path} 已成功删除")
        else:
            print(f"文件 {file_path} 不存在")
    except Exception as e:
        print(f"删除文件时发生错误: {e}")


def automatic(root_dir, analyse_obj, ignore=False, analyse_obj_ignore=False):
    usps_arr = []
    today = datetime.now()
    current_year = today.year
    current_month = today.month
    current_day = today.day
    current_time = f"{current_year}-{current_month}-{current_day}"
    is_morning = (datetime.now().hour) < 12
    gz_time = datetime.strptime(getYmd(), "%Y/%m/%d")
    # print(current_year, current_month, current_day)
    for dirpath, dirnames, filenames in os.walk(root_dir):
        dirnames.sort(key=natural_key)
        for dirname in dirnames:
            sun_dir_path = os.path.join(dirpath, dirname)
            parts = dirname.split(".")
            year, month = parts[0], parts[1]
            files = [f for f in os.listdir(sun_dir_path) if f.lower().endswith(('.xlsx', '.xls'))]
            files.sort(key=natural_key)
            for file in files:
                xlsx_path = os.path.join(sun_dir_path, file)
                match = re.search(r"创建时间(\d+)_", file)
                if match:
                    day = match.group(1)
                    current_times = f"{year}-{month}-{day}"
                    exceed = is_time_difference_exceed(current_time, current_times)
                    if exceed <= 15:
                        print(f"正在处理文件: {xlsx_path}")

                        if ignore:
                            # go(analyse_obj, xlsx_path)
                            usps_arr.append(xlsx_path)
                            continue

                        # ck_time = get_days_difference(xlsx_path)
                        ck_time =  datetime.strptime(get_days_difference(xlsx_path), "%Y/%m/%d")
                        # print(
                        #     f"gz_time: {gz_time} , ck_time: {ck_time} , type: {type(gz_time)}, typ2e: {type(ck_time)}")
                        interval_time = (gz_time - ck_time).days

                        if interval_time == 1 and is_morning:
                            # go(analyse_obj, xlsx_path)
                            usps_arr.append(xlsx_path)
                        else:
                            if is_morning:
                                continue

                            if (analyse_obj_ignore):
                                # go(analyse_obj, xlsx_path)
                                usps_arr.append(xlsx_path)
                                continue

                            if (check_and_add_courier_column(xlsx_path)):
                                # go(analyse_obj, xlsx_path)
                                usps_arr.append(xlsx_path)
                            else:
                                shipment_received_interval2_list = get_shipment_received_numbers(xlsx_path, getYmd())
                                change_shipment_received_count = len(shipment_received_interval2_list)

                                output_file = os.path.splitext(xlsx_path)[0] + "_去重.xlsx"
                                filter_tracking_numbers(xlsx_path, output_file)

                                # 同一单会有多个sku，多个sku会生成多行数据，分析sku的时候不能去重，其它的需要去重
                                total_count = remove_duplicates_by_column(output_file, output_file, RowName.Tracking_No)

                                patterns = {
                                    CourierStateMapValue.no_track: Pattern.no_track,
                                    CourierStateMapValue.delivered: Pattern.delivered,
                                    CourierStateMapValue.unpaid: Pattern.unpaid,
                                    CourierStateMapValue.not_yet: Pattern.not_yet,
                                    CourierStateMapValue.pre_ship: Pattern.pre_ship,
                                    CourierStateMapValue.irregular_no_tracking: Pattern.irregular_no_tracking,
                                    CourierStateMapValue.no_tracking: Pattern.no_tracking,
                                    CourierStateMapValue.tracking: Pattern.tracking
                                }

                                count_dict = count_pattern_and_tracking_with_sf_date(output_file, RowName.Courier,
                                                                                     RowName.SfDateInterval, patterns)

                                no_track_count = count_dict[CourierStateMapValue.no_track]
                                delivered_count = count_dict[CourierStateMapValue.delivered]
                                unpaid_count = count_dict[CourierStateMapValue.unpaid]
                                not_yet_count = count_dict[CourierStateMapValue.not_yet]
                                pre_ship_count = count_dict[CourierStateMapValue.pre_ship]
                                irregular_no_tracking_count = count_dict[CourierStateMapValue.irregular_no_tracking]
                                no_tracking_count = count_dict[CourierStateMapValue.no_tracking]
                                tracking_count = count_dict[CourierStateMapValue.tracking]
                                tracking_zero_count = count_dict["sfDateInterval"]

                                # 先进行一次计算，并缓存结果
                                total_count_int = int(total_count)
                                no_track_count_int = int(no_track_count)
                                track_count_int = total_count_int - no_track_count_int
                                tracking_zero_count_int = int(tracking_zero_count)
                                delivered_count_int = int(delivered_count)
                                unpaid_count_int = int(unpaid_count)
                                not_yet_count_int = int(not_yet_count)
                                pre_ship_count_int = int(pre_ship_count)
                                irregular_no_tracking_count_int = int(irregular_no_tracking_count)
                                no_tracking_count_int = int(no_tracking_count)
                                tracking_count_int = int(tracking_count)
                                # real_no_track_count = no_track_count_int + tracking_zero_count_int  # 真正的未上网数
                                # real_track_count = total_count_int - no_track_count  # 真正的上网数
                                # real_tracking_count = tracking_count_int - tracking_zero_count_int

                                if tracking_zero_count_int == 0 and delivered_count_int == 0 and unpaid_count_int == 0 \
                                        and not_yet_count_int == 0 and pre_ship_count_int == 0 and no_tracking_count_int == 0 \
                                        and tracking_count_int == 0:
                                    delete_file(output_file)
                                    usps_arr.append(xlsx_path)
                                    continue

                                # 计算百分比
                                swl = round2(100 - ((no_track_count_int) / total_count_int * 100))
                                wswl = round2(100 - swl)
                                qsl = round2((delivered_count_int / total_count_int) * 100)
                                unpaidl = round2((unpaid_count_int / total_count_int) * 100)
                                not_yetl = round2((not_yet_count_int / total_count_int) * 100)
                                pre_shipl = round2((pre_ship_count_int / total_count_int) * 100)
                                irregular_no_trackingl = round2(
                                    (irregular_no_tracking_count_int / total_count_int) * 100)
                                no_tracking_countl = round2((no_tracking_count_int / total_count_int) * 100)
                                tracking_countl = round2((tracking_count_int / total_count_int) * 100)
                                tracking_zero_countl = round2((tracking_zero_count_int / total_count_int) * 100)
                                change_shipment_received_countl = round2(
                                    (change_shipment_received_count / total_count_int) * 100)

                                delete_file(output_file)

                                print(
                                    f"swl:{swl}, change_shipment_received_count:{change_shipment_received_count}, unpaid_count:{unpaid_count}, exceed:{exceed}, qsl:{qsl},")
                                if swl < 99 or change_shipment_received_count >= 10 or unpaid_count > 0 or \
                                        (exceed >= 14 and qsl < 98):
                                    # go(analyse_obj, xlsx_path)
                                    usps_arr.append(xlsx_path)

    return usps_arr


def call2():
    zbw_usps_arr = automatic("/Users/zkp/Desktop/B&Y/轨迹统计/zbw/", ClientConstants.zbw, False, False)
    print(zbw_usps_arr)
    # sanrio_usps_arr = automatic("/Users/zkp/Desktop/B&Y/轨迹统计/sanrio/", ClientConstants.sanrio, False, False)
    # xyl_usps_arr = automatic("/Users/zkp/Desktop/B&Y/轨迹统计/xyl/", ClientConstants.xyl, False, True)
    # kaer_usps_arr = automatic("/Users/zkp/Desktop/B&Y/轨迹统计/kaer/", ClientConstants.kaer, False, True)


call2()
