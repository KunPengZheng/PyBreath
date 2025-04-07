import re
import os
from openpyxl import load_workbook

from xinshili.utils import get_filename_with_extension


def lenb_minus_len(text):
    """
    计算类似 Excel 公式 LENB(A1) - LEN(A1) 的值：
    计算字符串中所有中文字符和中文符号占用的字节数与字符数的差值
    """
    if not isinstance(text, str):  # 处理 None 或非字符串情况
        return 0

    # 匹配所有中文字符（基本汉字 + 符号）
    chinese_pattern = re.compile(r'[\u4e00-\u9fff]')  # 匹配汉字
    # 中文符号（全角标点）
    chinese_symbols = re.compile(r'[，。？！【】（）《》“”‘’：；￥— –]')  # 匹配常见中文符号，包括破折号（—）

    chinese_chars = chinese_pattern.findall(text) + chinese_symbols.findall(text)  # 获取所有中文字符+符号

    # 计算 UTF-8 编码后的字节数
    utf8_length = len(text.encode('utf-8'))
    # 计算普通字符长度
    char_length = len(text)
    # 计算额外的字节数（相当于 Excel 公式 LENB(A1) - LEN(A1)）
    extra_bytes = utf8_length - char_length
    return extra_bytes


def rule_replace(text):
    """
    替换中文符号为对应的英文符号，去除品牌名、敏感内容、非法字符，避免侵权和不适当的内容。
    """
    if not isinstance(text, str):  # 处理 None 或非字符串情况
        return ""

    # 定义符号映射（中文符号 -> 英文符号）
    symbol_map = {
        "，": ",", "。": ".", "？": "?", "！": "!", "【": "[", "】": "]",
        "（": "(", "）": ")", "《": "<", "》": ">", "“": "\"", "”": "\"",
        "‘": "'", "’": "'", "：": ":", "；": ";", "￥": "$", "—": "-", "–": "-",
        "m²": "", "｜": "", "®": "", "™": "", "©": "", "2024": "2025", "ö": "",
        "——": "-", "【": "[", "】": "]", "→": "", "↑": "", "↓": "", "＊": "*",
    }

    rule_words = {
        # 敏感内容词汇（包含政治、黄赌毒等敏感话题）
        "pornography", "gambling", "drugs", "violence", "terrorism", "extremism", "hate speech", "racism",
        "sexism", "corruption", "murder", "scam", "fraud", "politics", "war", "terrorist", "extremist",
        "hate", "abuse", "pedophile", "child abuse", "illegal", "mafia", "cartel", "gangs", "racist",
        "xenophobic", "homophobic", "anti-semitic", "rape", "sexual assault", "slavery", "trafficking",
        "extortion", "harassment", "bullying", "weapon", "death", "murderer", "kidnapping", "explosive"

        # 名人、演员、歌手、政治人物等
                                                                                            "Barack Obama",
        "Elvis Presley", "Michael Jackson", "Beyoncé", "Taylor Swift", "Madonna", "Ariana Grande",
        "Drake", "Justin Bieber", "Kanye West", "Bill Gates", "Steve Jobs", "Oprah Winfrey", "Leonardo DiCaprio",
        "Brad Pitt", "Angelina Jolie", "Robert Downey Jr.", "Scarlett Johansson", "Tom Hanks", "Will Smith",
        "Chris Hemsworth", "Johnny Depp", "Meryl Streep", "Dwayne Johnson", "Emma Watson", "Rihanna", "Lord",

        # 书名、小说
        "Harry Potter", "The Lord of the Rings", "The Catcher in the Rye", "1984", "To Kill a Mockingbird",
        "The Great Gatsby", "Pride and Prejudice", "The Da Vinci Code", "Moby-Dick", "War and Peace",
        "The Hobbit", "The Bible", "The Quran", "Fifty Shades of Grey", "The Hunger Games", "The Chronicles of Narnia",
        "The Road", "Brave New World",

        # 电影名、系列
        "Star Wars", "Toy Story", "Avengers", "Spider-Man", "Batman", "Superman", "Iron Man", "Thor", "Hulk",
        "Captain America", "Black Panther", "Wonder Woman", "The Matrix", "Inception", "Titanic",
        "The Shawshank Redemption", "Forrest Gump", "The Dark Knight", "Pulp Fiction", "The Godfather",
        "The Terminator",
        "Jurassic Park", "The Lion King", "Frozen", "Avatar", "Jurassic World", "The Avengers: Endgame",

        # 视频游戏、动漫等
        "Super Mario", "Minecraft", "Fortnite", "The Sims", "World of Warcraft", "Pokémon", "Dragon Ball", "Naruto",
        "One Piece", "Doraemon", "Attack on Titan", "Sailor Moon", "Bleach", "Fairy Tail", "My Hero Academia",
        "Yu-Gi-Oh", "Dragon Quest", "League of Legends", "Counter-Strike", "Grand Theft Auto", "The Witcher",
        "Call of Duty", "The Elder Scrolls",

        # 其他
        "Disney", "Nike", "Apple", "Samsung", "Microsoft", "Google", "Facebook", "Instagram", "Amazon", "YouTube",
        "WhatsApp", "Twitter", "TikTok", "CocaCola", "Pepsi", "Adidas", "Lamborghini", "Ferrari", "Rolls Royce",
        "Porsche", "McDonald's", "KFC", "Starbucks", "BMW", "Audi", "Mercedes", "Disneyland", "Disney", "Mickey",
        "2pcs", "2paces", "outfit", "Long sleeve", "shorts",
    }

    # 使用正则替换，忽略大小写
    for zh_symbol, en_symbol in symbol_map.items():
        text = re.sub(re.escape(zh_symbol), en_symbol, text, flags=re.IGNORECASE)

    for word in rule_words:
        text = re.sub(re.escape(word), "", text, flags=re.IGNORECASE)

    # 去除不合适的字符，避免非法字符影响文件名
    text = re.sub(r'[<>:"/\\|?*]', "", text)  # 删除文件名中不允许的字符
    text = re.sub(r'\s+', " ", text)  # 去除多余的空格
    text = re.sub(r'^\s+|\s+?$', '', text)  # 去除首尾空白字符
    text = re.sub(r'\.{2,}', ".", text)  # 将多个连续的句点（..）替换为一个句点
    # 清理表情符号及特殊字符
    text = re.sub(r'[^\x00-\x7F]+', '', text)  # 删除非ASCII字符（表情符号、特殊字符）
    # 去除重复的词汇或无意义的描述
    text = re.sub(r'\b(\w+)\s+\1\b', r'\1', text)  # 删除重复的单词，例如 "T-shirt T-shirt"
    # 增强文本一致性（避免无意义的修饰词）
    text = re.sub(r'(really|very|totally|extremely|incredibly)\s+', "", text)  # 去除多余的程度副词

    # 最后修剪文本，去掉多余的空格
    text = text.strip()

    return text


def check_cell_value(value):
    # 定义正则表达式
    pattern = r"^[A-Z]\d{2,4}$"

    # 使用正则匹配
    if re.match(pattern, value):
        return True  # 匹配成功
    else:
        return False  # 不匹配


def process_replace(filepath, column_a="A", column_b="B", column_c="C", women_big_size_suffix_flag=False):
    workbook = load_workbook(filename=filepath)
    sheet = workbook.active  # 选择默认工作表

    for row in range(1, sheet.max_row + 1):  # 遍历所有行
        cell = ""

        column_a_is_title = True
        cell_a = sheet[f"{column_a}{row}"]
        if (check_cell_value(cell_a)):  # a列是sku列
            cell = sheet[f"{column_b}{row}"]  # 获取b列
            column_a_is_title = False
        else:  # a列是标题列
            cell = cell_a
            column_a_is_title = True

        text = str(cell.value) if cell.value else ""  # 处理空值

        # 计算中文字符和符号的字节差值
        # difference = lenb_minus_len(text)
        # sheet[f"{result_column_b}{row}"] = difference

        if (women_big_size_suffix_flag):
            text += ", L-XXXXL Plus Size"
        replaced_text = rule_replace(text)

        if (column_a_is_title):
            sheet[f"{column_a}{row}"] = replaced_text
            sheet[f"{column_b}{row}"] = replaced_text
        else:
            sheet[f"{column_b}{row}"] = replaced_text
            sheet[f"{column_c}{row}"] = replaced_text

    # 保存 Excel
    workbook.save(filepath)
    print(f"处理完成，已更新 {filepath}")


def rename_files_in_folder(folder_path, prefix):
    # 获取文件夹中的所有文件
    files = sorted(os.listdir(folder_path))  # 按名称排序，避免乱序

    for file in files:
        old_path = os.path.join(folder_path, file)

        # 确保是文件，而不是文件夹
        if os.path.isfile(old_path):
            extension = get_filename_with_extension(file)
            new_name = f"{prefix}{extension}"  # 生成新文件名
            new_path = os.path.join(folder_path, new_name)
            os.rename(old_path, new_path)  # 重命名文件

# process_replace("/Users/zkp/Desktop/未命名文件夹/20250303黑色女装标题.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250304黑色女装标题.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250305白色女装标题J041-K080.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250306白色女装标题K081-K100.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250306黑色女装标题L001-L100.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250308白色女装标题.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250310黑色女装标题.xlsx")
# process_replace("/Users/zkp/Desktop/未命名文件夹/20250310黑色女装标题H301-H400.xlsx")

# rename_files_in_folder(r"/Users/zkp/Desktop/0315服装230/图片")
