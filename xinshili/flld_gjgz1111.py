from datetime import datetime
import os
import re

import pandas as pd
from openpyxl import load_workbook

from xinshili.fs_utils_plus import get_token, brief_sheet_value, ClientConstants
from xinshili.gjgz_plus111 import check_and_add_courier_column, RowName, extract_and_process_data, \
    update_courier_status, CourierStateMapKey, count_pattern_state, Pattern
from xinshili.utils import convert_csv_to_xlsx, delete_file, getYmd, round2


def extract_path_before_csv(file_path):
    # 判断文件路径是否以 .csv 结尾
    if file_path.endswith('.csv'):
        # 提取 .csv 前的所有字符
        base_path = file_path.rsplit('.csv', 1)[0]
        xlsx_ = base_path + ".xlsx"
        convert_csv_to_xlsx(file_path, xlsx_)
        delete_file(file_path)
        return xlsx_
    else:
        return file_path


def str_strip(filepath: str, column_name: str):
    data = pd.read_excel(filepath)
    data[column_name] = data[column_name].str.replace('\t', '', regex=False).str.strip()
    data.to_excel(filepath, index=False)


def get_unpaid_tracking_data(file_path, courier_column='Courier/快递', waybill_column='单号',
                             tracking_column='快递单号'):
    # 读取Excel文件
    data = pd.read_excel(file_path)

    # 确保必要的列存在
    if courier_column not in data.columns or waybill_column not in data.columns or tracking_column not in data.columns:
        raise ValueError(f"文件中缺少必要的列，请检查列名是否正确")

    # 筛选出 Courier/快递 列内容为 'unpaid' 的数据
    unpaid_data = data[data[courier_column].str.strip().str.lower() == 'unpaid']

    # 使用 map 存储结果，单号列作为 key，快递单号列作为 value
    result_map = dict(zip(unpaid_data[waybill_column], unpaid_data[tracking_column]))

    return result_map


def get_days_difference(file_path, column_name="打单时间"):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        # 获取表头
        headers = [cell.value for cell in sheet[1]]
        if column_name not in headers:
            raise ValueError(f"列名 '{column_name}' 不存在！")
        # 获取列索引
        column_index = headers.index(column_name) + 1
        # 获取第一条数据
        first_row_value = sheet.cell(row=2, column=column_index).value  # 假设数据从第二行开始
        if not first_row_value:
            raise ValueError(f"'{column_name}' 列的第一条数据为空！")

        # 获取当前年份
        current_year = datetime.now().year

        # 检查日期类型，如果是字符串并且没有年份，补充当前年份
        if isinstance(first_row_value, str):
            # 如果是字符串并且没有年份，补充当前年份
            if len(first_row_value.split('-')) == 2:  # 格式为 'MM-DD' 或 'DD-MM'
                first_row_value = f"{current_year}-{first_row_value}"
            # 尝试解析日期
            outbound_time = datetime.strptime(first_row_value, "%Y-%m-%d %H:%M")
        elif isinstance(first_row_value, datetime):
            # 如果是已经是 datetime 类型，直接处理
            outbound_time = first_row_value
        else:
            raise ValueError(f"无法解析日期: {first_row_value}")

        # 格式化为 "%Y/%m/%d" 格式
        formatted_date = outbound_time.strftime("%Y/%m/%d")
        return formatted_date
    except Exception as e:
        print(f"发生错误: {e}")
        return None


def merge_csv_files(file_paths, output_path):
    # 创建一个空的列表来存储所有的 DataFrame
    data_frames = []

    # 遍历所有文件路径，读取 CSV 文件
    for file_path in file_paths:
        # 读取当前 CSV 文件
        df = pd.read_csv(file_path)
        data_frames.append(df)

    # 使用 pandas.concat 合并所有 DataFrame
    merged_data = pd.concat(data_frames, ignore_index=True)

    # 将合并后的数据保存为一个新的 CSV 文件
    merged_data.to_csv(output_path, index=False)
    print(f"所有 CSV 文件已合并，结果保存至 {output_path}")


def merge_xlsx_files(file1, file2, output_file):
    # 读取两个 Excel 文件
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # 合并两个 DataFrame，默认按行合并
    merged_df = pd.concat([df1, df2], ignore_index=True)

    # 将合并后的数据保存到新的 Excel 文件
    merged_df.to_excel(output_file, index=False)

    print(f"两个文件已合并，结果保存到 {output_file}")


def go(input_path):
    if input_path is None:
        input_path = input("请输入文件的绝对路径：")
    xlsx_path = extract_path_before_csv(input_path)
    str_strip(xlsx_path, "快递单号")
    check_and_add_courier_column(xlsx_path)
    results = extract_and_process_data(xlsx_path, RowName.Courier, 100, "快递单号")

    update_courier_status(xlsx_path, results[CourierStateMapKey.not_yet_map], "快递单号")
    update_courier_status(xlsx_path, results[CourierStateMapKey.pre_ship_map], "快递单号")
    update_courier_status(xlsx_path, results[CourierStateMapKey.unpaid_map], "快递单号")
    update_courier_status(xlsx_path, results[CourierStateMapKey.delivered_map], "快递单号")
    update_courier_status(xlsx_path, results[CourierStateMapKey.no_tracking_map], "快递单号")
    update_courier_status(xlsx_path, results[CourierStateMapKey.tracking_map], "快递单号")

    total_count, no_track_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.no_track)
    track_count = total_count - no_track_count
    total_count2, delivered_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.delivered)
    total_count3, unpaid_count = count_pattern_state(xlsx_path, RowName.Courier, Pattern.unpaid)
    total_count4, not_yet_count = count_pattern_state(xlsx_path, RowName.Courier, r"not_yet")
    total_count5, pre_ship_count = count_pattern_state(xlsx_path, RowName.Courier, r"pre_ship")

    text = ""
    fs_text = ""
    ck_time = get_days_difference(xlsx_path)
    gz_time = getYmd()
    text += f"打单日期：{ck_time}"
    text += f"\n跟踪日期：{gz_time}"

    text += f"\n订单总数：{total_count}"
    fs_text += f"\n订单总数：{total_count}"

    swl = round2(100 - ((int(no_track_count) / int(total_count)) * 100))
    wswl = round2(100 - swl)
    text += f"\n上网：（{track_count}, {swl}）"
    fs_text += f"\n上网：（{track_count}, {swl}）"

    text += f"\n未上网：（{no_track_count}, {wswl}）"
    fs_text += f"\n未上网：（{no_track_count}, {wswl}）"

    text += f"\nnot_yet：{not_yet_count}"
    fs_text += f"\nnot_yet：{not_yet_count}"

    text += f"\npre_ship：{pre_ship_count}"
    fs_text += f"\npre_ship：{pre_ship_count}"

    text += f"\ndelivered：{delivered_count}"
    fs_text += f"\ndelivered：{delivered_count}"

    text += f"\nunpaid：{unpaid_count}"
    fs_text += f"\nunpaid：{unpaid_count}"

    unpaid_tracking_data = get_unpaid_tracking_data(xlsx_path)
    if (len(unpaid_tracking_data) > 0):
        text += f"\n-------unpaid详情-------"
        fs_text += f"\n-------unpaid详情-------"
        for key, value in unpaid_tracking_data.items():
            text += f"\n（单号：{key}, 快递单号：{value}）"
            fs_text += f"\n（单号：{key}, 快递单号：{value}）"
    print(text)

    tat = get_token()
    brief_sheet_value(tat, [fs_text], ck_time, gz_time, ClientConstants.md_flld)


def automatic(dir_path):
    for root, dirs, files in os.walk(dir_path):
        pattern = r"^打单时间\d+_\d+\.xlsx$"  # 正则表达式
        for ele in files:
            if re.match(pattern, ele):
                xlsx_path = f"{root}/{ele}"
                print(f"匹配的文件: {xlsx_path}")
                go(xlsx_path)


if __name__ == '__main__':
    # 示例调用
    # file_paths = [
    #     '/Users/zkp/Downloads/table_1.csv',
    #     '/Users/zkp/Downloads/table_12.csv',
    # ]  # 请替换为实际的文件路径
    # output_path = '/Users/zkp/Downloads/打单时间16_119.xlsx'  # 合并后的文件路径
    # merge_csv_files(file_paths, output_path)

    # file1 = '/Users/zkp/Downloads/table_1.xlsx'  # 请替换为您的第一个 Excel 文件路径
    # file2 = '/Users/zkp/Downloads/table_12.xlsx'  # 请替换为您的第二个 Excel 文件路径
    # output_file = '/Users/zkp/Downloads/merged_output.xlsx'  # 合并后的 Excel 文件路径
    # merge_xlsx_files(file1, file2, output_file)

    # go(None)

    automatic("/Users/zkp/Desktop/B&Y/轨迹统计/flld")
    # automatic("/Users/zkp/Desktop/B&Y/轨迹统计/flld/2025.2")
