from openpyxl import load_workbook


def fuzzy_match_and_delete(file1, file2, output_file, col1="F", col2="B"):
    try:
        # 加载表1
        wb1 = load_workbook(file1, data_only=True)
        ws1 = wb1.active

        # 加载表2
        wb2 = load_workbook(file2)
        ws2 = wb2.active

        # 获取表1 F列的所有数据
        data_col1 = [cell.value for cell in ws1[col1] if cell.value is not None]

        # 找到表2 B列中模糊匹配到的数据所在的行
        rows_to_delete = []
        for row in range(2, ws2.max_row + 1):  # 从第2行开始，跳过标题
            cell_value = ws2[f"{col2}{row}"].value
            if cell_value and any(keyword in str(cell_value) for keyword in data_col1):
                rows_to_delete.append(row)

        # 删除匹配到的行（倒序删除以避免索引混乱）
        for row in reversed(rows_to_delete):
            ws2.delete_rows(row)

        # 保存修改后的表2
        wb2.save(output_file)
        print(f"模糊匹配到的数据已从表2中删除，结果保存到: {output_file}")

    except Exception as e:
        print(f"执行过程中发生错误: {e}")


# 输入文件路径
file1_path = "/Users/zkp/Desktop/B&Y/孤品/销量分析SKU20250120-739892729682944000.xlsx"
file2_path = "/Users/zkp/Desktop/B&Y/孤品/table_1 (3).xlsx"
output_file_path = "/Users/zkp/Desktop/B&Y/孤品/滞销（1.13-1.20）.xlsx"

# 调用函数
fuzzy_match_and_delete(file1_path, file2_path, output_file_path)

